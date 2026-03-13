from __future__ import annotations

import json
import math
import os
import tomllib
import unicodedata
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import pandas as pd
from openpyxl import load_workbook


TEXT_CONFIG_SCHEMA_VERSION = 1
TEXT_CONFIG_DIR_NAME = "pipeline29_text"

METADATA_FILENAME = "metadata.toml"
DEFAULTS_FILENAME = "defaults.toml"
DATA_QUALITY_FILENAME = "data_quality.toml"
MAPPINGS_FILENAME = "mappings.toml"
INSTRUMENTS_FILENAME = "instruments.toml"
REPORTING_FILENAME = "reporting_rounding.toml"
PLOTS_FILENAME = "plots.toml"

DEFAULT_MAPPING_COLUMNS = ["key", "col_mean", "col_sd", "unit", "notes"]
DEFAULT_INSTRUMENT_COLUMNS = [
    "key",
    "component",
    "dist",
    "range_min",
    "range_max",
    "acc_abs",
    "acc_pct",
    "digits",
    "lsd",
    "resolution",
    "source",
    "notes",
    "setting_param",
    "setting_value",
]
DEFAULT_REPORTING_COLUMNS = ["key", "report_resolution", "report_digits", "rule", "notes"]
DEFAULT_PLOT_COLUMNS = [
    "enabled",
    "plot_type",
    "filename",
    "title",
    "x_col",
    "y_col",
    "yerr_col",
    "show_uncertainty",
    "x_label",
    "y_label",
    "x_min",
    "x_max",
    "x_step",
    "y_min",
    "y_max",
    "y_step",
    "y_tol_plus",
    "y_tol_minus",
    "filter_h2o_list",
    "label_variant",
    "notes",
]
DEFAULT_KEY_VALUE_COLUMNS = ["param", "value", "notes"]

REQUIRED_MAPPING_KEYS = {"power_kw", "fuel_kgh", "lhv_kj_kg"}


@dataclass
class Pipeline29ConfigBundle:
    mappings: Dict[str, Dict[str, str]]
    instruments_df: pd.DataFrame
    reporting_df: pd.DataFrame
    plots_df: pd.DataFrame
    data_quality_cfg: Dict[str, float]
    defaults_cfg: Dict[str, str]
    source_kind: str = "text"
    source_path: Optional[Path] = None
    text_dir: Optional[Path] = None


def default_text_config_dir(base_dir: Path) -> Path:
    return base_dir / "config" / TEXT_CONFIG_DIR_NAME


def default_app_state_dir() -> Path:
    return Path(os.environ.get("LOCALAPPDATA", str(Path.home()))) / "nanum_pipeline_29"


def default_gui_state_path() -> Path:
    return default_app_state_dir() / "config_gui_state.json"


def default_preset_dir() -> Path:
    return default_app_state_dir() / "presets"


def bundle_required_paths(config_dir: Path) -> Dict[str, Path]:
    return {
        "metadata": config_dir / METADATA_FILENAME,
        "defaults": config_dir / DEFAULTS_FILENAME,
        "data_quality": config_dir / DATA_QUALITY_FILENAME,
        "mappings": config_dir / MAPPINGS_FILENAME,
        "instruments": config_dir / INSTRUMENTS_FILENAME,
        "reporting": config_dir / REPORTING_FILENAME,
        "plots": config_dir / PLOTS_FILENAME,
    }


def text_config_exists(config_dir: Path) -> bool:
    paths = bundle_required_paths(config_dir)
    return all(path.exists() for path in paths.values())


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    try:
        if pd.isna(value):
            return True
    except Exception:
        pass
    return str(value).replace("\ufeff", "").strip() == ""


def _norm_text(value: Any) -> str:
    if _is_blank(value):
        return ""
    return str(value).replace("\ufeff", "").strip()


def _norm_key(value: Any) -> str:
    text = _norm_text(value).lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _to_builtin_scalar(value: Any) -> Any:
    if _is_blank(value):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and not isinstance(value, bool):
        return int(value)
    if isinstance(value, float):
        return None if not math.isfinite(value) else float(value)
    if hasattr(value, "item"):
        try:
            return _to_builtin_scalar(value.item())
        except Exception:
            pass
    txt = _norm_text(value)
    if txt == "":
        return None
    return txt


def _coerce_row_dict(raw: Dict[str, Any], field_order: Iterable[str]) -> Dict[str, Any]:
    out: Dict[str, Any] = {}
    for field in field_order:
        out[field] = raw.get(field, None)
    for key, value in raw.items():
        if key not in out:
            out[key] = value
    return out


def _records_to_dataframe(records: List[Dict[str, Any]], columns: List[str]) -> pd.DataFrame:
    normalized = [_coerce_row_dict(record, columns) for record in records]
    df = pd.DataFrame(normalized)
    for column in columns:
        if column not in df.columns:
            df[column] = pd.NA
    if df.empty:
        df = pd.DataFrame(columns=columns)
    return df[columns + [c for c in df.columns if c not in columns]].copy()


def _normalize_bundle_shapes(bundle: Pipeline29ConfigBundle) -> Pipeline29ConfigBundle:
    inst = _records_to_dataframe(bundle.instruments_df.to_dict(orient="records"), DEFAULT_INSTRUMENT_COLUMNS)
    rep = _records_to_dataframe(bundle.reporting_df.to_dict(orient="records"), DEFAULT_REPORTING_COLUMNS)
    plots = _records_to_dataframe(bundle.plots_df.to_dict(orient="records"), DEFAULT_PLOT_COLUMNS)
    defaults_cfg = {str(k).strip(): _norm_text(v) for k, v in bundle.defaults_cfg.items() if str(k).strip()}
    data_quality_cfg = {}
    for key, value in bundle.data_quality_cfg.items():
        key_txt = str(key).strip()
        if not key_txt:
            continue
        if _is_blank(value):
            continue
        try:
            data_quality_cfg[key_txt] = float(value)
        except Exception:
            continue
    return Pipeline29ConfigBundle(
        mappings={str(k).strip(): dict(v or {}) for k, v in bundle.mappings.items() if str(k).strip()},
        instruments_df=inst,
        reporting_df=rep,
        plots_df=plots,
        data_quality_cfg=data_quality_cfg,
        defaults_cfg=defaults_cfg,
        source_kind=bundle.source_kind,
        source_path=bundle.source_path,
        text_dir=bundle.text_dir,
    )


def validate_bundle(bundle: Pipeline29ConfigBundle) -> List[str]:
    errors: List[str] = []
    mapping_keys_norm = {_norm_key(key) for key in bundle.mappings.keys()}
    missing_keys = REQUIRED_MAPPING_KEYS - mapping_keys_norm
    if missing_keys:
        errors.append(f"Mappings sem chaves obrigatorias: {sorted(missing_keys)}")
    for column in DEFAULT_INSTRUMENT_COLUMNS:
        if column not in bundle.instruments_df.columns:
            errors.append(f"Instruments sem coluna obrigatoria: {column}")
    for column in DEFAULT_REPORTING_COLUMNS:
        if column not in bundle.reporting_df.columns:
            errors.append(f"Reporting_Rounding sem coluna obrigatoria: {column}")
    for column in DEFAULT_PLOT_COLUMNS:
        if column not in bundle.plots_df.columns:
            errors.append(f"Plots sem coluna obrigatoria: {column}")
    return errors


def _toml_scalar(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int) and not isinstance(value, bool):
        return str(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError("NaN/inf nao sao suportados em TOML")
        text = repr(float(value))
        if text == "-0.0":
            return "0.0"
        return text
    return json.dumps(str(value), ensure_ascii=False)


def _iter_clean_items(record: Dict[str, Any], preferred_order: Iterable[str]) -> List[tuple[str, Any]]:
    items: List[tuple[str, Any]] = []
    seen: set[str] = set()
    for key in preferred_order:
        if key not in record:
            continue
        value = _to_builtin_scalar(record.get(key))
        if value is None:
            continue
        items.append((key, value))
        seen.add(key)
    for key, value in record.items():
        if key in seen:
            continue
        clean = _to_builtin_scalar(value)
        if clean is None:
            continue
        items.append((key, clean))
    return items


def _write_toml_table_file(path: Path, table_name: str, values: Dict[str, Any]) -> None:
    lines = [f"schema_version = {TEXT_CONFIG_SCHEMA_VERSION}", ""]
    lines.append(f"[{table_name}]")
    for key, value in _iter_clean_items(values, values.keys()):
        lines.append(f"{json.dumps(str(key), ensure_ascii=False)} = {_toml_scalar(value)}")
    lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_toml_keyed_tables(path: Path, table_name: str, values: Dict[str, Dict[str, Any]], field_order: List[str]) -> None:
    lines = [f"schema_version = {TEXT_CONFIG_SCHEMA_VERSION}", ""]
    for key, row in values.items():
        row_clean = _iter_clean_items(row, field_order)
        if not row_clean:
            continue
        lines.append(f"[{table_name}.{json.dumps(str(key), ensure_ascii=False)}]")
        for field, value in row_clean:
            lines.append(f"{json.dumps(str(field), ensure_ascii=False)} = {_toml_scalar(value)}")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_toml_array_of_tables(path: Path, array_name: str, records: List[Dict[str, Any]], field_order: List[str]) -> None:
    lines = [f"schema_version = {TEXT_CONFIG_SCHEMA_VERSION}", ""]
    for record in records:
        clean_items = _iter_clean_items(record, field_order)
        if not clean_items:
            continue
        lines.append(f"[[{array_name}]]")
        for key, value in clean_items:
            lines.append(f"{json.dumps(str(key), ensure_ascii=False)} = {_toml_scalar(value)}")
        lines.append("")
    path.write_text("\n".join(lines), encoding="utf-8")


def _dataframe_records(df: pd.DataFrame, field_order: List[str]) -> List[Dict[str, Any]]:
    if df is None or df.empty:
        return []
    records = df.to_dict(orient="records")
    return [_coerce_row_dict(record, field_order) for record in records]


def save_text_config_bundle(
    bundle: Pipeline29ConfigBundle,
    config_dir: Path,
    *,
    bootstrapped_from: Optional[Path] = None,
) -> Pipeline29ConfigBundle:
    config_dir.mkdir(parents=True, exist_ok=True)
    normalized = _normalize_bundle_shapes(bundle)

    metadata = {
        "format": "pipeline29_text",
        "schema_version": TEXT_CONFIG_SCHEMA_VERSION,
        "updated_at_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat(),
    }
    if bootstrapped_from is not None:
        metadata["bootstrapped_from"] = str(bootstrapped_from)
    if normalized.source_kind:
        metadata["source_kind"] = normalized.source_kind
    if normalized.source_path is not None:
        metadata["source_path"] = str(normalized.source_path)

    paths = bundle_required_paths(config_dir)
    _write_toml_table_file(paths["metadata"], "metadata", metadata)
    _write_toml_table_file(paths["defaults"], "defaults", normalized.defaults_cfg)
    _write_toml_table_file(paths["data_quality"], "data_quality", normalized.data_quality_cfg)
    mapping_rows = {
        key: {
            "mean": spec.get("mean", ""),
            "sd": spec.get("sd", ""),
            "unit": spec.get("unit", ""),
            "notes": spec.get("notes", ""),
        }
        for key, spec in normalized.mappings.items()
    }
    _write_toml_keyed_tables(paths["mappings"], "mappings", mapping_rows, ["mean", "sd", "unit", "notes"])
    _write_toml_array_of_tables(
        paths["instruments"],
        "instruments",
        _dataframe_records(normalized.instruments_df, DEFAULT_INSTRUMENT_COLUMNS),
        DEFAULT_INSTRUMENT_COLUMNS,
    )
    _write_toml_array_of_tables(
        paths["reporting"],
        "reporting",
        _dataframe_records(normalized.reporting_df, DEFAULT_REPORTING_COLUMNS),
        DEFAULT_REPORTING_COLUMNS,
    )
    _write_toml_array_of_tables(
        paths["plots"],
        "plots",
        _dataframe_records(normalized.plots_df, DEFAULT_PLOT_COLUMNS),
        DEFAULT_PLOT_COLUMNS,
    )
    normalized.text_dir = config_dir
    normalized.source_kind = "text"
    normalized.source_path = config_dir
    return normalized


def _read_toml_file(path: Path) -> Dict[str, Any]:
    if not path.exists():
        return {}
    return tomllib.loads(path.read_text(encoding="utf-8"))


def load_text_config_bundle(config_dir: Path) -> Pipeline29ConfigBundle:
    paths = bundle_required_paths(config_dir)
    if not text_config_exists(config_dir):
        missing = [str(path.name) for path in paths.values() if not path.exists()]
        raise FileNotFoundError(f"Config textual incompleta em {config_dir}: faltam {missing}")

    defaults_doc = _read_toml_file(paths["defaults"])
    data_quality_doc = _read_toml_file(paths["data_quality"])
    mappings_doc = _read_toml_file(paths["mappings"])
    instruments_doc = _read_toml_file(paths["instruments"])
    reporting_doc = _read_toml_file(paths["reporting"])
    plots_doc = _read_toml_file(paths["plots"])

    defaults_cfg = {str(k): _norm_text(v) for k, v in defaults_doc.get("defaults", {}).items()}

    data_quality_cfg: Dict[str, float] = {}
    for key, value in data_quality_doc.get("data_quality", {}).items():
        if _is_blank(value):
            continue
        try:
            data_quality_cfg[str(key)] = float(value)
        except Exception:
            continue

    mappings: Dict[str, Dict[str, str]] = {}
    for key, spec in mappings_doc.get("mappings", {}).items():
        if not str(key).strip():
            continue
        spec = spec or {}
        mappings[str(key)] = {
            "mean": _norm_text(spec.get("mean", "")),
            "sd": _norm_text(spec.get("sd", "")),
            "unit": _norm_text(spec.get("unit", "")),
            "notes": _norm_text(spec.get("notes", "")),
        }

    instruments_df = _records_to_dataframe(instruments_doc.get("instruments", []) or [], DEFAULT_INSTRUMENT_COLUMNS)
    reporting_df = _records_to_dataframe(reporting_doc.get("reporting", []) or [], DEFAULT_REPORTING_COLUMNS)
    plots_df = _records_to_dataframe(plots_doc.get("plots", []) or [], DEFAULT_PLOT_COLUMNS)

    bundle = Pipeline29ConfigBundle(
        mappings=mappings,
        instruments_df=instruments_df,
        reporting_df=reporting_df,
        plots_df=plots_df,
        data_quality_cfg=data_quality_cfg,
        defaults_cfg=defaults_cfg,
        source_kind="text",
        source_path=config_dir,
        text_dir=config_dir,
    )
    return _normalize_bundle_shapes(bundle)


def _worksheet_rows(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=False)
    try:
        ws = None
        for candidate in wb.sheetnames:
            if candidate == sheet_name or str(candidate).strip().lower() == str(sheet_name).strip().lower():
                ws = wb[candidate]
                break
        if ws is None:
            return []
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [_norm_text(value) for value in rows[0]]
        out: List[Dict[str, Any]] = []
        for raw_row in rows[1:]:
            if raw_row is None:
                continue
            row = {header[idx]: raw_row[idx] for idx in range(min(len(header), len(raw_row))) if _norm_text(header[idx])}
            if not any(not _is_blank(value) for value in row.values()):
                continue
            out.append(row)
        return out
    finally:
        wb.close()


def _excel_rows_to_bundle(excel_path: Path) -> Pipeline29ConfigBundle:
    mapping_rows = _worksheet_rows(excel_path, "Mappings")
    mappings: Dict[str, Dict[str, str]] = {}
    for row in mapping_rows:
        key = _norm_text(row.get("key", ""))
        col_mean = _norm_text(row.get("col_mean", ""))
        if not key or "logical variable identifier" in key.lower():
            continue
        if "exact dataframe column name" in col_mean.lower():
            continue
        mappings[key] = {
            "mean": col_mean,
            "sd": _norm_text(row.get("col_sd", "")),
            "unit": _norm_text(row.get("unit", "")),
            "notes": _norm_text(row.get("notes", "")),
        }

    instruments_rows = _worksheet_rows(excel_path, "Instruments")
    reporting_rows = _worksheet_rows(excel_path, "Reporting_Rounding")
    if not reporting_rows:
        reporting_rows = _worksheet_rows(excel_path, "UPD_Rounding")
    defaults_rows = _worksheet_rows(excel_path, "Defaults")
    plots_rows = _worksheet_rows(excel_path, "Plots")
    data_quality_rows = _worksheet_rows(excel_path, "data quality assessment")

    defaults_cfg: Dict[str, str] = {}
    for row in defaults_rows:
        param = _norm_text(row.get("param", ""))
        if not param or "global parameter name" in param.lower():
            continue
        defaults_cfg[param] = _norm_text(row.get("value", ""))

    data_quality_cfg: Dict[str, float] = {}
    for row in data_quality_rows:
        param = _norm_text(row.get("param", ""))
        if not param:
            continue
        value = row.get("value", None)
        if _is_blank(value):
            continue
        try:
            data_quality_cfg[param] = float(value)
        except Exception:
            continue

    instruments_df = _records_to_dataframe(instruments_rows, DEFAULT_INSTRUMENT_COLUMNS)
    reporting_df = _records_to_dataframe(reporting_rows, DEFAULT_REPORTING_COLUMNS)
    plots_df = _records_to_dataframe(plots_rows, DEFAULT_PLOT_COLUMNS)
    if not plots_df.empty and "show_uncertainty" in plots_df.columns:
        for idx, row in plots_df.iterrows():
            yerr = _norm_text(row.get("yerr_col", ""))
            if _norm_text(row.get("show_uncertainty", "")):
                continue
            if yerr.lower() in {"off", "none", "disabled", "disable", "0", "na", "n/a"}:
                plots_df.at[idx, "show_uncertainty"] = "off"
            else:
                plots_df.at[idx, "show_uncertainty"] = "auto"

    bundle = Pipeline29ConfigBundle(
        mappings=mappings,
        instruments_df=instruments_df,
        reporting_df=reporting_df,
        plots_df=plots_df,
        data_quality_cfg=data_quality_cfg,
        defaults_cfg=defaults_cfg,
        source_kind="excel",
        source_path=excel_path,
        text_dir=None,
    )
    return _normalize_bundle_shapes(bundle)


def bootstrap_text_config_from_excel(excel_path: Path, config_dir: Path) -> Pipeline29ConfigBundle:
    bundle = _excel_rows_to_bundle(excel_path)
    return save_text_config_bundle(bundle, config_dir, bootstrapped_from=excel_path)


def bundle_to_preset_payload(bundle: Pipeline29ConfigBundle) -> Dict[str, Any]:
    normalized = _normalize_bundle_shapes(bundle)
    return {
        "schema_version": TEXT_CONFIG_SCHEMA_VERSION,
        "mappings": normalized.mappings,
        "instruments": _dataframe_records(normalized.instruments_df, DEFAULT_INSTRUMENT_COLUMNS),
        "reporting": _dataframe_records(normalized.reporting_df, DEFAULT_REPORTING_COLUMNS),
        "plots": _dataframe_records(normalized.plots_df, DEFAULT_PLOT_COLUMNS),
        "data_quality": normalized.data_quality_cfg,
        "defaults": normalized.defaults_cfg,
    }


def bundle_from_preset_payload(payload: Dict[str, Any]) -> Pipeline29ConfigBundle:
    return _normalize_bundle_shapes(
        Pipeline29ConfigBundle(
            mappings=dict(payload.get("mappings", {}) or {}),
            instruments_df=_records_to_dataframe(payload.get("instruments", []) or [], DEFAULT_INSTRUMENT_COLUMNS),
            reporting_df=_records_to_dataframe(payload.get("reporting", []) or [], DEFAULT_REPORTING_COLUMNS),
            plots_df=_records_to_dataframe(payload.get("plots", []) or [], DEFAULT_PLOT_COLUMNS),
            data_quality_cfg=dict(payload.get("data_quality", {}) or {}),
            defaults_cfg=dict(payload.get("defaults", {}) or {}),
            source_kind="preset",
            source_path=None,
            text_dir=None,
        )
    )


def save_bundle_preset(bundle: Pipeline29ConfigBundle, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = bundle_to_preset_payload(bundle)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def load_bundle_preset(path: Path) -> Pipeline29ConfigBundle:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return bundle_from_preset_payload(payload)


def save_gui_state(payload: Dict[str, Any], path: Optional[Path] = None) -> None:
    target = path or default_gui_state_path()
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def load_gui_state(path: Optional[Path] = None) -> Dict[str, Any]:
    target = path or default_gui_state_path()
    if not target.exists():
        return {}
    try:
        return json.loads(target.read_text(encoding="utf-8"))
    except Exception:
        return {}
