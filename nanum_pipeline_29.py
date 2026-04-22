from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
import unicodedata
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Optional, List, Dict, Tuple
from datetime import datetime
from math import sqrt
import difflib

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter

from pipeline29_config_backend import (
    DEFAULT_FUEL_PROPERTY_COLUMNS,
    Pipeline29ConfigBundle,
    bootstrap_text_config_from_excel,
    default_gui_state_path,
    load_excel_config_bundle,
    default_text_config_dir,
    load_gui_state,
    load_text_config_bundle,
    text_config_exists,
)

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except Exception:
    tk = None
    filedialog = None
    messagebox = None
    ttk = None

try:
    from PySide6.QtCore import Qt
    from PySide6.QtWidgets import (
        QApplication,
        QCheckBox,
        QDialog,
        QGridLayout,
        QHBoxLayout,
        QLabel,
        QMessageBox,
        QPushButton,
        QStyleFactory,
        QTableWidget,
        QTableWidgetItem,
        QVBoxLayout,
        QWidget,
        QHeaderView,
    )
except Exception:
    QApplication = None
    QCheckBox = None
    QDialog = None
    QGridLayout = None
    QHBoxLayout = None
    QLabel = None
    QMessageBox = None
    QPushButton = None
    QStyleFactory = None
    QTableWidget = None
    QTableWidgetItem = None
    QVBoxLayout = None
    QWidget = None
    QHeaderView = None
    Qt = None


# =========================
# Paths / constants
# =========================
BASE_DIR = Path(__file__).resolve().parent
DEFAULT_RAW_DIR = BASE_DIR / "raw"
DEFAULT_PROCESS_DIR = DEFAULT_RAW_DIR / "PROCESSAR"
DEFAULT_OUT_DIR = BASE_DIR / "out"
MESTRADO_ROOT = Path(r"D:\Drive\Faculdade\PUC\Mestrado")
RUNTIME_SETTINGS_DIR = Path(os.environ.get("LOCALAPPDATA", str(Path.home()))) / "nanum_pipeline_29"
RUNTIME_SETTINGS_PATH = RUNTIME_SETTINGS_DIR / "pipeline29_runtime_paths.json"

RAW_DIR = DEFAULT_RAW_DIR
PROCESS_DIR = DEFAULT_PROCESS_DIR
OUT_DIR = DEFAULT_OUT_DIR
PLOTS_DIR = OUT_DIR / "plots"
CFG_DIR = BASE_DIR / "config"

PREFERRED_SHEET_NAME = "labview"
B_ETANOL_COL_CANDIDATES = ["B_Etanol", "B_ETANOL", "B_ETANOL (kg)", "B_Etanol (kg)"]

SAMPLES_PER_WINDOW = 30
MIN_SAMPLES_PER_WINDOW = 30
DT_S = 1.0
TIME_DELTA_ERROR_THRESHOLD_S = 1.2
TIME_DELTA_PLOT_YMIN_S = 0.8
TIME_DELTA_PLOT_YMAX_S = 1.6
TIME_DELTA_PLOT_YSTEP_S = 0.1
DEFAULT_MAX_DELTA_BETWEEN_SAMPLES_MS = 1200.0
DEFAULT_MAX_ACT_CONTROL_ERROR_C = 5.0
DEFAULT_MAX_ECT_CONTROL_ERROR_C = 2.0
TIME_DIAG_PLOT_DPI = 150
TIME_DIAG_FILE_SCATTER_MAX_POINTS = 200
K_COVERAGE = 2.0
GUI_COMPARE_ENABLE_UNITARY_KEY = "GUI_COMPARE_ENABLE_UNITARY"


def _path_is_within(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except Exception:
        return False


def is_mestrado_runtime() -> bool:
    return _path_is_within(Path.cwd(), MESTRADO_ROOT)

FUEL_H2O_LEVELS = [6, 25, 35]  # â€œcombustÃ­veisâ€ por hidrataÃ§Ã£o
COMPOSITION_COLS = ["DIES_pct", "BIOD_pct", "EtOH_pct", "H2O_pct"]
FUEL_BLEND_DEFAULTS = {
    "D85B15": {
        "density_param": "FUEL_DENSITY_KG_M3_D85B15",
        "cost_param": "FUEL_COST_R_L_D85B15",
    },
    "E94H6": {
        "density_param": "FUEL_DENSITY_KG_M3_E94H6",
        "cost_param": "FUEL_COST_R_L_E94H6",
    },
    "E75H25": {
        "density_param": "FUEL_DENSITY_KG_M3_E75H25",
        "cost_param": "FUEL_COST_R_L_E75H25",
    },
    "E65H35": {
        "density_param": "FUEL_DENSITY_KG_M3_E65H35",
        "cost_param": "FUEL_COST_R_L_E65H35",
    },
}
FUEL_LABEL_BY_H2O_LEVEL = {
    0: "D85B15",
    6: "E94H6",
    25: "E75H25",
    35: "E65H35",
}
FUEL_H2O_LEVEL_BY_LABEL = {label: level for level, label in FUEL_LABEL_BY_H2O_LEVEL.items()}
SCENARIO_REFERENCE_FUEL_LABEL = "E94H6"
MACHINE_SCENARIO_SPECS = [
    {
        "key": "Colheitadeira",
        "label": "Colheitadeira",
        "hours_param": "MACHINE_HOURS_PER_YEAR_COLHEITADEIRA",
        "diesel_l_h_param": "MACHINE_DIESEL_L_H_COLHEITADEIRA",
        "color": "#1f77b4",
    },
    {
        "key": "Trator_Transbordo",
        "label": "Trator transbordo",
        "hours_param": "MACHINE_HOURS_PER_YEAR_TRATOR_TRANSBORDO",
        "diesel_l_h_param": "MACHINE_DIESEL_L_H_TRATOR_TRANSBORDO",
        "color": "#ff7f0e",
    },
    {
        "key": "Caminhao",
        "label": "Caminhao",
        "hours_param": "MACHINE_HOURS_PER_YEAR_CAMINHAO",
        "diesel_l_h_param": "MACHINE_DIESEL_L_H_CAMINHAO",
        "color": "#2ca02c",
    },
]

# =========================
# Airflow assumptions
# =========================
AFR_STOICH_E94H6 = 8.4
ETHANOL_FRAC_E94H6 = 0.94
AFR_STOICH_ETHANOL = 9.0
AFR_STOICH_BIODIESEL = 12.5
AFR_STOICH_DIESEL = 14.5
LAMBDA_DEFAULT = 1.0
R_AIR_DRY_J_KG_K = 287.058


# =========================
# Psychrometrics / cp models
# =========================
R_V_WATER = 461.5  # J/(kg*K)
CP_WATER_VAPOR_KJ_KG_K = 1.86  # kJ/(kg*K), engineering approximation


# =========================
# Exhaust emissions helpers
# =========================
MW_CO2_KG_KMOL = 44.0095
MW_CO_KG_KMOL = 28.0101
MW_O2_KG_KMOL = 31.9988
MW_N2_KG_KMOL = 28.0134
MW_H2O_KG_KMOL = 18.0153
MW_C3H8_KG_KMOL = 44.097
MW_NO_KG_KMOL = 30.006
MW_NO2_KG_KMOL = 46.0055

# Fuel surrogates used only to estimate hydrogen-derived water formation.
# They are explicit so we can swap them later if the project adopts measured
# elemental analyses for the bench fuels.
H_MASS_FRAC_DIESEL = 0.1385641540557986       # surrogate C12H23
H_MASS_FRAC_BIODIESEL = 0.12238992225838548   # surrogate C19H36O2
H_MASS_FRAC_ETHANOL = 0.1312813388612733      # ethanol C2H6O
THC_LOW_SIGNAL_WARN_PPM = 10.0


# =========================
# Excel helpers
# =========================
def _excel_engine_preferred() -> str:
    try:
        import python_calamine  # noqa: F401
        return "calamine"
    except Exception:
        return "openpyxl"


def _read_excel(path: Path, sheet_name: str | int | None = 0) -> pd.DataFrame:
    eng = _excel_engine_preferred()
    if eng == "calamine":
        try:
            return pd.read_excel(path, sheet_name=sheet_name, engine="calamine")
        except Exception as e:
            print(f"[WARN] read_excel calamine falhou em {path.name} (sheet={sheet_name}): {e}. Tentando openpyxl...")
    return pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")


def _excel_file(path: Path) -> pd.ExcelFile:
    eng = _excel_engine_preferred()
    if eng == "calamine":
        try:
            return pd.ExcelFile(path, engine="calamine")
        except Exception as e:
            print(f"[WARN] ExcelFile calamine falhou em {path.name}: {e}. Tentando openpyxl...")
    return pd.ExcelFile(path, engine="openpyxl")


# =========================
# Generic helpers
# =========================
def norm_key(x: object) -> str:
    return str(x).replace("\ufeff", "").strip().lower()


def _normalize_cols(cols: List[str]) -> List[str]:
    return [str(c).replace("\ufeff", "").strip() for c in cols]


def _canon_name(x: object) -> str:
    s = str(x).replace("\ufeff", "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", s)


def _normalize_repeated_stat_tokens_in_name(x: object) -> str:
    s = str(x).replace("\ufeff", "").strip()
    if not s:
        return s

    replacements = [
        ("_mean_mean_of_windows", "_mean_of_windows"),
        ("_mean_sd_of_windows", "_sd_of_windows"),
        ("_sd_mean_of_windows", "_sd_of_windows"),
        ("_sd_sd_of_windows", "_sd_of_windows"),
        ("_mean_mean", "_mean"),
        ("_sd_sd", "_sd"),
    ]

    prev = None
    while prev != s:
        prev = s
        for old, new in replacements:
            s = s.replace(old, new)
    s = re.sub(r"__+", "_", s)
    return s


def _coalesce_equivalent_columns(df: pd.DataFrame, context: str = "") -> pd.DataFrame:
    if df is None or df.empty:
        return df

    merged: Dict[str, pd.Series] = {}
    sources: Dict[str, List[str]] = {}
    for idx, raw_col in enumerate(df.columns):
        col = _normalize_repeated_stat_tokens_in_name(raw_col)
        series = df.iloc[:, idx].copy()
        series.name = col
        sources.setdefault(col, []).append(str(raw_col))
        if col in merged:
            merged[col] = merged[col].where(merged[col].notna(), series)
        else:
            merged[col] = series

    duplicates = {col: cols for col, cols in sources.items() if len(cols) > 1}
    if duplicates:
        preview = "; ".join(f"{col} <= {cols}" for col, cols in list(duplicates.items())[:5])
        where = f" em {context}" if context else ""
        print(f"[INFO] Consolidei colunas equivalentes{where}: {preview}")

    return pd.DataFrame(merged, index=df.index)


def resolve_col(df: pd.DataFrame, requested: str) -> str:
    requested = str(requested).replace("\ufeff", "").strip()
    if not requested:
        raise KeyError("Nome de coluna solicitado estÃ¡ vazio (verifique Mappings no config).")

    if requested in df.columns:
        return requested

    low_map = {str(c).lower().strip(): c for c in df.columns}
    req_low = requested.lower().strip()
    if req_low in low_map:
        return low_map[req_low]

    canon_map = {_canon_name(c): c for c in df.columns}
    req_canon = _canon_name(requested)
    if req_canon in canon_map:
        return canon_map[req_canon]

    req_stats_norm = _normalize_repeated_stat_tokens_in_name(requested)
    if req_stats_norm in df.columns:
        return req_stats_norm

    stats_norm_map: Dict[str, str] = {}
    for c in df.columns:
        c_norm = _normalize_repeated_stat_tokens_in_name(c)
        if c_norm not in stats_norm_map:
            stats_norm_map[c_norm] = c
    if req_stats_norm in stats_norm_map:
        return stats_norm_map[req_stats_norm]

    stats_norm_canon_map: Dict[str, str] = {}
    for c in df.columns:
        c_norm = _normalize_repeated_stat_tokens_in_name(c)
        c_norm_canon = _canon_name(c_norm)
        if c_norm_canon not in stats_norm_canon_map:
            stats_norm_canon_map[c_norm_canon] = c
    req_stats_canon = _canon_name(req_stats_norm)
    if req_stats_canon in stats_norm_canon_map:
        return stats_norm_canon_map[req_stats_canon]

    suggestion = difflib.get_close_matches(requested, list(df.columns), n=6)
    sug_txt = f" SugestÃµes: {suggestion}" if suggestion else ""
    raise KeyError(f"Coluna '{requested}' nÃ£o encontrada no dataframe.{sug_txt}")


def safe_to_excel(df: pd.DataFrame, path: Path) -> Path:
    try:
        df.to_excel(path, index=False)
        return path
    except PermissionError:
        ts = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        alt = path.with_name(f"{path.stem}_{ts}{path.suffix}")
        df.to_excel(alt, index=False)
        return alt


def clear_output_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)
    for child in path.iterdir():
        if child.is_dir():
            clear_output_dir(child)
            try:
                child.rmdir()
            except OSError:
                pass
            continue
        try:
            child.unlink()
        except PermissionError as e:
            raise SystemExit(
                f"NÃ£o consegui limpar o output porque '{child}' estÃ¡ em uso. "
                "Feche o arquivo ou o programa que o estÃ¡ usando e rode novamente."
            ) from e


def rect_to_std(limit: pd.Series | float) -> pd.Series:
    return pd.to_numeric(limit, errors="coerce") / sqrt(3)


def res_to_std(step: float) -> float:
    return step / sqrt(12) if step > 0 else 0.0


def _to_float(x: object, default: float = 0.0) -> float:
    if x is None:
        return default
    try:
        if pd.isna(x):
            return default
    except Exception:
        pass

    if isinstance(x, (int, float)):
        try:
            return float(x)
        except Exception:
            return default

    s = str(x).replace("\ufeff", "").strip()
    if s == "":
        return default
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return default


def _format_pct_for_label(value: object) -> str:
    numeric = _to_float(value, default=float("nan"))
    if not np.isfinite(numeric):
        return _to_str_or_empty(value)
    if abs(numeric - round(numeric)) <= 1e-9:
        return str(int(round(numeric)))
    return f"{numeric:g}"


def _fuel_label_from_components(
    dies_pct: object,
    biod_pct: object,
    etoh_pct: object,
    h2o_pct: object,
    tol: float = 0.6,
) -> str:
    dies = _to_float(dies_pct, default=float("nan"))
    biod = _to_float(biod_pct, default=float("nan"))
    etoh = _to_float(etoh_pct, default=float("nan"))
    h2o = _to_float(h2o_pct, default=float("nan"))

    def _near_zero(value: float) -> bool:
        return (not np.isfinite(value)) or abs(value) <= tol

    if np.isfinite(dies) and np.isfinite(biod) and abs(dies - 85.0) <= tol and abs(biod - 15.0) <= tol and _near_zero(etoh) and _near_zero(h2o):
        return "D85B15"
    if np.isfinite(etoh) and np.isfinite(h2o) and abs(etoh - 94.0) <= tol and abs(h2o - 6.0) <= tol and _near_zero(dies) and _near_zero(biod):
        return "E94H6"
    if np.isfinite(etoh) and np.isfinite(h2o) and abs(etoh - 75.0) <= tol and abs(h2o - 25.0) <= tol and _near_zero(dies) and _near_zero(biod):
        return "E75H25"
    if np.isfinite(etoh) and np.isfinite(h2o) and abs(etoh - 65.0) <= tol and abs(h2o - 35.0) <= tol and _near_zero(dies) and _near_zero(biod):
        return "E65H35"

    if np.isfinite(dies) and np.isfinite(biod) and _near_zero(etoh) and _near_zero(h2o):
        if abs(dies) <= tol:
            return f"B{_format_pct_for_label(biod)}"
        if abs(biod) <= tol:
            return f"D{_format_pct_for_label(dies)}"
        return f"D{_format_pct_for_label(dies)}B{_format_pct_for_label(biod)}"

    if np.isfinite(biod) and np.isfinite(etoh) and _near_zero(dies) and _near_zero(h2o):
        if abs(etoh) <= tol:
            return f"B{_format_pct_for_label(biod)}"
        if abs(biod) <= tol:
            return f"E{_format_pct_for_label(etoh)}"
        return f"B{_format_pct_for_label(biod)}E{_format_pct_for_label(etoh)}"

    if np.isfinite(dies) and np.isfinite(etoh) and _near_zero(biod) and _near_zero(h2o):
        if abs(etoh) <= tol:
            return f"D{_format_pct_for_label(dies)}"
        if abs(dies) <= tol:
            return f"E{_format_pct_for_label(etoh)}"
        return f"D{_format_pct_for_label(dies)}E{_format_pct_for_label(etoh)}"

    return ""


def _canon_unit_token(text: object) -> str:
    s = _canon_name(text).replace("º", "").replace("°", "")
    s = s.replace("/", "_").replace("-", "_")
    if not s:
        return ""
    aliases = {
        "mbar": "mbar",
        "mbars": "mbar",
        "millibar": "mbar",
        "millibars": "mbar",
        "kpa": "kpa",
        "pa": "pa",
        "bar": "bar",
        "c": "c",
        "degc": "c",
        "celsius": "c",
    }
    return aliases.get(s, s)


def _unit_scale_to_base(unit: str) -> Optional[float]:
    unit_norm = _canon_unit_token(unit)
    scales = {
        "pa": 1.0,
        "mbar": 100.0,
        "kpa": 1000.0,
        "bar": 100000.0,
        "c": 1.0,
    }
    return scales.get(unit_norm)


def _convert_unit_value(value: float, from_unit: str, to_unit: str) -> Optional[float]:
    from_scale = _unit_scale_to_base(from_unit)
    to_scale = _unit_scale_to_base(to_unit)
    if from_scale is None or to_scale is None:
        return None
    return float(value * from_scale / to_scale)


def _mapping_unit_for_y_col(y_col: str, mappings: dict) -> Optional[str]:
    y_text = _to_str_or_empty(y_col)
    if not y_text:
        return None
    for _key_norm, spec in mappings.items():
        col_mean_req = str(spec.get("mean", "")).strip()
        if not col_mean_req:
            continue
        if norm_key(col_mean_req) == norm_key(y_text):
            unit = _to_str_or_empty(spec.get("unit", ""))
            return unit or None
    return None


def _parse_axis_value(value: object, *, target_unit: Optional[str] = None, default: float = np.nan) -> float:
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except Exception:
        pass

    if isinstance(value, (int, float)):
        try:
            return float(value)
        except Exception:
            return default

    text = str(value).replace("\ufeff", "").strip()
    if not text:
        return default
    if text.lower() in {"auto", "nan", "none", "off", "disabled", "n/a", "na"}:
        return default

    text_num = text.replace(",", ".")
    try:
        return float(text_num)
    except Exception:
        pass

    match = re.fullmatch(r"\s*([+-]?\d+(?:\.\d+)?)\s*([A-Za-z°º/_-]+)\s*", text_num)
    if not match:
        return default

    number = float(match.group(1))
    unit = _canon_unit_token(match.group(2))
    if not unit:
        return number
    if not target_unit:
        return number

    converted = _convert_unit_value(number, unit, target_unit)
    if converted is None:
        return default
    return converted


def _safe_name(name: str) -> str:
    s = re.sub(r"[^A-Za-z0-9_]+", "_", str(name))
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _is_blank_cell(x: object) -> bool:
    if x is None:
        return True
    try:
        if pd.isna(x):
            return True
    except Exception:
        pass
    s = str(x).replace("\ufeff", "").strip()
    return s == "" or s.lower() == "nan"


def _to_str_or_empty(x: object) -> str:
    return "" if _is_blank_cell(x) else str(x).replace("\ufeff", "").strip()


def _format_load_kw_label(v: object) -> str:
    x = _to_float(v, default=float("nan"))
    if not np.isfinite(x):
        return ""
    if abs(x - round(x)) <= 1e-9:
        return f"{int(round(x))}"
    return f"{x:g}"


def _find_first_col_by_substrings(df: pd.DataFrame, substrings: List[str]) -> Optional[str]:
    cols = list(df.columns)
    for c in cols:
        cl = str(c).lower()
        ok = True
        for s in substrings:
            if str(s).lower() not in cl:
                ok = False
                break
        if ok:
            return c
    return None


def _basename_parts(basename: object) -> List[str]:
    return [str(p).strip() for p in str(basename).split("__") if str(p).strip()]


def _basename_source_folder_parts(basename: object) -> List[str]:
    parts = _basename_parts(basename)
    if len(parts) <= 1:
        return []
    return parts[:-1]


def _basename_source_folder_display(basename: object) -> str:
    return " / ".join(_basename_source_folder_parts(basename))


def _basename_source_file(basename: object) -> str:
    parts = _basename_parts(basename)
    if not parts:
        return ""
    return parts[-1]


def _basename_source_plot_dir(basename: object, root: Path | None = None) -> Path:
    base_root = PLOTS_DIR if root is None else root
    folder_parts = _basename_source_folder_parts(basename)
    if not folder_parts:
        return base_root
    return base_root.joinpath(*folder_parts)


def add_source_identity_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "BaseName" not in df.columns:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()

    out = df.copy()
    out["SourceFolder"] = out["BaseName"].map(_basename_source_folder_display)
    out["SourceFile"] = out["BaseName"].map(_basename_source_file)
    return out


def iter_source_plot_groups(df: pd.DataFrame, root: Path | None = None) -> List[Tuple[str, Path, pd.DataFrame]]:
    if df is None or df.empty:
        return []

    if "BaseName" not in df.columns:
        base_root = PLOTS_DIR if root is None else root
        return [("", base_root, df.copy())]

    tmp = add_source_identity_columns(df)
    groups: List[Tuple[str, Path, pd.DataFrame]] = []
    for source_folder, d in tmp.groupby("SourceFolder", dropna=False, sort=True):
        basename_example = d["BaseName"].iloc[0]
        plot_dir = _basename_source_plot_dir(basename_example, root=root)
        groups.append((str(source_folder or ""), plot_dir, d.copy()))
    return groups


def _source_folder_leaf(source_folder: object) -> str:
    s = str(source_folder or "").strip()
    if not s:
        return ""
    parts = [p.strip() for p in s.split("/") if p.strip()]
    return parts[-1] if parts else s


def _normalize_compare_series_name(source_folder: object) -> str:
    leaf = _source_folder_leaf(source_folder)
    if not leaf:
        return "origem_desconhecida"

    s = _canon_name(leaf).replace(" ", "_").replace("-", "_")
    s = re.sub(r"_+", "_", s).strip("_")
    s = re.sub(r"(^|_)subindo(?=_|$)", r"\1subida", s)
    s = re.sub(r"(^|_)descendo(?=_|$)", r"\1descida", s)
    if not s:
        return "origem_desconhecida"
    return s


def _safe_folder_name(name: object) -> str:
    s = str(name or "").strip()
    if not s:
        return "compare"
    s = re.sub(r'[<>:"/\\|?*]', "_", s)
    s = s.strip().rstrip(".")
    return s if s else "compare"


def _infer_source_direction_from_folder_name(source_folder: object) -> Optional[str]:
    s = _canon_name(source_folder).replace("_", " ").replace("-", " ")
    if "subindo" in s or "subida" in s or re.search(r"\bup\b", s):
        return "subindo"
    if "descendo" in s or "descida" in s or re.search(r"\bdown\b", s):
        return "descendo"
    return None


def _compare_group_key_from_source_folder(source_folder: object) -> str:
    s = str(source_folder or "").strip()
    if not s:
        return ""

    parts = [p.strip() for p in s.split("/") if p.strip()]
    clean_parts: List[str] = []
    for part in parts:
        t = _canon_name(part).replace("_", " ").replace("-", " ")
        t = re.sub(r"\b(subindo|subida|descendo|descida|up|down)\b", " ", t)
        t = re.sub(r"\s+", " ", t).strip()
        if t:
            clean_parts.append(t)

    if not clean_parts:
        return ""
    return "__".join(_safe_name(p) for p in clean_parts)


def iter_compare_plot_groups(df: pd.DataFrame, root: Path | None = None) -> List[Tuple[str, Path, pd.DataFrame]]:
    """
    Build compare groups combining subida/descida datasets for the same run key.
    Output path pattern: <root>/compare/<group_key>/...
    """
    if df is None or df.empty:
        return []

    tmp = add_source_identity_columns(df)
    if "SourceFolder" not in tmp.columns:
        return []

    tmp = tmp.copy()
    tmp["_COMPARE_GROUP"] = tmp["SourceFolder"].map(_compare_group_key_from_source_folder)
    tmp["_COMPARE_SERIES"] = tmp["SourceFolder"].map(_normalize_compare_series_name)
    tmp["_COMPARE_DIRECTION"] = tmp["SourceFolder"].map(_infer_source_direction_from_folder_name)
    tmp["_COMPARE_SERIES"] = tmp["_COMPARE_SERIES"].where(
        tmp["_COMPARE_SERIES"].map(lambda x: not _is_blank_cell(x)),
        "origem_desconhecida",
    )

    base_root = (PLOTS_DIR if root is None else root) / "compare"
    groups: List[Tuple[str, Path, pd.DataFrame]] = []
    for group_key, d in tmp.groupby("_COMPARE_GROUP", dropna=False, sort=True):
        gk = str(group_key or "").strip()
        if not gk:
            continue

        dirs = set(str(x).strip().lower() for x in d["_COMPARE_DIRECTION"].dropna().tolist() if str(x).strip())
        # Compare plots are only useful when both directions exist.
        if "subindo" not in dirs or "descendo" not in dirs:
            continue

        subida_vals = sorted(
            set(
                str(v).strip()
                for v in d.loc[d["_COMPARE_DIRECTION"].eq("subindo"), "_COMPARE_SERIES"].dropna().tolist()
                if str(v).strip()
            )
        )
        descida_vals = sorted(
            set(
                str(v).strip()
                for v in d.loc[d["_COMPARE_DIRECTION"].eq("descendo"), "_COMPARE_SERIES"].dropna().tolist()
                if str(v).strip()
            )
        )
        if subida_vals and descida_vals:
            compare_name = f"{subida_vals[0]} vs {descida_vals[0]}"
        else:
            uniq = sorted(
                set(str(v).strip() for v in d["_COMPARE_SERIES"].dropna().tolist() if str(v).strip())
            )
            compare_name = " vs ".join(uniq[:2]) if uniq else gk

        plot_dir = base_root / _safe_folder_name(compare_name)
        groups.append((gk, plot_dir, d.copy()))

    return groups


def _infer_sentido_carga_from_folder_parts(parts: List[str]) -> object:
    for part in reversed(parts):
        s = _canon_name(part).replace("_", " ").replace("-", " ")
        if "subindo" in s or "subida" in s or re.search(r"\bup\b", s):
            return "subida"
        if "descendo" in s or "descida" in s or re.search(r"\bdown\b", s):
            return "descida"
    return pd.NA


def _infer_iteracao_from_folder_parts(parts: List[str]) -> object:
    for part in reversed(parts):
        m = re.search(r"(\d+)\s*$", str(part))
        if m:
            return int(m.group(1))

    for part in reversed(parts):
        nums = re.findall(r"\d+", str(part))
        if nums:
            return int(nums[-1])

    return pd.NA


def _sentido_carga_rank(x: object) -> int:
    s = _canon_name(x).replace("_", " ").replace("-", " ")
    if "subida" in s or "subindo" in s or re.search(r"\bup\b", s):
        return 0
    if "descida" in s or "descendo" in s or re.search(r"\bdown\b", s):
        return 1
    return 9


def add_run_context_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty or "BaseName" not in df.columns:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()

    out = df.copy()
    folder_parts = out["BaseName"].map(_basename_source_folder_parts)
    out["Sentido_Carga"] = folder_parts.map(_infer_sentido_carga_from_folder_parts)
    out["Iteracao"] = pd.to_numeric(folder_parts.map(_infer_iteracao_from_folder_parts), errors="coerce").astype("Int64")
    return out


def _find_kibox_col_by_tokens(df: pd.DataFrame, tokens: List[str]) -> Optional[str]:
    want = [str(t).lower().replace("_", "").replace(" ", "") for t in tokens if str(t).strip()]
    if not want:
        return None

    for c in df.columns:
        cs = str(c)
        if not cs.startswith("KIBOX_"):
            continue
        canon = cs.lower().replace("_", "").replace(" ", "")
        ok = True
        for w in want:
            if w not in canon:
                ok = False
                break
        if ok:
            return c
    return None


def _parse_csv_list_ints(x: object) -> Optional[List[int]]:
    if _is_blank_cell(x):
        return None
    s = str(x).replace("\ufeff", "").strip()
    if not s:
        return None
    parts = [p.strip() for p in s.split(",") if p.strip()]
    out: List[int] = []
    for p in parts:
        if p == "":
            continue
        try:
            out.append(int(float(p.replace(",", "."))))
        except Exception:
            continue
    return out if out else None


def _parse_axis_spec(
    min_v: object,
    max_v: object,
    step_v: object,
    *,
    target_unit: Optional[str] = None,
) -> Optional[Tuple[float, float, float]]:
    a = _parse_axis_value(min_v, target_unit=target_unit, default=np.nan)
    b = _parse_axis_value(max_v, target_unit=target_unit, default=np.nan)
    c = _parse_axis_value(step_v, target_unit=target_unit, default=np.nan)
    if not (np.isfinite(a) and np.isfinite(b) and np.isfinite(c)):
        return None
    if c <= 0:
        return None
    return (float(a), float(b), float(c))


def _parse_axis_limits(
    min_v: object,
    max_v: object,
    *,
    target_unit: Optional[str] = None,
) -> Optional[Tuple[float, float]]:
    a = _parse_axis_value(min_v, target_unit=target_unit, default=np.nan)
    b = _parse_axis_value(max_v, target_unit=target_unit, default=np.nan)
    if not (np.isfinite(a) and np.isfinite(b)):
        return None
    if float(b) <= float(a):
        return None
    return (float(a), float(b))


# =========================
# Reporting rounding helpers (rev2 sheet Reporting_Rounding)
# =========================
def _round_half_up_to_resolution(x: pd.Series, res: float) -> pd.Series:
    s = pd.to_numeric(x, errors="coerce")
    if res <= 0:
        return s
    q = s / res

    pos = q.where(q >= 0)
    neg = q.where(q < 0)

    pos_r = np.floor(pos + 0.5)
    neg_r = np.ceil(neg - 0.5)

    out = q.copy()
    out = out.where(q.isna(), np.nan)
    out = out.where(q < 0, pos_r)
    out = out.where(q >= 0, neg_r)
    return out * res


# =========================
# Derived airflow channels (no row count change)
# =========================
def _ethanol_mass_fraction_from_etoh_pct(etoh_pct: pd.Series) -> pd.Series:
    return pd.to_numeric(etoh_pct, errors="coerce") / 100.0


def _nan_series(index: pd.Index) -> pd.Series:
    return pd.Series(np.nan, index=index, dtype="float64")


def _airflow_component_fraction(df: pd.DataFrame, column: str) -> pd.Series:
    return pd.to_numeric(df.get(column, _nan_series(df.index)), errors="coerce") / 100.0


def _ethanol_trial_mask(df: pd.DataFrame) -> pd.Series:
    etoh = pd.to_numeric(df.get("EtOH_pct", _nan_series(df.index)), errors="coerce")
    h2o = pd.to_numeric(df.get("H2O_pct", _nan_series(df.index)), errors="coerce")
    return etoh.gt(0) | h2o.gt(0)


def _diesel_like_no_ethanol_mask(df: pd.DataFrame) -> pd.Series:
    dies = pd.to_numeric(df.get("DIES_pct", _nan_series(df.index)), errors="coerce")
    biod = pd.to_numeric(df.get("BIOD_pct", _nan_series(df.index)), errors="coerce")
    ethanol_mask = _ethanol_trial_mask(df)
    return (dies.gt(0) | biod.gt(0)) & ~ethanol_mask


def _airflow_stoich_blend_from_composition(df: pd.DataFrame) -> pd.Series:
    dies_frac = _airflow_component_fraction(df, "DIES_pct")
    biod_frac = _airflow_component_fraction(df, "BIOD_pct")
    etoh_frac = _airflow_component_fraction(df, "EtOH_pct")
    h2o_frac = _airflow_component_fraction(df, "H2O_pct")

    valid_components = dies_frac.notna() | biod_frac.notna() | etoh_frac.notna() | h2o_frac.notna()
    blend_afr = (
        dies_frac.fillna(0.0) * AFR_STOICH_DIESEL
        + biod_frac.fillna(0.0) * AFR_STOICH_BIODIESEL
        + etoh_frac.fillna(0.0) * AFR_STOICH_ETHANOL
    )
    return blend_afr.where(valid_components & blend_afr.gt(0), pd.NA)


def _find_preferred_column(
    df: pd.DataFrame,
    *,
    preferred_names: List[str],
    include_tokens: List[str],
    exclude_tokens: Optional[List[str]] = None,
) -> Optional[str]:
    for requested in preferred_names:
        req = _to_str_or_empty(requested)
        if not req:
            continue
        try:
            return resolve_col(df, req)
        except Exception:
            continue

    exclude_tokens = exclude_tokens or []
    for column in df.columns:
        canon = _canon_name(column)
        if any(_canon_name(token) not in canon for token in include_tokens):
            continue
        if any(_canon_name(token) in canon for token in exclude_tokens):
            continue
        return column
    return None


def _resolve_airflow_lambda_col(df: pd.DataFrame, mappings: dict) -> Optional[str]:
    preferred_names: List[str] = []
    if "lambda" in mappings and mappings["lambda"].get("mean"):
        preferred_names.append(mappings["lambda"]["mean"])
    preferred_names.extend(
        [
            "Motec_Exhaust Lambda_mean_of_windows",
            "Exhaust Lambda_mean_of_windows",
            "Lambda_mean_of_windows",
        ]
    )
    return _find_preferred_column(
        df,
        preferred_names=preferred_names,
        include_tokens=["lambda", "mean"],
        exclude_tokens=["sd", "diagnostic", "normalised", "normalized"],
    )


def _resolve_airflow_maf_col(df: pd.DataFrame, defaults_cfg: Dict[str, str]) -> Optional[str]:
    preferred_name = _to_str_or_empty(defaults_cfg.get(norm_key("VOL_EFF_DIESEL_MAF_COL"), "")) or "MAF_mean_of_windows"
    return _find_preferred_column(
        df,
        preferred_names=[preferred_name, "MAF_mean_of_windows", "Motec_MAF_mean_of_windows"],
        include_tokens=["maf"],
        exclude_tokens=["sd"],
    )


def add_airflow_channels_inplace(df: pd.DataFrame, lambda_col: str | None = None) -> pd.DataFrame:
    out = df.copy()

    fuel_col = None
    for c in ["Consumo_kg_h_mean_of_windows", "Consumo_kg_h", "Fuel_kg_h", "fuel_kgh_mean_of_windows"]:
        if c in out.columns:
            fuel_col = c
            break
    if fuel_col is None:
        candidates = [c for c in out.columns if "consumo" in c.lower() and "mean_of_windows" in c.lower()]
        fuel_col = candidates[0] if candidates else None

    if fuel_col is None:
        print("[WARN] Airflow: nÃ£o achei coluna de consumo (kg/h). Pulando canais de ar.")
        return out

    fuel_mix_kg_h = pd.to_numeric(out[fuel_col], errors="coerce")

    x_etoh = _ethanol_mass_fraction_from_etoh_pct(out["EtOH_pct"])
    out["EtOH_pure_mass_frac"] = x_etoh

    out["Fuel_EtOH_pure_kg_h"] = fuel_mix_kg_h * x_etoh
    out["Fuel_E94H6_eq_kg_h"] = out["Fuel_EtOH_pure_kg_h"] / ETHANOL_FRAC_E94H6

    if lambda_col and lambda_col in out.columns:
        out["lambda_used"] = pd.to_numeric(out[lambda_col], errors="coerce")
    else:
        out["lambda_used"] = LAMBDA_DEFAULT

    out["AFR_stoich_E94H6"] = AFR_STOICH_E94H6
    out["AFR_real"] = out["lambda_used"] * out["AFR_stoich_E94H6"]

    out["Air_kg_h"] = out["AFR_real"] * out["Fuel_E94H6_eq_kg_h"]
    out["Air_kg_s"] = out["Air_kg_h"] / 3600.0
    out["Air_g_s"] = out["Air_kg_s"] * 1000.0

    return out


def _static_maf_mask_by_fuel(maf: pd.Series, fuel_labels: pd.Series, *, min_points: int = 4) -> Tuple[pd.Series, List[str]]:
    invalid_mask = pd.Series(False, index=maf.index, dtype="bool")
    static_labels: List[str] = []
    labels = fuel_labels.map(_to_str_or_empty).replace("", pd.NA)
    for label in sorted(v for v in labels.dropna().unique().tolist() if _to_str_or_empty(v)):
        label_mask = labels.eq(label)
        label_maf = pd.to_numeric(maf.where(label_mask), errors="coerce").dropna()
        if len(label_maf) < min_points:
            continue
        if _series_is_static(label_maf):
            invalid_mask = invalid_mask | (label_mask & maf.notna())
            static_labels.append(str(label))
    return invalid_mask, static_labels


def add_airflow_channels_prefer_maf_inplace(
    df: pd.DataFrame,
    lambda_col: str | None = None,
    maf_col: str | None = None,
    *,
    maf_min_kgh: float = 0.0,
    maf_max_kgh: float = 300.0,
) -> pd.DataFrame:
    out = df.copy()
    nan_series = _nan_series(out.index)

    fuel_col = None
    for c in ["Consumo_kg_h_mean_of_windows", "Consumo_kg_h", "Fuel_kg_h", "fuel_kgh_mean_of_windows"]:
        if c in out.columns:
            fuel_col = c
            break
    if fuel_col is None and not out.empty:
        candidates = [c for c in out.columns if "consumo" in c.lower() and "mean_of_windows" in c.lower()]
        fuel_col = candidates[0] if candidates else None

    fuel_mix_kg_h = pd.to_numeric(out[fuel_col], errors="coerce") if fuel_col else nan_series.copy()

    x_etoh = _ethanol_mass_fraction_from_etoh_pct(out.get("EtOH_pct", nan_series))
    out["EtOH_pure_mass_frac"] = x_etoh
    out["Fuel_EtOH_pure_kg_h"] = fuel_mix_kg_h * x_etoh
    out["Fuel_E94H6_eq_kg_h"] = out["Fuel_EtOH_pure_kg_h"] / ETHANOL_FRAC_E94H6

    lambda_measured = pd.to_numeric(out[lambda_col], errors="coerce") if lambda_col and lambda_col in out.columns else nan_series.copy()
    lambda_valid = lambda_measured.gt(0)
    out["lambda_used"] = lambda_measured.where(lambda_valid, LAMBDA_DEFAULT)
    out["LAMBDA_SOURCE"] = pd.Series("default_1.0", index=out.index, dtype="object")
    out.loc[lambda_valid, "LAMBDA_SOURCE"] = "measured"

    out["AFR_stoich_blend"] = _airflow_stoich_blend_from_composition(out)
    out["AFR_stoich_E94H6"] = AFR_STOICH_E94H6
    out["AFR_real"] = out["lambda_used"] * out["AFR_stoich_blend"]

    out["Air_kg_h_from_Fuel_Lambda"] = (
        pd.to_numeric(out["AFR_real"], errors="coerce") * fuel_mix_kg_h
    ).where(fuel_mix_kg_h.gt(0) & pd.to_numeric(out["AFR_real"], errors="coerce").gt(0), pd.NA)

    maf = pd.to_numeric(out[maf_col], errors="coerce") if maf_col and maf_col in out.columns else nan_series.copy()
    fuel_labels = out.get("Fuel_Label", pd.Series(pd.NA, index=out.index, dtype="object"))
    fuel_labels = fuel_labels.where(fuel_labels.notna(), _fuel_blend_labels(out))
    ethanol_mask = _ethanol_trial_mask(out)
    diesel_like_mask = _diesel_like_no_ethanol_mask(out)

    ignored_ethanol_maf_mask = ethanol_mask & maf.notna()
    if bool(ignored_ethanol_maf_mask.any()):
        ignored_labels = sorted(
            {
                str(label).strip()
                for label in fuel_labels.where(ignored_ethanol_maf_mask).dropna().tolist()
                if str(label).strip()
            }
        )
        ignored_txt = ", ".join(ignored_labels) if ignored_labels else "combustiveis com etanol"
        print(
            f"[INFO] Airflow: MAF ignorado em {int(ignored_ethanol_maf_mask.sum())} ponto(s) com etanol "
            f"({ignored_txt}); vou usar consumo+lambda por regra."
        )

    static_maf_mask, static_maf_labels = _static_maf_mask_by_fuel(maf.where(diesel_like_mask), fuel_labels.where(diesel_like_mask))
    if static_maf_labels:
        static_txt = ", ".join(static_maf_labels)
        print(f"[WARN] Airflow: ignorei MAF estatico para {static_txt}. Vou usar fuel+lambda nesses combustiveis.")
    maf_valid = diesel_like_mask & maf.gt(0) & maf.between(maf_min_kgh, maf_max_kgh, inclusive="both") & ~static_maf_mask
    invalid_maf_mask = diesel_like_mask & maf.notna() & ~maf_valid & ~static_maf_mask
    if bool(invalid_maf_mask.any()):
        print(
            f"[WARN] Airflow: {int(invalid_maf_mask.sum())} ponto(s) com MAF fora de {maf_min_kgh:g}..{maf_max_kgh:g} kg/h; "
            "vou usar fuel+lambda nesses pontos."
        )
    out["Air_kg_h_from_MAF"] = maf.where(maf_valid, pd.NA)

    out["Air_kg_h"] = out["Air_kg_h_from_MAF"].where(out["Air_kg_h_from_MAF"].notna(), out["Air_kg_h_from_Fuel_Lambda"])
    out["Air_kg_s"] = out["Air_kg_h"] / 3600.0
    out["Air_g_s"] = out["Air_kg_s"] * 1000.0

    out["Airflow_Method"] = pd.Series("unavailable", index=out.index, dtype="object")
    out.loc[out["Air_kg_h_from_MAF"].notna(), "Airflow_Method"] = "MAF"
    fuel_lambda_mask = out["Air_kg_h_from_MAF"].isna() & out["Air_kg_h_from_Fuel_Lambda"].notna()
    out.loc[fuel_lambda_mask & out["LAMBDA_SOURCE"].eq("measured"), "Airflow_Method"] = "fuel_lambda"
    out.loc[fuel_lambda_mask & out["LAMBDA_SOURCE"].ne("measured"), "Airflow_Method"] = "fuel_lambda_default"

    if fuel_col is None and not bool(maf_valid.any()):
        print("[WARN] Airflow: nao achei consumo em kg/h nem MAF valido. Canais de ar ficaram vazios.")

    return out


def _resolve_existing_column(df: pd.DataFrame, preferred_name: str, fallback_tokens: List[str]) -> Optional[str]:
    preferred = str(preferred_name or "").strip()
    if preferred and preferred in df.columns:
        return preferred
    return _find_first_col_by_substrings(df, fallback_tokens)


def _series_is_static(values: pd.Series, tol: float = 1e-9) -> bool:
    numeric = pd.to_numeric(values, errors="coerce").dropna()
    if numeric.empty:
        return False
    if numeric.nunique(dropna=True) <= 1:
        return True
    try:
        return float(numeric.max() - numeric.min()) <= tol
    except Exception:
        return False


def add_volumetric_efficiency_channels_inplace(df: pd.DataFrame, defaults_cfg: Dict[str, str]) -> pd.DataFrame:
    out = df.copy()

    displacement_l = _to_float(defaults_cfg.get(norm_key("ENGINE_DISPLACEMENT_L"), ""), default=3.992)
    ref_pressure_kpa = _to_float(defaults_cfg.get(norm_key("VOL_EFF_REF_PRESSURE_kPa"), ""), default=101.3)
    rpm_col_name = _to_str_or_empty(defaults_cfg.get(norm_key("VOL_EFF_RPM_COL"), "")) or "Rotação_mean_of_windows"
    maf_col_name = _to_str_or_empty(defaults_cfg.get(norm_key("VOL_EFF_DIESEL_MAF_COL"), "")) or "MAF_mean_of_windows"
    maf_min_kgh = _to_float(defaults_cfg.get(norm_key("VOL_EFF_DIESEL_MAF_MIN_KGH"), ""), default=0.0)
    maf_max_kgh = _to_float(defaults_cfg.get(norm_key("VOL_EFF_DIESEL_MAF_MAX_KGH"), ""), default=300.0)

    for column in (
        "VOL_EFF_AIR_kg_h_USED",
        "VOL_EFF_THEORETICAL_AIR_kg_h",
        "VOL_EFF_RHO_REF_kg_m3",
        "VOL_EFF_RPM_USED",
        "ETA_V",
        "ETA_V_pct",
    ):
        if column not in out.columns:
            out[column] = pd.NA
    out["VOL_EFF_AIR_SOURCE"] = pd.Series(pd.NA, index=out.index, dtype="object")
    out["VOL_EFF_REF_PRESSURE_kPa"] = ref_pressure_kpa if np.isfinite(ref_pressure_kpa) else pd.NA

    if not np.isfinite(displacement_l) or displacement_l <= 0:
        print("[WARN] ENGINE_DISPLACEMENT_L invalida; nao calculei eficiencia volumetrica.")
        return out
    if not np.isfinite(ref_pressure_kpa) or ref_pressure_kpa <= 0:
        print("[WARN] VOL_EFF_REF_PRESSURE_kPa invalida; nao calculei eficiencia volumetrica.")
        return out

    t_col = _resolve_existing_column(out, "T_ADMISSAO_mean_of_windows", ["t", "admiss"])
    rpm_col = _resolve_existing_column(out, rpm_col_name, ["rotação", "rotacao", "rpm motor", "rpm"])
    if t_col is None or rpm_col is None:
        print(f"[WARN] Nao calculei eficiencia volumetrica: t_col={t_col}, rpm_col={rpm_col}.")
        return out

    displacement_m3 = displacement_l / 1000.0
    intake_t_k = pd.to_numeric(out[t_col], errors="coerce") + 273.15
    rpm = pd.to_numeric(out[rpm_col], errors="coerce")
    rho_ref = (ref_pressure_kpa * 1000.0) / (R_AIR_DRY_J_KG_K * intake_t_k)
    theoretical_air_kg_h = rho_ref * displacement_m3 * (rpm / 2.0) * 60.0

    out["VOL_EFF_RHO_REF_kg_m3"] = rho_ref
    out["VOL_EFF_THEORETICAL_AIR_kg_h"] = theoretical_air_kg_h
    out["VOL_EFF_RPM_USED"] = rpm

    fuel_labels = out.get("Fuel_Label", pd.Series(pd.NA, index=out.index, dtype="object"))
    fuel_labels = fuel_labels.where(fuel_labels.notna(), _fuel_blend_labels(out))
    diesel_mask = _diesel_like_no_ethanol_mask(out)

    air_used = pd.to_numeric(out.get("Air_kg_h", pd.NA), errors="coerce")
    out.loc[air_used.gt(0), "VOL_EFF_AIR_SOURCE"] = "Air_kg_h"

    maf_col = _resolve_existing_column(out, maf_col_name, ["maf"])
    if maf_col is not None and bool(diesel_mask.any()):
        maf = pd.to_numeric(out[maf_col], errors="coerce")
        diesel_maf = maf.where(diesel_mask)
        if _series_is_static(diesel_maf):
            print("[WARN] MAF diesel-like sem etanol ficou estatico; cancelei eficiencia volumetrica nesses pontos.")
            air_used.loc[diesel_mask] = pd.NA
            out.loc[diesel_mask, "VOL_EFF_AIR_SOURCE"] = pd.NA
        else:
            diesel_maf_valid = diesel_maf.between(maf_min_kgh, maf_max_kgh, inclusive="both") & diesel_maf.gt(0)
            invalid_diesel_maf = diesel_mask & diesel_maf.notna() & ~diesel_maf_valid
            if bool(invalid_diesel_maf.any()):
                print(
                    f"[WARN] {int(invalid_diesel_maf.sum())} ponto(s) diesel com MAF fora de {maf_min_kgh:g}..{maf_max_kgh:g} kg/h; cancelei eficiencia volumetrica nesses pontos."
                )
            air_used.loc[diesel_mask] = pd.NA
            air_used.loc[diesel_mask & diesel_maf_valid] = diesel_maf.loc[diesel_mask & diesel_maf_valid]
            out.loc[diesel_mask, "VOL_EFF_AIR_SOURCE"] = pd.NA
            out.loc[diesel_mask & diesel_maf_valid, "VOL_EFF_AIR_SOURCE"] = "MAF"
    elif bool(diesel_mask.any()):
        print("[WARN] Nao achei coluna de MAF para pontos diesel-like sem etanol; cancelei eficiencia volumetrica nesses pontos.")
        air_used.loc[diesel_mask] = pd.NA
        out.loc[diesel_mask, "VOL_EFF_AIR_SOURCE"] = pd.NA

    valid = air_used.gt(0) & theoretical_air_kg_h.gt(0) & intake_t_k.gt(0) & rpm.gt(0)
    eta_v = (air_used / theoretical_air_kg_h).where(valid, pd.NA)

    out["VOL_EFF_AIR_kg_h_USED"] = air_used
    out["ETA_V"] = eta_v
    out["ETA_V_pct"] = eta_v * 100.0
    return out


def add_volumetric_efficiency_from_airflow_method_inplace(df: pd.DataFrame, defaults_cfg: Dict[str, str]) -> pd.DataFrame:
    out = df.copy()

    displacement_l = _to_float(defaults_cfg.get(norm_key("ENGINE_DISPLACEMENT_L"), ""), default=3.992)
    ref_pressure_kpa = _to_float(defaults_cfg.get(norm_key("VOL_EFF_REF_PRESSURE_kPa"), ""), default=101.3)
    rpm_col_name = _to_str_or_empty(defaults_cfg.get(norm_key("VOL_EFF_RPM_COL"), "")) or "Rotação_mean_of_windows"

    for column in (
        "VOL_EFF_AIR_kg_h_USED",
        "VOL_EFF_THEORETICAL_AIR_kg_h",
        "VOL_EFF_RHO_REF_kg_m3",
        "VOL_EFF_RPM_USED",
        "ETA_V",
        "ETA_V_pct",
    ):
        if column not in out.columns:
            out[column] = pd.NA
    out["VOL_EFF_AIR_SOURCE"] = pd.Series(pd.NA, index=out.index, dtype="object")
    out["VOL_EFF_REF_PRESSURE_kPa"] = ref_pressure_kpa if np.isfinite(ref_pressure_kpa) else pd.NA

    if not np.isfinite(displacement_l) or displacement_l <= 0:
        print("[WARN] ENGINE_DISPLACEMENT_L invalida; nao calculei eficiencia volumetrica.")
        return out
    if not np.isfinite(ref_pressure_kpa) or ref_pressure_kpa <= 0:
        print("[WARN] VOL_EFF_REF_PRESSURE_kPa invalida; nao calculei eficiencia volumetrica.")
        return out

    t_col = _resolve_existing_column(out, "T_ADMISSAO_mean_of_windows", ["t", "admiss"])
    rpm_col = _resolve_existing_column(out, rpm_col_name, ["rotação", "rotacao", "rpm motor", "rpm"])
    if t_col is None or rpm_col is None:
        print(f"[WARN] Nao calculei eficiencia volumetrica: t_col={t_col}, rpm_col={rpm_col}.")
        return out

    displacement_m3 = displacement_l / 1000.0
    intake_t_k = pd.to_numeric(out[t_col], errors="coerce") + 273.15
    rpm = pd.to_numeric(out[rpm_col], errors="coerce")
    rho_ref = (ref_pressure_kpa * 1000.0) / (R_AIR_DRY_J_KG_K * intake_t_k)
    theoretical_air_kg_h = rho_ref * displacement_m3 * (rpm / 2.0) * 60.0

    out["VOL_EFF_RHO_REF_kg_m3"] = rho_ref
    out["VOL_EFF_THEORETICAL_AIR_kg_h"] = theoretical_air_kg_h
    out["VOL_EFF_RPM_USED"] = rpm

    air_used = pd.to_numeric(out.get("Air_kg_h", pd.NA), errors="coerce")
    airflow_method = out.get("Airflow_Method", pd.Series(pd.NA, index=out.index, dtype="object"))
    valid_air_mask = air_used.gt(0)
    out.loc[valid_air_mask, "VOL_EFF_AIR_SOURCE"] = airflow_method.loc[valid_air_mask]
    out.loc[valid_air_mask & out["VOL_EFF_AIR_SOURCE"].isna(), "VOL_EFF_AIR_SOURCE"] = "Air_kg_h"

    valid = air_used.gt(0) & theoretical_air_kg_h.gt(0) & intake_t_k.gt(0) & rpm.gt(0)
    eta_v = (air_used / theoretical_air_kg_h).where(valid, pd.NA)

    out["VOL_EFF_AIR_kg_h_USED"] = air_used
    out["ETA_V"] = eta_v
    out["ETA_V_pct"] = eta_v * 100.0
    return out


# =========================
# Psychrometrics helpers
# =========================
def _psat_water_pa_magnus(T_C: pd.Series) -> pd.Series:
    T = pd.to_numeric(T_C, errors="coerce")
    es_hpa = 6.112 * np.exp((17.62 * T) / (243.12 + T))
    return es_hpa * 100.0  # Pa


def _humidity_ratio_w_from_rh(T_C: pd.Series, RH_pct: pd.Series, P_kPa_abs: pd.Series) -> pd.Series:
    T = pd.to_numeric(T_C, errors="coerce")
    RH = pd.to_numeric(RH_pct, errors="coerce") / 100.0
    P_pa = pd.to_numeric(P_kPa_abs, errors="coerce") * 1000.0

    psat = _psat_water_pa_magnus(T)
    pv = RH.clip(lower=0.0, upper=1.0) * psat
    pv = pv.where((pv.notna()) & (P_pa.notna()) & (pv < 0.99 * P_pa), pd.NA)

    w = 0.62198 * pv / (P_pa - pv)
    return pd.to_numeric(w, errors="coerce")


def _absolute_humidity_g_m3(T_C: pd.Series, RH_pct: pd.Series) -> pd.Series:
    T = pd.to_numeric(T_C, errors="coerce")
    RH = pd.to_numeric(RH_pct, errors="coerce") / 100.0

    T_K = T + 273.15
    psat = _psat_water_pa_magnus(T)
    pv = RH.clip(lower=0.0, upper=1.0) * psat

    rho_v_kg_m3 = pv / (R_V_WATER * T_K)
    return rho_v_kg_m3 * 1000.0  # g/m^3


def _cp_air_dry_kj_kgk(T_C: pd.Series) -> pd.Series:
    T = pd.to_numeric(T_C, errors="coerce")
    return 1.005 + 0.0001 * (T - 25.0)


def _cp_moist_air_kj_kgk(T_C: pd.Series, RH_pct: pd.Series, P_kPa_abs: pd.Series) -> pd.Series:
    w = _humidity_ratio_w_from_rh(T_C, RH_pct, P_kPa_abs)
    yv = w / (1.0 + w)
    cp_dry = _cp_air_dry_kj_kgk(T_C)
    cp_mix = (1.0 - yv) * cp_dry + yv * CP_WATER_VAPOR_KJ_KG_K
    return pd.to_numeric(cp_mix, errors="coerce")


def _as_numeric_series(value: object, index: pd.Index) -> pd.Series:
    if isinstance(value, pd.Series):
        return pd.to_numeric(value.reindex(index), errors="coerce")
    return pd.to_numeric(pd.Series(value, index=index, dtype="float64"), errors="coerce")


def _percent_to_fraction(value: object, index: pd.Index) -> pd.Series:
    return _as_numeric_series(value, index) / 100.0


def _ppm_to_fraction(value: object, index: pd.Index) -> pd.Series:
    return _as_numeric_series(value, index) / 1_000_000.0


def _clip_fraction(value: pd.Series, *, lower: float = 0.0, upper: float = 1.0) -> pd.Series:
    s = pd.to_numeric(value, errors="coerce")
    return s.clip(lower=lower, upper=upper)


def specific_emissions_from_analyzer(
    *,
    air_kg_h: object,
    fuel_kg_h: object,
    power_kW: object,
    co2_pct_dry: object,
    co_pct_dry: object,
    o2_pct_dry: object,
    nox_ppm_dry: object,
    thc_ppm_dry: object,
    h2o_wet_frac: object,
    nox_basis: str = "NO",
) -> pd.DataFrame:
    for candidate in (power_kW, air_kg_h, fuel_kg_h, co2_pct_dry, co_pct_dry, o2_pct_dry, nox_ppm_dry, thc_ppm_dry, h2o_wet_frac):
        if isinstance(candidate, pd.Series):
            index = candidate.index
            break
    else:
        raise ValueError("specific_emissions_from_analyzer precisa de pelo menos uma Series para inferir o index.")

    air = _as_numeric_series(air_kg_h, index)
    fuel = _as_numeric_series(fuel_kg_h, index)
    power = _as_numeric_series(power_kW, index)
    co2_dry = _percent_to_fraction(co2_pct_dry, index)
    co_dry = _percent_to_fraction(co_pct_dry, index)
    o2_dry = _percent_to_fraction(o2_pct_dry, index)
    nox_dry = _ppm_to_fraction(nox_ppm_dry, index)
    thc_dry = _ppm_to_fraction(thc_ppm_dry, index)
    h2o_frac = _clip_fraction(_as_numeric_series(h2o_wet_frac, index), lower=0.0, upper=0.999)

    nox_basis_norm = _to_str_or_empty(nox_basis).upper()
    if nox_basis_norm not in {"NO", "NO2"}:
        raise ValueError(f"nox_basis invalido: {nox_basis}")
    mw_nox = MW_NO_KG_KMOL if nox_basis_norm == "NO" else MW_NO2_KG_KMOL

    co2_dry_mix = _clip_fraction(co2_dry)
    co_dry_mix = _clip_fraction(co_dry)
    o2_dry_mix = _clip_fraction(o2_dry)
    nox_dry_mix = _clip_fraction(nox_dry)
    thc_dry_mix = _clip_fraction(thc_dry, lower=0.0)

    dry_known_sum = co2_dry_mix + co_dry_mix + o2_dry_mix + nox_dry_mix + thc_dry_mix
    n2_dry = (1.0 - dry_known_sum).clip(lower=0.0)

    mw_dry = (
        co2_dry_mix * MW_CO2_KG_KMOL
        + co_dry_mix * MW_CO_KG_KMOL
        + o2_dry_mix * MW_O2_KG_KMOL
        + nox_dry_mix * mw_nox
        + thc_dry_mix * MW_C3H8_KG_KMOL
        + n2_dry * MW_N2_KG_KMOL
    )

    exhaust_kg_h = air + fuel
    exhaust_dry_kg_h = exhaust_kg_h * (1.0 - h2o_frac)
    dry_kmol_h = exhaust_dry_kg_h / mw_dry
    wet_kmol_h = dry_kmol_h / (1.0 - h2o_frac)
    h2o_kmol_h = wet_kmol_h * h2o_frac
    mw_wet = (1.0 - h2o_frac) * mw_dry + h2o_frac * MW_H2O_KG_KMOL

    co2_wet = co2_dry * (1.0 - h2o_frac)
    co_wet = co_dry * (1.0 - h2o_frac)
    o2_wet = o2_dry * (1.0 - h2o_frac)
    nox_wet = nox_dry * (1.0 - h2o_frac)
    thc_wet = thc_dry * (1.0 - h2o_frac)
    n2_wet = n2_dry * (1.0 - h2o_frac)

    def _mass_fraction(x_wet: pd.Series, mw_species: float) -> pd.Series:
        return (x_wet * mw_species / mw_wet).where(mw_wet.gt(0), pd.NA)

    co2_mass_frac = _mass_fraction(co2_wet, MW_CO2_KG_KMOL)
    co_mass_frac = _mass_fraction(co_wet, MW_CO_KG_KMOL)
    nox_mass_frac = _mass_fraction(nox_wet, mw_nox)
    thc_mass_frac = _mass_fraction(thc_wet, MW_C3H8_KG_KMOL)

    dry_valid = exhaust_kg_h.gt(0) & mw_dry.gt(0)
    valid_specific = dry_valid & power.gt(0) & mw_wet.gt(0) & h2o_frac.notna()

    def _g_h(mass_fraction: pd.Series) -> pd.Series:
        return (mass_fraction * exhaust_kg_h * 1000.0).where(valid_specific, pd.NA)

    co2_g_h = _g_h(co2_mass_frac)
    co_g_h = _g_h(co_mass_frac)
    nox_g_h = _g_h(nox_mass_frac)
    thc_g_h = _g_h(thc_mass_frac)

    out = pd.DataFrame(index=index)
    out["CO2_dry_frac"] = co2_dry
    out["CO_dry_frac"] = co_dry
    out["O2_dry_frac"] = o2_dry
    out["NOx_dry_frac"] = nox_dry
    out["THC_dry_frac"] = thc_dry
    out["N2_dry_frac"] = n2_dry
    out["CO2_wet_frac"] = co2_wet
    out["CO_wet_frac"] = co_wet
    out["O2_wet_frac"] = o2_wet
    out["NOx_wet_frac"] = nox_wet
    out["THC_wet_frac"] = thc_wet
    out["N2_wet_frac"] = n2_wet
    out["MW_dry_kg_kmol"] = mw_dry.where(dry_valid, pd.NA)
    out["MW_wet_kg_kmol"] = mw_wet.where(valid_specific, pd.NA)
    out["Exhaust_kg_h"] = exhaust_kg_h.where(exhaust_kg_h.gt(0), pd.NA)
    out["Exhaust_Dry_kg_h"] = exhaust_dry_kg_h.where(dry_valid, pd.NA)
    out["Exhaust_Dry_kmol_h"] = dry_kmol_h.where(dry_valid, pd.NA)
    out["Exhaust_H2O_kmol_h"] = h2o_kmol_h.where(valid_specific, pd.NA)
    out["CO2_g_h"] = co2_g_h
    out["CO_g_h"] = co_g_h
    out["NOx_g_h"] = nox_g_h
    out["THC_g_h"] = thc_g_h
    out["CO2_g_kWh"] = (co2_g_h / power).where(valid_specific, pd.NA)
    out["CO_g_kWh"] = (co_g_h / power).where(valid_specific, pd.NA)
    out["NOx_g_kWh"] = (nox_g_h / power).where(valid_specific, pd.NA)
    out["THC_g_kWh"] = (thc_g_h / power).where(valid_specific, pd.NA)
    return out


def _resolve_intake_humidity_ratio_for_emissions(
    df: pd.DataFrame,
    defaults_cfg: Optional[Dict[str, str]] = None,
) -> Tuple[pd.Series, str]:
    t_amb_col = _resolve_existing_column(df, "T_AMBIENTE_mean_of_windows", ["t_ambiente", "amb"])
    rh_col = _resolve_existing_column(df, "UMIDADE_mean_of_windows", ["umidade"])
    p_baro_col = _resolve_existing_column(df, "P_BARO_mean_of_windows", ["p_baro", "baro"])

    default_pressure_kpa = _to_float(
        (defaults_cfg or {}).get(norm_key("VOL_EFF_REF_PRESSURE_kPa"), ""),
        default=101.3,
    )
    pressure = (
        pd.to_numeric(df[p_baro_col], errors="coerce")
        if p_baro_col
        else pd.Series(default_pressure_kpa, index=df.index, dtype="float64")
    )
    pressure_source = "P_BARO"
    pressure_valid = pressure.dropna()
    if not pressure_valid.empty and float(pressure_valid.median()) > 200.0:
        pressure = pressure / 10.0
        pressure_source = "P_BARO_mbar->kPa"
    if t_amb_col and rh_col:
        return _humidity_ratio_w_from_rh(df[t_amb_col], df[rh_col], pressure), f"T_AMBIENTE+UMIDADE+{pressure_source}"
    if rh_col:
        t_e_comp_col = _resolve_existing_column(df, "T_E_COMP_mean_of_windows", ["t_e_comp"])
        if t_e_comp_col:
            return _humidity_ratio_w_from_rh(df[t_e_comp_col], df[rh_col], pressure), f"T_E_COMP+UMIDADE+{pressure_source}"
    return pd.Series(0.0, index=df.index, dtype="float64"), "fallback_seco"


def add_specific_emissions_channels_inplace(
    df: pd.DataFrame,
    *,
    power_kW: pd.Series,
    fuel_kg_h: pd.Series,
    defaults_cfg: Optional[Dict[str, str]] = None,
) -> pd.DataFrame:
    out = df.copy()
    idx = out.index

    air_kg_h = pd.to_numeric(out.get("Air_kg_h", pd.Series(pd.NA, index=idx)), errors="coerce")
    fuel_kg_h = pd.to_numeric(fuel_kg_h.reindex(idx), errors="coerce")
    power_kW = pd.to_numeric(power_kW.reindex(idx), errors="coerce")

    dies_frac = _airflow_component_fraction(out, "DIES_pct").fillna(0.0)
    biod_frac = _airflow_component_fraction(out, "BIOD_pct").fillna(0.0)
    etoh_frac = _airflow_component_fraction(out, "EtOH_pct").fillna(0.0)
    fuel_h2o_frac = _airflow_component_fraction(out, "H2O_pct").fillna(0.0)

    intake_humidity_ratio, humidity_source = _resolve_intake_humidity_ratio_for_emissions(out, defaults_cfg)
    intake_humidity_ratio = pd.to_numeric(intake_humidity_ratio, errors="coerce").clip(lower=0.0)
    air_h2o_kg_h = (air_kg_h * intake_humidity_ratio / (1.0 + intake_humidity_ratio)).where(air_kg_h.gt(0), pd.NA)
    fuel_h2o_kg_h = (fuel_kg_h * fuel_h2o_frac).where(fuel_kg_h.ge(0), pd.NA)

    fuel_h_mass_frac = (
        dies_frac * H_MASS_FRAC_DIESEL
        + biod_frac * H_MASS_FRAC_BIODIESEL
        + etoh_frac * H_MASS_FRAC_ETHANOL
    )
    fuel_h_kg_h = (fuel_kg_h * fuel_h_mass_frac).where(fuel_kg_h.ge(0), pd.NA)
    combustion_h2o_kg_h = fuel_h_kg_h * 9.0

    exhaust_kg_h = air_kg_h + fuel_kg_h
    exhaust_h2o_kg_h = air_h2o_kg_h + fuel_h2o_kg_h + combustion_h2o_kg_h
    exhaust_dry_kg_h = exhaust_kg_h - exhaust_h2o_kg_h

    out["Intake_Humidity_Ratio_kgkg"] = intake_humidity_ratio
    out["Intake_Air_H2O_kg_h"] = air_h2o_kg_h
    out["Fuel_H2O_kg_h"] = fuel_h2o_kg_h
    out["Fuel_H_kg_h"] = fuel_h_kg_h
    out["Combustion_H2O_kg_h"] = combustion_h2o_kg_h
    out["Exhaust_H2O_kg_h"] = exhaust_h2o_kg_h
    out["Exhaust_Dry_kg_h"] = exhaust_dry_kg_h

    co2_col = _resolve_existing_column(out, "CO2_mean_of_windows", ["co2"])
    co_col = _resolve_existing_column(out, "CO_mean_of_windows", ["co"])
    o2_col = _resolve_existing_column(out, "O2_mean_of_windows", ["o2"])
    nox_col = _resolve_existing_column(out, "NOX_mean_of_windows", ["nox"])
    thc_col = _resolve_existing_column(out, "THC_mean_of_windows", ["thc"])

    co2 = pd.to_numeric(out.get(co2_col, pd.Series(pd.NA, index=idx)), errors="coerce")
    co = pd.to_numeric(out.get(co_col, pd.Series(pd.NA, index=idx)), errors="coerce")
    o2 = pd.to_numeric(out.get(o2_col, pd.Series(pd.NA, index=idx)), errors="coerce")
    nox = pd.to_numeric(out.get(nox_col, pd.Series(pd.NA, index=idx)), errors="coerce")
    thc = pd.to_numeric(out.get(thc_col, pd.Series(pd.NA, index=idx)), errors="coerce")

    low_thc_mask = thc.abs().lt(THC_LOW_SIGNAL_WARN_PPM)
    negative_thc_mask = thc.lt(0)
    out["THC_LOW_SIGNAL_FLAG"] = low_thc_mask.astype("Int64")
    out["THC_NEGATIVE_FLAG"] = negative_thc_mask.astype("Int64")

    emissions_ref = specific_emissions_from_analyzer(
        air_kg_h=air_kg_h,
        fuel_kg_h=fuel_kg_h,
        power_kW=power_kW,
        co2_pct_dry=co2,
        co_pct_dry=co,
        o2_pct_dry=o2,
        nox_ppm_dry=nox,
        thc_ppm_dry=thc,
        h2o_wet_frac=pd.Series(np.nan, index=idx, dtype="float64"),
        nox_basis="NO",
    )

    dry_valid = pd.to_numeric(emissions_ref["MW_dry_kg_kmol"], errors="coerce").gt(0)
    h2o_wet_frac = (
        pd.to_numeric(exhaust_h2o_kg_h, errors="coerce") / MW_H2O_KG_KMOL
    ) / (
        (pd.to_numeric(exhaust_h2o_kg_h, errors="coerce") / MW_H2O_KG_KMOL)
        + (pd.to_numeric(exhaust_dry_kg_h, errors="coerce") / pd.to_numeric(emissions_ref["MW_dry_kg_kmol"], errors="coerce"))
    )
    h2o_wet_frac = h2o_wet_frac.where(dry_valid & exhaust_dry_kg_h.gt(0) & exhaust_h2o_kg_h.ge(0), pd.NA)
    h2o_wet_frac = h2o_wet_frac.clip(lower=0.0, upper=0.999)

    out["H2O_wet_frac"] = h2o_wet_frac

    emissions_no = specific_emissions_from_analyzer(
        air_kg_h=air_kg_h,
        fuel_kg_h=fuel_kg_h,
        power_kW=power_kW,
        co2_pct_dry=co2,
        co_pct_dry=co,
        o2_pct_dry=o2,
        nox_ppm_dry=nox,
        thc_ppm_dry=thc,
        h2o_wet_frac=h2o_wet_frac,
        nox_basis="NO",
    )
    emissions_no2 = specific_emissions_from_analyzer(
        air_kg_h=air_kg_h,
        fuel_kg_h=fuel_kg_h,
        power_kW=power_kW,
        co2_pct_dry=co2,
        co_pct_dry=co,
        o2_pct_dry=o2,
        nox_ppm_dry=nox,
        thc_ppm_dry=thc,
        h2o_wet_frac=h2o_wet_frac,
        nox_basis="NO2",
    )

    shared_cols = [
        "CO2_dry_frac",
        "CO_dry_frac",
        "O2_dry_frac",
        "NOx_dry_frac",
        "THC_dry_frac",
        "N2_dry_frac",
        "CO2_wet_frac",
        "CO_wet_frac",
        "O2_wet_frac",
        "NOx_wet_frac",
        "THC_wet_frac",
        "N2_wet_frac",
        "MW_dry_kg_kmol",
        "MW_wet_kg_kmol",
        "Exhaust_kg_h",
        "Exhaust_Dry_kg_h",
        "Exhaust_Dry_kmol_h",
        "Exhaust_H2O_kmol_h",
        "CO2_g_h",
        "CO_g_h",
        "THC_g_h",
        "CO2_g_kWh",
        "CO_g_kWh",
        "THC_g_kWh",
    ]
    for column in shared_cols:
        out[column] = emissions_no[column]

    out["MW_dry_kg_kmol_as_NO"] = emissions_no["MW_dry_kg_kmol"]
    out["MW_wet_kg_kmol_as_NO"] = emissions_no["MW_wet_kg_kmol"]
    out["MW_dry_kg_kmol_as_NO2"] = emissions_no2["MW_dry_kg_kmol"]
    out["MW_wet_kg_kmol_as_NO2"] = emissions_no2["MW_wet_kg_kmol"]
    out["NOx_g_h_as_NO"] = emissions_no["NOx_g_h"]
    out["NOx_g_kWh_as_NO"] = emissions_no["NOx_g_kWh"]
    out["NOx_g_h_as_NO2"] = emissions_no2["NOx_g_h"]
    out["NOx_g_kWh_as_NO2"] = emissions_no2["NOx_g_kWh"]
    out["NOx_as_NO_g_kWh"] = emissions_no["NOx_g_kWh"]
    out["NOx_as_NO2_g_kWh"] = emissions_no2["NOx_g_kWh"]
    out["EMISSIONS_H2O_SOURCE"] = humidity_source
    return out


# =========================
# File meta
# =========================
@dataclass(frozen=True)
class FileMeta:
    path: Path
    basename: str
    source_type: str  # "LABVIEW" or "KIBOX" or "MOTEC"
    load_kw: Optional[float]
    dies_pct: Optional[float]
    biod_pct: Optional[float]
    etoh_pct: Optional[float]
    h2o_pct: Optional[float]
    load_parse: str = ""
    composition_parse: str = ""


def _to_pct_or_none(x: object) -> Optional[float]:
    if x is None:
        return None
    try:
        v = float(str(x).replace(",", "."))
    except Exception:
        return None
    if not np.isfinite(v):
        return None
    return v


def _parse_filename_composition(stem: str) -> Tuple[Optional[float], Optional[float], Optional[float], Optional[float], str]:
    m_eh = re.search(r"E(\d+)\s*H(\d+)", stem, flags=re.IGNORECASE)
    if m_eh:
        return None, None, _to_pct_or_none(m_eh.group(1)), _to_pct_or_none(m_eh.group(2)), "filename_ethanol"

    dies_pct = None
    biod_pct = None
    etoh_pct = None
    h2o_pct = None

    m_db = re.search(r"(?:^|[^A-Za-z0-9])D(\d+(?:[.,]\d+)?)\s*B(\d+(?:[.,]\d+)?)(?:$|[^A-Za-z0-9])", stem, flags=re.IGNORECASE)
    if m_db:
        dies_pct = _to_pct_or_none(m_db.group(1))
        biod_pct = _to_pct_or_none(m_db.group(2))
        return dies_pct, biod_pct, None, None, "filename_diesel"

    m_bd = re.search(r"(?:^|[^A-Za-z0-9])B(\d+(?:[.,]\d+)?)\s*D(\d+(?:[.,]\d+)?)(?:$|[^A-Za-z0-9])", stem, flags=re.IGNORECASE)
    if m_bd:
        biod_pct = _to_pct_or_none(m_bd.group(1))
        dies_pct = _to_pct_or_none(m_bd.group(2))
        return dies_pct, biod_pct, None, None, "filename_diesel_reversed"

    m_be = re.search(r"(?:^|[^A-Za-z0-9])B(\d+(?:[.,]\d+)?)\s*E(\d+(?:[.,]\d+)?)(?:$|[^A-Za-z0-9])", stem, flags=re.IGNORECASE)
    if m_be:
        biod_pct = _to_pct_or_none(m_be.group(1))
        etoh_pct = _to_pct_or_none(m_be.group(2))
        return 0.0, biod_pct, etoh_pct, 0.0, "filename_biodiesel_ethanol"

    m_eb = re.search(r"(?:^|[^A-Za-z0-9])E(\d+(?:[.,]\d+)?)\s*B(\d+(?:[.,]\d+)?)(?:$|[^A-Za-z0-9])", stem, flags=re.IGNORECASE)
    if m_eb:
        etoh_pct = _to_pct_or_none(m_eb.group(1))
        biod_pct = _to_pct_or_none(m_eb.group(2))
        return 0.0, biod_pct, etoh_pct, 0.0, "filename_ethanol_biodiesel"

    m_de = re.search(r"(?:^|[^A-Za-z0-9])D(\d+(?:[.,]\d+)?)\s*E(\d+(?:[.,]\d+)?)(?:$|[^A-Za-z0-9])", stem, flags=re.IGNORECASE)
    if m_de:
        dies_pct = _to_pct_or_none(m_de.group(1))
        etoh_pct = _to_pct_or_none(m_de.group(2))
        return dies_pct, 0.0, etoh_pct, 0.0, "filename_diesel_ethanol"

    m_ed = re.search(r"(?:^|[^A-Za-z0-9])E(\d+(?:[.,]\d+)?)\s*D(\d+(?:[.,]\d+)?)(?:$|[^A-Za-z0-9])", stem, flags=re.IGNORECASE)
    if m_ed:
        etoh_pct = _to_pct_or_none(m_ed.group(1))
        dies_pct = _to_pct_or_none(m_ed.group(2))
        return dies_pct, 0.0, etoh_pct, 0.0, "filename_ethanol_diesel"

    m_b = re.match(r"^B(\d+(?:[.,]\d+)?)(?:$|[^A-Za-z0-9])", stem, flags=re.IGNORECASE)
    if m_b:
        biod_pct = _to_pct_or_none(m_b.group(1))
        if biod_pct is not None:
            return 0.0, biod_pct, 0.0, 0.0, "filename_biodiesel"

    m_d = re.match(r"^D(\d+(?:[.,]\d+)?)(?:$|[^A-Za-z0-9])", stem, flags=re.IGNORECASE)
    if m_d:
        dies_pct = _to_pct_or_none(m_d.group(1))
        if dies_pct is not None:
            return dies_pct, 0.0, 0.0, 0.0, "filename_diesel_only"

    m_dies = re.search(r"(?:dies_pct|diesel|dies)\s*[-_ ]*(\d+(?:[.,]\d+)?)", stem, flags=re.IGNORECASE)
    if m_dies:
        dies_pct = _to_pct_or_none(m_dies.group(1))

    m_biod = re.search(r"(?:biod_pct|biodiesel|biod)\s*[-_ ]*(\d+(?:[.,]\d+)?)", stem, flags=re.IGNORECASE)
    if m_biod:
        biod_pct = _to_pct_or_none(m_biod.group(1))

    if dies_pct is None and biod_pct is not None and 0.0 <= biod_pct <= 100.0:
        dies_pct = 100.0 - biod_pct
        return dies_pct, biod_pct, None, None, "filename_diesel_inferred"

    if biod_pct is None and dies_pct is not None and 0.0 <= dies_pct <= 100.0:
        biod_pct = 100.0 - dies_pct
        return dies_pct, biod_pct, None, None, "filename_diesel_inferred"

    if dies_pct is not None or biod_pct is not None:
        return dies_pct, biod_pct, None, None, "filename_diesel"

    return None, None, None, None, "missing_filename"


def parse_meta(path: Path) -> FileMeta:
    try:
        rel = path.relative_to(PROCESS_DIR)
        basename = "__".join(rel.with_suffix("").parts)
    except Exception:
        try:
            rel = path.relative_to(RAW_DIR)
            basename = "__".join(rel.with_suffix("").parts)
        except Exception:
            basename = "__".join((path.parent.name, path.stem))

    stem_lower = path.stem.lower()
    if stem_lower.endswith("_i"):
        source_type = "KIBOX"
    elif stem_lower.endswith("_m"):
        source_type = "MOTEC"
    else:
        source_type = "LABVIEW"

    load_tokens = re.findall(r"(\d+(?:[.,]\d+)?)\s*[-_ ]?\s*kw", path.stem, flags=re.IGNORECASE)
    if not load_tokens:
        bare_num = re.fullmatch(r"\s*(\d+(?:[.,]\d+)?)\s*", path.stem)
        if bare_num:
            load_tokens = [bare_num.group(1)]

    load_candidates: List[float] = []
    for tok in load_tokens:
        val = float(str(tok).replace(",", "."))
        if val not in load_candidates:
            load_candidates.append(val)

    if len(load_candidates) == 1:
        load_kw = load_candidates[0]
        load_parse = "filename"
    elif len(load_candidates) > 1:
        load_kw = None
        load_parse = "ambiguous_filename"
    else:
        load_kw = None
        load_parse = "missing_filename"

    dies_pct, biod_pct, etoh_pct, h2o_pct, composition_parse = _parse_filename_composition(path.stem)

    return FileMeta(
        path=path,
        basename=basename,
        source_type=source_type,
        load_kw=load_kw,
        dies_pct=dies_pct,
        biod_pct=biod_pct,
        etoh_pct=etoh_pct,
        h2o_pct=h2o_pct,
        load_parse=load_parse,
        composition_parse=composition_parse,
    )


# =========================
# LabVIEW read
# =========================
def list_sheet_names_xlsx(path: Path) -> List[str]:
    xf = _excel_file(path)
    return list(xf.sheet_names)


def choose_labview_sheet(path: Path) -> str:
    sheets = list_sheet_names_xlsx(path)
    if not sheets:
        raise ValueError(f"Nenhuma aba encontrada em {path.name}")

    for s in sheets:
        if s.strip().lower() == PREFERRED_SHEET_NAME.lower():
            return s
    for s in sheets:
        if "labview" in s.strip().lower():
            return s
    if len(sheets) == 1:
        return sheets[0]

    raise ValueError(f"NÃ£o encontrei aba '{PREFERRED_SHEET_NAME}' e existem mÃºltiplas abas em {path.name}: {sheets}.")


def find_b_etanol_col(df: pd.DataFrame) -> str:
    for c in B_ETANOL_COL_CANDIDATES:
        if c in df.columns:
            return c
    raise KeyError(
        f"NÃ£o encontrei coluna de balanÃ§a. Procurei: {B_ETANOL_COL_CANDIDATES}. "
        f"Colunas (primeiras 40): {list(df.columns)[:40]}"
    )


def _sanitize_labview_invalid_pressure_sentinels(df: pd.DataFrame, *, basename: str = "") -> pd.DataFrame:
    out = df.copy()
    sentinel = -1000.0
    hits: List[Tuple[str, int]] = []

    for col in out.columns:
        col_text = str(col).strip()
        if not col_text.startswith("P_"):
            continue

        series = pd.to_numeric(out[col], errors="coerce")
        bad_mask = series.eq(sentinel)
        n_bad = int(bad_mask.sum())
        if n_bad <= 0:
            continue

        out.loc[bad_mask, col] = pd.NA
        hits.append((col_text, n_bad))

    if hits:
        label = basename or "<unknown>"
        parts = ", ".join(f"{col}={count}" for col, count in hits)
        print(f"[WARN] Sentinela invalido '-1000' removido em {label}: {parts}")
    return out


def _infer_load_series_from_signal(df: pd.DataFrame) -> Optional[pd.Series]:
    load_col = "Carga (kW)" if "Carga (kW)" in df.columns else _find_first_col_by_substrings(df, ["carga", "kw"])
    if not load_col:
        return None

    v = pd.to_numeric(df[load_col], errors="coerce")
    if v.notna().sum() == 0:
        return None

    # Quantize the measured load to the expected 0.5 kW steps to keep grouping stable.
    return pd.Series(np.round(v * 2.0) / 2.0, index=df.index)


def _infer_single_load_from_signal(df: pd.DataFrame) -> Optional[float]:
    inferred = _infer_load_series_from_signal(df)
    if inferred is None:
        return None

    vals = sorted(pd.unique(pd.to_numeric(inferred, errors="coerce").dropna()))
    if len(vals) != 1:
        return None
    return float(vals[0])


def read_labview_xlsx(meta: FileMeta) -> pd.DataFrame:
    sheet = choose_labview_sheet(meta.path)
    df = _read_excel(meta.path, sheet_name=sheet)

    df.columns = _normalize_cols(list(df.columns))
    df = df.loc[:, ~pd.Series(df.columns).astype(str).str.startswith("Unnamed").values].copy()
    df = _sanitize_labview_invalid_pressure_sentinels(df, basename=meta.basename)

    df = df.reset_index(drop=True)
    df["Index"] = range(len(df))
    df["WindowID"] = df["Index"] // SAMPLES_PER_WINDOW

    load_series = pd.Series(meta.load_kw, index=df.index, dtype="float64")
    load_signal_series = pd.Series(np.nan, index=df.index, dtype="float64")
    dies_series = pd.Series(meta.dies_pct, index=df.index, dtype="float64")
    biod_series = pd.Series(meta.biod_pct, index=df.index, dtype="float64")
    etoh_series = pd.Series(meta.etoh_pct, index=df.index, dtype="float64")
    h2o_series = pd.Series(meta.h2o_pct, index=df.index, dtype="float64")
    inferred_load = _infer_load_series_from_signal(df)
    inferred_single = _infer_single_load_from_signal(df)

    if inferred_load is not None and inferred_load.notna().any():
        load_signal_series = pd.to_numeric(inferred_load, errors="coerce")

    if meta.load_kw is None or meta.load_parse == "ambiguous_filename":
        if inferred_load is not None and inferred_load.notna().any():
            load_series = pd.to_numeric(inferred_load, errors="coerce")
            print(f"[INFO] Load_kW inferido pela coluna de carga para {meta.basename} ({meta.load_parse}).")
        elif meta.load_kw is None:
            print(f"[WARN] Load_kW nÃ£o identificado para {meta.basename}; a coluna ficarÃ¡ vazia no output.")
    elif inferred_single is not None and abs(inferred_single - float(meta.load_kw)) > 0.75:
        print(
            f"[WARN] Load_kW do nome ({meta.load_kw:g}) difere da carga medida ({inferred_single:g}) "
            f"em {meta.basename}. Vou preservar o load nominal do nome e guardar a carga inferida em Load_Signal_kW."
        )

    df = df.assign(
        BaseName=meta.basename,
        Load_kW=load_series,
        Load_Signal_kW=load_signal_series,
        DIES_pct=dies_series,
        BIOD_pct=biod_series,
        EtOH_pct=etoh_series,
        H2O_pct=h2o_series,
    )

    first_cols = ["BaseName", "Load_kW", "Load_Signal_kW", "DIES_pct", "BIOD_pct", "EtOH_pct", "H2O_pct", "Index", "WindowID"]
    rest = [c for c in df.columns if c not in first_cols]
    return df[first_cols + rest].copy()


def _read_motec_metadata(path: Path, delim: str = ",") -> Dict[str, float]:
    meta: Dict[str, float] = {}
    with path.open("r", encoding="latin-1", errors="ignore", newline="") as fh:
        reader = csv.reader(fh, delimiter=delim)
        for i, row in enumerate(reader, start=1):
            if i > 14:
                break
            if not row:
                continue
            key = str(row[0]).replace("\ufeff", "").strip().strip('"')
            key_norm = norm_key(key)
            if key_norm == "sample rate" and len(row) > 1:
                meta["Motec_SampleRate_Hz"] = _to_float(row[1], default=np.nan)
            elif key_norm == "duration" and len(row) > 1:
                meta["Motec_Duration_s"] = _to_float(row[1], default=np.nan)
    return meta


def read_motec_csv(meta: FileMeta) -> pd.DataFrame:
    text = meta.path.read_text(encoding="latin-1", errors="ignore")
    sample = "\n".join(text.splitlines()[:20])
    delim = _sniff_delimiter(sample)

    try:
        df = pd.read_csv(meta.path, sep=delim, engine="python", encoding="utf-8-sig", skiprows=14)
    except UnicodeDecodeError:
        df = pd.read_csv(meta.path, sep=delim, engine="python", encoding="latin-1", skiprows=14)
    df.columns = _normalize_cols(list(df.columns))
    df = df.loc[:, ~pd.Series(df.columns).astype(str).str.startswith("Unnamed").values].copy()
    if len(df) < 1:
        raise ValueError(f"Arquivo MOTEC sem linhas de dados apos o cabecalho: {meta.path.name}")

    # Row 16 in the source file contains units. Data starts on row 17.
    df = df.iloc[1:].reset_index(drop=True).copy()
    motec_cols = []
    for i, c in enumerate(df.columns):
        clean = str(c).replace("\ufeff", "").strip()
        if not clean:
            clean = f"Col_{i + 1}"
        motec_cols.append(f"Motec_{clean}")
    df.columns = motec_cols

    meta_cols = _read_motec_metadata(meta.path, delim=delim)
    for key, value in meta_cols.items():
        df[key] = value

    time_col = next((c for c in df.columns if norm_key(c) == norm_key("Motec_Time")), "")
    if time_col:
        t = pd.to_numeric(df[time_col], errors="coerce")
        df["Motec_Time_Delta_s"] = t.diff()

    df = df.reset_index(drop=True)
    df["Index"] = range(len(df))
    df["WindowID"] = df["Index"] // SAMPLES_PER_WINDOW
    df["BaseName"] = meta.basename
    df["Load_kW"] = pd.Series(meta.load_kw, index=df.index, dtype="float64")
    df["DIES_pct"] = pd.Series(meta.dies_pct, index=df.index, dtype="float64")
    df["BIOD_pct"] = pd.Series(meta.biod_pct, index=df.index, dtype="float64")
    df["EtOH_pct"] = pd.Series(meta.etoh_pct, index=df.index, dtype="float64")
    df["H2O_pct"] = pd.Series(meta.h2o_pct, index=df.index, dtype="float64")

    first_cols = ["BaseName", "Load_kW", "DIES_pct", "BIOD_pct", "EtOH_pct", "H2O_pct", "Index", "WindowID"]
    rest = [c for c in df.columns if c not in first_cols]
    return df[first_cols + rest].copy()


def _parse_time_series(s: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(s):
        return pd.to_datetime(s, errors="coerce")

    dt = pd.to_datetime(s, errors="coerce")
    if dt.notna().any():
        return dt

    v = pd.to_numeric(s, errors="coerce")
    if v.notna().sum() == 0:
        return pd.Series(pd.NaT, index=s.index, dtype="datetime64[ns]")

    # Fallback for Excel serial date/time values.
    return pd.to_datetime(v, unit="D", origin="1899-12-30", errors="coerce")


def build_time_diagnostics(
    lv_raw: pd.DataFrame,
    time_col: str = "Time",
    quality_cfg: Optional[Dict[str, float]] = None,
) -> pd.DataFrame:
    quality_cfg = quality_cfg or {}
    max_delta_ms = _to_float(
        quality_cfg.get("MAX_DELTA_BETWEEN_SAMPLES_ms", DEFAULT_MAX_DELTA_BETWEEN_SAMPLES_MS),
        DEFAULT_MAX_DELTA_BETWEEN_SAMPLES_MS,
    )
    max_delta_s = max_delta_ms / 1000.0
    max_act_error_c = _to_float(
        quality_cfg.get("MAX_ACT_CONTROL_ERROR", DEFAULT_MAX_ACT_CONTROL_ERROR_C),
        DEFAULT_MAX_ACT_CONTROL_ERROR_C,
    )
    max_ect_error_c = _to_float(
        quality_cfg.get("MAX_ECT_CONTROL_ERROR", DEFAULT_MAX_ECT_CONTROL_ERROR_C),
        DEFAULT_MAX_ECT_CONTROL_ERROR_C,
    )

    try:
        time_col = resolve_col(lv_raw, time_col)
    except Exception:
        try:
            time_col = resolve_col(lv_raw, "TIME")
        except Exception:
            return pd.DataFrame()

    if time_col not in lv_raw.columns:
        return pd.DataFrame()

    base_cols = [c for c in ["BaseName", "Load_kW", "DIES_pct", "BIOD_pct", "EtOH_pct", "H2O_pct", "Index"] if c in lv_raw.columns]
    out = lv_raw[base_cols + [time_col]].copy()
    out = add_source_identity_columns(out)
    out = add_run_context_columns(out)

    t = _parse_time_series(out[time_col])
    base_name_series = out.get("BaseName", pd.Series([pd.NA] * len(out), index=out.index))
    out["TIME_PARSED"] = t
    out["TIME_HOUR"] = t.dt.hour.astype("Int64")
    out["TIME_MINUTE"] = t.dt.minute.astype("Int64")
    out["TIME_SECOND"] = t.dt.second.astype("Int64")
    out["TIME_MILLISECOND"] = (t.dt.microsecond // 1000).astype("Int64")

    prev_t = t.groupby(base_name_series, dropna=False, sort=False).shift(1)
    next_t = t.groupby(base_name_series, dropna=False, sort=False).shift(-1)

    delta_from_prev_s = (t - prev_t).dt.total_seconds()
    delta_to_next_s = (next_t - t).dt.total_seconds()
    out["TIME_DELTA_FROM_PREV_s"] = delta_from_prev_s
    out["TIME_DELTA_TO_NEXT_s"] = delta_to_next_s
    out["TIME_DELTA_TO_NEXT_ms"] = delta_to_next_s * 1000.0

    typical_dt = delta_to_next_s.groupby(base_name_series, dropna=False, sort=False).transform("median")
    out["TIME_DELTA_REFERENCE_s"] = typical_dt
    out["TIME_DELTA_ERROR_ms"] = (delta_to_next_s - typical_dt) * 1000.0
    out["MAX_DELTA_BETWEEN_SAMPLES_ms"] = max_delta_ms
    out["TIME_DELTA_LIMIT_s"] = max_delta_s
    out["TIME_DELTA_LIMIT_ms"] = max_delta_ms
    out["TIME_DELTA_ERROR_FLAG"] = delta_to_next_s > max_delta_s
    out["TIME_SAMPLE_GLOBAL"] = np.arange(len(out), dtype=int)

    t_adm_col = "T_ADMISSAO" if "T_ADMISSAO" in lv_raw.columns else _find_first_col_by_substrings(lv_raw, ["t", "admiss"])
    dem_act_col = (
        "DEM ACT AQUECEDOR"
        if "DEM ACT AQUECEDOR" in lv_raw.columns
        else _find_first_col_by_substrings(lv_raw, ["dem", "act"])
    )
    out["MAX_ACT_CONTROL_ERROR"] = max_act_error_c
    out["ACT_CTRL_ACTUAL_C"] = pd.NA
    out["ACT_CTRL_TARGET_C"] = pd.NA
    out["ACT_CTRL_ERROR_C"] = pd.NA
    out["ACT_CTRL_ERROR_ABS_C"] = pd.NA
    out["ACT_CTRL_ERROR_FLAG"] = pd.NA
    if t_adm_col and dem_act_col:
        act_actual = pd.to_numeric(lv_raw[t_adm_col], errors="coerce")
        act_target = pd.to_numeric(lv_raw[dem_act_col], errors="coerce")
        act_err = act_actual - act_target
        out["ACT_CTRL_ACTUAL_C"] = act_actual
        out["ACT_CTRL_TARGET_C"] = act_target
        out["ACT_CTRL_ERROR_C"] = act_err
        out["ACT_CTRL_ERROR_ABS_C"] = act_err.abs()
        out["ACT_CTRL_ERROR_FLAG"] = act_err.abs() > max_act_error_c

    t_s_agua_col = None
    for cand in ["T_S_AGUA", "T_S_ÃGUA", "T_S AGUA", "T_S ÃGUA"]:
        if cand in lv_raw.columns:
            t_s_agua_col = cand
            break
    if t_s_agua_col is None:
        t_s_agua_col = _find_first_col_by_substrings(lv_raw, ["t_s", "agua"])
    if t_s_agua_col is None:
        t_s_agua_col = _find_first_col_by_substrings(lv_raw, ["t_s", "Ã¡gua"])

    dem_th2o_col = None
    for cand in ["DEM_TH2O", "DEM TH2O"]:
        if cand in lv_raw.columns:
            dem_th2o_col = cand
            break
    if dem_th2o_col is None:
        dem_th2o_col = _find_first_col_by_substrings(lv_raw, ["dem", "th2o"])

    out["MAX_ECT_CONTROL_ERROR"] = max_ect_error_c
    out["ECT_CTRL_ACTUAL_C"] = pd.NA
    out["ECT_CTRL_TARGET_C"] = pd.NA
    out["ECT_CTRL_LIMIT_LOW_C"] = pd.NA
    out["ECT_CTRL_LIMIT_HIGH_C"] = pd.NA
    out["ECT_CTRL_ERROR_C"] = pd.NA
    out["ECT_CTRL_ERROR_ABS_C"] = pd.NA
    out["ECT_CTRL_ERROR_FLAG"] = pd.NA
    if t_s_agua_col and dem_th2o_col:
        ect_actual = pd.to_numeric(lv_raw[t_s_agua_col], errors="coerce")
        ect_target = pd.to_numeric(lv_raw[dem_th2o_col], errors="coerce")
        ect_err = ect_actual - ect_target
        out["ECT_CTRL_ACTUAL_C"] = ect_actual
        out["ECT_CTRL_TARGET_C"] = ect_target
        out["ECT_CTRL_LIMIT_LOW_C"] = ect_target - max_ect_error_c
        out["ECT_CTRL_LIMIT_HIGH_C"] = ect_target + max_ect_error_c
        out["ECT_CTRL_ERROR_C"] = ect_err
        out["ECT_CTRL_ERROR_ABS_C"] = ect_err.abs()
        out["ECT_CTRL_ERROR_FLAG"] = ect_err.abs() > max_ect_error_c

    return out


def _time_diag_load_title(load_kw: object) -> str:
    v = pd.to_numeric(pd.Series([load_kw]), errors="coerce").iloc[0]
    if pd.isna(v):
        return "carga_desconhecida"
    v = float(v)
    txt = f"{int(v)}" if v.is_integer() else f"{v:g}".replace(".", ",")
    return f"{txt} kW"


def _time_diag_load_slug(load_kw: object) -> str:
    v = pd.to_numeric(pd.Series([load_kw]), errors="coerce").iloc[0]
    if pd.isna(v):
        return "carga_desconhecida"
    v = float(v)
    txt = f"{int(v)}" if v.is_integer() else f"{v:g}".replace(".", "p")
    return f"{txt}kW"


def _time_diag_has_sampling_error(dt_next: pd.Series, threshold_s: float = TIME_DELTA_ERROR_THRESHOLD_S) -> bool:
    dt_num = pd.to_numeric(dt_next, errors="coerce")
    if dt_num.notna().sum() == 0:
        return False
    return bool((dt_num > threshold_s).any())


def _time_diag_status_from_flags(flags: pd.Series) -> str:
    s = pd.Series(flags)
    valid = s.dropna()
    if valid.empty:
        return "NA"
    return "ERRO" if bool(valid.astype(bool).any()) else "OK"


def _first_last_transient_times(
    flags: pd.Series,
    time_parsed: pd.Series,
) -> Tuple[object, object]:
    mask = pd.Series(flags).fillna(False).astype(bool)
    if mask.sum() == 0:
        return pd.NA, pd.NA

    times = pd.to_datetime(pd.Series(time_parsed), errors="coerce")
    flagged_times = times[mask].dropna()
    if flagged_times.empty:
        return pd.NA, pd.NA

    return (
        flagged_times.iloc[0].strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
        flagged_times.iloc[-1].strftime("%Y-%m-%d %H:%M:%S.%f")[:-3],
    )


def _apply_time_delta_axis_format(ax: plt.Axes) -> None:
    ax.set_ylim(TIME_DELTA_PLOT_YMIN_S, TIME_DELTA_PLOT_YMAX_S)
    ax.set_yticks(
        np.arange(
            TIME_DELTA_PLOT_YMIN_S,
            TIME_DELTA_PLOT_YMAX_S + (TIME_DELTA_PLOT_YSTEP_S * 0.5),
            TIME_DELTA_PLOT_YSTEP_S,
        )
    )


def summarize_time_diagnostics(time_df: pd.DataFrame) -> pd.DataFrame:
    if time_df is None or time_df.empty:
        return pd.DataFrame()

    rows: List[dict] = []
    for basename, d in time_df.groupby("BaseName", dropna=False, sort=False):
        dt_next = pd.to_numeric(d["TIME_DELTA_TO_NEXT_s"], errors="coerce")
        err_ms = pd.to_numeric(d["TIME_DELTA_ERROR_ms"], errors="coerce")
        t_parsed = pd.to_datetime(d["TIME_PARSED"], errors="coerce")
        time_limit_ms = _to_float(
            d.get("MAX_DELTA_BETWEEN_SAMPLES_ms", pd.Series([DEFAULT_MAX_DELTA_BETWEEN_SAMPLES_MS])).iloc[0],
            DEFAULT_MAX_DELTA_BETWEEN_SAMPLES_MS,
        )
        time_limit_s = time_limit_ms / 1000.0
        time_flag = d.get("TIME_DELTA_ERROR_FLAG", pd.Series([pd.NA] * len(d)))
        smp_status = _time_diag_status_from_flags(time_flag)
        act_flag = d.get("ACT_CTRL_ERROR_FLAG", pd.Series([pd.NA] * len(d)))
        act_status = _time_diag_status_from_flags(act_flag)
        ect_flag = d.get("ECT_CTRL_ERROR_FLAG", pd.Series([pd.NA] * len(d)))
        ect_status = _time_diag_status_from_flags(ect_flag)
        dq_status = "ERRO" if "ERRO" in {smp_status, act_status, ect_status} else ("OK" if {smp_status, act_status, ect_status} <= {"OK"} else "NA")
        time_error_n = int(pd.Series(time_flag).fillna(False).astype(bool).sum()) if smp_status != "NA" else 0
        act_error_n = int(pd.Series(act_flag).fillna(False).astype(bool).sum()) if act_status != "NA" else 0
        ect_error_n = int(pd.Series(ect_flag).fillna(False).astype(bool).sum()) if ect_status != "NA" else 0
        act_abs = pd.to_numeric(d.get("ACT_CTRL_ERROR_ABS_C", pd.Series([pd.NA] * len(d))), errors="coerce")
        ect_abs = pd.to_numeric(d.get("ECT_CTRL_ERROR_ABS_C", pd.Series([pd.NA] * len(d))), errors="coerce")
        act_transient_status = act_status
        act_transient_t_on, act_transient_t_off = _first_last_transient_times(
            act_flag,
            d.get("TIME_PARSED", pd.Series([pd.NA] * len(d))),
        )
        ect_transient_status = ect_status
        ect_transient_t_on, ect_transient_t_off = _first_last_transient_times(
            ect_flag,
            d.get("TIME_PARSED", pd.Series([pd.NA] * len(d))),
        )
        max_act_error = _to_float(
            d.get("MAX_ACT_CONTROL_ERROR", pd.Series([DEFAULT_MAX_ACT_CONTROL_ERROR_C])).iloc[0],
            DEFAULT_MAX_ACT_CONTROL_ERROR_C,
        )
        max_ect_error = _to_float(
            d.get("MAX_ECT_CONTROL_ERROR", pd.Series([DEFAULT_MAX_ECT_CONTROL_ERROR_C])).iloc[0],
            DEFAULT_MAX_ECT_CONTROL_ERROR_C,
        )

        rows.append(
            {
                "Smp_ERROR": smp_status,
                "ACT_CTRL_ERRO": act_status,
                "ACT_CTRL_ERRO_TRANSIENTE": act_transient_status,
                "ACT_CTRL_ERRO_TRANSIENTE_t_on": act_transient_t_on,
                "ACT_CTRL_ERRO_TRANSIENTE_t_off": act_transient_t_off,
                "ECT_CTRL_ERRO": ect_status,
                "ECT_CTRL_ERRO_TRANSIENTE": ect_transient_status,
                "ECT_CTRL_ERRO_TRANSIENTE_t_on": ect_transient_t_on,
                "ECT_CTRL_ERRO_TRANSIENTE_t_off": ect_transient_t_off,
                "DQ_ERROR": dq_status,
                "BaseName": basename,
                "SourceFolder": d.get("SourceFolder", pd.Series([""])).iloc[0],
                "SourceFile": d.get("SourceFile", pd.Series([basename])).iloc[0],
                "Iteracao": pd.to_numeric(d.get("Iteracao", pd.Series([pd.NA])).iloc[0], errors="coerce"),
                "Sentido_Carga": d.get("Sentido_Carga", pd.Series([pd.NA])).iloc[0],
                "Load_kW": pd.to_numeric(d.get("Load_kW", pd.Series([pd.NA])).iloc[0], errors="coerce"),
                "DIES_pct": pd.to_numeric(d.get("DIES_pct", pd.Series([pd.NA])).iloc[0], errors="coerce"),
                "BIOD_pct": pd.to_numeric(d.get("BIOD_pct", pd.Series([pd.NA])).iloc[0], errors="coerce"),
                "EtOH_pct": pd.to_numeric(d.get("EtOH_pct", pd.Series([pd.NA])).iloc[0], errors="coerce"),
                "H2O_pct": pd.to_numeric(d.get("H2O_pct", pd.Series([pd.NA])).iloc[0], errors="coerce"),
                "N_samples": int(len(d)),
                "TIME_START": t_parsed.min(),
                "TIME_END": t_parsed.max(),
                "MAX_DELTA_BETWEEN_SAMPLES_ms": time_limit_ms,
                "TIME_DELTA_ERROR_N": time_error_n,
                "TIME_DELTA_ERROR_PCT": (time_error_n / len(d)) * 100.0 if len(d) > 0 else np.nan,
                "TIME_DELTA_MEDIAN_s": dt_next.median(),
                "TIME_DELTA_MEAN_s": dt_next.mean(),
                "TIME_DELTA_MIN_s": dt_next.min(),
                "TIME_DELTA_MAX_s": dt_next.max(),
                "TIME_DELTA_LIMIT_s": time_limit_s,
                "TIME_DELTA_STD_ms": dt_next.std(ddof=1) * 1000.0,
                "TIME_DELTA_MAX_ABS_ERROR_ms": err_ms.abs().max(),
                "TIME_DELTA_NONPOSITIVE_N": int((dt_next <= 0).fillna(False).sum()),
                "TIME_DELTA_MISSING_N": int(dt_next.isna().sum()),
                "MAX_ACT_CONTROL_ERROR": max_act_error,
                "ACT_CTRL_ERROR_N": act_error_n,
                "ACT_CTRL_ERROR_PCT": (act_error_n / len(d)) * 100.0 if len(d) > 0 else np.nan,
                "ACT_CTRL_ERROR_MEAN_ABS_C": act_abs.mean(),
                "ACT_CTRL_ERROR_MAX_ABS_C": act_abs.max(),
                "MAX_ECT_CONTROL_ERROR": max_ect_error,
                "ECT_CTRL_ERROR_N": ect_error_n,
                "ECT_CTRL_ERROR_PCT": (ect_error_n / len(d)) * 100.0 if len(d) > 0 else np.nan,
                "ECT_CTRL_ERROR_MEAN_ABS_C": ect_abs.mean(),
                "ECT_CTRL_ERROR_MAX_ABS_C": ect_abs.max(),
            }
        )

    out = pd.DataFrame(rows)
    if "Iteracao" in out.columns:
        out["Iteracao"] = pd.to_numeric(out["Iteracao"], errors="coerce").astype("Int64")
    return out


def plot_time_delta_all_samples(
    time_df: pd.DataFrame,
    filename: str = "time_delta_to_next_all_samples.png",
    plot_dir: Optional[Path] = None,
) -> Dict[str, object]:
    if time_df is None or time_df.empty:
        print("[WARN] Sem dados para plot de delta T do TIME.")
        return

    d = time_df.sort_values(["BaseName", "Index"]).copy()
    x = pd.to_numeric(d["TIME_SAMPLE_GLOBAL"], errors="coerce")
    y = pd.to_numeric(d["TIME_DELTA_TO_NEXT_s"], errors="coerce")
    valid = x.notna() & y.notna()
    if valid.sum() == 0:
        print("[WARN] Sem delta T vÃ¡lido para plotar.")
        return

    fig, ax = plt.subplots(figsize=(14, 5))
    ax.plot(x[valid], y[valid], "-", linewidth=0.8, color="tab:blue", alpha=0.85)
    valid_idx = np.flatnonzero(valid.to_numpy(dtype=bool))
    if len(valid_idx) > 0:
        step = max(len(valid_idx) // TIME_DIAG_FILE_SCATTER_MAX_POINTS, 1)
        scatter_idx = valid_idx[::step]
        ax.scatter(x.iloc[scatter_idx], y.iloc[scatter_idx], s=8, color="tab:blue", alpha=0.35)

    median_dt = float(y[valid].median())
    ax.axhline(median_dt, color="tab:red", linestyle="--", linewidth=1.0, label=f"median={median_dt:.6f} s")
    time_limit_s = _to_float(
        d.get("TIME_DELTA_LIMIT_s", pd.Series([TIME_DELTA_ERROR_THRESHOLD_S])).iloc[0],
        TIME_DELTA_ERROR_THRESHOLD_S,
    )
    ax.axhline(
        time_limit_s,
        color="tab:orange",
        linestyle=":",
        linewidth=1.2,
        label=f"limite erro={time_limit_s:.3f} s",
    )

    ax.set_xlabel("Global sample index")
    ax.set_ylabel("Delta T to next sample (s)")
    ax.set_title("TIME delta between consecutive samples")
    _apply_time_delta_axis_format(ax)
    ax.grid(True, which="both", linestyle="--", linewidth=0.5)
    table_rows = [("", xi, yi) for xi, yi in zip(x[valid].tolist(), y[valid].tolist())]
    _add_xy_value_table(ax, table_rows)
    ax.legend()

    target_dir = PLOTS_DIR if plot_dir is None else plot_dir
    target_dir.mkdir(parents=True, exist_ok=True)
    outpath = target_dir / filename
    fig.tight_layout()
    fig.savefig(outpath, dpi=TIME_DIAG_PLOT_DPI)
    plt.close(fig)
    print(f"[OK] Salvei {outpath}")


def plot_time_delta_by_file(time_df: pd.DataFrame, plot_dir: Optional[Path] = None) -> None:
    if time_df is None or time_df.empty:
        print("[WARN] Sem dados para plots individuais de delta T do TIME.")
        return

    base_dir = PLOTS_DIR if plot_dir is None else plot_dir
    out_dir = base_dir / "time_delta_by_file"
    out_dir.mkdir(parents=True, exist_ok=True)

    n_ok = 0
    n_skip = 0
    for basename, d in time_df.groupby("BaseName", dropna=False, sort=True):
        d = d.sort_values("Index").copy()
        x = pd.to_numeric(d["Index"], errors="coerce")
        y = pd.to_numeric(d["TIME_DELTA_TO_NEXT_s"], errors="coerce")
        valid = x.notna() & y.notna()
        if valid.sum() == 0:
            n_skip += 1
            skipped_items.append((plot_label, "plot_type vazio"))
            continue

        source_folder = str(d.get("SourceFolder", pd.Series([""])).iloc[0] or "")
        source_file = str(d.get("SourceFile", pd.Series([basename])).iloc[0] or basename)
        load_kw = d.get("Load_kW", pd.Series([pd.NA])).iloc[0]
        load_title = _time_diag_load_title(load_kw)
        load_slug = _time_diag_load_slug(load_kw)
        time_limit_s = _to_float(
            d.get("TIME_DELTA_LIMIT_s", pd.Series([TIME_DELTA_ERROR_THRESHOLD_S])).iloc[0],
            TIME_DELTA_ERROR_THRESHOLD_S,
        )
        error_mask = valid & (y > time_limit_s)
        has_sampling_error = bool(error_mask.any())

        fig, ax = plt.subplots(figsize=(12, 4.5))
        ax.plot(x[valid], y[valid], "-", linewidth=0.9, color="tab:blue", alpha=0.9)
        valid_idx = np.flatnonzero(valid.to_numpy(dtype=bool))
        if len(valid_idx) > 0:
            step = max(len(valid_idx) // TIME_DIAG_FILE_SCATTER_MAX_POINTS, 1)
            scatter_idx = valid_idx[::step]
            ax.scatter(x.iloc[scatter_idx], y.iloc[scatter_idx], s=10, color="tab:blue", alpha=0.35)
        if has_sampling_error:
            ax.scatter(x[error_mask], y[error_mask], s=18, color="tab:red", alpha=0.95, label=f"delta T > {time_limit_s:.3f} s")
            ax.set_facecolor("#fff4f4")
            for spine in ax.spines.values():
                spine.set_color("tab:red")
                spine.set_linewidth(1.2)

        median_dt = float(y[valid].median())
        ax.axhline(median_dt, color="tab:red", linestyle="--", linewidth=1.0, label=f"median={median_dt:.6f} s")
        ax.axhline(
            time_limit_s,
            color="tab:orange",
            linestyle=":",
            linewidth=1.2,
            label=f"limite erro={time_limit_s:.3f} s",
        )

        title_parts = ["TIME delta entre amostras", source_file]
        if has_sampling_error:
            title_parts.insert(0, "ERRO")
        if source_folder:
            title_parts.append(source_folder)
        title_parts.append(load_title)

        ax.set_xlabel("Sample index in file")
        ax.set_ylabel("Delta T to next sample (s)")
        ax.set_title(" | ".join(title_parts))
        _apply_time_delta_axis_format(ax)
        ax.grid(True, which="both", linestyle="--", linewidth=0.5)
        table_rows = [("", xi, yi) for xi, yi in zip(x[valid].tolist(), y[valid].tolist())]
        _add_xy_value_table(ax, table_rows)
        ax.legend()

        error_prefix = "ERRO_" if has_sampling_error else ""
        filename_stem = f"{error_prefix}time_delta_to_next_{source_folder}_{load_slug}_{source_file}"
        outpath = out_dir / f"{_safe_name(filename_stem)}.png"
        fig.tight_layout()
        fig.savefig(outpath, dpi=TIME_DIAG_PLOT_DPI)
        plt.close(fig)
        print(f"[OK] Salvei {outpath}")
        n_ok += 1

    print(f"[OK] Plots TIME por arquivo: {n_ok} gerados; {n_skip} pulados.")


# =========================
# Kibox robust parsing
# =========================
def _sniff_delimiter(sample: str) -> str:
    candidates = [",", ";", "\t", "|"]
    counts = {d: sample.count(d) for d in candidates}
    return max(counts, key=counts.get)


def _find_header_row(lines: List[str], delim: str, min_cols: int = 6) -> int:
    best_i = 0
    best_cols = 0
    for i, ln in enumerate(lines[:80]):
        cols = ln.split(delim)
        ncols = len(cols)
        if ncols > best_cols:
            best_cols = ncols
            best_i = i
        if ncols >= min_cols and any(ch.isalpha() for ch in ln):
            return i
    return best_i


_num_regex = re.compile(r"[-+]?(\d{1,3}(\.\d{3})+|\d+)([.,]\d+)?")


def _coerce_numeric_series(s: pd.Series) -> pd.Series:
    if pd.api.types.is_numeric_dtype(s):
        return pd.to_numeric(s, errors="coerce")

    x = s.astype(str).str.replace("\ufeff", "", regex=False).str.strip()
    x = x.str.replace("\u00A0", " ", regex=False).str.replace(" ", "", regex=False)

    extracted = x.str.extract(_num_regex)[0]
    extracted = extracted.where(extracted.notna(), None)

    def fix_num(v: Optional[str]) -> Optional[str]:
        if v is None:
            return None
        v = str(v)
        if "," in v and "." in v:
            if v.rfind(",") > v.rfind("."):
                v = v.replace(".", "").replace(",", ".")
            else:
                v = v.replace(",", "")
            return v
        if "," in v and "." not in v:
            return v.replace(",", ".")
        return v

    fixed = extracted.map(fix_num)
    return pd.to_numeric(fixed, errors="coerce")


def read_kibox_csv_robust(path: Path) -> pd.DataFrame:
    text = path.read_text(encoding="utf-8-sig", errors="ignore")
    sample = "\n".join(text.splitlines()[:50])
    delim = _sniff_delimiter(sample)
    lines = text.splitlines()
    header_row = _find_header_row(lines, delim=delim, min_cols=6)

    df = pd.read_csv(path, sep=delim, engine="python", encoding="utf-8-sig", skiprows=header_row)
    df.columns = _normalize_cols(list(df.columns))
    df = df.loc[:, ~pd.Series(df.columns).astype(str).str.startswith("Unnamed").values].copy()
    df = _coalesce_equivalent_columns(df, context=path.name)
    return df


def kibox_mean_row(meta: FileMeta) -> pd.DataFrame:
    df_raw = read_kibox_csv_robust(meta.path)
    num_df = pd.DataFrame({c: _coerce_numeric_series(df_raw[c]) for c in df_raw.columns})

    keep_cols = [c for c in num_df.columns if num_df[c].notna().mean() >= 0.2]
    if not keep_cols:
        fill = sorted([(c, float(num_df[c].notna().mean())) for c in num_df.columns], key=lambda x: x[1], reverse=True)
        keep_cols = [c for c, _ in fill[:30]]

    means = num_df[keep_cols].mean(numeric_only=True)

    row = {f"KIBOX_{c}": float(means[c]) if pd.notna(means[c]) else pd.NA for c in means.index}
    row.update(
        {
            "SourceFolder": _basename_source_folder_display(meta.basename),
            "Load_kW": meta.load_kw,
            "DIES_pct": meta.dies_pct,
            "BIOD_pct": meta.biod_pct,
            "EtOH_pct": meta.etoh_pct,
            "H2O_pct": meta.h2o_pct,
        }
    )
    return pd.DataFrame([row])


def kibox_aggregate(kibox_files: List[FileMeta]) -> pd.DataFrame:
    rows: List[pd.DataFrame] = []
    for m in kibox_files:
        has_diesel = m.dies_pct is not None or m.biod_pct is not None
        has_ethanol = m.etoh_pct is not None or m.h2o_pct is not None
        if m.load_kw is None or (not has_diesel and not has_ethanol):
            print(f"[WARN] Kibox sem composicao valida no nome (nao vou agregar): {m.path.name}")
            continue
        try:
            rows.append(kibox_mean_row(m))
        except Exception as e:
            print(f"[ERROR] Kibox {m.path.name}: {e}")

    if not rows:
        return pd.DataFrame(columns=["Load_kW", "DIES_pct", "BIOD_pct", "EtOH_pct", "H2O_pct"])

    allk = pd.concat(rows, ignore_index=True)
    key_cols = ["SourceFolder", "Load_kW", "DIES_pct", "BIOD_pct", "EtOH_pct", "H2O_pct"]
    value_cols = [c for c in allk.columns if c.startswith("KIBOX_")]

    agg = allk.groupby(key_cols, dropna=False, sort=True)[value_cols].mean(numeric_only=True).reset_index()
    cnt = allk.groupby(key_cols, dropna=False, sort=True).size().reset_index(name="KIBOX_N_files")
    return agg.merge(cnt, on=key_cols, how="left")


# =========================
# Config / LHV / Instruments rev3
# =========================
def _choose_text_config_dir(config_dir: Optional[Path] = None) -> Path:
    if config_dir is None:
        return default_text_config_dir(BASE_DIR)
    return Path(config_dir).expanduser().resolve()


def _choose_config_path() -> Path:
    p = CFG_DIR / "config_incertezas_rev3.xlsx"
    if p.exists():
        return p
    raise FileNotFoundError(f"Nao encontrei {p.name} em {CFG_DIR.resolve()}")


def _prepare_config_bundle_for_pipeline(bundle: Pipeline29ConfigBundle) -> Pipeline29ConfigBundle:
    mappings_prepared: Dict[str, Dict[str, str]] = {}
    for key, spec in (bundle.mappings or {}).items():
        key_norm = norm_key(key)
        if not key_norm:
            continue
        mappings_prepared[key_norm] = {
            "mean": _to_str_or_empty((spec or {}).get("mean", "")),
            "sd": _to_str_or_empty((spec or {}).get("sd", "")),
            "unit": _to_str_or_empty((spec or {}).get("unit", "")),
            "notes": _to_str_or_empty((spec or {}).get("notes", "")),
        }

    defaults_prepared = {
        norm_key(key): _to_str_or_empty(value)
        for key, value in (bundle.defaults_cfg or {}).items()
        if norm_key(key)
    }

    instruments_df = bundle.instruments_df.copy() if bundle.instruments_df is not None else pd.DataFrame()
    if "key" not in instruments_df.columns:
        instruments_df["key"] = pd.NA
    instruments_df["key_norm"] = instruments_df["key"].map(norm_key)

    reporting_df = bundle.reporting_df.copy() if bundle.reporting_df is not None else pd.DataFrame()
    if "key" not in reporting_df.columns:
        reporting_df["key"] = pd.NA
    reporting_df["key_norm"] = reporting_df["key"].map(norm_key)

    plots_df = bundle.plots_df.copy() if bundle.plots_df is not None else pd.DataFrame()
    compare_df = bundle.compare_df.copy() if bundle.compare_df is not None else pd.DataFrame()
    fuel_properties_df = bundle.fuel_properties_df.copy() if bundle.fuel_properties_df is not None else pd.DataFrame()

    return Pipeline29ConfigBundle(
        mappings=mappings_prepared,
        instruments_df=instruments_df,
        reporting_df=reporting_df,
        plots_df=plots_df,
        compare_df=compare_df,
        fuel_properties_df=fuel_properties_df,
        data_quality_cfg=dict(bundle.data_quality_cfg or {}),
        defaults_cfg=defaults_prepared,
        source_kind=bundle.source_kind,
        source_path=bundle.source_path,
        text_dir=bundle.text_dir,
    )


def load_pipeline29_config_bundle(
    *,
    config_source: str = "auto",
    text_config_dir: Optional[Path] = None,
    rebuild_text_config: bool = False,
) -> Pipeline29ConfigBundle:
    source_mode = _to_str_or_empty(config_source).lower() or "auto"
    if source_mode not in {"auto", "text", "excel"}:
        raise ValueError(f"config_source invalido: {config_source}")

    text_dir = _choose_text_config_dir(text_config_dir)

    if source_mode in {"auto", "text"}:
        if rebuild_text_config or not text_config_exists(text_dir):
            excel_path = _choose_config_path()
            print(f"[INFO] Gerando config textual do pipeline29 em: {text_dir}")
            bootstrap_text_config_from_excel(excel_path, text_dir)
        if text_config_exists(text_dir):
            bundle = load_text_config_bundle(text_dir)
            bundle.text_dir = text_dir
            bundle.source_kind = "text"
            bundle.source_path = text_dir
            return _prepare_config_bundle_for_pipeline(bundle)
        if source_mode == "text":
            raise FileNotFoundError(f"Nao encontrei config textual completa em {text_dir}")

    excel_path = _choose_config_path()
    bundle = load_excel_config_bundle(excel_path)
    bundle.source_kind = "excel"
    bundle.source_path = excel_path
    bundle.text_dir = text_dir if text_config_exists(text_dir) else None
    return _prepare_config_bundle_for_pipeline(bundle)


def _try_read_sheet(xlsx_path: Path, sheet: str) -> Optional[pd.DataFrame]:
    try:
        xf = _excel_file(xlsx_path)
        selected_sheet = None
        for s in xf.sheet_names:
            if s == sheet or str(s).strip().lower() == str(sheet).strip().lower():
                selected_sheet = s
                break
        if selected_sheet is None:
            return None
        return _read_excel(xlsx_path, sheet_name=selected_sheet)
    except Exception:
        return None


def _load_data_quality_config(xlsx_path: Path) -> Dict[str, float]:
    cfg = {
        "MAX_DELTA_BETWEEN_SAMPLES_ms": DEFAULT_MAX_DELTA_BETWEEN_SAMPLES_MS,
        "MAX_ACT_CONTROL_ERROR": DEFAULT_MAX_ACT_CONTROL_ERROR_C,
        "MAX_ECT_CONTROL_ERROR": DEFAULT_MAX_ECT_CONTROL_ERROR_C,
    }

    dqa = _try_read_sheet(xlsx_path, "data quality assessment")
    if dqa is None or dqa.empty:
        return cfg

    dqa.columns = _normalize_cols(list(dqa.columns))
    param_col = "param" if "param" in dqa.columns else (dqa.columns[0] if len(dqa.columns) >= 1 else None)
    value_col = "value" if "value" in dqa.columns else (dqa.columns[1] if len(dqa.columns) >= 2 else None)
    if param_col is None or value_col is None:
        return cfg

    for _, row in dqa.iterrows():
        param = str(row.get(param_col, "")).replace("\ufeff", "").strip()
        if not param:
            continue
        param_norm = norm_key(param)
        if param_norm == norm_key("MAX_DELTA_BETWEEN_SAMPLES_ms"):
            cfg["MAX_DELTA_BETWEEN_SAMPLES_ms"] = _to_float(
                row.get(value_col, DEFAULT_MAX_DELTA_BETWEEN_SAMPLES_MS),
                DEFAULT_MAX_DELTA_BETWEEN_SAMPLES_MS,
            )
        elif param_norm == norm_key("MAX_ACT_CONTROL_ERROR"):
            cfg["MAX_ACT_CONTROL_ERROR"] = _to_float(
                row.get(value_col, DEFAULT_MAX_ACT_CONTROL_ERROR_C),
                DEFAULT_MAX_ACT_CONTROL_ERROR_C,
            )
        elif param_norm == norm_key("MAX_ECT_CONTROL_ERROR"):
            cfg["MAX_ECT_CONTROL_ERROR"] = _to_float(
                row.get(value_col, DEFAULT_MAX_ECT_CONTROL_ERROR_C),
                DEFAULT_MAX_ECT_CONTROL_ERROR_C,
            )

    return cfg


def _load_defaults_config(xlsx_path: Path) -> Dict[str, str]:
    cfg: Dict[str, str] = {}

    defaults_df = _try_read_sheet(xlsx_path, "Defaults")
    if defaults_df is None or defaults_df.empty:
        return cfg

    defaults_df.columns = _normalize_cols(list(defaults_df.columns))
    param_col = "param" if "param" in defaults_df.columns else (defaults_df.columns[0] if len(defaults_df.columns) >= 1 else None)
    value_col = "value" if "value" in defaults_df.columns else (defaults_df.columns[1] if len(defaults_df.columns) >= 2 else None)
    if param_col is None or value_col is None:
        return cfg

    for _, row in defaults_df.iterrows():
        param = str(row.get(param_col, "")).replace("\ufeff", "").strip()
        if not param or "global parameter name" in param.lower():
            continue
        value = row.get(value_col, "")
        if pd.isna(value):
            value = ""
        cfg[norm_key(param)] = str(value).replace("\ufeff", "").strip()

    return cfg


def _resolve_runtime_dir(value: object, default_path: Path) -> Path:
    raw = str(value or "").replace("\ufeff", "").strip().strip('"').strip("'")
    if not raw:
        return default_path

    p = Path(raw).expanduser()
    if not p.is_absolute():
        p = (BASE_DIR / p).resolve()
    return p


def _prepare_output_dir(path: Path) -> bool:
    try:
        path.mkdir(parents=True, exist_ok=True)
        return True
    except Exception:
        return False


def _load_runtime_path_settings() -> Dict[str, str]:
    try:
        if not RUNTIME_SETTINGS_PATH.exists():
            return {}
        return json.loads(RUNTIME_SETTINGS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _save_runtime_path_settings(input_dir: Path, out_dir: Path) -> None:
    RUNTIME_SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
    payload = {
        "raw_input_dir": str(input_dir),
        "out_dir": str(out_dir),
    }
    RUNTIME_SETTINGS_PATH.write_text(
        json.dumps(payload, indent=2, ensure_ascii=True),
        encoding="utf-8",
    )


def _best_existing_dir(*candidates: object) -> Path:
    for candidate in candidates:
        raw = str(candidate or "").strip().strip('"').strip("'")
        if not raw:
            continue
        try:
            p = Path(raw).expanduser()
        except Exception:
            continue
        if p.exists() and p.is_dir():
            return p
        if p.parent.exists() and p.parent.is_dir():
            return p.parent
    return BASE_DIR


def _run_windows_folder_dialog(*, title: str, initial_dir: Path) -> Optional[Path]:
    initial_dir = _best_existing_dir(initial_dir)
    escaped_title = title.replace("'", "''")
    escaped_initial = str(initial_dir).replace("'", "''")
    ps_script = f"""
Add-Type -AssemblyName System.Windows.Forms
$dialog = New-Object System.Windows.Forms.FolderBrowserDialog
$dialog.Description = '{escaped_title}'
$dialog.ShowNewFolderButton = $true
if (Test-Path -LiteralPath '{escaped_initial}') {{
    $dialog.SelectedPath = '{escaped_initial}'
}}
$result = $dialog.ShowDialog()
if ($result -eq [System.Windows.Forms.DialogResult]::OK -and $dialog.SelectedPath) {{
    [Console]::OutputEncoding = [System.Text.Encoding]::UTF8
    Write-Output $dialog.SelectedPath
    exit 0
}}
exit 2
"""

    completed = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
        capture_output=True,
        text=True,
    )
    stdout_lines = [line.strip() for line in completed.stdout.splitlines() if line.strip()]
    stdout = stdout_lines[0] if stdout_lines else ""
    stderr = completed.stderr.strip()
    if completed.returncode == 0 and stdout:
        return Path(stdout).expanduser().resolve()
    if completed.returncode == 2:
        return None
    raise RuntimeError(
        "Falha ao abrir o seletor nativo de pasta no Windows. "
        f"stdout={stdout!r} stderr={stderr!r} code={completed.returncode}"
    )


def _prompt_runtime_dirs_via_windows_dialog(initial_input_dir: Path, initial_out_dir: Path) -> Tuple[Path, Path]:
    print("[INFO] Abrindo seletor nativo do Windows para o diretorio de entrada...")
    input_dir = _run_windows_folder_dialog(
        title="Selecione o diretorio de entrada do pipeline",
        initial_dir=initial_input_dir,
    )
    if input_dir is None:
        raise SystemExit("Execucao cancelada pelo usuario na selecao do diretorio de entrada.")

    print("[INFO] Abrindo seletor nativo do Windows para o diretorio de saida...")
    out_dir = _run_windows_folder_dialog(
        title="Selecione o diretorio de saida do pipeline",
        initial_dir=initial_out_dir,
    )
    if out_dir is None:
        raise SystemExit("Execucao cancelada pelo usuario na selecao do diretorio de saida.")
    return input_dir, out_dir


def _prompt_runtime_dirs_via_tk_dialog(initial_input_dir: Path, initial_out_dir: Path) -> Tuple[Path, Path]:
    if tk is None or ttk is None or filedialog is None or messagebox is None:
        raise RuntimeError(
            "Tkinter nao esta disponivel neste Python. O pipeline29 agora exige popup Windows "
            "para selecionar RAW_INPUT_DIR e OUT_DIR."
        )

    root = tk.Tk()
    root.title("Pipeline 29 - Diretorios de execucao")
    root.resizable(False, False)
    root.attributes("-topmost", True)

    input_var = tk.StringVar(master=root, value=str(initial_input_dir))
    out_var = tk.StringVar(master=root, value=str(initial_out_dir))
    result: dict[str, Path] = {}

    root.columnconfigure(1, weight=1)

    ttk.Label(
        root,
        text="Selecione o diretorio de entrada do pipeline e o diretorio de saida para esta execucao.",
    ).grid(row=0, column=0, columnspan=3, sticky="w", padx=12, pady=(12, 10))

    ttk.Label(root, text="Input dir").grid(row=1, column=0, sticky="w", padx=(12, 8), pady=6)
    input_entry = ttk.Entry(root, textvariable=input_var, width=90)
    input_entry.grid(row=1, column=1, sticky="ew", padx=(0, 8), pady=6)

    def browse_input() -> None:
        selected = filedialog.askdirectory(
            parent=root,
            title="Selecione o diretorio de entrada do pipeline",
            initialdir=str(_best_existing_dir(input_var.get(), initial_input_dir)),
        )
        if selected:
            input_var.set(selected)

    ttk.Button(root, text="Browse...", command=browse_input).grid(row=1, column=2, sticky="e", padx=(0, 12), pady=6)

    ttk.Label(root, text="Out dir").grid(row=2, column=0, sticky="w", padx=(12, 8), pady=6)
    out_entry = ttk.Entry(root, textvariable=out_var, width=90)
    out_entry.grid(row=2, column=1, sticky="ew", padx=(0, 8), pady=6)

    def browse_output() -> None:
        selected = filedialog.askdirectory(
            parent=root,
            title="Selecione o diretorio de saida",
            initialdir=str(_best_existing_dir(out_var.get(), initial_out_dir)),
        )
        if selected:
            out_var.set(selected)

    ttk.Button(root, text="Browse...", command=browse_output).grid(row=2, column=2, sticky="e", padx=(0, 12), pady=6)

    ttk.Label(
        root,
        text="A ultima selecao fica salva localmente e volta preenchida na proxima abertura.",
    ).grid(row=3, column=0, columnspan=3, sticky="w", padx=12, pady=(4, 10))

    def confirm() -> None:
        raw_input = input_var.get().strip()
        out_input = out_var.get().strip()
        if not raw_input:
            messagebox.showerror("Pipeline 29", "Selecione o diretorio de entrada.", parent=root)
            return
        if not out_input:
            messagebox.showerror("Pipeline 29", "Selecione o diretorio de saida.", parent=root)
            return

        input_dir = Path(raw_input).expanduser().resolve()
        out_dir = Path(out_input).expanduser().resolve()

        if not input_dir.exists():
            messagebox.showerror("Pipeline 29", f"Input dir nao existe:\n{input_dir}", parent=root)
            return
        if not input_dir.is_dir():
            messagebox.showerror("Pipeline 29", f"Input dir nao e diretorio:\n{input_dir}", parent=root)
            return
        try:
            out_dir.mkdir(parents=True, exist_ok=True)
        except Exception as exc:
            messagebox.showerror(
                "Pipeline 29",
                f"Nao consegui preparar o diretorio de saida:\n{out_dir}\n\n{exc}",
                parent=root,
            )
            return

        result["input_dir"] = input_dir
        result["out_dir"] = out_dir
        root.destroy()

    def cancel() -> None:
        root.destroy()

    button_row = ttk.Frame(root)
    button_row.grid(row=4, column=0, columnspan=3, sticky="e", padx=12, pady=(0, 12))
    ttk.Button(button_row, text="Cancelar", command=cancel).pack(side="right")
    ttk.Button(button_row, text="Confirmar", command=confirm).pack(side="right", padx=(0, 8))

    root.protocol("WM_DELETE_WINDOW", cancel)
    root.bind("<Return>", lambda _event: confirm())
    root.bind("<Escape>", lambda _event: cancel())
    input_entry.focus_set()

    root.update_idletasks()
    width = max(root.winfo_reqwidth(), 900)
    height = root.winfo_reqheight()
    screen_w = root.winfo_screenwidth()
    screen_h = root.winfo_screenheight()
    pos_x = max((screen_w - width) // 2, 0)
    pos_y = max((screen_h - height) // 3, 0)
    root.geometry(f"{width}x{height}+{pos_x}+{pos_y}")
    root.deiconify()
    root.lift()
    try:
        root.focus_force()
    except Exception:
        pass
    root.after(400, lambda: root.attributes("-topmost", False))
    root.mainloop()

    input_dir = result.get("input_dir")
    out_dir = result.get("out_dir")
    if input_dir is None or out_dir is None:
        raise SystemExit("Execucao cancelada pelo usuario na selecao de diretorios.")
    return input_dir, out_dir


def _prompt_runtime_dirs_via_cli(initial_input_dir: Path, initial_out_dir: Path) -> Tuple[Path, Path]:
    print("[WARN] GUI indisponivel. Caindo para entrada manual no terminal.")
    raw_prompt = f"RAW_INPUT_DIR [{initial_input_dir}]: "
    out_prompt = f"OUT_DIR [{initial_out_dir}]: "

    raw_input = input(raw_prompt).strip()
    out_input = input(out_prompt).strip()

    input_dir = Path(raw_input or str(initial_input_dir)).expanduser().resolve()
    out_dir = Path(out_input or str(initial_out_dir)).expanduser().resolve()
    return input_dir, out_dir


def _prompt_open_config_gui_via_windows_dialog() -> Optional[bool]:
    ps_script = r"""
Add-Type -AssemblyName System.Windows.Forms
$result = [System.Windows.Forms.MessageBox]::Show(
    'Abrir a GUI de configuracao do pipeline29 antes de rodar?',
    'Pipeline 29',
    [System.Windows.Forms.MessageBoxButtons]::YesNo,
    [System.Windows.Forms.MessageBoxIcon]::Question
)
if ($result -eq [System.Windows.Forms.DialogResult]::Yes) { exit 10 }
if ($result -eq [System.Windows.Forms.DialogResult]::No) { exit 11 }
exit 2
"""
    completed = subprocess.run(
        ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps_script],
        capture_output=True,
        text=True,
    )
    if completed.returncode == 10:
        return True
    if completed.returncode == 11:
        return False
    if completed.returncode == 2:
        return None
    raise RuntimeError(
        "Falha ao abrir prompt de GUI do pipeline29. "
        f"stderr={completed.stderr.strip()!r} code={completed.returncode}"
    )


def _prompt_open_config_gui_via_tk_dialog() -> Optional[bool]:
    if tk is None or messagebox is None:
        return None
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    try:
        answer = messagebox.askyesno(
            "Pipeline 29",
            "Abrir a GUI de configuracao do pipeline29 antes de rodar?",
            parent=root,
        )
        return bool(answer)
    finally:
        root.destroy()


def _prompt_open_config_gui_via_cli() -> bool:
    raw = input("Abrir GUI de configuracao do pipeline29 antes de rodar? [y/N]: ").strip().lower()
    return raw in {"y", "yes", "s", "sim", "1", "true", "on"}


def _prompt_open_config_gui() -> bool:
    if os.name == "nt":
        try:
            answer = _prompt_open_config_gui_via_windows_dialog()
            if answer is not None:
                return answer
        except Exception as exc:
            print(f"[WARN] Prompt nativo da GUI de configuracao falhou: {exc}")

    try:
        answer = _prompt_open_config_gui_via_tk_dialog()
        if answer is not None:
            return answer
    except Exception as exc:
        print(f"[WARN] Prompt Tkinter da GUI de configuracao falhou: {exc}")

    return _prompt_open_config_gui_via_cli()


def _prompt_runtime_dirs(initial_input_dir: Path, initial_out_dir: Path) -> Tuple[Path, Path]:
    if os.name == "nt":
        try:
            return _prompt_runtime_dirs_via_windows_dialog(initial_input_dir, initial_out_dir)
        except SystemExit:
            raise
        except Exception as exc:
            print(f"[WARN] Seletor nativo do Windows falhou: {exc}")

    try:
        return _prompt_runtime_dirs_via_tk_dialog(initial_input_dir, initial_out_dir)
    except SystemExit:
        raise
    except Exception as exc:
        print(f"[WARN] Popup Tkinter falhou: {exc}")

    return _prompt_runtime_dirs_via_cli(initial_input_dir, initial_out_dir)


def _write_runtime_dirs_to_defaults_excel(xlsx_path: Path, input_dir: Path, out_dir: Path) -> None:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        print(f"[WARN] Nao consegui importar openpyxl para atualizar o Excel de Defaults: {exc}")
        return

    wb = load_workbook(xlsx_path)
    ws = None
    for sheet_name in wb.sheetnames:
        if str(sheet_name).strip().lower() == "defaults":
            ws = wb[sheet_name]
            break
    if ws is None:
        print(f"[WARN] Aba 'Defaults' nao encontrada em {xlsx_path}.")
        return

    header_map: dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        header = norm_key(cell.value)
        if header:
            header_map[header] = idx

    param_col = header_map.get("param", 1)
    value_col = header_map.get("value", 2)

    replacements = {
        norm_key("RAW_INPUT_DIR"): str(input_dir),
        norm_key("OUT_DIR"): str(out_dir),
    }
    updated = set()

    for row_idx in range(1, ws.max_row + 1):
        param_value = norm_key(ws.cell(row=row_idx, column=param_col).value)
        if param_value in replacements:
            ws.cell(row=row_idx, column=value_col).value = replacements[param_value]
            updated.add(param_value)

    next_row = ws.max_row + 1
    for param_key, value in replacements.items():
        if param_key in updated:
            continue
        ws.cell(row=next_row, column=param_col).value = "RAW_INPUT_DIR" if param_key == norm_key("RAW_INPUT_DIR") else "OUT_DIR"
        ws.cell(row=next_row, column=value_col).value = value
        next_row += 1

    wb.save(xlsx_path)


def _choose_runtime_dirs(defaults_cfg: Dict[str, str]) -> Tuple[Path, Path]:
    saved_cfg = _load_runtime_path_settings()

    raw_cfg = saved_cfg.get("raw_input_dir") or defaults_cfg.get(norm_key("RAW_INPUT_DIR"), "")
    out_cfg = saved_cfg.get("out_dir") or defaults_cfg.get(norm_key("OUT_DIR"), "")

    initial_input_dir = _resolve_runtime_dir(raw_cfg, DEFAULT_PROCESS_DIR)
    initial_out_dir = _resolve_runtime_dir(out_cfg, DEFAULT_OUT_DIR)

    use_defaults_env = norm_key(
        os.environ.get("PIPELINE29_USE_DEFAULT_RUNTIME_DIRS", "")
        or os.environ.get("PIPELINE28_USE_DEFAULT_RUNTIME_DIRS", "")
    )
    if use_defaults_env in {"1", "true", "yes", "on"}:
        print("[INFO] PIPELINE29_USE_DEFAULT_RUNTIME_DIRS ativo; usando RAW_INPUT_DIR/OUT_DIR sem popup.")
        defaults_cfg[norm_key("RAW_INPUT_DIR")] = str(initial_input_dir)
        defaults_cfg[norm_key("OUT_DIR")] = str(initial_out_dir)
        return initial_input_dir, initial_out_dir

    print("[INFO] Abrindo popup para selecionar RAW_INPUT_DIR e OUT_DIR...")
    input_dir, out_dir = _prompt_runtime_dirs(initial_input_dir, initial_out_dir)
    _save_runtime_path_settings(input_dir, out_dir)
    defaults_cfg[norm_key("RAW_INPUT_DIR")] = str(input_dir)
    defaults_cfg[norm_key("OUT_DIR")] = str(out_dir)
    print(f"[INFO] RAW_INPUT_DIR (GUI): {input_dir}")
    print(f"[INFO] OUT_DIR (GUI): {out_dir}")
    print(f"[INFO] Ultima selecao salva em: {RUNTIME_SETTINGS_PATH}")
    return input_dir, out_dir


def _sync_runtime_dirs_to_config_source(
    config_bundle: Optional[Pipeline29ConfigBundle],
    input_dir: Path,
    out_dir: Path,
) -> None:
    if config_bundle is None:
        return
    if config_bundle.source_kind == "excel" and config_bundle.source_path is not None:
        _write_runtime_dirs_to_defaults_excel(config_bundle.source_path, input_dir, out_dir)
        print(f"[INFO] Aba Defaults sincronizada apenas para RAW_INPUT_DIR/OUT_DIR em: {config_bundle.source_path}")
        return
    if config_bundle.source_kind == "text" and config_bundle.source_path is not None:
        print(
            "[INFO] Runtime dirs do pipeline29 ficaram salvos localmente em "
            f"{RUNTIME_SETTINGS_PATH}; a config textual em {config_bundle.source_path} nao foi alterada."
        )


def apply_runtime_path_overrides(
    defaults_cfg: Dict[str, str],
    config_bundle: Optional[Pipeline29ConfigBundle] = None,
) -> None:
    global RAW_DIR, PROCESS_DIR, OUT_DIR, PLOTS_DIR

    input_dir, out_dir = _choose_runtime_dirs(defaults_cfg)
    _sync_runtime_dirs_to_config_source(config_bundle, input_dir, out_dir)

    if not input_dir.exists():
        raise FileNotFoundError(f"Nao encontrei o diretorio selecionado para RAW_INPUT_DIR: {input_dir}")
    if not input_dir.is_dir():
        raise NotADirectoryError(f"RAW_INPUT_DIR selecionado nao aponta para um diretorio: {input_dir}")

    if not _prepare_output_dir(out_dir):
        raise FileNotFoundError(
            f"Nao consegui preparar o diretorio de saida selecionado em OUT_DIR: {out_dir}"
        )

    RAW_DIR = input_dir.parent
    PROCESS_DIR = input_dir
    OUT_DIR = out_dir
    PLOTS_DIR = OUT_DIR / "plots"


def _plot_point_fuel_labels(df: pd.DataFrame) -> pd.Series:
    idx = df.index
    labels = df.get("Fuel_Label", pd.Series(pd.NA, index=idx, dtype="object")).copy()
    labels = labels.where(labels.map(lambda x: not _is_blank_cell(x)), _fuel_blend_labels(df))

    h2o = pd.to_numeric(df.get("H2O_pct", pd.Series(pd.NA, index=idx)), errors="coerce")
    fallback = pd.Series(pd.NA, index=idx, dtype="object")
    for level, label in FUEL_LABEL_BY_H2O_LEVEL.items():
        fallback = fallback.mask(h2o.sub(float(level)).abs() <= 0.6, label)

    labels = labels.where(labels.map(lambda x: not _is_blank_cell(x)), fallback)
    labels = labels.where(labels.map(lambda x: not _is_blank_cell(x)), pd.NA)
    return labels


PLOT_POINT_FILTER_STATE_PATH = RUNTIME_SETTINGS_DIR / "plot_point_filter_last.json"


def _normalize_plot_point_key(fuel_label: object, load_kw: object) -> Optional[Tuple[str, float]]:
    label = str(fuel_label or "").strip()
    if not label:
        return None
    try:
        load = round(float(load_kw), 6)
    except Exception:
        return None
    if not np.isfinite(load):
        return None
    return (label, load)


def _plot_point_keys_to_jsonable(points: set[Tuple[str, float]]) -> List[Dict[str, object]]:
    out: List[Dict[str, object]] = []
    for fuel_label, load_kw in sorted(points, key=lambda item: (_canon_name(item[0]), item[1])):
        out.append({"fuel_label": fuel_label, "load_kw": round(float(load_kw), 6)})
    return out


def _load_plot_point_filter_state() -> Optional[Dict[str, object]]:
    try:
        if not PLOT_POINT_FILTER_STATE_PATH.exists():
            return None
        payload = json.loads(PLOT_POINT_FILTER_STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return None

    selected_points: set[Tuple[str, float]] = set()
    for row in payload.get("selected_points", []) or []:
        if not isinstance(row, dict):
            continue
        key = _normalize_plot_point_key(row.get("fuel_label", ""), row.get("load_kw", None))
        if key is not None:
            selected_points.add(key)

    available_points: set[Tuple[str, float]] = set()
    for row in payload.get("available_points", []) or []:
        if not isinstance(row, dict):
            continue
        key = _normalize_plot_point_key(row.get("fuel_label", ""), row.get("load_kw", None))
        if key is not None:
            available_points.add(key)

    return {
        "selected_points": selected_points,
        "available_points": available_points,
        "saved_at": str(payload.get("saved_at", "")).strip(),
    }


def _save_plot_point_filter_state(selected_points: set[Tuple[str, float]], available_points: set[Tuple[str, float]]) -> None:
    try:
        RUNTIME_SETTINGS_DIR.mkdir(parents=True, exist_ok=True)
        payload = {
            "saved_at": datetime.now().isoformat(timespec="seconds"),
            "selected_points": _plot_point_keys_to_jsonable(selected_points),
            "available_points": _plot_point_keys_to_jsonable(available_points),
        }
        PLOT_POINT_FILTER_STATE_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception as exc:
        print(f"[WARN] Nao consegui salvar a ultima selecao do filtro de plots: {exc}")


def _resolve_plot_point_initial_selection(
    available_points: set[Tuple[str, float]],
) -> Tuple[Dict[Tuple[str, float], bool], str]:
    defaults = {key: True for key in available_points}
    state = _load_plot_point_filter_state()
    if state is None:
        return defaults, "Sem ultima selecao salva. Todos os pontos vieram marcados."

    saved_available = set(state.get("available_points", set()) or set())
    saved_selected = set(state.get("selected_points", set()) or set())
    matched_known_points = 0
    for key in sorted(available_points):
        if key in saved_available:
            defaults[key] = key in saved_selected
            matched_known_points += 1
    if matched_known_points == 0:
        return defaults, "Ultima selecao salva nao combinou com este conjunto. Todos os pontos vieram marcados."

    new_points = available_points - saved_available
    selected_count = sum(1 for selected in defaults.values() if selected)
    message = f"Ultima selecao carregada automaticamente: {selected_count} / {len(available_points)} ponto(s) marcados."
    if new_points:
        message += f" {len(new_points)} ponto(s) novo(s) ficaram selecionados por padrao."
    return defaults, message


def _normalized_plot_point_loads(df: pd.DataFrame) -> pd.Series:
    loads = pd.to_numeric(df.get("Load_kW", pd.Series(pd.NA, index=df.index)), errors="coerce")
    return loads.round(6)


def _fuel_label_from_composition_values(
    dies_pct: object,
    biod_pct: object,
    etoh_pct: object,
    h2o_pct: object,
    tol: float = 0.6,
) -> str:
    return _fuel_label_from_components(dies_pct, biod_pct, etoh_pct, h2o_pct, tol=tol)


def _preferred_fuel_label_order(labels: List[str]) -> List[str]:
    preferred = ["D85B15", "E94H6", "E75H25", "E65H35"]
    uniq: List[str] = []
    seen: set[str] = set()
    for value in labels:
        label = str(value).strip()
        if not label or label in seen:
            continue
        uniq.append(label)
        seen.add(label)

    ordered = [label for label in preferred if label in seen]
    extras = sorted([label for label in uniq if label not in ordered], key=_canon_name)
    return ordered + extras


def _build_plot_point_catalog(df: pd.DataFrame) -> Tuple[List[str], List[float], Dict[Tuple[str, float], int]]:
    if df is None or df.empty:
        return [], [], {}

    labels = _plot_point_fuel_labels(df)
    loads = _normalized_plot_point_loads(df)
    tmp = pd.DataFrame({"Fuel_Label": labels, "Load_kW": loads}, index=df.index).dropna(subset=["Fuel_Label", "Load_kW"])
    if tmp.empty:
        return [], [], {}

    counts_df = (
        tmp.groupby(["Fuel_Label", "Load_kW"], dropna=False, sort=True)
        .size()
        .reset_index(name="N_points")
    )
    counts: Dict[Tuple[str, float], int] = {}
    for _, row in counts_df.iterrows():
        key = (str(row["Fuel_Label"]).strip(), float(row["Load_kW"]))
        counts[key] = int(row["N_points"])

    fuel_labels = _preferred_fuel_label_order(counts_df["Fuel_Label"].astype(str).tolist())
    load_values = sorted(float(v) for v in counts_df["Load_kW"].dropna().unique().tolist())
    return fuel_labels, load_values, counts


def _build_plot_point_catalog_from_metas(metas: List["FileMeta"]) -> Tuple[List[str], List[float], Dict[Tuple[str, float], int]]:
    rows: List[Tuple[str, float]] = []
    for meta in metas:
        label = _fuel_label_from_composition_values(meta.dies_pct, meta.biod_pct, meta.etoh_pct, meta.h2o_pct)
        if not label:
            continue
        if meta.load_kw is None or not np.isfinite(meta.load_kw):
            continue
        rows.append((label, round(float(meta.load_kw), 6)))

    if not rows:
        return [], [], {}

    counts: Dict[Tuple[str, float], int] = {}
    for key in rows:
        counts[key] = counts.get(key, 0) + 1

    fuel_labels = _preferred_fuel_label_order([fuel_label for fuel_label, _ in counts.keys()])
    load_values = sorted({float(load_kw) for _, load_kw in counts.keys()})
    return fuel_labels, load_values, counts


def _ensure_qt_application() -> Tuple[object, bool]:
    if QApplication is None:
        raise RuntimeError("PySide6 nao esta disponivel.")

    app = QApplication.instance()
    owns_app = False
    if app is None:
        app = QApplication(["pipeline29"])
        owns_app = True
        if QStyleFactory is not None:
            try:
                if "Fusion" in QStyleFactory.keys():
                    app.setStyle("Fusion")
            except Exception:
                pass
    return app, owns_app


def _prompt_plot_point_filter_catalog_via_qt(
    fuel_labels: List[str],
    load_values: List[float],
    counts: Dict[Tuple[str, float], int],
) -> Optional[set[Tuple[str, float]]]:
    if QApplication is None or QDialog is None or QTableWidget is None or Qt is None:
        raise RuntimeError("PySide6 nao esta disponivel.")
    if not fuel_labels or not load_values or not counts:
        return None

    app, owns_app = _ensure_qt_application()
    _ = app
    dialog = QDialog()
    dialog.setWindowTitle("Pipeline 29 - filtro de pontos para plots")
    dialog.setModal(True)
    dialog.resize(1120, 760)

    main_layout = QVBoxLayout(dialog)
    title = QLabel("Selecione os pontos que entram nos graficos. Os calculos e o lv_kpis_clean.xlsx continuam completos.")
    title.setWordWrap(True)
    title.setStyleSheet("font-size: 15px; font-weight: 600;")
    subtitle = QLabel("Colunas = combustiveis | Linhas = cargas nominais | Tudo vem selecionado por padrao.")
    subtitle.setStyleSheet("color: #5f6b76;")
    main_layout.addWidget(title)
    main_layout.addWidget(subtitle)

    available_points = {key for key, count in counts.items() if count > 0}
    initial_selection, initial_message = _resolve_plot_point_initial_selection(available_points)

    toolbar = QHBoxLayout()
    btn_select_all = QPushButton("Selecionar tudo")
    btn_clear_all = QPushButton("Limpar tudo")
    btn_load_last = QPushButton("Carregar ultima")
    btn_save_last = QPushButton("Salvar atual")
    info_label = QLabel("Numero pequeno = quantidade de linhas/iteracoes do ponto.")
    info_label.setStyleSheet("color: #5f6b76;")
    status_label = QLabel()
    status_label.setStyleSheet("font-weight: 600;")
    toolbar.addWidget(btn_select_all)
    toolbar.addWidget(btn_clear_all)
    toolbar.addWidget(btn_load_last)
    toolbar.addWidget(btn_save_last)
    toolbar.addSpacing(8)
    toolbar.addWidget(info_label)
    toolbar.addStretch(1)
    toolbar.addWidget(status_label)
    main_layout.addLayout(toolbar)

    selection_info_label = QLabel(initial_message)
    selection_info_label.setWordWrap(True)
    selection_info_label.setStyleSheet("color: #5f6b76;")
    main_layout.addWidget(selection_info_label)

    table = QTableWidget(len(load_values), len(fuel_labels))
    table.setHorizontalHeaderLabels(fuel_labels)
    table.setVerticalHeaderLabels([_format_load_kw_label(v) for v in load_values])
    table.setShowGrid(True)
    table.setAlternatingRowColors(True)
    table.setSelectionMode(QTableWidget.NoSelection)
    table.setEditTriggers(QTableWidget.NoEditTriggers)
    table.setFocusPolicy(Qt.NoFocus)
    table.setStyleSheet(
        """
        QTableWidget {
            gridline-color: #d7dce1;
            background: #ffffff;
            alternate-background-color: #fbfcfd;
            border: 1px solid #d7dce1;
        }
        QHeaderView::section {
            background: #f3f5f7;
            color: #1f2933;
            padding: 6px;
            border: 1px solid #d7dce1;
            font-weight: 600;
        }
        """
    )
    table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
    table.verticalHeader().setSectionResizeMode(QHeaderView.Fixed)
    table.verticalHeader().setDefaultSectionSize(38)
    table.horizontalHeader().setMinimumSectionSize(120)
    main_layout.addWidget(table, stretch=1)

    checkbox_map: Dict[Tuple[str, float], object] = {}

    def refresh_status() -> None:
        selected = sum(1 for cb in checkbox_map.values() if bool(cb.isChecked()))
        status_label.setText(f"Pontos selecionados: {selected} / {len(checkbox_map)}")

    def set_all(value: bool) -> None:
        for checkbox in checkbox_map.values():
            checkbox.setChecked(value)
        refresh_status()

    def selected_points_now() -> set[Tuple[str, float]]:
        return {key for key, checkbox in checkbox_map.items() if bool(checkbox.isChecked())}

    def load_last_selection() -> None:
        defaults, message = _resolve_plot_point_initial_selection(available_points)
        for key, checkbox in checkbox_map.items():
            checkbox.setChecked(bool(defaults.get(key, True)))
        refresh_status()
        selection_info_label.setText(message)

    def save_current_selection() -> None:
        selected = selected_points_now()
        _save_plot_point_filter_state(selected, available_points)
        selection_info_label.setText(
            f"Selecao atual salva como ultima: {len(selected)} / {len(available_points)} ponto(s) marcados."
        )

    for row_idx, load_kw in enumerate(load_values):
        for col_idx, fuel_label in enumerate(fuel_labels):
            key = (fuel_label, float(load_kw))
            count = counts.get(key, 0)
            if count <= 0:
                item = QTableWidgetItem("—")
                item.setTextAlignment(int(Qt.AlignCenter))
                item.setFlags(Qt.ItemIsEnabled)
                table.setItem(row_idx, col_idx, item)
                continue

            checkbox = QCheckBox()
            checkbox.setChecked(bool(initial_selection.get(key, True)))
            checkbox.setStyleSheet("QCheckBox::indicator { width: 14px; height: 14px; }")
            checkbox.stateChanged.connect(lambda _state, _refresh=refresh_status: _refresh())
            count_label = QLabel("" if count == 1 else f"{count}x")
            count_label.setAlignment(Qt.AlignCenter)
            count_label.setStyleSheet("color: #5f6b76; font-size: 10px;")

            cell_widget = QWidget()
            cell_layout = QVBoxLayout(cell_widget)
            cell_layout.setContentsMargins(0, 0, 0, 0)
            cell_layout.setSpacing(0)
            cell_layout.addWidget(checkbox, alignment=Qt.AlignCenter)
            cell_layout.addWidget(count_label, alignment=Qt.AlignCenter)

            table.setCellWidget(row_idx, col_idx, cell_widget)
            checkbox_map[key] = checkbox

    refresh_status()

    buttons_layout = QHBoxLayout()
    buttons_layout.addStretch(1)
    btn_cancel = QPushButton("Cancelar")
    btn_run = QPushButton("Gerar graficos")
    btn_run.setDefault(True)
    buttons_layout.addWidget(btn_cancel)
    buttons_layout.addWidget(btn_run)
    main_layout.addLayout(buttons_layout)

    btn_select_all.clicked.connect(lambda: set_all(True))
    btn_clear_all.clicked.connect(lambda: set_all(False))
    btn_load_last.clicked.connect(load_last_selection)
    btn_save_last.clicked.connect(save_current_selection)

    selected_result: dict[str, object] = {"selected": None}

    def accept_selection() -> None:
        selected = selected_points_now()
        if not selected:
            QMessageBox.critical(dialog, "Pipeline 29", "Selecione pelo menos um ponto para gerar os graficos.")
            return
        _save_plot_point_filter_state(selected, available_points)
        selected_result["selected"] = selected
        dialog.accept()

    btn_run.clicked.connect(accept_selection)
    btn_cancel.clicked.connect(dialog.reject)

    if dialog.exec() != QDialog.Accepted:
        raise SystemExit("Execucao cancelada pelo usuario na selecao de pontos para plot.")

    selected = selected_result.get("selected")
    if selected is None:
        raise SystemExit("Execucao cancelada pelo usuario na selecao de pontos para plot.")
    return set(selected)


def _prompt_plot_point_filter_catalog_via_tk(
    fuel_labels: List[str],
    load_values: List[float],
    counts: Dict[Tuple[str, float], int],
) -> Optional[set[Tuple[str, float]]]:
    if tk is None or ttk is None or messagebox is None:
        raise RuntimeError("Tkinter nao esta disponivel.")
    if not fuel_labels or not load_values or not counts:
        return None

    result: dict[str, object] = {"selected": None}
    available_points = {key for key, count in counts.items() if count > 0}
    initial_selection, initial_message = _resolve_plot_point_initial_selection(available_points)
    root = tk.Tk()
    root.title("Pipeline 29 - filtro de pontos para plots")
    root.withdraw()
    root.resizable(True, True)
    root.attributes("-topmost", True)

    ttk.Label(
        root,
        text="Selecione os pontos que entram nos graficos. Os calculos e o lv_kpis_clean.xlsx continuam completos.",
        wraplength=1100,
        justify="left",
    ).grid(row=0, column=0, columnspan=3, sticky="w", padx=12, pady=(12, 4))
    ttk.Label(
        root,
        text="Colunas = combustiveis | Linhas = cargas nominais. Tudo vem selecionado por padrao.",
    ).grid(row=1, column=0, columnspan=3, sticky="w", padx=12, pady=(0, 8))
    selection_info_var = tk.StringVar(master=root, value=initial_message)
    ttk.Label(
        root,
        textvariable=selection_info_var,
        wraplength=1100,
        justify="left",
    ).grid(row=2, column=0, columnspan=3, sticky="w", padx=12, pady=(0, 8))

    toolbar = ttk.Frame(root)
    toolbar.grid(row=3, column=0, columnspan=3, sticky="we", padx=12, pady=(0, 8))
    toolbar.columnconfigure(5, weight=1)

    body = ttk.Frame(root)
    body.grid(row=4, column=0, columnspan=3, sticky="nsew", padx=12, pady=0)
    root.columnconfigure(0, weight=1)
    root.rowconfigure(4, weight=1)
    body.columnconfigure(0, weight=1)
    body.rowconfigure(0, weight=1)

    canvas = tk.Canvas(body, highlightthickness=0)
    vbar = ttk.Scrollbar(body, orient="vertical", command=canvas.yview)
    hbar = ttk.Scrollbar(body, orient="horizontal", command=canvas.xview)
    canvas.configure(yscrollcommand=vbar.set, xscrollcommand=hbar.set)
    canvas.grid(row=0, column=0, sticky="nsew")
    vbar.grid(row=0, column=1, sticky="ns")
    hbar.grid(row=1, column=0, sticky="ew")

    grid_frame = ttk.Frame(canvas)
    canvas_window = canvas.create_window((0, 0), window=grid_frame, anchor="nw")

    def _sync_canvas(_event: object = None) -> None:
        canvas.configure(scrollregion=canvas.bbox("all"))
        canvas.itemconfigure(canvas_window, width=max(canvas.winfo_width(), grid_frame.winfo_reqwidth()))

    grid_frame.bind("<Configure>", _sync_canvas)
    canvas.bind("<Configure>", _sync_canvas)

    header_bg = "#f4f6f8"
    cell_border = "#d7dce1"

    def make_cell(row: int, column: int, *, bg: str = "white") -> tk.Frame:
        cell = tk.Frame(
            grid_frame,
            bg=bg,
            highlightbackground=cell_border,
            highlightthickness=1,
            bd=0,
            padx=3,
            pady=0,
        )
        cell.grid(row=row, column=column, sticky="nsew")
        return cell

    header_cell = make_cell(0, 0, bg=header_bg)
    ttk.Label(header_cell, text="Carga (kW)", anchor="center").pack(fill="both", expand=True)
    cell_vars: Dict[Tuple[str, float], tk.BooleanVar] = {}

    for col_idx, fuel_label in enumerate(fuel_labels, start=1):
        header_cell = make_cell(0, col_idx, bg=header_bg)
        ttk.Label(header_cell, text=fuel_label, anchor="center", justify="center").pack(fill="both", expand=True)
        grid_frame.columnconfigure(col_idx, weight=1)

    for row_idx, load_kw in enumerate(load_values, start=1):
        load_cell = make_cell(row_idx, 0, bg=header_bg)
        ttk.Label(load_cell, text=_format_load_kw_label(load_kw), anchor="center").pack(fill="both", expand=True)
        for col_idx, fuel_label in enumerate(fuel_labels, start=1):
            key = (fuel_label, float(load_kw))
            count = counts.get(key, 0)
            if count <= 0:
                empty_cell = make_cell(row_idx, col_idx)
                ttk.Label(empty_cell, text="-", anchor="center").pack(fill="both", expand=True)
                continue

            var = tk.BooleanVar(value=True)
            var.set(bool(initial_selection.get(key, True)))
            cell_vars[key] = var
            point_cell = make_cell(row_idx, col_idx)
            inner = ttk.Frame(point_cell)
            inner.pack(fill="both", expand=True)
            ttk.Checkbutton(inner, variable=var).pack(anchor="center", pady=0)
            ttk.Label(inner, text="" if count == 1 else f"{count}x", anchor="center", justify="center").pack(anchor="center")

    status_var = tk.StringVar()

    def refresh_status() -> None:
        selected = sum(1 for var in cell_vars.values() if bool(var.get()))
        status_var.set(f"Pontos selecionados para plot: {selected} / {len(cell_vars)}")

    for var in cell_vars.values():
        var.trace_add("write", lambda *_args: refresh_status())

    def set_all(value: bool) -> None:
        for var in cell_vars.values():
            var.set(value)

    def load_last_selection() -> None:
        defaults, message = _resolve_plot_point_initial_selection(available_points)
        for key, var in cell_vars.items():
            var.set(bool(defaults.get(key, True)))
        selection_info_var.set(message)

    def save_current_selection() -> None:
        selected = {key for key, var in cell_vars.items() if bool(var.get())}
        _save_plot_point_filter_state(selected, available_points)
        selection_info_var.set(
            f"Selecao atual salva como ultima: {len(selected)} / {len(available_points)} ponto(s) marcados."
        )

    def confirm() -> None:
        selected = {key for key, var in cell_vars.items() if bool(var.get())}
        if not selected:
            messagebox.showerror("Pipeline 29", "Selecione pelo menos um ponto para gerar os graficos.", parent=root)
            return
        _save_plot_point_filter_state(selected, available_points)
        result["selected"] = selected
        root.destroy()

    def cancel() -> None:
        root.destroy()

    ttk.Button(toolbar, text="Selecionar tudo", command=lambda: set_all(True)).grid(row=0, column=0, padx=(0, 8), pady=0)
    ttk.Button(toolbar, text="Limpar tudo", command=lambda: set_all(False)).grid(row=0, column=1, padx=(0, 8), pady=0)
    ttk.Button(toolbar, text="Carregar ultima", command=load_last_selection).grid(row=0, column=2, padx=(0, 8), pady=0)
    ttk.Button(toolbar, text="Salvar atual", command=save_current_selection).grid(row=0, column=3, padx=(0, 8), pady=0)
    ttk.Label(toolbar, text="Numero no checkbox = quantidade de linhas/iteracoes para o ponto.").grid(
        row=0,
        column=4,
        sticky="w",
    )
    ttk.Label(toolbar, textvariable=status_var).grid(row=0, column=5, sticky="e")
    refresh_status()

    buttons = ttk.Frame(root)
    buttons.grid(row=5, column=0, columnspan=3, sticky="e", padx=12, pady=(8, 12))
    ttk.Button(buttons, text="Cancelar", command=cancel).pack(side="right")
    ttk.Button(buttons, text="Gerar graficos", command=confirm).pack(side="right", padx=(0, 8))

    root.protocol("WM_DELETE_WINDOW", cancel)
    root.bind("<Return>", lambda _event: confirm())
    root.bind("<Escape>", lambda _event: cancel())

    root.update_idletasks()
    width = min(max(root.winfo_reqwidth(), 1000), max(root.winfo_screenwidth() - 80, 1000))
    height = min(max(root.winfo_reqheight(), 600), max(root.winfo_screenheight() - 120, 600))
    pos_x = max((root.winfo_screenwidth() - width) // 2, 0)
    pos_y = max((root.winfo_screenheight() - height) // 4, 0)
    root.geometry(f"{width}x{height}+{pos_x}+{pos_y}")
    root.deiconify()
    root.lift()
    try:
        root.focus_force()
    except Exception:
        pass
    root.after(400, lambda: root.attributes("-topmost", False))
    root.mainloop()

    selected = result.get("selected")
    if selected is None:
        raise SystemExit("Execucao cancelada pelo usuario na selecao de pontos para plot.")
    return set(selected)


def prompt_plot_point_filter(df: pd.DataFrame) -> Optional[set[Tuple[str, float]]]:
    fuel_labels, load_values, counts = _build_plot_point_catalog(df)
    if not fuel_labels or not load_values or not counts:
        print("[WARN] Nao encontrei pontos com Fuel_Label e Load_kW para abrir o filtro de plots. Vou usar todos.")
        return None

    if QApplication is not None:
        try:
            return _prompt_plot_point_filter_catalog_via_qt(fuel_labels, load_values, counts)
        except SystemExit:
            raise
        except Exception as exc:
            print(f"[WARN] GUI PySide6 de filtro de pontos falhou: {exc}. Tentando fallback...")

    if os.name == "nt" or (tk is not None and ttk is not None):
        try:
            return _prompt_plot_point_filter_catalog_via_tk(fuel_labels, load_values, counts)
        except SystemExit:
            raise
        except Exception as exc:
            print(f"[WARN] GUI de filtro de pontos falhou: {exc}. Vou usar todos os pontos.")
            return None

    print("[WARN] GUI de filtro de pontos indisponivel neste ambiente. Vou usar todos os pontos.")
    return None


def prompt_plot_point_filter_from_metas(metas: List["FileMeta"]) -> Optional[set[Tuple[str, float]]]:
    valid_meta_count = 0
    for meta in metas:
        label = _fuel_label_from_composition_values(meta.dies_pct, meta.biod_pct, meta.etoh_pct, meta.h2o_pct)
        if label and meta.load_kw is not None and np.isfinite(meta.load_kw):
            valid_meta_count += 1
    if valid_meta_count < len(metas):
        print(
            "[INFO] Alguns pontos dependem de inferencia posterior de carga/composicao; "
            "vou abrir o filtro de plots depois do processamento completo."
        )
        return None

    fuel_labels, load_values, counts = _build_plot_point_catalog_from_metas(metas)
    if not fuel_labels or not load_values or not counts:
        return None

    if QApplication is not None:
        try:
            return _prompt_plot_point_filter_catalog_via_qt(fuel_labels, load_values, counts)
        except SystemExit:
            raise
        except Exception as exc:
            print(f"[WARN] GUI PySide6 de filtro de pontos falhou: {exc}. Tentando fallback...")

    if os.name == "nt" or (tk is not None and ttk is not None):
        try:
            return _prompt_plot_point_filter_catalog_via_tk(fuel_labels, load_values, counts)
        except SystemExit:
            raise
        except Exception as exc:
            print(f"[WARN] GUI de filtro de pontos falhou: {exc}. Vou usar todos os pontos.")
            return None

    return None


def _apply_plot_point_filter(df: pd.DataFrame, selected_points: Optional[set[Tuple[str, float]]]) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame()
    if df.empty or selected_points is None:
        return df.copy()

    fuel_labels = _plot_point_fuel_labels(df)
    loads = _normalized_plot_point_loads(df)
    mask = pd.Series(False, index=df.index, dtype="bool")

    for fuel_label, load_kw in selected_points:
        if not fuel_label or not np.isfinite(load_kw):
            continue
        mask = mask | (fuel_labels.eq(fuel_label) & loads.eq(round(float(load_kw), 6)))

    kept = int(mask.sum())
    print(f"[INFO] Filtro de plots: {kept} linha(s) mantida(s) para os graficos.")
    return df.loc[mask].copy()


def load_config_excel(xlsx_path: Optional[Path] = None) -> Tuple[dict, pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict[str, float], Dict[str, str]]:
    p = _choose_config_path() if xlsx_path is None else xlsx_path

    m = _read_excel(p, sheet_name="Mappings")
    m.columns = _normalize_cols(list(m.columns))

    mappings: dict = {}
    for _, row in m.iterrows():
        k = norm_key(row.get("key", ""))
        col_mean_raw = str(row.get("col_mean", "")).replace("\ufeff", "").strip()
        if "logical variable identifier" in k or "exact dataframe column name" in col_mean_raw.lower():
            continue
        if not k:
            continue
        mappings[k] = {
            "mean": col_mean_raw,
            "sd": str(row.get("col_sd", "")).replace("\ufeff", "").strip(),
        }

    required = {"power_kw", "fuel_kgh", "lhv_kj_kg"}
    missing = required - set(mappings.keys())
    if missing:
        raise KeyError(f"Faltam keys em Mappings: {missing}. Keys lidas: {sorted(mappings.keys())}")

    ins = _try_read_sheet(p, "Instruments")
    if ins is None:
        ins = pd.DataFrame()

    ins.columns = _normalize_cols(list(ins.columns))
    ins_cols_low = {c.lower().strip(): c for c in ins.columns}

    is_rev2 = "acc_abs" in ins_cols_low and "acc_pct" in ins_cols_low
    if not is_rev2 and not ins.empty:
        def get_col(name: str) -> Optional[str]:
            return ins_cols_low.get(name.lower().strip())

        key_c = get_col("key")
        dist_c = get_col("dist")
        pct_c = get_col("percent")
        dig_c = get_col("digits")
        lsd_c = get_col("lsd")
        abs_c = get_col("abs")
        res_c = get_col("resolution")
        model_c = get_col("model")

        rows = []
        for _, r in ins.iterrows():
            k = str(r.get(key_c, "")).strip()
            if not k:
                continue
            dist = str(r.get(dist_c, "rect")).strip() if dist_c else "rect"
            pct = _to_float(r.get(pct_c, 0.0), 0.0) if pct_c else 0.0
            dig = _to_float(r.get(dig_c, 0.0), 0.0) if dig_c else 0.0
            lsd = _to_float(r.get(lsd_c, 0.0), 0.0) if lsd_c else 0.0
            absv = _to_float(r.get(abs_c, 0.0), 0.0) if abs_c else 0.0
            res = _to_float(r.get(res_c, 0.0), 0.0) if res_c else 0.0
            model = str(r.get(model_c, "")).strip() if model_c else ""

            rows.append(
                {
                    "key": k,
                    "component": f"{k}_spec",
                    "dist": dist if dist else "rect",
                    "range_min": pd.NA,
                    "range_max": pd.NA,
                    "acc_abs": absv,
                    "acc_pct": pct,
                    "digits": dig,
                    "lsd": lsd,
                    "resolution": res,
                    "source": "",
                    "notes": f"migrated_from_model={model}",
                }
            )
        instruments_df = pd.DataFrame(rows)
    else:
        instruments_df = ins.copy()
        for c in [
            "key",
            "component",
            "dist",
            "range_min",
            "range_max",
            "acc_abs",
            "acc_pct",
            "digits",
            "lsd",
            "resolution",
            "source",
            "notes",
            "setting_param",
            "setting_value",
        ]:
            if c not in instruments_df.columns:
                instruments_df[c] = pd.NA

    instruments_df["key_norm"] = instruments_df["key"].map(norm_key)

    rep = _try_read_sheet(p, "Reporting_Rounding")
    if rep is None:
        rep = _try_read_sheet(p, "UPD_Rounding")
    if rep is None:
        rep = pd.DataFrame(columns=["key", "report_resolution", "report_digits", "rule", "notes"])

    rep.columns = _normalize_cols(list(rep.columns))
    if "key" not in rep.columns:
        rep["key"] = pd.NA
    if "report_resolution" not in rep.columns:
        rep["report_resolution"] = pd.NA
    if "rule" not in rep.columns:
        rep["rule"] = "round_half_up"
    rep["key_norm"] = rep["key"].map(norm_key)

    plots = _try_read_sheet(p, "Plots")
    if plots is None:
        plots = pd.DataFrame()

    if not plots.empty:
        plots.columns = _normalize_cols(list(plots.columns))
    else:
        plots = pd.DataFrame(
            columns=[
                "enabled",
                "with_uncertainty",
                "without_uncertainty",
                "plot_type",
                "filename",
                "title",
                "x_col",
                "y_col",
                "yerr_col",
                "show_uncertainty",
                "x_label",
                "y_label",
                "x_min",
                "x_max",
                "x_step",
                "y_min",
                "y_max",
                "y_step",
                "y_tol_plus",
                "y_tol_minus",
                "filter_h2o_list",
                "label_variant",
                "notes",
            ]
        )
    for column in ("with_uncertainty", "without_uncertainty", "show_uncertainty"):
        if column not in plots.columns:
            plots[column] = pd.NA
    if "show_uncertainty" not in plots.columns:
        plots["show_uncertainty"] = pd.NA
    if "yerr_col" in plots.columns:
        for idx, value in plots["show_uncertainty"].items():
            if not _is_blank_cell(value):
                continue
            yerr_value = _to_str_or_empty(plots.at[idx, "yerr_col"])
            plots.at[idx, "show_uncertainty"] = "off" if _yerr_disabled_token(yerr_value) else "auto"
    for idx, row in plots.iterrows():
        with_raw = _to_str_or_empty(row.get("with_uncertainty", ""))
        without_raw = _to_str_or_empty(row.get("without_uncertainty", ""))
        mode = _to_str_or_empty(row.get("show_uncertainty", "auto")).lower()
        with_flag = with_raw in {"1", "true", "yes", "on", "y", "checked"}
        without_flag = without_raw in {"1", "true", "yes", "on", "y", "checked"}
        with_defined = with_raw in {"1", "0", "true", "false", "yes", "no", "on", "off", "y", "n", "checked", "unchecked"}
        without_defined = without_raw in {"1", "0", "true", "false", "yes", "no", "on", "off", "y", "n", "checked", "unchecked"}
        if not with_defined and not without_defined:
            if mode == "off":
                with_flag, without_flag = False, True
            elif mode in {"both", "all", "dual", "on_off"}:
                with_flag, without_flag = True, True
            else:
                with_flag, without_flag = True, False
        elif not with_flag and not without_flag:
            with_flag = True
        plots.at[idx, "with_uncertainty"] = "1" if with_flag else "0"
        plots.at[idx, "without_uncertainty"] = "1" if without_flag else "0"
        if with_flag and without_flag:
            plots.at[idx, "show_uncertainty"] = "both"
        elif with_flag:
            plots.at[idx, "show_uncertainty"] = "on"
        else:
            plots.at[idx, "show_uncertainty"] = "off"

    data_quality_cfg = _load_data_quality_config(p)
    defaults_cfg = _load_defaults_config(p)

    return mappings, instruments_df, rep, plots, data_quality_cfg, defaults_cfg


def load_lhv_lookup() -> pd.DataFrame:
    p = CFG_DIR / "lhv.csv"
    if not p.exists():
        raise FileNotFoundError(f"NÃ£o encontrei {p}.")

    df = pd.read_csv(p, sep=None, engine="python", encoding="utf-8-sig")
    df.columns = _normalize_cols(list(df.columns))

    colmap: Dict[str, str] = {}
    for c in df.columns:
        cl = c.lower().strip()
        if cl in {"dies_pct", "dies", "diesel_pct", "diesel"}:
            colmap[c] = "DIES_pct"
        elif cl in {"biod_pct", "biod", "biodiesel_pct", "biodiesel"}:
            colmap[c] = "BIOD_pct"
        elif cl in {"etoh_pct", "etoh", "e_pct", "e"}:
            colmap[c] = "EtOH_pct"
        elif cl in {"h2o_pct", "h2o", "h20_pct", "h20", "h_pct", "h"}:
            colmap[c] = "H2O_pct"
        elif cl in {"lhv_kj_kg", "lhv", "pci_kj_kg", "pci"}:
            colmap[c] = "LHV_kJ_kg"
    df = df.rename(columns=colmap)

    if "LHV_kJ_kg" not in df.columns:
        raise KeyError(f"lhv.csv precisa da coluna LHV_kJ_kg. Colunas atuais: {list(df.columns)}")

    for c in COMPOSITION_COLS:
        if c not in df.columns:
            df[c] = pd.NA

    for c in COMPOSITION_COLS:
        df[c] = pd.to_numeric(df[c], errors="coerce").astype("Float64")
    df["LHV_kJ_kg"] = pd.to_numeric(df["LHV_kJ_kg"], errors="coerce")
    return df


def _normalize_fuel_properties_df(df: Optional[pd.DataFrame]) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=DEFAULT_FUEL_PROPERTY_COLUMNS)

    out = df.copy()
    rename_map: Dict[str, str] = {}
    for column in out.columns:
        cl = norm_key(column)
        if cl in {"fuel_label", "label"}:
            rename_map[column] = "Fuel_Label"
        elif cl in {"dies_pct", "dies", "diesel_pct", "diesel"}:
            rename_map[column] = "DIES_pct"
        elif cl in {"biod_pct", "biod", "biodiesel_pct", "biodiesel"}:
            rename_map[column] = "BIOD_pct"
        elif cl in {"etoh_pct", "etoh", "e_pct", "e"}:
            rename_map[column] = "EtOH_pct"
        elif cl in {"h2o_pct", "h2o", "h20_pct", "h20", "h_pct", "h"}:
            rename_map[column] = "H2O_pct"
        elif cl in {"lhv_kj_kg", "lhv", "pci_kj_kg", "pci"}:
            rename_map[column] = "LHV_kJ_kg"
        elif cl in {"fuel_density_kg_m3", "density_kg_m3", "density"}:
            rename_map[column] = "Fuel_Density_kg_m3"
        elif cl in {"fuel_cost_r_l", "cost_r_l", "cost"}:
            rename_map[column] = "Fuel_Cost_R_L"
        elif cl in {"reference", "source"}:
            rename_map[column] = "reference"
        elif cl in {"notes", "note"}:
            rename_map[column] = "notes"
    out = out.rename(columns=rename_map)

    for column in DEFAULT_FUEL_PROPERTY_COLUMNS:
        if column not in out.columns:
            out[column] = pd.NA

    for column in COMPOSITION_COLS:
        out[column] = pd.to_numeric(out[column], errors="coerce").astype("Float64")
    for column in ("LHV_kJ_kg", "Fuel_Density_kg_m3", "Fuel_Cost_R_L"):
        out[column] = pd.to_numeric(out[column], errors="coerce")
    if "Fuel_Label" in out.columns:
        missing_label = out["Fuel_Label"].isna() | out["Fuel_Label"].map(lambda value: not _to_str_or_empty(value).strip())
        if bool(missing_label.any()):
            inferred = out.apply(
                lambda row: _fuel_label_from_components(
                    row.get("DIES_pct", pd.NA),
                    row.get("BIOD_pct", pd.NA),
                    row.get("EtOH_pct", pd.NA),
                    row.get("H2O_pct", pd.NA),
                ),
                axis=1,
            )
            inferred = inferred.map(lambda value: value or pd.NA).astype("object")
            out.loc[missing_label, "Fuel_Label"] = inferred.loc[missing_label]
    for column in ("Fuel_Label", "reference", "notes"):
        out[column] = out[column].map(lambda value: _to_str_or_empty(value) or pd.NA)

    return out[DEFAULT_FUEL_PROPERTY_COLUMNS].copy()


def _fill_fuel_property_defaults(fuel_properties_df: pd.DataFrame, defaults_cfg: Dict[str, str]) -> pd.DataFrame:
    if fuel_properties_df is None or fuel_properties_df.empty:
        return pd.DataFrame(columns=DEFAULT_FUEL_PROPERTY_COLUMNS)

    out = _normalize_fuel_properties_df(fuel_properties_df)
    if out.empty:
        return out

    for idx, row in out.iterrows():
        label = _to_str_or_empty(row.get("Fuel_Label", "")).strip()
        if not label:
            continue
        density = _to_float(row.get("Fuel_Density_kg_m3", pd.NA), default=float("nan"))
        cost = _to_float(row.get("Fuel_Cost_R_L", pd.NA), default=float("nan"))
        if not np.isfinite(density) or density <= 0:
            density_default = _to_float(defaults_cfg.get(norm_key(f"FUEL_DENSITY_KG_M3_{label}"), ""), default=float("nan"))
            if np.isfinite(density_default) and density_default > 0:
                out.at[idx, "Fuel_Density_kg_m3"] = float(density_default)
        if not np.isfinite(cost) or cost <= 0:
            cost_default = _to_float(defaults_cfg.get(norm_key(f"FUEL_COST_R_L_{label}"), ""), default=float("nan"))
            if np.isfinite(cost_default) and cost_default > 0:
                out.at[idx, "Fuel_Cost_R_L"] = float(cost_default)
    return out


def load_lhv_lookup(defaults_cfg: Optional[Dict[str, str]] = None) -> pd.DataFrame:
    p = CFG_DIR / "lhv.csv"
    if not p.exists():
        raise FileNotFoundError(f"NÃƒÂ£o encontrei {p}.")

    df = pd.read_csv(p, sep=None, engine="python", encoding="utf-8-sig")
    df = _normalize_fuel_properties_df(df)
    if "LHV_kJ_kg" not in df.columns:
        raise KeyError(f"lhv.csv precisa da coluna LHV_kJ_kg. Colunas atuais: {list(df.columns)}")
    if defaults_cfg:
        df = _fill_fuel_property_defaults(df, defaults_cfg)
    return df


def load_fuel_properties_lookup(
    config_bundle: Optional[Pipeline29ConfigBundle],
    defaults_cfg: Dict[str, str],
) -> pd.DataFrame:
    configured = pd.DataFrame(columns=DEFAULT_FUEL_PROPERTY_COLUMNS)
    if config_bundle is not None and getattr(config_bundle, "fuel_properties_df", None) is not None:
        configured = _fill_fuel_property_defaults(config_bundle.fuel_properties_df, defaults_cfg)
        if not configured.empty and pd.to_numeric(configured.get("LHV_kJ_kg", pd.Series(dtype="float64")), errors="coerce").notna().any():
            return configured

    try:
        legacy = load_lhv_lookup(defaults_cfg=defaults_cfg)
    except Exception:
        legacy = pd.DataFrame(columns=DEFAULT_FUEL_PROPERTY_COLUMNS)

    if configured.empty:
        return legacy
    if legacy.empty:
        return configured

    configured_keys = {
        (
            _to_float(row.get("DIES_pct", pd.NA), default=float("nan")),
            _to_float(row.get("BIOD_pct", pd.NA), default=float("nan")),
            _to_float(row.get("EtOH_pct", pd.NA), default=float("nan")),
            _to_float(row.get("H2O_pct", pd.NA), default=float("nan")),
        )
        for _, row in configured.iterrows()
    }
    missing_legacy_rows: List[Dict[str, Any]] = []
    for _, row in legacy.iterrows():
        key = (
            _to_float(row.get("DIES_pct", pd.NA), default=float("nan")),
            _to_float(row.get("BIOD_pct", pd.NA), default=float("nan")),
            _to_float(row.get("EtOH_pct", pd.NA), default=float("nan")),
            _to_float(row.get("H2O_pct", pd.NA), default=float("nan")),
        )
        if key in configured_keys:
            continue
        missing_legacy_rows.append(row.to_dict())
    if not missing_legacy_rows:
        return configured
    combined = pd.concat([configured, pd.DataFrame(missing_legacy_rows)], ignore_index=True)
    return _normalize_fuel_properties_df(combined)


def _lookup_lhv_for_blend(
    lhv_df: pd.DataFrame,
    *,
    etoh_pct: float,
    h2o_pct: float,
    tol: float = 0.6,
) -> float:
    if lhv_df is None or lhv_df.empty:
        return float("nan")
    if "LHV_kJ_kg" not in lhv_df.columns:
        return float("nan")

    etoh = pd.to_numeric(lhv_df.get("EtOH_pct", pd.Series(pd.NA, index=lhv_df.index)), errors="coerce")
    h2o = pd.to_numeric(lhv_df.get("H2O_pct", pd.Series(pd.NA, index=lhv_df.index)), errors="coerce")
    m = (etoh.sub(etoh_pct).abs() <= tol) & (h2o.sub(h2o_pct).abs() <= tol)
    if not bool(m.any()):
        return float("nan")

    vals = pd.to_numeric(lhv_df.loc[m, "LHV_kJ_kg"], errors="coerce").dropna()
    if vals.empty:
        return float("nan")
    return float(vals.iloc[0])


def _fuel_blend_labels(df: pd.DataFrame, tol: float = 0.6) -> pd.Series:
    idx = df.index
    comps = pd.DataFrame(
        {
            "DIES_pct": pd.to_numeric(df.get("DIES_pct", pd.Series(pd.NA, index=idx)), errors="coerce"),
            "BIOD_pct": pd.to_numeric(df.get("BIOD_pct", pd.Series(pd.NA, index=idx)), errors="coerce"),
            "EtOH_pct": pd.to_numeric(df.get("EtOH_pct", pd.Series(pd.NA, index=idx)), errors="coerce"),
            "H2O_pct": pd.to_numeric(df.get("H2O_pct", pd.Series(pd.NA, index=idx)), errors="coerce"),
        },
        index=idx,
    )
    labels = comps.apply(
        lambda row: _fuel_label_from_components(
            row["DIES_pct"],
            row["BIOD_pct"],
            row["EtOH_pct"],
            row["H2O_pct"],
            tol=tol,
        ),
        axis=1,
    )
    return labels.map(lambda value: value or pd.NA).astype("object")


def _fuel_default_lookup_series(
    df: pd.DataFrame,
    defaults_cfg: Dict[str, str],
    *,
    field: str,
) -> Tuple[pd.Series, List[str]]:
    labels = _fuel_blend_labels(df)
    values = pd.Series(np.nan, index=df.index, dtype="float64")
    missing: List[str] = []

    field_prefix = {
        "density_param": "FUEL_DENSITY_KG_M3_",
        "cost_param": "FUEL_COST_R_L_",
    }.get(field, "")
    if not field_prefix:
        raise KeyError(f"Campo de lookup de combustivel nao suportado: {field}")

    unique_labels = sorted({str(label).strip() for label in labels.dropna().tolist() if str(label).strip()})
    for label in unique_labels:
        mask = labels.eq(label)
        if not bool(mask.any()):
            continue

        param_name = f"{field_prefix}{label}"
        param_value = _to_float(defaults_cfg.get(norm_key(param_name), ""), default=float("nan"))
        if np.isfinite(param_value) and (param_value > 0):
            values.loc[mask] = float(param_value)
        else:
            missing.append(f"{label} -> {param_name}")

    return values, missing


def _aggregate_metric_with_uncertainty(
    df: pd.DataFrame,
    *,
    group_cols: List[str],
    value_col: str,
    uA_col: str,
    uB_col: str,
    uc_col: str,
    U_col: str,
    value_name: str,
) -> pd.DataFrame:
    out_cols = group_cols + [
        value_name,
        f"uA_{value_name}",
        f"uB_{value_name}",
        f"uc_{value_name}",
        f"U_{value_name}",
        "n_points",
    ]
    if df is None or df.empty:
        return pd.DataFrame(columns=out_cols)

    tmp = df.copy()
    required_cols = group_cols + [value_col, uA_col, uB_col, uc_col, U_col]
    for c in required_cols:
        if c not in tmp.columns:
            tmp[c] = pd.NA

    tmp = tmp.dropna(subset=group_cols).copy()
    for c in [value_col, uA_col, uB_col, uc_col, U_col]:
        tmp[c] = pd.to_numeric(tmp[c], errors="coerce")
    tmp = tmp.dropna(subset=[value_col]).copy()
    if tmp.empty:
        return pd.DataFrame(columns=out_cols)

    g = (
        tmp.groupby(group_cols, dropna=False, sort=True)
        .agg(
            **{
                value_name: (value_col, "mean"),
                "n_points": (value_col, "count"),
                "_uA_rss": (uA_col, _rss_or_na),
                "_uB_rss": (uB_col, _rss_or_na),
                "_uc_rss": (uc_col, _rss_or_na),
                "_U_rss": (U_col, _rss_or_na),
            }
        )
        .reset_index()
    )

    n = pd.to_numeric(g["n_points"], errors="coerce").replace(0, np.nan)
    g[f"uA_{value_name}"] = g["_uA_rss"] / n
    g[f"uB_{value_name}"] = g["_uB_rss"] / n
    g[f"uc_{value_name}"] = (
        pd.to_numeric(g[f"uA_{value_name}"], errors="coerce") ** 2
        + pd.to_numeric(g[f"uB_{value_name}"], errors="coerce") ** 2
    ) ** 0.5
    g[f"uc_{value_name}"] = g[f"uc_{value_name}"].where(
        g[f"uc_{value_name}"].notna(),
        g["_uc_rss"] / n,
    )
    g[f"U_{value_name}"] = K_COVERAGE * pd.to_numeric(g[f"uc_{value_name}"], errors="coerce")
    g[f"U_{value_name}"] = g[f"U_{value_name}"].where(
        g[f"U_{value_name}"].notna(),
        g["_U_rss"] / n,
    )

    return g[out_cols].copy()


def _attach_diesel_cost_delta_metrics(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    out = df.copy()
    idx = out.index
    fuel_labels = out.get("Fuel_Label", pd.Series(pd.NA, index=idx, dtype="object"))
    fuel_labels = fuel_labels.where(fuel_labels.notna(), _fuel_blend_labels(out))
    out["Fuel_Label"] = fuel_labels

    load_key = pd.to_numeric(out.get("Load_kW", pd.Series(pd.NA, index=idx)), errors="coerce").round(6)
    out["_diesel_baseline_load_key"] = load_key

    diesel_points = out[fuel_labels.eq("D85B15")].copy()
    diesel_points = diesel_points[diesel_points["_diesel_baseline_load_key"].notna()].copy()

    baseline_ref_cols = [
        "Diesel_Baseline_Custo_R_h",
        "uA_Diesel_Baseline_Custo_R_h",
        "uB_Diesel_Baseline_Custo_R_h",
        "uc_Diesel_Baseline_Custo_R_h",
        "U_Diesel_Baseline_Custo_R_h",
        "Diesel_Baseline_N_points",
    ]
    delta_cols = [
        "Razao_Custo_vs_Diesel",
        "Economia_vs_Diesel_R_h",
        "uA_Economia_vs_Diesel_R_h",
        "uB_Economia_vs_Diesel_R_h",
        "uc_Economia_vs_Diesel_R_h",
        "U_Economia_vs_Diesel_R_h",
        "Economia_vs_Diesel_pct",
        "uA_Economia_vs_Diesel_pct",
        "uB_Economia_vs_Diesel_pct",
        "uc_Economia_vs_Diesel_pct",
        "U_Economia_vs_Diesel_pct",
        "delta_over_U_Economia_vs_Diesel_pct",
        "Interpretacao_Economia_vs_Diesel",
    ]
    for c in delta_cols:
        if c not in out.columns:
            out[c] = pd.NA

    if diesel_points.empty:
        print("[WARN] Nao encontrei pontos Diesel D85B15 para calcular economia vs diesel.")
        for c in baseline_ref_cols:
            if c not in out.columns:
                out[c] = pd.NA
        return out.drop(columns=["_diesel_baseline_load_key"], errors="ignore")

    diesel_baseline = _aggregate_metric_with_uncertainty(
        diesel_points,
        group_cols=["_diesel_baseline_load_key"],
        value_col="Custo_R_h",
        uA_col="uA_Custo_R_h",
        uB_col="uB_Custo_R_h",
        uc_col="uc_Custo_R_h",
        U_col="U_Custo_R_h",
        value_name="Diesel_Baseline_Custo_R_h",
    )
    if diesel_baseline.empty:
        print("[WARN] Nao consegui agregar o baseline Diesel por carga para economia vs diesel.")
        for c in baseline_ref_cols:
            if c not in out.columns:
                out[c] = pd.NA
        return out.drop(columns=["_diesel_baseline_load_key"], errors="ignore")

    diesel_baseline = diesel_baseline.rename(columns={"n_points": "Diesel_Baseline_N_points"})
    out = out.drop(columns=baseline_ref_cols, errors="ignore")
    out = out.merge(diesel_baseline, on="_diesel_baseline_load_key", how="left", suffixes=("", "_drop"))
    out = out.drop(columns=[c for c in out.columns if c.endswith("_drop")], errors="ignore")

    custo_atual = pd.to_numeric(out.get("Custo_R_h", pd.NA), errors="coerce")
    custo_diesel = pd.to_numeric(out.get("Diesel_Baseline_Custo_R_h", pd.NA), errors="coerce")
    valid_delta = custo_atual.notna() & custo_diesel.gt(0)

    ua_atual = pd.to_numeric(out.get("uA_Custo_R_h", pd.NA), errors="coerce")
    ub_atual = pd.to_numeric(out.get("uB_Custo_R_h", pd.NA), errors="coerce")
    uc_atual = pd.to_numeric(out.get("uc_Custo_R_h", pd.NA), errors="coerce")

    ua_diesel = pd.to_numeric(out.get("uA_Diesel_Baseline_Custo_R_h", pd.NA), errors="coerce")
    ub_diesel = pd.to_numeric(out.get("uB_Diesel_Baseline_Custo_R_h", pd.NA), errors="coerce")
    uc_diesel = pd.to_numeric(out.get("uc_Diesel_Baseline_Custo_R_h", pd.NA), errors="coerce")

    out["Razao_Custo_vs_Diesel"] = (custo_atual / custo_diesel).where(valid_delta, pd.NA)
    out["Economia_vs_Diesel_R_h"] = (custo_atual - custo_diesel).where(valid_delta, pd.NA)
    out["uA_Economia_vs_Diesel_R_h"] = ((ua_atual**2 + ua_diesel**2) ** 0.5).where(valid_delta, pd.NA)
    out["uB_Economia_vs_Diesel_R_h"] = ((ub_atual**2 + ub_diesel**2) ** 0.5).where(valid_delta, pd.NA)
    out["uc_Economia_vs_Diesel_R_h"] = ((uc_atual**2 + uc_diesel**2) ** 0.5).where(valid_delta, pd.NA)
    out["U_Economia_vs_Diesel_R_h"] = (K_COVERAGE * pd.to_numeric(out["uc_Economia_vs_Diesel_R_h"], errors="coerce")).where(valid_delta, pd.NA)

    out["Economia_vs_Diesel_pct"] = (100.0 * (pd.to_numeric(out["Razao_Custo_vs_Diesel"], errors="coerce") - 1.0)).where(valid_delta, pd.NA)

    d_pct_d_custo = 100.0 / custo_diesel
    d_pct_d_diesel = -100.0 * custo_atual / (custo_diesel**2)
    ua_pct_from_atual = d_pct_d_custo.abs() * ua_atual
    ua_pct_from_diesel = d_pct_d_diesel.abs() * ua_diesel
    ub_pct_from_atual = d_pct_d_custo.abs() * ub_atual
    ub_pct_from_diesel = d_pct_d_diesel.abs() * ub_diesel
    uc_pct_from_atual = d_pct_d_custo.abs() * uc_atual
    uc_pct_from_diesel = d_pct_d_diesel.abs() * uc_diesel

    out["uA_Economia_vs_Diesel_pct"] = ((ua_pct_from_atual**2 + ua_pct_from_diesel**2) ** 0.5).where(valid_delta, pd.NA)
    out["uB_Economia_vs_Diesel_pct"] = ((ub_pct_from_atual**2 + ub_pct_from_diesel**2) ** 0.5).where(valid_delta, pd.NA)
    out["uc_Economia_vs_Diesel_pct"] = ((uc_pct_from_atual**2 + uc_pct_from_diesel**2) ** 0.5).where(valid_delta, pd.NA)
    out["U_Economia_vs_Diesel_pct"] = (K_COVERAGE * pd.to_numeric(out["uc_Economia_vs_Diesel_pct"], errors="coerce")).where(valid_delta, pd.NA)
    out["delta_over_U_Economia_vs_Diesel_pct"] = (
        pd.to_numeric(out["Economia_vs_Diesel_pct"], errors="coerce")
        / pd.to_numeric(out["U_Economia_vs_Diesel_pct"], errors="coerce")
    ).where(valid_delta, pd.NA)

    diesel_mask = out["Fuel_Label"].astype("string").eq("D85B15") & valid_delta
    out.loc[diesel_mask, "Razao_Custo_vs_Diesel"] = 1.0
    out.loc[diesel_mask, "Economia_vs_Diesel_R_h"] = 0.0
    out.loc[diesel_mask, "Economia_vs_Diesel_pct"] = 0.0
    out.loc[diesel_mask, "delta_over_U_Economia_vs_Diesel_pct"] = 0.0

    interpret = pd.Series(pd.NA, index=out.index, dtype="object")
    economia_pct = pd.to_numeric(out["Economia_vs_Diesel_pct"], errors="coerce")
    interpret.loc[economia_pct.lt(0)] = "economia_vs_diesel"
    interpret.loc[economia_pct.gt(0)] = "piora_vs_diesel"
    interpret.loc[economia_pct.eq(0)] = "igual_ao_diesel"
    out["Interpretacao_Economia_vs_Diesel"] = interpret

    return out.drop(columns=["_diesel_baseline_load_key"], errors="ignore")


def _scenario_machine_col(machine_key: str, suffix: str) -> str:
    return f"Scenario_{machine_key}_{suffix}"


def _resolve_machine_scenario_inputs(
    defaults_cfg: Dict[str, str],
    spec: Dict[str, str],
) -> Tuple[float, float, bool]:
    hours = _to_float(defaults_cfg.get(norm_key(spec["hours_param"]), ""), default=float("nan"))
    diesel_l_h = _to_float(defaults_cfg.get(norm_key(spec["diesel_l_h_param"]), ""), default=float("nan"))
    swapped = False

    if np.isfinite(hours) and np.isfinite(diesel_l_h):
        likely_swapped = (
            (hours < 100.0 and diesel_l_h > 200.0)
            or (hours < 200.0 and diesel_l_h > 1000.0)
        )
        if likely_swapped:
            hours, diesel_l_h = diesel_l_h, hours
            swapped = True
            print(
                f"[WARN] Parametros de maquina parecem invertidos em {spec['label']}: "
                f"{spec['hours_param']}={_to_float(defaults_cfg.get(norm_key(spec['hours_param']), ''), default=float('nan'))}, "
                f"{spec['diesel_l_h_param']}={_to_float(defaults_cfg.get(norm_key(spec['diesel_l_h_param']), ''), default=float('nan'))}. "
                f"Vou usar hours/ano={hours:g} e diesel_L_h={diesel_l_h:g}."
            )

    return hours, diesel_l_h, swapped


def _attach_e94h6_machine_scenario_metrics(
    df: pd.DataFrame,
    defaults_cfg: Dict[str, str],
) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    out = df.copy()
    idx = out.index
    fuel_labels = out.get("Fuel_Label", pd.Series(pd.NA, index=idx, dtype="object"))
    fuel_labels = fuel_labels.where(fuel_labels.notna(), _fuel_blend_labels(out))
    out["Fuel_Label"] = fuel_labels

    scenario_suffixes = [
        "Hours_Ano",
        "Diesel_L_h",
        "Diesel_L_ano",
        "Diesel_Custo_R_h",
        "Diesel_Custo_R_ano",
        "E94H6_L_h",
        "U_E94H6_L_h",
        "E94H6_L_ano",
        "U_E94H6_L_ano",
        "E94H6_Custo_R_h",
        "U_E94H6_Custo_R_h",
        "E94H6_Custo_R_ano",
        "U_E94H6_Custo_R_ano",
        "Economia_R_h",
        "U_Economia_R_h",
        "Economia_R_ano",
        "U_Economia_R_ano",
    ]
    for spec in MACHINE_SCENARIO_SPECS:
        for suffix in scenario_suffixes:
            col = _scenario_machine_col(spec["key"], suffix)
            if col not in out.columns:
                out[col] = pd.NA

    ref_mask = fuel_labels.eq(SCENARIO_REFERENCE_FUEL_LABEL)
    if not bool(ref_mask.any()):
        print(f"[WARN] Nao encontrei pontos {SCENARIO_REFERENCE_FUEL_LABEL} para os cenarios de maquinas.")
        return out

    diesel_cost_l = _to_float(defaults_cfg.get(norm_key("FUEL_COST_R_L_D85B15"), ""), default=float("nan"))
    ethanol_cost_l = _to_float(defaults_cfg.get(norm_key("FUEL_COST_R_L_E94H6"), ""), default=float("nan"))
    if not (np.isfinite(diesel_cost_l) and diesel_cost_l > 0):
        print("[WARN] FUEL_COST_R_L_D85B15 invalido no Defaults; cenarios de maquinas ficarao vazios.")
        return out
    if not (np.isfinite(ethanol_cost_l) and ethanol_cost_l > 0):
        print("[WARN] FUEL_COST_R_L_E94H6 invalido no Defaults; cenarios de maquinas ficarao vazios.")
        return out

    economia_pct = pd.to_numeric(out.get("Economia_vs_Diesel_pct", pd.NA), errors="coerce")
    U_economia_pct = pd.to_numeric(out.get("U_Economia_vs_Diesel_pct", pd.NA), errors="coerce")
    valid_ref = ref_mask & economia_pct.notna()

    missing_params: List[str] = []
    for spec in MACHINE_SCENARIO_SPECS:
        hours, diesel_l_h, _swapped = _resolve_machine_scenario_inputs(defaults_cfg, spec)
        if not (np.isfinite(hours) and hours > 0):
            missing_params.append(spec["hours_param"])
            continue
        if not (np.isfinite(diesel_l_h) and diesel_l_h > 0):
            missing_params.append(spec["diesel_l_h_param"])
            continue

        ratio_ethanol_vs_diesel = 1.0 + (economia_pct / 100.0)
        valid = valid_ref & ratio_ethanol_vs_diesel.gt(0)
        if not bool(valid.any()):
            continue

        diesel_cost_h = diesel_l_h * diesel_cost_l
        diesel_l_ano = diesel_l_h * hours
        diesel_cost_ano = diesel_cost_h * hours

        ethanol_cost_h = diesel_cost_h * ratio_ethanol_vs_diesel
        U_ethanol_cost_h = diesel_cost_h * (U_economia_pct.abs() / 100.0)
        ethanol_l_h = ethanol_cost_h / ethanol_cost_l
        U_ethanol_l_h = U_ethanol_cost_h / ethanol_cost_l
        ethanol_l_ano = ethanol_l_h * hours
        U_ethanol_l_ano = U_ethanol_l_h * hours
        ethanol_cost_ano = ethanol_cost_h * hours
        U_ethanol_cost_ano = U_ethanol_cost_h * hours
        economia_r_h = ethanol_cost_h - diesel_cost_h
        economia_r_ano = ethanol_cost_ano - diesel_cost_ano

        const_pairs = {
            "Hours_Ano": hours,
            "Diesel_L_h": diesel_l_h,
            "Diesel_L_ano": diesel_l_ano,
            "Diesel_Custo_R_h": diesel_cost_h,
            "Diesel_Custo_R_ano": diesel_cost_ano,
        }
        for suffix, value in const_pairs.items():
            out.loc[valid, _scenario_machine_col(spec["key"], suffix)] = value

        value_pairs = {
            "E94H6_L_h": ethanol_l_h,
            "U_E94H6_L_h": U_ethanol_l_h,
            "E94H6_L_ano": ethanol_l_ano,
            "U_E94H6_L_ano": U_ethanol_l_ano,
            "E94H6_Custo_R_h": ethanol_cost_h,
            "U_E94H6_Custo_R_h": U_ethanol_cost_h,
            "E94H6_Custo_R_ano": ethanol_cost_ano,
            "U_E94H6_Custo_R_ano": U_ethanol_cost_ano,
            "Economia_R_h": economia_r_h,
            "U_Economia_R_h": U_ethanol_cost_h,
            "Economia_R_ano": economia_r_ano,
            "U_Economia_R_ano": U_ethanol_cost_ano,
        }
        for suffix, series in value_pairs.items():
            out.loc[valid, _scenario_machine_col(spec["key"], suffix)] = pd.to_numeric(series, errors="coerce").where(valid, pd.NA)

    if missing_params:
        print(
            "[WARN] Defaults ausentes/invalidos para cenarios de maquinas: "
            + ", ".join(sorted(set(missing_params)))
            + ". As colunas desses cenarios ficarao vazias."
        )

    return out


def _fuel_label_for_group(df: pd.DataFrame) -> str:
    labels = _fuel_blend_labels(df).dropna()
    if labels.empty:
        return ""
    return str(labels.iloc[0]).strip()


def _expand_legacy_all_fuels_filter(df: pd.DataFrame, fuels_override: Optional[List[int]]) -> Optional[List[int]]:
    if fuels_override is None:
        return None

    try:
        normalized = sorted({int(float(v)) for v in fuels_override})
    except Exception:
        return fuels_override

    if 0 in normalized:
        return normalized
    if set(normalized) != set(FUEL_H2O_LEVELS):
        return normalized

    h2o = pd.to_numeric(df.get("H2O_pct", pd.Series(pd.NA, index=df.index)), errors="coerce")
    if not bool((h2o.abs() <= 0.6).any()):
        return normalized

    return [0] + normalized


# =========================
# Instruments rev2: uB computation
# =========================
def _defaults_text_value(defaults_cfg: Optional[Dict[str, str]], param: object, fallback: str = "") -> str:
    if defaults_cfg is None:
        return fallback
    if _is_blank_cell(param):
        return fallback
    p = norm_key(param)
    if not p:
        return fallback
    raw = defaults_cfg.get(p, fallback)
    return _to_str_or_empty(raw) or fallback


def _split_setting_values(raw: object) -> List[str]:
    txt = _to_str_or_empty(raw)
    if not txt:
        return []
    return [norm_key(part) for part in re.split(r"[|,;/]+", txt) if norm_key(part)]


def _filter_instrument_rows_by_defaults(
    rows: pd.DataFrame,
    defaults_cfg: Optional[Dict[str, str]] = None,
) -> pd.DataFrame:
    if rows is None or rows.empty or defaults_cfg is None:
        return rows

    if "setting_param" not in rows.columns or "setting_value" not in rows.columns:
        return rows

    keep_mask = pd.Series(True, index=rows.index, dtype="bool")
    for idx, row in rows.iterrows():
        setting_param = _to_str_or_empty(row.get("setting_param", ""))
        if not setting_param:
            continue

        expected_values = _split_setting_values(row.get("setting_value", ""))
        if not expected_values or any(v in {"*", "any"} for v in expected_values):
            continue

        actual_value = norm_key(_defaults_text_value(defaults_cfg, setting_param, ""))
        if actual_value not in expected_values:
            keep_mask.loc[idx] = False

    return rows.loc[keep_mask].copy()


def _instrument_rows_for_key(
    instruments_df: pd.DataFrame,
    key_norm: str,
    defaults_cfg: Optional[Dict[str, str]] = None,
) -> pd.DataFrame:
    if instruments_df is None or instruments_df.empty:
        return pd.DataFrame()
    if "key_norm" not in instruments_df.columns:
        return pd.DataFrame()

    rows = instruments_df[instruments_df["key_norm"].eq(key_norm)].copy()
    if rows.empty:
        fallback_key = {
            "t_e_comp_c": "t_s_agua_c",
            "t_s_comp_c": "t_s_agua_c",
        }.get(key_norm, "")
        if fallback_key:
            rows = instruments_df[instruments_df["key_norm"].eq(fallback_key)].copy()
    if rows.empty:
        return rows
    return _filter_instrument_rows_by_defaults(rows, defaults_cfg=defaults_cfg)


def _has_instrument_key(
    instruments_df: pd.DataFrame,
    key_norm: str,
    defaults_cfg: Optional[Dict[str, str]] = None,
) -> bool:
    if instruments_df is None or instruments_df.empty:
        return False
    if "key_norm" not in instruments_df.columns:
        return False
    rows = _instrument_rows_for_key(instruments_df, key_norm=key_norm, defaults_cfg=defaults_cfg)
    return not rows.empty


def _get_resolution_for_key(
    instruments_df: pd.DataFrame,
    key_norm: str,
    defaults_cfg: Optional[Dict[str, str]] = None,
) -> Optional[float]:
    if not _has_instrument_key(instruments_df, key_norm, defaults_cfg=defaults_cfg):
        return None
    rows = _instrument_rows_for_key(instruments_df, key_norm=key_norm, defaults_cfg=defaults_cfg)
    if rows.empty:
        return None
    res = pd.to_numeric(rows.get("resolution", pd.Series([], dtype="float64")), errors="coerce").abs()
    if res.dropna().empty:
        return None
    return float(res.dropna().max())


def uB_from_instruments_rev2(
    x: pd.Series,
    key_norm: str,
    instruments_df: pd.DataFrame,
    defaults_cfg: Optional[Dict[str, str]] = None,
) -> pd.Series:
    if instruments_df is None or instruments_df.empty:
        return pd.Series([pd.NA] * len(x), index=x.index)

    if not _has_instrument_key(instruments_df, key_norm, defaults_cfg=defaults_cfg):
        return pd.Series([pd.NA] * len(x), index=x.index)

    rows = _instrument_rows_for_key(instruments_df, key_norm=key_norm, defaults_cfg=defaults_cfg)
    if rows.empty:
        return pd.Series([pd.NA] * len(x), index=x.index)

    xv = pd.to_numeric(x, errors="coerce")
    u2 = pd.Series(0.0, index=xv.index, dtype="float64")

    for _, r in rows.iterrows():
        dist = str(r.get("dist", "rect")).strip().lower() or "rect"

        rmin = r.get("range_min", pd.NA)
        rmax = r.get("range_max", pd.NA)
        rmin_v = _to_float(rmin, default=np.nan)
        rmax_v = _to_float(rmax, default=np.nan)

        mask = pd.Series(True, index=xv.index)
        if np.isfinite(rmin_v):
            mask = mask & (xv >= rmin_v)
        if np.isfinite(rmax_v):
            mask = mask & (xv <= rmax_v)

        acc_abs = _to_float(r.get("acc_abs", 0.0), 0.0)
        acc_pct = _to_float(r.get("acc_pct", 0.0), 0.0)
        digits = _to_float(r.get("digits", 0.0), 0.0)
        lsd = _to_float(r.get("lsd", 0.0), 0.0)
        resolution = _to_float(r.get("resolution", 0.0), 0.0)

        limit = xv.abs() * acc_pct + acc_abs + abs(digits) * abs(lsd)
        limit = limit.where(mask, 0.0)

        if dist == "normal":
            u_acc = limit
        else:
            u_acc = rect_to_std(limit)

        u_res = res_to_std(abs(resolution))
        u_comp = (u_acc**2 + (u_res**2)) ** 0.5

        u2 = u2 + (pd.to_numeric(u_comp, errors="coerce").fillna(0.0) ** 2)

    u = (u2**0.5).where(xv.notna(), pd.NA)
    return u


# =========================
# LabVIEW stats (trechos / ponto)
# =========================
def compute_trechos_stats(lv_raw: pd.DataFrame, instruments_df: pd.DataFrame) -> pd.DataFrame:
    bcol = find_b_etanol_col(lv_raw)

    group_cols = ["BaseName", "Load_kW", "DIES_pct", "BIOD_pct", "EtOH_pct", "H2O_pct", "WindowID"]
    ignore_cols = set(group_cols + ["Index"])
    candidate_cols = [c for c in lv_raw.columns if c not in ignore_cols]

    lv = lv_raw.copy()
    if candidate_cols:
        lv[candidate_cols] = lv[candidate_cols].apply(pd.to_numeric, errors="coerce")

    g = lv.groupby(group_cols, dropna=False, sort=True)
    n_df = g.size().reset_index(name="N_samples")
    valid_groups = n_df[n_df["N_samples"] >= MIN_SAMPLES_PER_WINDOW][group_cols].copy()
    if valid_groups.empty:
        return pd.DataFrame(columns=group_cols + ["N_samples", "Consumo_kg_h", "uB_Consumo_kg_h"])

    lv_valid = lv.merge(valid_groups, on=group_cols, how="inner")
    gv = lv_valid.groupby(group_cols, dropna=False, sort=True)

    means = gv[candidate_cols].mean(numeric_only=True).add_suffix("_mean").copy()
    first = gv[bcol].first().rename("BEtanol_start")
    last = gv[bcol].last().rename("BEtanol_end")
    n2 = gv.size().rename("N_samples")

    out = pd.concat([means, first, last, n2], axis=1).reset_index().copy()

    out["Delta_BEtanol"] = out["BEtanol_start"] - out["BEtanol_end"]
    out["DeltaT_s"] = (out["N_samples"] - 1) * DT_S
    out["Consumo_kg_h"] = (out["Delta_BEtanol"] / out["DeltaT_s"]) * 3600.0
    out.loc[out["DeltaT_s"] <= 0, "Consumo_kg_h"] = pd.NA

    bal_key = "balance_kg"
    if _has_instrument_key(instruments_df, bal_key):
        res_kg = _get_resolution_for_key(instruments_df, bal_key)
        if res_kg is None or not np.isfinite(res_kg) or res_kg <= 0:
            out["uB_Consumo_kg_h"] = pd.NA
            print("[WARN] balance_kg existe em Instruments, mas 'resolution' estÃ¡ vazio/ invÃ¡lido. uB_Consumo_kg_h ficou NA.")
        else:
            u_read = res_to_std(res_kg)  # kg
            u_delta = sqrt(2) * u_read   # kg
            out["uB_Consumo_kg_h"] = (u_delta / out["DeltaT_s"]) * 3600.0
            out.loc[out["DeltaT_s"] <= 0, "uB_Consumo_kg_h"] = pd.NA
    else:
        out["uB_Consumo_kg_h"] = pd.NA

    keep = group_cols + [c for c in out.columns if c.endswith("_mean")] + ["Consumo_kg_h", "uB_Consumo_kg_h", "N_samples"]
    return out[keep].copy()


def compute_ponto_stats(trechos: pd.DataFrame) -> pd.DataFrame:
    if trechos.empty:
        return pd.DataFrame()

    group_cols = ["BaseName", "Load_kW", "DIES_pct", "BIOD_pct", "EtOH_pct", "H2O_pct"]
    value_cols = [c for c in trechos.columns if c not in group_cols and c != "WindowID"]

    tre = trechos.copy()
    if value_cols:
        tre[value_cols] = tre[value_cols].apply(pd.to_numeric, errors="coerce")

    g = tre.groupby(group_cols, dropna=False, sort=True)

    mean_of_windows = g[value_cols].mean(numeric_only=True).add_suffix("_mean_of_windows").copy()
    sd_of_windows = g[value_cols].std(ddof=1, numeric_only=True).add_suffix("_sd_of_windows").copy()
    mean_of_windows.columns = [_normalize_repeated_stat_tokens_in_name(c) for c in mean_of_windows.columns]
    sd_of_windows.columns = [_normalize_repeated_stat_tokens_in_name(c) for c in sd_of_windows.columns]
    n_trechos = g.size().rename("N_trechos_validos")

    out = pd.concat([mean_of_windows, sd_of_windows, n_trechos], axis=1).reset_index().copy()

    uB_col = "uB_Consumo_kg_h"
    if uB_col in tre.columns:
        tmp = tre[group_cols + [uB_col]].copy()
        tmp[uB_col] = pd.to_numeric(tmp[uB_col], errors="coerce")

        sum_u2_df = (
            tmp.groupby(group_cols, dropna=False, sort=True)[uB_col]
            .apply(lambda s: float((s**2).sum()))
            .reset_index(name="sum_u2")
        )
        out = out.merge(sum_u2_df, on=group_cols, how="left").copy()

        N = pd.to_numeric(out["N_trechos_validos"], errors="coerce")
        out["uB_Consumo_kg_h_mean_of_windows"] = (pd.to_numeric(out["sum_u2"], errors="coerce") ** 0.5) / N
        out.drop(columns=["sum_u2"], inplace=True)
    else:
        out["uB_Consumo_kg_h_mean_of_windows"] = pd.NA

    return out.copy()


def compute_motec_trechos_stats(motec_raw: pd.DataFrame) -> pd.DataFrame:
    if motec_raw.empty:
        return pd.DataFrame()

    group_cols = ["BaseName", "Load_kW", "DIES_pct", "BIOD_pct", "EtOH_pct", "H2O_pct", "WindowID"]
    ignore_cols = set(group_cols + ["Index"])
    candidate_cols = [c for c in motec_raw.columns if c not in ignore_cols]

    mot = motec_raw.copy()
    if candidate_cols:
        mot[candidate_cols] = mot[candidate_cols].apply(pd.to_numeric, errors="coerce")

    g = mot.groupby(group_cols, dropna=False, sort=True)
    n_df = g.size().reset_index(name="Motec_N_samples")
    valid_groups = n_df[n_df["Motec_N_samples"] >= MIN_SAMPLES_PER_WINDOW][group_cols].copy()
    if valid_groups.empty:
        return pd.DataFrame(columns=group_cols + ["Motec_N_samples"])

    mot_valid = mot.merge(valid_groups, on=group_cols, how="inner")
    gv = mot_valid.groupby(group_cols, dropna=False, sort=True)

    means = gv[candidate_cols].mean(numeric_only=True).add_suffix("_mean").copy()
    n2 = gv.size().rename("Motec_N_samples")

    out = pd.concat([means, n2], axis=1).reset_index().copy()
    keep = group_cols + [c for c in out.columns if c.endswith("_mean")] + ["Motec_N_samples"]
    return out[keep].copy()


def compute_motec_ponto_stats(motec_trechos: pd.DataFrame) -> pd.DataFrame:
    if motec_trechos.empty:
        return pd.DataFrame(columns=["Load_kW", "DIES_pct", "BIOD_pct", "EtOH_pct", "H2O_pct"])

    group_cols = ["Load_kW", "DIES_pct", "BIOD_pct", "EtOH_pct", "H2O_pct"]
    value_cols = [c for c in motec_trechos.columns if c not in set(group_cols + ["BaseName", "WindowID", "Motec_N_samples"])]

    mot = motec_trechos.copy()
    if value_cols:
        mot[value_cols] = mot[value_cols].apply(pd.to_numeric, errors="coerce")

    g = mot.groupby(group_cols, dropna=False, sort=True)
    mean_of_windows = g[value_cols].mean(numeric_only=True).add_suffix("_mean_of_windows").copy()
    sd_of_windows = g[value_cols].std(ddof=1, numeric_only=True).add_suffix("_sd_of_windows").copy()
    mean_of_windows.columns = [_normalize_repeated_stat_tokens_in_name(c) for c in mean_of_windows.columns]
    sd_of_windows.columns = [_normalize_repeated_stat_tokens_in_name(c) for c in sd_of_windows.columns]
    n_trechos = g.size().rename("Motec_N_trechos_validos")
    n_files = g["BaseName"].nunique().rename("Motec_N_files")
    mean_samples = g["Motec_N_samples"].mean().rename("Motec_N_samples_mean_of_windows")

    out = pd.concat([mean_of_windows, sd_of_windows, n_trechos, n_files, mean_samples], axis=1).reset_index().copy()
    return out


# =========================
# Uncertainty workflow (generic, mapping-driven)
# =========================
def _prefix_from_key_norm(key_norm: str) -> str:
    if key_norm == "power_kw":
        return "P_kw"
    if key_norm == "fuel_kgh":
        return "Consumo_kg_h"
    if key_norm == "lhv_kj_kg":
        return "LHV_kJ_kg"
    return key_norm.upper()


def add_uncertainties_from_mappings(
    df: pd.DataFrame,
    mappings: dict,
    instruments_df: pd.DataFrame,
    N: pd.Series,
    defaults_cfg: Optional[Dict[str, str]] = None,
) -> pd.DataFrame:
    out = df.copy()

    for key_norm, spec in mappings.items():
        col_mean_req = str(spec.get("mean", "")).strip()
        if not col_mean_req:
            continue

        try:
            col_mean = resolve_col(out, col_mean_req)
        except Exception as e:
            print(f"[WARN] Uncertainty: key='{key_norm}' col_mean '{col_mean_req}' nÃ£o encontrada no output. Pulando. ({e})")
            continue

        col_sd_req = str(spec.get("sd", "")).strip()
        col_sd = None
        if col_sd_req:
            try:
                col_sd = resolve_col(out, col_sd_req)
            except Exception:
                col_sd = None

        prefix = _prefix_from_key_norm(key_norm)

        if col_sd is not None and col_sd in out.columns:
            out[f"uA_{prefix}"] = pd.to_numeric(out[col_sd], errors="coerce") / (pd.to_numeric(N, errors="coerce") ** 0.5)
        else:
            out[f"uA_{prefix}"] = pd.NA

        out[f"uB_{prefix}"] = uB_from_instruments_rev2(
            pd.to_numeric(out[col_mean], errors="coerce"),
            key_norm=key_norm,
            instruments_df=instruments_df,
            defaults_cfg=defaults_cfg,
        )

        ua = pd.to_numeric(out[f"uA_{prefix}"], errors="coerce")
        ub = pd.to_numeric(out[f"uB_{prefix}"], errors="coerce")
        out[f"uc_{prefix}"] = (ua**2 + ub**2) ** 0.5
        out[f"U_{prefix}"] = K_COVERAGE * out[f"uc_{prefix}"]

    return out


def _combine_average_temperature_uncertainties(
    df: pd.DataFrame,
    *,
    mean_cols: List[str],
    source_prefixes: List[str],
    target_mean_col: str,
    target_prefix: str,
) -> pd.DataFrame:
    out = df.copy()
    existing_mean_cols = [c for c in mean_cols if c in out.columns]
    if not existing_mean_cols:
        out[target_mean_col] = pd.NA
        for suffix in ("uA", "uB", "uc", "U"):
            out[f"{suffix}_{target_prefix}"] = pd.NA
        return out

    mean_df = out[existing_mean_cols].apply(pd.to_numeric, errors="coerce")
    out[target_mean_col] = mean_df.mean(axis=1)
    n_valid = mean_df.notna().sum(axis=1).astype("float64")
    n_valid = n_valid.where(n_valid > 0, np.nan)

    for prefix_kind in ("uA", "uB"):
        cols = [f"{prefix_kind}_{p}" for p in source_prefixes if f"{prefix_kind}_{p}" in out.columns]
        if not cols:
            out[f"{prefix_kind}_{target_prefix}"] = pd.NA
            continue
        comp = out[cols].apply(pd.to_numeric, errors="coerce")
        out[f"{prefix_kind}_{target_prefix}"] = ((comp**2).sum(axis=1) ** 0.5) / n_valid
        out.loc[n_valid.isna(), f"{prefix_kind}_{target_prefix}"] = pd.NA

    ua = pd.to_numeric(out.get(f"uA_{target_prefix}", pd.NA), errors="coerce")
    ub = pd.to_numeric(out.get(f"uB_{target_prefix}", pd.NA), errors="coerce")
    out[f"uc_{target_prefix}"] = (ua**2 + ub**2) ** 0.5
    out[f"U_{target_prefix}"] = K_COVERAGE * pd.to_numeric(out[f"uc_{target_prefix}"], errors="coerce")
    return out


def _combine_delta_temperature_uncertainties(
    df: pd.DataFrame,
    *,
    minuend_col: str,
    subtrahend_col: str,
    minuend_prefix: str,
    subtrahend_prefix: str,
    target_value_col: str,
    target_prefix: str,
) -> pd.DataFrame:
    out = df.copy()
    if minuend_col not in out.columns or subtrahend_col not in out.columns:
        out[target_value_col] = pd.NA
        for suffix in ("uA", "uB", "uc", "U"):
            out[f"{suffix}_{target_prefix}"] = pd.NA
        return out

    minuend = pd.to_numeric(out[minuend_col], errors="coerce")
    subtrahend = pd.to_numeric(out[subtrahend_col], errors="coerce")
    out[target_value_col] = minuend - subtrahend

    for prefix_kind in ("uA", "uB"):
        a = pd.to_numeric(out.get(f"{prefix_kind}_{minuend_prefix}", pd.NA), errors="coerce")
        b = pd.to_numeric(out.get(f"{prefix_kind}_{subtrahend_prefix}", pd.NA), errors="coerce")
        out[f"{prefix_kind}_{target_prefix}"] = (a**2 + b**2) ** 0.5

    ua = pd.to_numeric(out.get(f"uA_{target_prefix}", pd.NA), errors="coerce")
    ub = pd.to_numeric(out.get(f"uB_{target_prefix}", pd.NA), errors="coerce")
    out[f"uc_{target_prefix}"] = (ua**2 + ub**2) ** 0.5
    out[f"U_{target_prefix}"] = K_COVERAGE * pd.to_numeric(out[f"uc_{target_prefix}"], errors="coerce")
    return out


def _apply_reporting_rounding(df: pd.DataFrame, mappings: dict, reporting_df: pd.DataFrame) -> pd.DataFrame:
    if reporting_df is None or reporting_df.empty:
        return df

    out = df.copy()
    for _, r in reporting_df.iterrows():
        key_norm = norm_key(r.get("key", ""))
        if not key_norm:
            continue
        if key_norm not in mappings:
            continue

        col_mean = str(mappings[key_norm].get("mean", "")).strip()
        if not col_mean:
            continue
        if col_mean not in out.columns:
            try:
                col_mean = resolve_col(out, col_mean)
            except Exception:
                continue

        res = _to_float(r.get("report_resolution", 0.0), 0.0)
        if res <= 0:
            continue

        rule = str(r.get("rule", "round_half_up")).strip().lower()
        new_col = f"{col_mean}_report"
        if new_col in out.columns:
            continue

        if rule == "round_half_up":
            out[new_col] = _round_half_up_to_resolution(out[col_mean], res)
        else:
            v = pd.to_numeric(out[col_mean], errors="coerce")
            out[new_col] = (np.round(v / res) * res)

    return out


def _normalized_composition_keys(df: pd.DataFrame) -> pd.DataFrame:
    idx = df.index
    out = pd.DataFrame(index=idx)

    dies = pd.to_numeric(df.get("DIES_pct", pd.Series(pd.NA, index=idx)), errors="coerce")
    biod = pd.to_numeric(df.get("BIOD_pct", pd.Series(pd.NA, index=idx)), errors="coerce")
    etoh = pd.to_numeric(df.get("EtOH_pct", pd.Series(pd.NA, index=idx)), errors="coerce")
    h2o = pd.to_numeric(df.get("H2O_pct", pd.Series(pd.NA, index=idx)), errors="coerce")

    has_diesel = dies.notna() | biod.notna()
    has_ethanol = etoh.notna() | h2o.notna()

    out["DIES_pct"] = dies.where(has_diesel, np.where(has_ethanol, 0.0, np.nan))
    out["BIOD_pct"] = biod.where(has_diesel, np.where(has_ethanol, 0.0, np.nan))
    out["EtOH_pct"] = etoh.where(has_ethanol, np.where(has_diesel, 0.0, np.nan))
    out["H2O_pct"] = h2o.where(has_ethanol, np.where(has_diesel, 0.0, np.nan))
    return out


def _normalized_extra_merge_key(df: pd.DataFrame, col: str) -> pd.Series:
    idx = df.index
    raw = df.get(col, pd.Series(pd.NA, index=idx))
    numeric = pd.to_numeric(raw, errors="coerce")
    if numeric.notna().any():
        return numeric
    return raw.map(_canon_name)


def _left_merge_on_fuel_keys(left: pd.DataFrame, right: pd.DataFrame, extra_on: Optional[List[str]] = None) -> pd.DataFrame:
    extra = extra_on or []

    l = left.copy()
    r = right.copy()

    l_norm = _normalized_composition_keys(l)
    r_norm = _normalized_composition_keys(r)

    merge_cols: List[str] = []
    for c in extra + COMPOSITION_COLS:
        tmp = f"__merge_{c}"
        if c in extra:
            l[tmp] = _normalized_extra_merge_key(l, c)
            r[tmp] = _normalized_extra_merge_key(r, c)
        else:
            l[tmp] = pd.to_numeric(l_norm[c], errors="coerce")
            r[tmp] = pd.to_numeric(r_norm[c], errors="coerce")
        merge_cols.append(tmp)

    right_payload = r.drop(columns=[c for c in extra + COMPOSITION_COLS if c in r.columns]).copy()
    for tmp in merge_cols:
        right_payload[tmp] = r[tmp]

    out = l.merge(right_payload, on=merge_cols, how="left")
    out.drop(columns=merge_cols, inplace=True)
    return out


def _guess_plot_uncertainty_col(out_df: pd.DataFrame, y_col: str, mappings: dict) -> Optional[str]:
    candidates: List[str] = []

    direct = f"U_{y_col}"
    if direct not in candidates:
        candidates.append(direct)

    for key_norm, spec in mappings.items():
        col_mean_req = str(spec.get("mean", "")).strip()
        if not col_mean_req:
            continue
        try:
            mapped_mean = resolve_col(out_df, col_mean_req)
        except Exception:
            continue
        if mapped_mean == y_col:
            cand = f"U_{_prefix_from_key_norm(key_norm)}"
            if cand not in candidates:
                candidates.append(cand)

    for cand in candidates:
        if cand not in out_df.columns:
            continue
        vals = pd.to_numeric(out_df[cand], errors="coerce")
        if vals.notna().any():
            return cand
    return None


def _compare_metric_uncertainty_cols(out_df: pd.DataFrame, metric_col: str, mappings: dict) -> Tuple[str, str, str, str]:
    U_col = _guess_plot_uncertainty_col(out_df, metric_col, mappings) or f"U_{metric_col}"
    suffix = U_col[2:] if U_col.startswith("U_") else metric_col
    return f"uA_{suffix}", f"uB_{suffix}", f"uc_{suffix}", U_col


# =========================
# Final table
# =========================
def build_final_table(
    ponto: pd.DataFrame,
    fuel_properties: pd.DataFrame,
    kibox_agg: pd.DataFrame,
    motec_ponto: pd.DataFrame,
    mappings: dict,
    instruments_df: pd.DataFrame,
    reporting_df: pd.DataFrame,
    defaults_cfg: Dict[str, str],
) -> pd.DataFrame:
    df = add_source_identity_columns(ponto)
    df = _left_merge_on_fuel_keys(df, fuel_properties)
    if kibox_agg is not None and not kibox_agg.empty:
        df = _left_merge_on_fuel_keys(df, kibox_agg, extra_on=["SourceFolder", "Load_kW"])
    if motec_ponto is not None and not motec_ponto.empty:
        df = _left_merge_on_fuel_keys(df, motec_ponto, extra_on=["Load_kW"])

    kibox_bug_cols = ["KIBOX_MBF_10_90_1", "KIBOX_MBF_10_90_AVG_1"]
    drop_now = [c for c in kibox_bug_cols if c in df.columns]
    if drop_now:
        df = df.drop(columns=drop_now)

    ai90_col = _find_kibox_col_by_tokens(df, ["ai", "90"])
    ai10_col = _find_kibox_col_by_tokens(df, ["ai", "10"])
    if ai90_col and ai10_col:
        df["MFB_10_90"] = pd.to_numeric(df[ai90_col], errors="coerce") - pd.to_numeric(df[ai10_col], errors="coerce")
    else:
        df["MFB_10_90"] = pd.NA
        if not ai90_col or not ai10_col:
            print(f"[WARN] NÃ£o calculei MFB_10_90: ai90_col={ai90_col}, ai10_col={ai10_col}")

    N = pd.to_numeric(df["N_trechos_validos"], errors="coerce")

    df = add_uncertainties_from_mappings(
        df,
        mappings=mappings,
        instruments_df=instruments_df,
        N=N,
        defaults_cfg=defaults_cfg,
    )

    if "uB_Consumo_kg_h" in df.columns:
        df["uB_Consumo_kg_h_instrument"] = df["uB_Consumo_kg_h"]
    else:
        df["uB_Consumo_kg_h_instrument"] = pd.NA

    df["uB_Consumo_kg_h"] = pd.to_numeric(df.get("uB_Consumo_kg_h_mean_of_windows", pd.NA), errors="coerce")

    if "uA_Consumo_kg_h" in df.columns:
        df["uc_Consumo_kg_h"] = (pd.to_numeric(df["uA_Consumo_kg_h"], errors="coerce") ** 2 + pd.to_numeric(df["uB_Consumo_kg_h"], errors="coerce") ** 2) ** 0.5
        df["U_Consumo_kg_h"] = K_COVERAGE * df["uc_Consumo_kg_h"]
    else:
        df["uc_Consumo_kg_h"] = pd.NA
        df["U_Consumo_kg_h"] = pd.NA

    merged_fuel_label = df.get("Fuel_Label", pd.Series(pd.NA, index=df.index, dtype="object"))
    fallback_fuel_label = _fuel_blend_labels(df)
    df["Fuel_Label"] = merged_fuel_label.where(merged_fuel_label.notna(), fallback_fuel_label)

    default_density, _missing_density_defaults = _fuel_default_lookup_series(
        df,
        defaults_cfg,
        field="density_param",
    )
    merged_density = pd.to_numeric(df.get("Fuel_Density_kg_m3", pd.NA), errors="coerce")
    df["Fuel_Density_kg_m3"] = merged_density.where(merged_density.gt(0), default_density)

    default_cost, _missing_cost_defaults = _fuel_default_lookup_series(
        df,
        defaults_cfg,
        field="cost_param",
    )
    merged_cost = pd.to_numeric(df.get("Fuel_Cost_R_L", pd.NA), errors="coerce")
    df["Fuel_Cost_R_L"] = merged_cost.where(merged_cost.gt(0), default_cost)

    missing_density = sorted(
        {
            str(label)
            for label in df.loc[pd.to_numeric(df["Fuel_Density_kg_m3"], errors="coerce").le(0) | pd.to_numeric(df["Fuel_Density_kg_m3"], errors="coerce").isna(), "Fuel_Label"].dropna()
            if str(label).strip()
        }
    )
    missing_cost = sorted(
        {
            str(label)
            for label in df.loc[pd.to_numeric(df["Fuel_Cost_R_L"], errors="coerce").le(0) | pd.to_numeric(df["Fuel_Cost_R_L"], errors="coerce").isna(), "Fuel_Label"].dropna()
            if str(label).strip()
        }
    )
    if missing_density:
        print(
            "[WARN] Densidade ausente/invalida em Fuel Properties/Defaults para: "
            + ", ".join(sorted(set(missing_density)))
            + ". Consumo_L_h ficara vazio nesses combustiveis."
        )
    if missing_cost:
        print(
            "[WARN] Custo por litro ausente/invalido em Fuel Properties/Defaults para: "
            + ", ".join(sorted(set(missing_cost)))
            + ". Custo_R_h ficara vazio nesses combustiveis."
        )

    P_mean = resolve_col(df, mappings["power_kw"]["mean"])
    F_mean = resolve_col(df, mappings["fuel_kgh"]["mean"])
    L_col = resolve_col(df, mappings["lhv_kj_kg"]["mean"])

    PkW = pd.to_numeric(df[P_mean], errors="coerce")
    Fkgh = pd.to_numeric(df[F_mean], errors="coerce")
    fuel_density = pd.to_numeric(df["Fuel_Density_kg_m3"], errors="coerce")
    fuel_cost = pd.to_numeric(df["Fuel_Cost_R_L"], errors="coerce")
    mdot = Fkgh / 3600.0
    LHVv = pd.to_numeric(df[L_col], errors="coerce")
    lhv_e94h6_kj_kg = _lookup_lhv_for_blend(fuel_properties, etoh_pct=94.0, h2o_pct=6.0)

    # Generic alias for the measured UPD power used by runtime-specific plots.
    df["UPD_Power_kW"] = PkW
    df["UPD_Power_Bin_kW"] = PkW.round(1).where(PkW.notna(), pd.NA)
    df["LHV_E94H6_kJ_kg"] = lhv_e94h6_kj_kg if np.isfinite(lhv_e94h6_kj_kg) else pd.NA
    if not np.isfinite(lhv_e94h6_kj_kg):
        print("[WARN] LHV E94H6 (94/6) nao encontrado em Fuel Properties; n_th_E94H6_eq_flow ficara vazio.")

    df["n_th"] = PkW / (mdot * LHVv)
    df.loc[(PkW <= 0) | (mdot <= 0) | (LHVv <= 0), "n_th"] = pd.NA
    df["n_th_pct"] = df["n_th"] * 100.0

    ucP = pd.to_numeric(df.get("uc_P_kw", pd.NA), errors="coerce")
    ucF = pd.to_numeric(df.get("uc_Consumo_kg_h", pd.NA), errors="coerce")
    uBL = pd.to_numeric(df.get("uB_LHV_kJ_kg", pd.NA), errors="coerce")

    rel_uc = ((ucP / PkW) ** 2 + (ucF / Fkgh) ** 2 + (uBL / LHVv) ** 2) ** 0.5
    df["uc_n_th"] = df["n_th"] * rel_uc
    df["U_n_th"] = K_COVERAGE * df["uc_n_th"]
    df["U_n_th_pct"] = df["U_n_th"] * 100.0

    volume_factor = 1000.0 / fuel_density
    valid_volumetric = Fkgh.notna() & fuel_density.gt(0)
    df["Consumo_L_h"] = (Fkgh * volume_factor).where(valid_volumetric, pd.NA)
    for src_col, dst_col in [
        ("uA_Consumo_kg_h", "uA_Consumo_L_h"),
        ("uB_Consumo_kg_h", "uB_Consumo_L_h"),
        ("uc_Consumo_kg_h", "uc_Consumo_L_h"),
        ("U_Consumo_kg_h", "U_Consumo_L_h"),
    ]:
        src = pd.to_numeric(df.get(src_col, pd.NA), errors="coerce")
        df[dst_col] = (src * volume_factor).where(valid_volumetric, pd.NA)

    consumo_l_h = pd.to_numeric(df["Consumo_L_h"], errors="coerce")
    valid_cost = consumo_l_h.notna() & fuel_cost.gt(0)
    df["Custo_R_h"] = (consumo_l_h * fuel_cost).where(valid_cost, pd.NA)
    for src_col, dst_col in [
        ("uA_Consumo_L_h", "uA_Custo_R_h"),
        ("uB_Consumo_L_h", "uB_Custo_R_h"),
        ("uc_Consumo_L_h", "uc_Custo_R_h"),
        ("U_Consumo_L_h", "U_Custo_R_h"),
    ]:
        src = pd.to_numeric(df.get(src_col, pd.NA), errors="coerce")
        df[dst_col] = (src * fuel_cost).where(valid_cost, pd.NA)

    df = _attach_diesel_cost_delta_metrics(df)
    df = _attach_e94h6_machine_scenario_metrics(df, defaults_cfg)

    # Specific fuel consumption (g/kWh): BSFC = 1000 * fuel_kg_h / power_kW.
    bsfc = (Fkgh * 1000.0) / PkW
    invalid_bsfc = (PkW <= 0) | (Fkgh <= 0)
    df["BSFC_g_kWh"] = bsfc.where(~invalid_bsfc, pd.NA)

    uA_P = pd.to_numeric(df.get("uA_P_kw", pd.NA), errors="coerce")
    uB_P = pd.to_numeric(df.get("uB_P_kw", pd.NA), errors="coerce")
    uA_F = pd.to_numeric(df.get("uA_Consumo_kg_h", pd.NA), errors="coerce")
    uB_F = pd.to_numeric(df.get("uB_Consumo_kg_h", pd.NA), errors="coerce")

    rel_uA_bsfc = ((uA_F / Fkgh) ** 2 + (uA_P / PkW) ** 2) ** 0.5
    rel_uB_bsfc = ((uB_F / Fkgh) ** 2 + (uB_P / PkW) ** 2) ** 0.5

    df["uA_BSFC_g_kWh"] = pd.to_numeric(df["BSFC_g_kWh"], errors="coerce") * rel_uA_bsfc
    df["uB_BSFC_g_kWh"] = pd.to_numeric(df["BSFC_g_kWh"], errors="coerce") * rel_uB_bsfc
    ua_bsfc = pd.to_numeric(df["uA_BSFC_g_kWh"], errors="coerce")
    ub_bsfc = pd.to_numeric(df["uB_BSFC_g_kWh"], errors="coerce")
    df["uc_BSFC_g_kWh"] = (ua_bsfc**2 + ub_bsfc**2) ** 0.5
    df["U_BSFC_g_kWh"] = K_COVERAGE * df["uc_BSFC_g_kWh"]

    df.loc[invalid_bsfc, ["uA_BSFC_g_kWh", "uB_BSFC_g_kWh", "uc_BSFC_g_kWh", "U_BSFC_g_kWh"]] = pd.NA

    lambda_col = _resolve_airflow_lambda_col(df, mappings)
    maf_col = _resolve_airflow_maf_col(df, defaults_cfg)
    maf_min_kgh = _to_float(defaults_cfg.get(norm_key("VOL_EFF_DIESEL_MAF_MIN_KGH"), ""), default=0.0)
    maf_max_kgh = _to_float(defaults_cfg.get(norm_key("VOL_EFF_DIESEL_MAF_MAX_KGH"), ""), default=300.0)

    if lambda_col:
        print(f"[INFO] Airflow: lambda da MoTeC = '{lambda_col}'.")
    else:
        print("[INFO] Airflow: lambda da MoTeC nao encontrada; vou priorizar MAF e usar lambda default=1.0 apenas no fallback fuel+lambda.")
    if maf_col:
        print(f"[INFO] Airflow: MAF = '{maf_col}' (faixa valida {maf_min_kgh:g}..{maf_max_kgh:g} kg/h).")
    else:
        print("[INFO] Airflow: MAF nao encontrado; airflow ficara no modo fuel+lambda.")

    df = add_airflow_channels_prefer_maf_inplace(
        df,
        lambda_col=lambda_col,
        maf_col=maf_col,
        maf_min_kgh=maf_min_kgh,
        maf_max_kgh=maf_max_kgh,
    )

    airflow_methods = df.get("Airflow_Method", pd.Series(pd.NA, index=df.index, dtype="object")).fillna("unavailable")
    airflow_counts = airflow_methods.value_counts(dropna=False)
    airflow_summary = ", ".join(
        f"{label}={int(count)}"
        for label, count in [
            ("MAF", airflow_counts.get("MAF", 0)),
            ("fuel+lambda", airflow_counts.get("fuel_lambda", 0)),
            ("fuel+lambda_default", airflow_counts.get("fuel_lambda_default", 0)),
            ("indisponivel", airflow_counts.get("unavailable", 0)),
        ]
        if int(count) > 0
    )
    lambda_source_counts = df.get("LAMBDA_SOURCE", pd.Series(pd.NA, index=df.index, dtype="object")).fillna("default_1.0").value_counts(dropna=False)
    print(
        "[INFO] Airflow por ponto: "
        + (airflow_summary if airflow_summary else "nenhum ponto valido")
        + f" | lambda medida={int(lambda_source_counts.get('measured', 0))}, default_1.0={int(lambda_source_counts.get('default_1.0', 0))}"
    )

    df = add_specific_emissions_channels_inplace(
        df,
        power_kW=PkW,
        fuel_kg_h=Fkgh,
        defaults_cfg=defaults_cfg,
    )

    humidity_source_counts = (
        df.get("EMISSIONS_H2O_SOURCE", pd.Series(pd.NA, index=df.index, dtype="object"))
        .fillna("indisponivel")
        .value_counts(dropna=False)
    )
    humidity_source_summary = ", ".join(f"{label}={int(count)}" for label, count in humidity_source_counts.items() if int(count) > 0)
    print(
        "[INFO] Emissoes g/kWh: H2O_wet_frac indireto via "
        + (humidity_source_summary if humidity_source_summary else "indisponivel")
        + " | agua no escape = agua do ar + agua no combustivel + agua formada pelo H do combustivel."
    )

    thc_low_mask = pd.to_numeric(df.get("THC_LOW_SIGNAL_FLAG", pd.NA), errors="coerce").fillna(0).gt(0)
    thc_neg_mask = pd.to_numeric(df.get("THC_NEGATIVE_FLAG", pd.NA), errors="coerce").fillna(0).gt(0)
    if bool(thc_low_mask.any()):
        print(
            f"[WARN] THC muito baixo em {int(thc_low_mask.sum())} ponto(s) (|THC| < {THC_LOW_SIGNAL_WARN_PPM:g} ppm); "
            "vou calcular e plotar THC_g_kWh mesmo assim."
        )
    if bool(thc_neg_mask.any()):
        print(
            f"[WARN] THC negativo em {int(thc_neg_mask.sum())} ponto(s); "
            "vou calcular e plotar THC_g_kWh mesmo assim para preservar o diagnostico do analisador."
        )

    # Thermal efficiency based on E94H6 equivalent flow:
    # n_th_E94H6_eq_flow = P / (m_dot_eq_E94H6 * LHV_E94H6)
    F_eq_kgh = pd.to_numeric(df.get("Fuel_E94H6_eq_kg_h", pd.NA), errors="coerce")
    mdot_eq = F_eq_kgh / 3600.0
    lhv_e94_series = pd.to_numeric(df.get("LHV_E94H6_kJ_kg", pd.NA), errors="coerce")
    qdot_mix_lhv = mdot * LHVv
    qdot_eq_e94 = mdot_eq * lhv_e94_series

    df["Qdot_fuel_LHV_mix_kW"] = qdot_mix_lhv
    df["Qdot_fuel_E94H6_eq_kW"] = qdot_eq_e94
    df["n_th_E94H6_eq_flow"] = PkW / qdot_eq_e94
    df.loc[(PkW <= 0) | (mdot_eq <= 0) | (lhv_e94_series <= 0), "n_th_E94H6_eq_flow"] = pd.NA
    df["n_th_E94H6_eq_flow_pct"] = df["n_th_E94H6_eq_flow"] * 100.0

    t_cil_cols = [
        "T_S_CIL_1_mean_of_windows",
        "T_S_CIL_2_mean_of_windows",
        "T_S_CIL_3_mean_of_windows",
        "T_S_CIL_4_mean_of_windows",
    ]
    df = _combine_average_temperature_uncertainties(
        df,
        mean_cols=t_cil_cols,
        source_prefixes=[
            "T_S_CIL_1_C",
            "T_S_CIL_2_C",
            "T_S_CIL_3_C",
            "T_S_CIL_4_C",
        ],
        target_mean_col="T_E_CIL_AVG_mean_of_windows",
        target_prefix="T_E_CIL_AVG_C",
    )

    t_e_comp_col = _resolve_existing_column(df, "T_E_COMP_mean_of_windows", ["t_e_comp"])
    t_adm_col = _find_first_col_by_substrings(df, ["t", "admiss"])
    p_col = _find_first_col_by_substrings(df, ["p", "coletor"])
    rh_col = _find_first_col_by_substrings(df, ["umidade"])

    if t_e_comp_col and rh_col:
        df["UMIDADE_ABS_g_m3"] = _absolute_humidity_g_m3(df[t_e_comp_col], df[rh_col])
    else:
        df["UMIDADE_ABS_g_m3"] = pd.NA

    if t_adm_col and rh_col and p_col:
        df["cp_air_dry_kJ_kgK"] = _cp_air_dry_kj_kgk(df[t_adm_col])
        df["cp_air_moist_kJ_kgK"] = _cp_moist_air_kj_kgk(df[t_adm_col], df[rh_col], df[p_col])
        df["hum_ratio_w_kgkg"] = _humidity_ratio_w_from_rh(df[t_adm_col], df[rh_col], df[p_col])
    else:
        df["cp_air_dry_kJ_kgK"] = pd.NA
        df["cp_air_moist_kJ_kgK"] = pd.NA
        df["hum_ratio_w_kgkg"] = pd.NA

    if t_adm_col and "T_E_CIL_AVG_mean_of_windows" in df.columns:
        df = _combine_delta_temperature_uncertainties(
            df,
            minuend_col="T_E_CIL_AVG_mean_of_windows",
            subtrahend_col=t_adm_col,
            minuend_prefix="T_E_CIL_AVG_C",
            subtrahend_prefix="T_ADMISSAO_C",
            target_value_col="DT_ADMISSAO_TO_T_E_CIL_AVG_C",
            target_prefix="DT_ADMISSAO_TO_T_E_CIL_AVG_C",
        )
    else:
        df["DT_ADMISSAO_TO_T_E_CIL_AVG_C"] = pd.NA
        for suffix in ("uA", "uB", "uc", "U"):
            df[f"{suffix}_DT_ADMISSAO_TO_T_E_CIL_AVG_C"] = pd.NA

    if "Air_kg_h" in df.columns and t_adm_col and "T_E_CIL_AVG_mean_of_windows" in df.columns:
        mdot_air = pd.to_numeric(df["Air_kg_h"], errors="coerce") / 3600.0
        dT = pd.to_numeric(df["DT_ADMISSAO_TO_T_E_CIL_AVG_C"], errors="coerce")

        cp_used = pd.to_numeric(df["cp_air_moist_kJ_kgK"], errors="coerce")
        cp_fallback = pd.to_numeric(df["cp_air_dry_kJ_kgK"], errors="coerce")
        cp_used = cp_used.where(cp_used.notna(), cp_fallback)
        cp_used = cp_used.where(cp_used.notna(), 1.005)

        df["Q_EVAP_NET_kW"] = mdot_air * cp_used * dT
    else:
        df["Q_EVAP_NET_kW"] = pd.NA

    df = add_volumetric_efficiency_from_airflow_method_inplace(df, defaults_cfg)

    # ECT control error sign convention:
    # positive error => coolant outlet temperature hotter than commanded setpoint.
    t_s_agua_col = None
    for cand in [
        "T_S_AGUA_mean_of_windows",
        "T_S_AGUA",
        "T_S_ÃGUA",
    ]:
        if cand in df.columns:
            t_s_agua_col = cand
            break
    if t_s_agua_col is None:
        t_s_agua_col = _find_first_col_by_substrings(df, ["t_s", "agua", "mean_of_windows"])
    if t_s_agua_col is None:
        t_s_agua_col = _find_first_col_by_substrings(df, ["t_s", "agua"])
    if t_s_agua_col is None:
        t_s_agua_col = _find_first_col_by_substrings(df, ["t_s", "Ã¡gua"])

    dem_th2o_col = None
    for cand in [
        "DEM_TH2O_mean_of_windows",
        "DEM TH2O_mean_of_windows",
        "DEM_TH2O",
        "DEM TH2O",
    ]:
        if cand in df.columns:
            dem_th2o_col = cand
            break
    if dem_th2o_col is None:
        dem_th2o_col = _find_first_col_by_substrings(df, ["dem", "th2o", "mean_of_windows"])
    if dem_th2o_col is None:
        dem_th2o_col = _find_first_col_by_substrings(df, ["dem", "th2o"])

    if t_s_agua_col and dem_th2o_col:
        ect_actual = pd.to_numeric(df[t_s_agua_col], errors="coerce")
        ect_target = pd.to_numeric(df[dem_th2o_col], errors="coerce")
        df["ECT_CTRL_ACTUAL_C"] = ect_actual
        df["ECT_CTRL_TARGET_C"] = ect_target
        df["ECT_CTRL_ERROR_C"] = ect_actual - ect_target
        df["ECT_CTRL_ERROR_ABS_C"] = pd.to_numeric(df["ECT_CTRL_ERROR_C"], errors="coerce").abs()
    else:
        df["ECT_CTRL_ACTUAL_C"] = pd.NA
        df["ECT_CTRL_TARGET_C"] = pd.NA
        df["ECT_CTRL_ERROR_C"] = pd.NA
        df["ECT_CTRL_ERROR_ABS_C"] = pd.NA

    # Ignition delay (absolute delta in crank angle):
    # - MoTeC ignition timing is positive for BTDC.
    # - KIBOX AI05 is positive for ATDC.
    # Convert both to a common ATDC-oriented axis by flipping MoTeC sign.
    # Therefore: delay_abs = abs(AI05_ATDC - (-Ignition_BTDC)) = abs(AI05 + Ignition).
    motec_ign_col = "Motec_Ignition Timing_mean_of_windows"
    kibox_ai05_col = "KIBOX_AI05_1"
    motec_ign = pd.to_numeric(
        df.get(motec_ign_col, pd.Series(pd.NA, index=df.index, dtype="Float64")),
        errors="coerce",
    )
    kibox_ai05 = pd.to_numeric(
        df.get(kibox_ai05_col, pd.Series(pd.NA, index=df.index, dtype="Float64")),
        errors="coerce",
    )
    delay_abs = (kibox_ai05 + motec_ign).abs()
    df["Ignition_Delay_abs_degCA"] = delay_abs.where(motec_ign.notna() & kibox_ai05.notna(), pd.NA)

    df = add_run_context_columns(df)
    df = _apply_reporting_rounding(df, mappings=mappings, reporting_df=reporting_df)

    # Keep run context columns at the beginning of the final spreadsheet.
    first_cols = [c for c in ["Iteracao", "Sentido_Carga"] if c in df.columns]
    if first_cols:
        rest_cols = [c for c in df.columns if c not in first_cols]
        df = df[first_cols + rest_cols].copy()

    return df


# =========================
# Plot primitives
# =========================
def _fuel_plot_groups(df: pd.DataFrame, fuels_override: Optional[List[int]] = None) -> List[Tuple[Optional[str], pd.DataFrame]]:
    idx = df.index
    h2o = pd.to_numeric(df.get("H2O_pct", pd.Series(pd.NA, index=idx)), errors="coerce")
    dies = pd.to_numeric(df.get("DIES_pct", pd.Series(pd.NA, index=idx)), errors="coerce")
    biod = pd.to_numeric(df.get("BIOD_pct", pd.Series(pd.NA, index=idx)), errors="coerce")
    fuel_labels = df.get("Fuel_Label", pd.Series(pd.NA, index=idx, dtype="object"))
    fuel_labels = fuel_labels.where(fuel_labels.notna(), _fuel_blend_labels(df))
    fuel_labels = fuel_labels.map(lambda value: _to_str_or_empty(value) or pd.NA).astype("object")

    fuels = _expand_legacy_all_fuels_filter(df, fuels_override)
    selected_h2o_levels: Optional[List[float]] = None
    if fuels is not None:
        try:
            selected_h2o_levels = [float(v) for v in fuels]
        except Exception:
            selected_h2o_levels = None

    labeled_fuels = _preferred_fuel_label_order(fuel_labels.dropna().astype(str).tolist())
    groups: List[Tuple[Optional[str], pd.DataFrame]] = []

    if labeled_fuels:
        selected_labels = labeled_fuels
        if selected_h2o_levels is not None:
            selected_labels = []
            for label in labeled_fuels:
                mapped_level = FUEL_H2O_LEVEL_BY_LABEL.get(label)
                if mapped_level is not None:
                    if any(abs(float(mapped_level) - float(level)) <= 0.6 for level in selected_h2o_levels):
                        selected_labels.append(label)
                    continue

                label_mask = fuel_labels.eq(label)
                label_h2o = h2o.where(label_mask)
                label_dies = dies.where(label_mask)
                label_biod = biod.where(label_mask)

                # Diesel/biodiesel blends do not carry H2O_pct, so legacy
                # "all fuels" filters that include 0 must still keep them.
                is_diesel_like = bool(label_dies.gt(0).any() or label_biod.gt(0).any())
                if is_diesel_like and any(abs(float(level)) <= 0.6 for level in selected_h2o_levels):
                    selected_labels.append(label)
                    continue

                matches_level = any(bool((label_h2o.sub(level).abs() <= 0.6).any()) for level in selected_h2o_levels)
                if matches_level:
                    selected_labels.append(label)

        for label in selected_labels:
            d = df[fuel_labels.eq(label)].copy()
            if not d.empty:
                groups.append((label, d))

    unlabeled = df[fuel_labels.isna()].copy()
    if unlabeled.empty:
        return groups or [(None, df.copy())]

    unlabeled_h2o = pd.to_numeric(unlabeled.get("H2O_pct", pd.Series(pd.NA, index=unlabeled.index)), errors="coerce")
    fallback_levels = selected_h2o_levels
    if fallback_levels is None:
        fallback_levels = sorted(float(v) for v in unlabeled_h2o.dropna().unique())

    if not fallback_levels:
        return groups or [(None, df.copy())]

    for h in fallback_levels:
        hv = float(h)
        d = unlabeled[unlabeled_h2o.sub(hv).abs() <= 0.6].copy()
        if d.empty:
            continue
        label = _fuel_label_for_group(d)
        if not label:
            label = f"H2O={int(hv)}%" if hv.is_integer() else f"H2O={hv:g}%"
        groups.append((label, d))

    return groups or [(None, df.copy())]


def _series_fuel_plot_groups(
    df: pd.DataFrame,
    fuels_override: Optional[List[int]] = None,
    series_col: Optional[str] = None,
) -> List[Tuple[Optional[str], pd.DataFrame]]:
    if not series_col or series_col not in df.columns:
        return _fuel_plot_groups(df, fuels_override=fuels_override)

    sv = df[series_col].map(_to_str_or_empty)
    sv = sv.where(sv.ne(""), "origem_desconhecida")

    groups: List[Tuple[Optional[str], pd.DataFrame]] = []
    for serie in sorted(sv.dropna().unique().tolist()):
        d_series = df[sv.eq(serie)].copy()
        if d_series.empty:
            continue
        for fuel_label, d in _fuel_plot_groups(d_series, fuels_override=fuels_override):
            if d.empty:
                continue
            label = str(serie)
            if fuel_label:
                label = f"{serie} | {fuel_label}"
            groups.append((label, d))

    if groups:
        return groups
    return _fuel_plot_groups(df, fuels_override=fuels_override)


def _normalize_tol_value(v: object) -> float:
    x = _to_float(v, 0.0)
    if not np.isfinite(x):
        return 0.0
    return abs(float(x))


def _add_y_tolerance_guides(ax: plt.Axes, y_tol_plus: object, y_tol_minus: object) -> int:
    tp = _normalize_tol_value(y_tol_plus)
    tm = _normalize_tol_value(y_tol_minus)
    n = 0
    if tp > 0:
        ax.axhline(tp, color="red", linestyle="--", linewidth=1.2, label=f"limite +{tp:g}")
        n += 1
    if tm > 0:
        ax.axhline(-tm, color="red", linestyle="--", linewidth=1.2, label=f"limite -{tm:g}")
        n += 1
    return n


def _fmt_table_number(v: object) -> str:
    x = _to_float(v, default=np.nan)
    if not np.isfinite(x):
        return ""
    if abs(x) >= 1000 or (abs(x) > 0 and abs(x) < 0.01):
        return f"{x:.3e}"
    return f"{x:.3f}"


def _add_xy_value_table(
    ax: plt.Axes,
    rows: List[Tuple[str, object, object]],
    max_rows: int = 12,
) -> None:
    # Tabelas embutidas nos plots foram desativadas por decisao de usabilidade.
    return


def _apply_y_tick_step(ax: plt.Axes, y_tick_step: Optional[float]) -> None:
    step = _to_float(y_tick_step, default=np.nan)
    if not np.isfinite(step) or step <= 0:
        return

    ymin, ymax = ax.get_ylim()
    if not (np.isfinite(ymin) and np.isfinite(ymax)):
        return

    eps = abs(step) * 1e-9
    snapped_min = np.floor((ymin + eps) / step) * step
    snapped_max = np.ceil((ymax - eps) / step) * step
    if not (np.isfinite(snapped_min) and np.isfinite(snapped_max)) or snapped_max <= snapped_min:
        return

    ticks = np.arange(snapped_min, snapped_max + (step * 0.5), step).tolist()
    if not ticks:
        return

    ax.set_yticks(ticks)
    ax.set_ylim(snapped_min, snapped_max)


def _blend_mask(df: pd.DataFrame, *, etoh_pct: float, h2o_pct: float, tol: float = 0.6) -> pd.Series:
    etoh = pd.to_numeric(df.get("EtOH_pct", pd.Series(pd.NA, index=df.index)), errors="coerce")
    h2o = pd.to_numeric(df.get("H2O_pct", pd.Series(pd.NA, index=df.index)), errors="coerce")
    return (etoh.sub(etoh_pct).abs() <= tol) & (h2o.sub(h2o_pct).abs() <= tol)


def _diesel_campaign_from_basename(basename: object) -> str:
    s = _canon_name(basename).replace(" ", "_").replace("-", "_")
    if not s:
        return ""
    if ("baseline_1" in s) or ("bl_1" in s) or ("baseline" in s):
        return "baseline"
    if ("aditivado_1" in s) or ("adtv_1" in s) or ("aditivado" in s) or ("adtv" in s):
        return "aditivado"
    return ""


def _diesel_sentido_from_row(row: pd.Series) -> str:
    sent = _canon_name(row.get("Sentido_Carga", ""))
    if "subida" in sent or "subindo" in sent or re.search(r"\bup\b", sent):
        return "subida"
    if "descida" in sent or "descendo" in sent or re.search(r"\bdown\b", sent):
        return "descida"

    base = _canon_name(row.get("BaseName", ""))
    if "subindo" in base or "subida" in base:
        return "subida"
    if "descendo" in base or "descida" in base:
        return "descida"
    return ""


def _rss_or_na(values: pd.Series) -> float:
    v = pd.to_numeric(values, errors="coerce").dropna()
    if v.empty:
        return float("nan")
    return float(np.sqrt(np.sum(np.square(v.to_numpy(dtype=float)))))


def _find_consumo_plot_col(df: pd.DataFrame) -> Optional[str]:
    for c in ["Consumo_kg_h_mean_of_windows", "Consumo_kg_h", "Fuel_kg_h", "fuel_kgh_mean_of_windows"]:
        if c in df.columns:
            return c
    for c in df.columns:
        cl = str(c).lower()
        if ("consumo" in cl) and ("mean_of_windows" in cl):
            return c
    return None


def _prepare_diesel_bl_adtv_points(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()

    if "BaseName" not in df.columns:
        print("[WARN] compare iteracoes BL vs ADTV: coluna BaseName ausente. Pulei.")
        return pd.DataFrame()

    consumo_col = _find_consumo_plot_col(df)
    if not consumo_col:
        print("[WARN] compare iteracoes BL vs ADTV: coluna de consumo nao encontrada. Pulei.")
        return pd.DataFrame()

    out = df.copy()
    out["_campaign_bl_adtv"] = out["BaseName"].map(_diesel_campaign_from_basename)
    out["_sentido_plot"] = out.apply(_diesel_sentido_from_row, axis=1)

    if ("DIES_pct" in out.columns) or ("BIOD_pct" in out.columns):
        dies = pd.to_numeric(out.get("DIES_pct", pd.Series(pd.NA, index=out.index)), errors="coerce")
        biod = pd.to_numeric(out.get("BIOD_pct", pd.Series(pd.NA, index=out.index)), errors="coerce")
        diesel_mask = dies.gt(0) | biod.gt(0)
        if bool(diesel_mask.any()):
            out = out[diesel_mask].copy()

    out["Load_kW"] = pd.to_numeric(out.get("Load_kW", pd.NA), errors="coerce")
    out["_consumo"] = pd.to_numeric(out[consumo_col], errors="coerce")
    out["_uA"] = pd.to_numeric(out.get("uA_Consumo_kg_h", pd.NA), errors="coerce")
    out["_uB"] = pd.to_numeric(out.get("uB_Consumo_kg_h", pd.NA), errors="coerce")
    out["_uc"] = pd.to_numeric(out.get("uc_Consumo_kg_h", pd.NA), errors="coerce")
    out["_U"] = pd.to_numeric(out.get("U_Consumo_kg_h", pd.NA), errors="coerce")

    out = out[
        out["_campaign_bl_adtv"].isin(["baseline", "aditivado"])
        & out["_sentido_plot"].isin(["subida", "descida"])
    ].copy()
    out = out.dropna(subset=["Load_kW", "_consumo"]).copy()
    return out


def _aggregate_consumo_with_uncertainty(d: pd.DataFrame, group_cols: List[str]) -> pd.DataFrame:
    if d is None or d.empty:
        return pd.DataFrame(
            columns=group_cols
            + [
                "consumo_kg_h",
                "uA_kg_h",
                "uB_kg_h",
                "uc_kg_h",
                "U_kg_h",
                "uA_consumo_kg_h",
                "uB_consumo_kg_h",
                "uc_consumo_kg_h",
                "U_consumo_kg_h",
                "n_points",
            ]
        )

    g = (
        d.groupby(group_cols, dropna=False, sort=True)
        .agg(
            consumo_kg_h=("_consumo", "mean"),
            n_points=("_consumo", "count"),
            uA_rss=("_uA", _rss_or_na),
            uB_rss=("_uB", _rss_or_na),
            uc_rss=("_uc", _rss_or_na),
            U_rss=("_U", _rss_or_na),
        )
        .reset_index()
    )

    n = pd.to_numeric(g["n_points"], errors="coerce")
    g["uA_kg_h"] = g["uA_rss"] / n
    g["uB_kg_h"] = g["uB_rss"] / n

    g["uc_kg_h"] = (pd.to_numeric(g["uA_kg_h"], errors="coerce") ** 2 + pd.to_numeric(g["uB_kg_h"], errors="coerce") ** 2) ** 0.5
    g["uc_kg_h"] = g["uc_kg_h"].where(g["uc_kg_h"].notna(), g["uc_rss"] / n)

    g["U_kg_h"] = K_COVERAGE * pd.to_numeric(g["uc_kg_h"], errors="coerce")
    g["U_kg_h"] = g["U_kg_h"].where(g["U_kg_h"].notna(), g["U_rss"] / n)

    # Alias names to match the generic compare_iter metric pipeline.
    g["uA_consumo_kg_h"] = g["uA_kg_h"]
    g["uB_consumo_kg_h"] = g["uB_kg_h"]
    g["uc_consumo_kg_h"] = g["uc_kg_h"]
    g["U_consumo_kg_h"] = g["U_kg_h"]

    return g[
        group_cols
        + [
            "consumo_kg_h",
            "uA_kg_h",
            "uB_kg_h",
            "uc_kg_h",
            "U_kg_h",
            "uA_consumo_kg_h",
            "uB_consumo_kg_h",
            "uc_consumo_kg_h",
            "U_consumo_kg_h",
            "n_points",
        ]
    ].copy()


def _mean_subida_descida_per_campaign(d: pd.DataFrame) -> pd.DataFrame:
    if d is None or d.empty:
        return pd.DataFrame(columns=["_campaign_bl_adtv", "Load_kW", "consumo_kg_h", "uA_kg_h", "uB_kg_h", "uc_kg_h", "U_kg_h", "n_points"])

    sub = d[d["_sentido_plot"].eq("subida")].copy()
    des = d[d["_sentido_plot"].eq("descida")].copy()
    if sub.empty or des.empty:
        return pd.DataFrame(columns=["_campaign_bl_adtv", "Load_kW", "consumo_kg_h", "uA_kg_h", "uB_kg_h", "uc_kg_h", "U_kg_h", "n_points"])

    m = sub.merge(
        des,
        on=["_campaign_bl_adtv", "Load_kW"],
        how="inner",
        suffixes=("_sub", "_des"),
    )
    if m.empty:
        return pd.DataFrame(columns=["_campaign_bl_adtv", "Load_kW", "consumo_kg_h", "uA_kg_h", "uB_kg_h", "uc_kg_h", "U_kg_h", "n_points"])

    out = pd.DataFrame()
    out["_campaign_bl_adtv"] = m["_campaign_bl_adtv"]
    out["Load_kW"] = pd.to_numeric(m["Load_kW"], errors="coerce")
    out["consumo_kg_h"] = (
        pd.to_numeric(m["consumo_kg_h_sub"], errors="coerce") + pd.to_numeric(m["consumo_kg_h_des"], errors="coerce")
    ) / 2.0

    ua_sub = pd.to_numeric(m["uA_kg_h_sub"], errors="coerce")
    ua_des = pd.to_numeric(m["uA_kg_h_des"], errors="coerce")
    ub_sub = pd.to_numeric(m["uB_kg_h_sub"], errors="coerce")
    ub_des = pd.to_numeric(m["uB_kg_h_des"], errors="coerce")
    uc_sub = pd.to_numeric(m["uc_kg_h_sub"], errors="coerce")
    uc_des = pd.to_numeric(m["uc_kg_h_des"], errors="coerce")
    U_sub = pd.to_numeric(m["U_kg_h_sub"], errors="coerce")
    U_des = pd.to_numeric(m["U_kg_h_des"], errors="coerce")

    out["uA_kg_h"] = (ua_sub**2 + ua_des**2) ** 0.5 / 2.0
    out["uB_kg_h"] = (ub_sub**2 + ub_des**2) ** 0.5 / 2.0
    out["uc_kg_h"] = (out["uA_kg_h"] ** 2 + out["uB_kg_h"] ** 2) ** 0.5
    out["uc_kg_h"] = out["uc_kg_h"].where(out["uc_kg_h"].notna(), (uc_sub**2 + uc_des**2) ** 0.5 / 2.0)
    out["U_kg_h"] = K_COVERAGE * out["uc_kg_h"]
    out["U_kg_h"] = out["U_kg_h"].where(out["U_kg_h"].notna(), (U_sub**2 + U_des**2) ** 0.5 / 2.0)
    out["n_points"] = pd.to_numeric(m["n_points_sub"], errors="coerce").fillna(0) + pd.to_numeric(m["n_points_des"], errors="coerce").fillna(0)

    return out.sort_values("Load_kW").copy()


def _campaign_label(campaign: str) -> str:
    if campaign == "baseline":
        return "BL (baseline_1)"
    if campaign == "aditivado":
        return "ADTV (aditivado_1)"
    return campaign


COMPARE_ITER_SERIES_META: Dict[str, Dict[str, str]] = {
    "baseline_media": {"label": "Baseline media", "slug": "baseline_media"},
    "baseline_subida": {"label": "Baseline subida", "slug": "baseline_subida"},
    "baseline_descida": {"label": "Baseline descida", "slug": "baseline_descida"},
    "aditivado_media": {"label": "Aditivado media", "slug": "aditivado_media"},
    "aditivado_subida": {"label": "Aditivado subida", "slug": "aditivado_subida"},
    "aditivado_descida": {"label": "Aditivado descida", "slug": "aditivado_descida"},
}


COMPARE_ITER_METRIC_SPECS: List[Dict[str, str]] = [
    {
        "metric_id": "consumo",
        "metric_col": "__consumo__",
        "value_name": "consumo_kg_h",
        "title": "Consumo absoluto",
        "y_label": "Consumo absoluto (kg/h)",
        "filename_slug": "consumo_abs",
    },
    {
        "metric_id": "co2",
        "metric_col": "CO2_mean_of_windows",
        "value_name": "co2_medido",
        "title": "CO2 medido",
        "y_label": "CO2 medido (%)",
        "filename_slug": "co2_medido",
    },
    {
        "metric_id": "co",
        "metric_col": "CO_mean_of_windows",
        "value_name": "co_medido",
        "title": "CO medido",
        "y_label": "CO medido (ppm)",
        "filename_slug": "co_medido",
    },
    {
        "metric_id": "o2",
        "metric_col": "O2_mean_of_windows",
        "value_name": "o2_medido",
        "title": "O2 medido",
        "y_label": "O2 medido (%)",
        "filename_slug": "o2_medido",
    },
    {
        "metric_id": "nox",
        "metric_col": "NOX_mean_of_windows",
        "value_name": "nox_medido",
        "title": "NOX medido",
        "y_label": "NOX medido (ppm)",
        "filename_slug": "nox_medido",
    },
    {
        "metric_id": "thc",
        "metric_col": "THC_mean_of_windows",
        "value_name": "thc_medido",
        "title": "THC medido",
        "y_label": "THC medido (ppm)",
        "filename_slug": "thc_medido",
    },
]
COMPARE_ITER_METRIC_SPECS_BY_ID: Dict[str, Dict[str, str]] = {
    str(spec.get("metric_id", "")).strip().lower(): spec
    for spec in COMPARE_ITER_METRIC_SPECS
    if str(spec.get("metric_id", "")).strip()
}


def _default_compare_iter_pairs() -> List[Tuple[str, str]]:
    return [
        ("baseline_media", "aditivado_media"),
        ("baseline_subida", "aditivado_subida"),
        ("baseline_descida", "aditivado_descida"),
        ("baseline_subida", "baseline_descida"),
        ("aditivado_subida", "aditivado_descida"),
    ]


def _cross_all_compare_iter_pairs() -> List[Tuple[str, str]]:
    left_ids = ["baseline_media", "baseline_subida", "baseline_descida"]
    right_ids = ["aditivado_media", "aditivado_subida", "aditivado_descida"]
    out: List[Tuple[str, str]] = []
    for left_id in left_ids:
        for right_id in right_ids:
            out.append((left_id, right_id))
    return out


def _all_compare_iter_pairs() -> List[Tuple[str, str]]:
    out = _cross_all_compare_iter_pairs()
    out.extend(
        [
            ("baseline_subida", "baseline_descida"),
            ("aditivado_subida", "aditivado_descida"),
        ]
    )
    return out


def _parse_compare_iter_pair_token(token: str) -> Optional[Tuple[str, str]]:
    raw = str(token or "").strip().lower()
    if not raw:
        return None
    for sep in (":", ">", "|", "=>"):
        if sep in raw:
            left_id, right_id = [part.strip() for part in raw.split(sep, 1)]
            break
    else:
        return None
    if left_id not in COMPARE_ITER_SERIES_META or right_id not in COMPARE_ITER_SERIES_META:
        return None
    if left_id == right_id:
        return None
    return left_id, right_id


def _resolve_compare_iter_pairs(cli_value: object) -> List[Tuple[str, str]]:
    raw = _to_str_or_empty(cli_value)
    if not raw:
        raw = _to_str_or_empty(os.environ.get("PIPELINE29_COMPARE_ITER_PAIRS", ""))
    if not raw:
        return _default_compare_iter_pairs()

    pairs: List[Tuple[str, str]] = []
    for token in [part.strip() for part in raw.split(",") if part.strip()]:
        token_norm = token.lower()
        if token_norm == "default":
            pairs.extend(_default_compare_iter_pairs())
            continue
        if token_norm == "cross_all":
            pairs.extend(_cross_all_compare_iter_pairs())
            continue
        if token_norm == "all":
            pairs.extend(_all_compare_iter_pairs())
            continue
        parsed = _parse_compare_iter_pair_token(token_norm)
        if parsed is None:
            print(
                "[WARN] compare_iteracoes: token invalido em --compare-iter-pairs: "
                f"'{token}'. Vou ignorar."
            )
            continue
        pairs.append(parsed)

    deduped: List[Tuple[str, str]] = []
    seen: set[Tuple[str, str]] = set()
    for pair in pairs:
        if pair in seen:
            continue
        seen.add(pair)
        deduped.append(pair)
    return deduped or _default_compare_iter_pairs()


def _build_default_compare_iter_requests(compare_iter_pairs: List[Tuple[str, str]]) -> List[Dict[str, object]]:
    requests: List[Dict[str, object]] = []
    dedupe: set[Tuple[str, str, str, str]] = set()
    for metric_spec in COMPARE_ITER_METRIC_SPECS:
        metric_id = str(metric_spec.get("metric_id", "")).strip().lower()
        if not metric_id:
            continue
        for left_id, right_id in compare_iter_pairs:
            key = (left_id, right_id, metric_id, "with_uncertainty")
            if key in dedupe:
                continue
            dedupe.add(key)
            requests.append(
                {
                    "left_id": left_id,
                    "right_id": right_id,
                    "metric_id": metric_id,
                    "variant_key": "with_uncertainty",
                    "show_uncertainty": "on",
                    "dual_variant": False,
                    "source": "fallback_pairs",
                }
            )
    return requests


def _resolve_compare_iter_requests(
    compare_df: Optional[pd.DataFrame],
    *,
    fallback_pairs: Optional[List[Tuple[str, str]]] = None,
    force_fallback_pairs: bool = False,
) -> Tuple[List[Dict[str, object]], str]:
    pairs = fallback_pairs or _default_compare_iter_pairs()
    if force_fallback_pairs:
        return _build_default_compare_iter_requests(pairs), "forced_fallback_pairs"
    if compare_df is None or compare_df.empty:
        return _build_default_compare_iter_requests(pairs), "fallback_pairs"

    rows = compare_df.to_dict(orient="records")
    enabled_rows = [row for row in rows if _row_enabled((row or {}).get("enabled", ""))]
    if not enabled_rows:
        return [], "gui_empty"

    requests: List[Dict[str, object]] = []
    dedupe: set[Tuple[str, str, str, str]] = set()
    for row_idx, row in enumerate(enabled_rows, start=1):
        left_id = _to_str_or_empty((row or {}).get("left_series", "")).lower()
        right_id = _to_str_or_empty((row or {}).get("right_series", "")).lower()
        metric_id = _to_str_or_empty((row or {}).get("metric_id", "")).lower()

        if left_id not in COMPARE_ITER_SERIES_META or right_id not in COMPARE_ITER_SERIES_META:
            print(
                f"[WARN] compare_iteracoes: linha {row_idx} na aba Compare com series invalidas "
                f"('{left_id}' vs '{right_id}'). Vou ignorar."
            )
            continue
        if left_id == right_id:
            print(
                f"[WARN] compare_iteracoes: linha {row_idx} na aba Compare usa series iguais "
                f"('{left_id}'). Vou ignorar."
            )
            continue
        if metric_id not in COMPARE_ITER_METRIC_SPECS_BY_ID:
            print(
                f"[WARN] compare_iteracoes: linha {row_idx} na aba Compare com metrica invalida "
                f"('{metric_id}'). Vou ignorar."
            )
            continue

        variants = _plot_uncertainty_variants(pd.Series(row))
        for variant_key, show_uncertainty, dual_variant in variants:
            dedupe_key = (left_id, right_id, metric_id, variant_key)
            if dedupe_key in dedupe:
                continue
            dedupe.add(dedupe_key)
            requests.append(
                {
                    "left_id": left_id,
                    "right_id": right_id,
                    "metric_id": metric_id,
                    "variant_key": variant_key,
                    "show_uncertainty": show_uncertainty,
                    "dual_variant": bool(dual_variant),
                    "source": "gui_compare_tab",
                }
            )

    if not requests:
        return [], "gui_invalid"
    return requests, "gui_compare_tab"


def _compare_iter_pair_context(left_id: str, right_id: str) -> Dict[str, str]:
    left_meta = COMPARE_ITER_SERIES_META[left_id]
    right_meta = COMPARE_ITER_SERIES_META[right_id]
    return {
        "left_label": left_meta["label"],
        "right_label": right_meta["label"],
        "pair_slug": f"{left_meta['slug']}_vs_{right_meta['slug']}",
        "pair_title": f"{left_meta['label']} vs {right_meta['label']}",
        "line_label": f"{right_meta['label']} vs {left_meta['label']}",
        "note_text": f"Negativo = {right_meta['label']} menor; Positivo = {right_meta['label']} maior",
        "interpret_neg": f"{right_meta['slug']}_menor",
        "interpret_pos": f"{right_meta['slug']}_maior",
    }


def _plot_bl_adtv_consumo_absolute(
    baseline: pd.DataFrame,
    aditivado: pd.DataFrame,
    *,
    title: str,
    filename: str,
    target_dir: Path,
    label_bl: str = "BL (baseline_1)",
    label_adtv: str = "ADTV (aditivado_1)",
) -> None:
    if (baseline is None or baseline.empty) and (aditivado is None or aditivado.empty):
        print(f"[WARN] compare iteracoes BL vs ADTV: sem dados para {filename}.")
        return

    plt.figure()
    any_curve = False
    specs = [
        (label_bl, baseline, "#1f77b4"),
        (label_adtv, aditivado, "#d62728"),
    ]
    for label, d, color in specs:
        if d is None or d.empty:
            print(f"[WARN] compare iteracoes BL vs ADTV: sem dados de consumo para {label} em {filename}.")
            continue

        x = pd.to_numeric(d["Load_kW"], errors="coerce")
        y = pd.to_numeric(d["consumo_kg_h"], errors="coerce")
        yerr = pd.to_numeric(d.get("U_kg_h", pd.NA), errors="coerce")
        p = pd.DataFrame({"x": x, "y": y, "yerr": yerr}).dropna(subset=["x", "y"]).sort_values("x")
        if p.empty:
            continue

        any_curve = True
        if p["yerr"].notna().any():
            plt.errorbar(p["x"], p["y"], yerr=p["yerr"], fmt="o-", capsize=3, linewidth=1.8, markersize=4.5, color=color, label=label)
        else:
            plt.plot(p["x"], p["y"], "o-", linewidth=1.8, markersize=4.5, color=color, label=label)

    if not any_curve:
        plt.close()
        print(f"[WARN] compare iteracoes BL vs ADTV: curvas vazias para {filename}.")
        return

    plt.xlabel("Carga nominal (kW)")
    plt.ylabel("Consumo absoluto (kg/h)")
    plt.title(title)
    plt.grid(True, which="both", linestyle="--", linewidth=0.5)
    plt.legend()
    plt.gcf().text(
        0.01,
        0.01,
        "Barras: U = 2*sqrt(uA^2 + uB^2), uA=desvio padrao, uB=balanca",
        fontsize=8,
        alpha=0.8,
    )
    outpath = target_dir / filename
    plt.tight_layout()
    plt.savefig(outpath, dpi=200)
    plt.close()
    print(f"[OK] Salvei {outpath}")


def _build_bl_adtv_delta_table(
    baseline: pd.DataFrame,
    aditivado: pd.DataFrame,
    *,
    label_bl: str = "baseline",
    label_adtv: str = "aditivado",
    interpret_neg: str = "economia_aditivado",
    interpret_pos: str = "piora_aditivado",
) -> pd.DataFrame:
    if baseline is None or baseline.empty or aditivado is None or aditivado.empty:
        return pd.DataFrame()

    base_cols = ["Load_kW", "consumo_kg_h", "uA_kg_h", "uB_kg_h", "uc_kg_h", "U_kg_h", "n_points"]

    b = baseline.copy()
    a = aditivado.copy()
    for c in base_cols:
        if c not in b.columns:
            b[c] = pd.NA
        if c not in a.columns:
            a[c] = pd.NA

    b = b[base_cols].rename(
        columns={
            "consumo_kg_h": "cons_bl_kg_h",
            "uA_kg_h": "uA_bl_kg_h",
            "uB_kg_h": "uB_bl_kg_h",
            "uc_kg_h": "uc_bl_kg_h",
            "U_kg_h": "U_bl_kg_h",
            "n_points": "n_points_bl",
        }
    )
    a = a[base_cols].rename(
        columns={
            "consumo_kg_h": "cons_adtv_kg_h",
            "uA_kg_h": "uA_adtv_kg_h",
            "uB_kg_h": "uB_adtv_kg_h",
            "uc_kg_h": "uc_adtv_kg_h",
            "U_kg_h": "U_adtv_kg_h",
            "n_points": "n_points_adtv",
        }
    )

    m = b.merge(a, on="Load_kW", how="inner")
    if m.empty:
        return pd.DataFrame()

    numeric_cols = [
        "Load_kW",
        "cons_bl_kg_h",
        "uA_bl_kg_h",
        "uB_bl_kg_h",
        "uc_bl_kg_h",
        "U_bl_kg_h",
        "n_points_bl",
        "cons_adtv_kg_h",
        "uA_adtv_kg_h",
        "uB_adtv_kg_h",
        "uc_adtv_kg_h",
        "U_adtv_kg_h",
        "n_points_adtv",
    ]
    for c in numeric_cols:
        m[c] = pd.to_numeric(m[c], errors="coerce")

    m = m.dropna(subset=["Load_kW", "cons_bl_kg_h", "cons_adtv_kg_h"]).copy()
    m = m[(m["cons_bl_kg_h"] > 0) & (m["cons_adtv_kg_h"] > 0)].copy()
    if m.empty:
        return pd.DataFrame()

    m["delta_abs_kg_h"] = m["cons_adtv_kg_h"] - m["cons_bl_kg_h"]
    m["ratio_adtv_over_bl"] = m["cons_adtv_kg_h"] / m["cons_bl_kg_h"]
    m["delta_pct"] = 100.0 * (m["ratio_adtv_over_bl"] - 1.0)

    # delta_pct = 100 * (cons_adtv / cons_bl - 1)
    m["d_delta_d_cons_adtv_pct_per_kgh"] = 100.0 / m["cons_bl_kg_h"]
    m["d_delta_d_cons_bl_pct_per_kgh"] = -100.0 * m["cons_adtv_kg_h"] / (m["cons_bl_kg_h"] ** 2)

    m["uA_contrib_from_adtv_pct"] = m["d_delta_d_cons_adtv_pct_per_kgh"].abs() * m["uA_adtv_kg_h"]
    m["uA_contrib_from_bl_pct"] = m["d_delta_d_cons_bl_pct_per_kgh"].abs() * m["uA_bl_kg_h"]
    m["uA_delta_pct"] = (m["uA_contrib_from_adtv_pct"] ** 2 + m["uA_contrib_from_bl_pct"] ** 2) ** 0.5

    m["uB_contrib_from_adtv_pct"] = m["d_delta_d_cons_adtv_pct_per_kgh"].abs() * m["uB_adtv_kg_h"]
    m["uB_contrib_from_bl_pct"] = m["d_delta_d_cons_bl_pct_per_kgh"].abs() * m["uB_bl_kg_h"]
    m["uB_delta_pct"] = (m["uB_contrib_from_adtv_pct"] ** 2 + m["uB_contrib_from_bl_pct"] ** 2) ** 0.5

    m["uc_delta_pct"] = (m["uA_delta_pct"] ** 2 + m["uB_delta_pct"] ** 2) ** 0.5
    m["U_delta_pct"] = K_COVERAGE * m["uc_delta_pct"]

    rel_uc_ratio = ((m["uc_adtv_kg_h"] / m["cons_adtv_kg_h"]) ** 2 + (m["uc_bl_kg_h"] / m["cons_bl_kg_h"]) ** 2) ** 0.5
    m["uc_delta_pct_from_uc_direct"] = 100.0 * m["ratio_adtv_over_bl"].abs() * rel_uc_ratio
    m["U_delta_pct_from_uc_direct"] = K_COVERAGE * m["uc_delta_pct_from_uc_direct"]

    m["delta_over_U"] = m["delta_pct"] / m["U_delta_pct"]
    m["label_bl"] = label_bl
    m["label_adtv"] = label_adtv
    m["interpretacao"] = np.where(
        m["delta_pct"] < 0,
        interpret_neg,
        interpret_pos,
    )
    m["significancia_95pct"] = np.where(
        m["delta_pct"].abs() > m["U_delta_pct"],
        "diferenca_maior_que_U",
        "diferenca_dentro_de_U",
    )

    return m.sort_values("Load_kW").copy()


def _plot_bl_adtv_delta_pct(
    baseline: pd.DataFrame,
    aditivado: pd.DataFrame,
    *,
    title: str,
    filename: str,
    target_dir: Path,
    label_line: str = "ADTV vs BL",
    note_text: str = "Negativo = economia no aditivado; Positivo = piora",
    label_bl: str = "baseline",
    label_adtv: str = "aditivado",
    interpret_neg: str = "economia_aditivado",
    interpret_pos: str = "piora_aditivado",
) -> None:
    m = _build_bl_adtv_delta_table(
        baseline,
        aditivado,
        label_bl=label_bl,
        label_adtv=label_adtv,
        interpret_neg=interpret_neg,
        interpret_pos=interpret_pos,
    )
    if m.empty:
        print(f"[WARN] compare iteracoes BL vs ADTV: sem pares validos para {filename}.")
        return

    plt.figure()
    if m["U_delta_pct"].notna().any():
        plt.errorbar(
            m["Load_kW"],
            m["delta_pct"],
            yerr=m["U_delta_pct"],
            fmt="o-",
            capsize=3,
            linewidth=1.8,
            markersize=4.5,
            color="#2ca02c",
            label=label_line,
        )
    else:
        plt.plot(m["Load_kW"], m["delta_pct"], "o-", linewidth=1.8, markersize=4.5, color="#2ca02c", label=label_line)

    plt.axhline(0.0, color="gray", linestyle="--", linewidth=1.0, label="0%")
    plt.xlabel("Carga nominal (kW)")
    plt.ylabel("Delta percentual de consumo (%)")
    plt.title(title)
    plt.grid(True, which="both", linestyle="--", linewidth=0.5)
    plt.legend()
    plt.gcf().text(
        0.01,
        0.01,
        note_text,
        fontsize=8,
        alpha=0.85,
    )
    outpath = target_dir / filename
    plt.tight_layout()
    plt.savefig(outpath, dpi=200)
    plt.close()
    print(f"[OK] Salvei {outpath}")


def _export_compare_iteracoes_bl_adtv_excel(
    *,
    b_med: pd.DataFrame,
    a_med: pd.DataFrame,
    b_sub: pd.DataFrame,
    a_sub: pd.DataFrame,
    b_des: pd.DataFrame,
    a_des: pd.DataFrame,
    target_dir: Path,
) -> None:
    chunks: List[pd.DataFrame] = []
    specs = [
        ("bl_vs_adtv_media_subida_descida", b_med, a_med, "baseline", "aditivado", "economia_aditivado", "piora_aditivado"),
        ("bl_vs_adtv_subida", b_sub, a_sub, "baseline", "aditivado", "economia_aditivado", "piora_aditivado"),
        ("bl_vs_adtv_descida", b_des, a_des, "baseline", "aditivado", "economia_aditivado", "piora_aditivado"),
        (
            "baseline_subida_vs_descida",
            b_sub,
            b_des,
            "baseline_subida",
            "baseline_descida",
            "descida_menor_que_subida",
            "descida_maior_que_subida",
        ),
        (
            "aditivado_subida_vs_descida",
            a_sub,
            a_des,
            "aditivado_subida",
            "aditivado_descida",
            "descida_menor_que_subida",
            "descida_maior_que_subida",
        ),
    ]
    for comp_name, b_df, a_df, lbl_bl, lbl_adtv, interp_neg, interp_pos in specs:
        t = _build_bl_adtv_delta_table(
            b_df,
            a_df,
            label_bl=lbl_bl,
            label_adtv=lbl_adtv,
            interpret_neg=interp_neg,
            interpret_pos=interp_pos,
        )
        if t.empty:
            print(f"[WARN] compare iteracoes BL vs ADTV: sem dados para export Excel em '{comp_name}'.")
            continue
        t = t.copy()
        t.insert(0, "Comparacao", comp_name)
        chunks.append(t)

    if not chunks:
        print("[WARN] compare iteracoes BL vs ADTV: sem dados para exportar Excel.")
        return

    out_df = pd.concat(chunks, ignore_index=True)
    out_df["Load_kW"] = pd.to_numeric(out_df["Load_kW"], errors="coerce")
    out_df = out_df.sort_values(["Comparacao", "Load_kW"]).copy()

    outpath = safe_to_excel(out_df, target_dir / "compare_iteracoes_bl_vs_adtv_consumo_incertezas.xlsx")
    print(f"[OK] Salvei {outpath}")


def _prepare_compare_metric_points(
    df: pd.DataFrame,
    *,
    metric_col: str,
    mappings: dict,
) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()

    if "BaseName" not in df.columns:
        print("[WARN] compare iteracoes: coluna BaseName ausente. Pulei.")
        return pd.DataFrame()

    if metric_col not in df.columns:
        print(f"[WARN] compare iteracoes: coluna '{metric_col}' nao encontrada no output. Pulei.")
        return pd.DataFrame()

    out = df.copy()
    out["_campaign_bl_adtv"] = out["BaseName"].map(_diesel_campaign_from_basename)
    out["_sentido_plot"] = out.apply(_diesel_sentido_from_row, axis=1)

    if ("DIES_pct" in out.columns) or ("BIOD_pct" in out.columns):
        dies = pd.to_numeric(out.get("DIES_pct", pd.Series(pd.NA, index=out.index)), errors="coerce")
        biod = pd.to_numeric(out.get("BIOD_pct", pd.Series(pd.NA, index=out.index)), errors="coerce")
        diesel_mask = dies.gt(0) | biod.gt(0)
        if bool(diesel_mask.any()):
            out = out[diesel_mask].copy()

    uA_col, uB_col, uc_col, U_col = _compare_metric_uncertainty_cols(out, metric_col, mappings)
    out["Load_kW"] = pd.to_numeric(out.get("Load_kW", pd.NA), errors="coerce")
    out["_metric"] = pd.to_numeric(out.get(metric_col, pd.NA), errors="coerce")
    out["_uA"] = pd.to_numeric(out.get(uA_col, pd.NA), errors="coerce")
    out["_uB"] = pd.to_numeric(out.get(uB_col, pd.NA), errors="coerce")
    out["_uc"] = pd.to_numeric(out.get(uc_col, pd.NA), errors="coerce")
    out["_U"] = pd.to_numeric(out.get(U_col, pd.NA), errors="coerce")

    out = out[
        out["_campaign_bl_adtv"].isin(["baseline", "aditivado"])
        & out["_sentido_plot"].isin(["subida", "descida"])
    ].copy()
    out = out.dropna(subset=["Load_kW", "_metric"]).copy()
    return out


def _aggregate_compare_metric_with_uncertainty(d: pd.DataFrame, *, value_name: str) -> pd.DataFrame:
    return _aggregate_metric_with_uncertainty(
        d,
        group_cols=["_campaign_bl_adtv", "_sentido_plot", "Load_kW"],
        value_col="_metric",
        uA_col="_uA",
        uB_col="_uB",
        uc_col="_uc",
        U_col="_U",
        value_name=value_name,
    )


def _mean_subida_descida_per_campaign_metric(d: pd.DataFrame, *, value_name: str) -> pd.DataFrame:
    out_cols = ["_campaign_bl_adtv", "Load_kW", value_name, f"uA_{value_name}", f"uB_{value_name}", f"uc_{value_name}", f"U_{value_name}", "n_points"]
    if d is None or d.empty:
        return pd.DataFrame(columns=out_cols)

    sub = d[d["_sentido_plot"].eq("subida")].copy()
    des = d[d["_sentido_plot"].eq("descida")].copy()
    if sub.empty or des.empty:
        return pd.DataFrame(columns=out_cols)

    m = sub.merge(
        des,
        on=["_campaign_bl_adtv", "Load_kW"],
        how="inner",
        suffixes=("_sub", "_des"),
    )
    if m.empty:
        return pd.DataFrame(columns=out_cols)

    value_sub = pd.to_numeric(m.get(f"{value_name}_sub", pd.Series(pd.NA, index=m.index)), errors="coerce")
    value_des = pd.to_numeric(m.get(f"{value_name}_des", pd.Series(pd.NA, index=m.index)), errors="coerce")
    ua_sub = pd.to_numeric(m.get(f"uA_{value_name}_sub", pd.Series(pd.NA, index=m.index)), errors="coerce")
    ua_des = pd.to_numeric(m.get(f"uA_{value_name}_des", pd.Series(pd.NA, index=m.index)), errors="coerce")
    ub_sub = pd.to_numeric(m.get(f"uB_{value_name}_sub", pd.Series(pd.NA, index=m.index)), errors="coerce")
    ub_des = pd.to_numeric(m.get(f"uB_{value_name}_des", pd.Series(pd.NA, index=m.index)), errors="coerce")
    uc_sub = pd.to_numeric(m.get(f"uc_{value_name}_sub", pd.Series(pd.NA, index=m.index)), errors="coerce")
    uc_des = pd.to_numeric(m.get(f"uc_{value_name}_des", pd.Series(pd.NA, index=m.index)), errors="coerce")
    U_sub = pd.to_numeric(m.get(f"U_{value_name}_sub", pd.Series(pd.NA, index=m.index)), errors="coerce")
    U_des = pd.to_numeric(m.get(f"U_{value_name}_des", pd.Series(pd.NA, index=m.index)), errors="coerce")

    out = pd.DataFrame()
    out["_campaign_bl_adtv"] = m["_campaign_bl_adtv"]
    out["Load_kW"] = pd.to_numeric(m["Load_kW"], errors="coerce")
    out[value_name] = (value_sub + value_des) / 2.0
    out[f"uA_{value_name}"] = (ua_sub**2 + ua_des**2) ** 0.5 / 2.0
    out[f"uB_{value_name}"] = (ub_sub**2 + ub_des**2) ** 0.5 / 2.0
    out[f"uc_{value_name}"] = (out[f"uA_{value_name}"] ** 2 + out[f"uB_{value_name}"] ** 2) ** 0.5
    out[f"uc_{value_name}"] = out[f"uc_{value_name}"].where(
        out[f"uc_{value_name}"].notna(),
        (uc_sub**2 + uc_des**2) ** 0.5 / 2.0,
    )
    out[f"U_{value_name}"] = K_COVERAGE * out[f"uc_{value_name}"]
    out[f"U_{value_name}"] = out[f"U_{value_name}"].where(
        out[f"U_{value_name}"].notna(),
        (U_sub**2 + U_des**2) ** 0.5 / 2.0,
    )
    out["n_points"] = pd.to_numeric(m["n_points_sub"], errors="coerce").fillna(0) + pd.to_numeric(m["n_points_des"], errors="coerce").fillna(0)
    return out.sort_values("Load_kW").copy()


def _plot_compare_metric_absolute(
    left_df: pd.DataFrame,
    right_df: pd.DataFrame,
    *,
    value_name: str,
    y_label: str,
    title: str,
    filename: str,
    target_dir: Path,
    label_left: str,
    label_right: str,
    include_uncertainty: bool = True,
) -> bool:
    if (left_df is None or left_df.empty) and (right_df is None or right_df.empty):
        print(f"[WARN] compare iteracoes: sem dados para {filename}.")
        return False

    plt.figure()
    any_curve = False
    specs = [
        (label_left, left_df, "#1f77b4"),
        (label_right, right_df, "#d62728"),
    ]
    for label, d, color in specs:
        if d is None or d.empty:
            print(f"[WARN] compare iteracoes: sem dados para {label} em {filename}.")
            continue
        x = pd.to_numeric(d["Load_kW"], errors="coerce")
        y = pd.to_numeric(d[value_name], errors="coerce")
        yerr = pd.to_numeric(d.get(f"U_{value_name}", pd.NA), errors="coerce")
        p = pd.DataFrame({"x": x, "y": y, "yerr": yerr}).dropna(subset=["x", "y"]).sort_values("x")
        if p.empty:
            continue
        any_curve = True
        if include_uncertainty and p["yerr"].notna().any():
            plt.errorbar(p["x"], p["y"], yerr=p["yerr"], fmt="o-", capsize=3, linewidth=1.8, markersize=4.5, color=color, label=label)
        else:
            plt.plot(p["x"], p["y"], "o-", linewidth=1.8, markersize=4.5, color=color, label=label)

    if not any_curve:
        plt.close()
        print(f"[WARN] compare iteracoes: curvas vazias para {filename}.")
        return False

    plt.xlabel("Carga nominal (kW)")
    plt.ylabel(y_label)
    plt.title(title)
    plt.grid(True, which="both", linestyle="--", linewidth=0.5)
    plt.legend()
    outpath = target_dir / filename
    plt.tight_layout()
    plt.savefig(outpath, dpi=200)
    plt.close()
    print(f"[OK] Salvei {outpath}")
    return True


def _build_compare_metric_delta_table(
    left_df: pd.DataFrame,
    right_df: pd.DataFrame,
    *,
    value_name: str,
    label_left: str,
    label_right: str,
    interpret_neg: str,
    interpret_pos: str,
) -> pd.DataFrame:
    if left_df is None or left_df.empty or right_df is None or right_df.empty:
        return pd.DataFrame()

    base_cols = ["Load_kW", value_name, f"uA_{value_name}", f"uB_{value_name}", f"uc_{value_name}", f"U_{value_name}", "n_points"]
    left = left_df.copy()
    right = right_df.copy()
    for c in base_cols:
        if c not in left.columns:
            left[c] = pd.NA
        if c not in right.columns:
            right[c] = pd.NA

    left = left[base_cols].rename(
        columns={
            value_name: "value_left",
            f"uA_{value_name}": "uA_left",
            f"uB_{value_name}": "uB_left",
            f"uc_{value_name}": "uc_left",
            f"U_{value_name}": "U_left",
            "n_points": "n_points_left",
        }
    )
    right = right[base_cols].rename(
        columns={
            value_name: "value_right",
            f"uA_{value_name}": "uA_right",
            f"uB_{value_name}": "uB_right",
            f"uc_{value_name}": "uc_right",
            f"U_{value_name}": "U_right",
            "n_points": "n_points_right",
        }
    )

    m = left.merge(right, on="Load_kW", how="inner")
    if m.empty:
        return pd.DataFrame()

    numeric_cols = [
        "Load_kW",
        "value_left",
        "uA_left",
        "uB_left",
        "uc_left",
        "U_left",
        "n_points_left",
        "value_right",
        "uA_right",
        "uB_right",
        "uc_right",
        "U_right",
        "n_points_right",
    ]
    for c in numeric_cols:
        m[c] = pd.to_numeric(m[c], errors="coerce")

    m = m.dropna(subset=["Load_kW", "value_left", "value_right"]).copy()
    m = m[(m["value_left"] != 0) & (m["value_right"] != 0)].copy()
    if m.empty:
        return pd.DataFrame()

    m["delta_abs"] = m["value_right"] - m["value_left"]
    m["ratio_right_over_left"] = m["value_right"] / m["value_left"]
    m["delta_pct"] = 100.0 * (m["ratio_right_over_left"] - 1.0)
    m["d_delta_d_right"] = 100.0 / m["value_left"]
    m["d_delta_d_left"] = -100.0 * m["value_right"] / (m["value_left"] ** 2)
    m["uA_delta_pct"] = ((m["d_delta_d_right"].abs() * m["uA_right"]) ** 2 + (m["d_delta_d_left"].abs() * m["uA_left"]) ** 2) ** 0.5
    m["uB_delta_pct"] = ((m["d_delta_d_right"].abs() * m["uB_right"]) ** 2 + (m["d_delta_d_left"].abs() * m["uB_left"]) ** 2) ** 0.5
    m["uc_delta_pct"] = (m["uA_delta_pct"] ** 2 + m["uB_delta_pct"] ** 2) ** 0.5
    m["U_delta_pct"] = K_COVERAGE * m["uc_delta_pct"]
    m["delta_over_U"] = m["delta_pct"] / m["U_delta_pct"]
    m["label_left"] = label_left
    m["label_right"] = label_right
    m["interpretacao"] = np.where(m["delta_pct"] < 0, interpret_neg, interpret_pos)
    m["significancia_95pct"] = np.where(
        m["delta_pct"].abs() > m["U_delta_pct"],
        "diferenca_maior_que_U",
        "diferenca_dentro_de_U",
    )
    return m.sort_values("Load_kW").copy()


def _plot_compare_metric_delta_pct(
    left_df: pd.DataFrame,
    right_df: pd.DataFrame,
    *,
    value_name: str,
    title: str,
    filename: str,
    target_dir: Path,
    label_left: str,
    label_right: str,
    label_line: str,
    note_text: str,
    interpret_neg: str,
    interpret_pos: str,
    include_uncertainty: bool = True,
) -> pd.DataFrame:
    m = _build_compare_metric_delta_table(
        left_df,
        right_df,
        value_name=value_name,
        label_left=label_left,
        label_right=label_right,
        interpret_neg=interpret_neg,
        interpret_pos=interpret_pos,
    )
    if m.empty:
        print(f"[WARN] compare iteracoes: sem pares validos para {filename}.")
        return pd.DataFrame()

    plt.figure()
    if include_uncertainty and m["U_delta_pct"].notna().any():
        plt.errorbar(
            m["Load_kW"],
            m["delta_pct"],
            yerr=m["U_delta_pct"],
            fmt="o-",
            capsize=3,
            linewidth=1.8,
            markersize=4.5,
            color="#2ca02c",
            label=label_line,
        )
    else:
        plt.plot(m["Load_kW"], m["delta_pct"], "o-", linewidth=1.8, markersize=4.5, color="#2ca02c", label=label_line)

    plt.axhline(0.0, color="gray", linestyle="--", linewidth=1.0, label="0%")
    plt.xlabel("Carga nominal (kW)")
    plt.ylabel("Delta percentual (%)")
    plt.title(title)
    plt.grid(True, which="both", linestyle="--", linewidth=0.5)
    plt.legend()
    plt.gcf().text(0.01, 0.01, note_text, fontsize=8, alpha=0.85)
    outpath = target_dir / filename
    plt.tight_layout()
    plt.savefig(outpath, dpi=200)
    plt.close()
    print(f"[OK] Salvei {outpath}")
    return m


def _build_compare_iter_series_frames(agg: pd.DataFrame, *, value_name: str) -> Dict[str, pd.DataFrame]:
    subida = agg[agg["_sentido_plot"].eq("subida")].copy()
    descida = agg[agg["_sentido_plot"].eq("descida")].copy()
    media_sd = _mean_subida_descida_per_campaign_metric(agg, value_name=value_name)
    return {
        "baseline_subida": subida[subida["_campaign_bl_adtv"].eq("baseline")].copy(),
        "baseline_descida": descida[descida["_campaign_bl_adtv"].eq("baseline")].copy(),
        "baseline_media": media_sd[media_sd["_campaign_bl_adtv"].eq("baseline")].copy(),
        "aditivado_subida": subida[subida["_campaign_bl_adtv"].eq("aditivado")].copy(),
        "aditivado_descida": descida[descida["_campaign_bl_adtv"].eq("aditivado")].copy(),
        "aditivado_media": media_sd[media_sd["_campaign_bl_adtv"].eq("aditivado")].copy(),
    }


def _export_compare_iteracoes_metricas_excel(rows: List[pd.DataFrame], target_dir: Path) -> None:
    if not rows:
        print("[WARN] compare iteracoes: sem dados para exportar Excel consolidado.")
        return
    out_df = pd.concat(rows, ignore_index=True)
    out_df["Load_kW"] = pd.to_numeric(out_df["Load_kW"], errors="coerce")
    out_df = out_df.sort_values(["Metrica", "Comparacao", "Load_kW"]).copy()
    outpath = safe_to_excel(out_df, target_dir / "compare_iteracoes_metricas_incertezas.xlsx")
    print(f"[OK] Salvei {outpath}")


def _plot_compare_iteracoes_bl_vs_adtv(
    df: pd.DataFrame,
    *,
    root_plot_dir: Optional[Path] = None,
    mappings: Optional[dict] = None,
    compare_iter_pairs: Optional[List[Tuple[str, str]]] = None,
    compare_requests: Optional[List[Dict[str, object]]] = None,
) -> None:
    base_root = PLOTS_DIR if root_plot_dir is None else root_plot_dir
    target_dir = base_root / "compare_iteracoes_bl_vs_adtv"
    target_dir.mkdir(parents=True, exist_ok=True)
    mappings = mappings or {}
    compare_iter_pairs = compare_iter_pairs or _default_compare_iter_pairs()
    requests = compare_requests if compare_requests is not None else _build_default_compare_iter_requests(compare_iter_pairs)
    if not requests:
        print("[INFO] compare_iteracoes: nenhuma linha habilitada na aba Compare. Pulei compare_iteracoes/.")
        return

    requests_by_metric: Dict[str, List[Dict[str, object]]] = {}
    for request in requests:
        metric_id = _to_str_or_empty(request.get("metric_id", "")).lower()
        if metric_id not in COMPARE_ITER_METRIC_SPECS_BY_ID:
            print(f"[WARN] compare iteracoes: metrica invalida no request '{metric_id}'. Vou ignorar.")
            continue
        requests_by_metric.setdefault(metric_id, []).append(request)

    if not requests_by_metric:
        print("[WARN] compare iteracoes: requests sem metricas validas. Pulei compare_iteracoes/.")
        return

    export_rows: List[pd.DataFrame] = []
    generated_any = False

    for metric_id, metric_requests in requests_by_metric.items():
        spec = COMPARE_ITER_METRIC_SPECS_BY_ID[metric_id]
        value_name = spec["value_name"]
        if metric_id == "consumo":
            pts = _prepare_diesel_bl_adtv_points(df)
            agg = _aggregate_consumo_with_uncertainty(pts, ["_campaign_bl_adtv", "_sentido_plot", "Load_kW"]) if not pts.empty else pd.DataFrame()
            frames = _build_compare_iter_series_frames(agg, value_name="consumo_kg_h") if not agg.empty else {}
        else:
            pts = _prepare_compare_metric_points(df, metric_col=spec["metric_col"], mappings=mappings)
            agg = _aggregate_compare_metric_with_uncertainty(pts, value_name=value_name) if not pts.empty else pd.DataFrame()
            frames = _build_compare_iter_series_frames(agg, value_name=value_name) if not agg.empty else {}

        if not frames:
            print(f"[WARN] compare iteracoes: sem dados para a metrica '{metric_id}'.")
            continue

        for request in metric_requests:
            left_id = _to_str_or_empty(request.get("left_id", "")).lower()
            right_id = _to_str_or_empty(request.get("right_id", "")).lower()
            if left_id not in COMPARE_ITER_SERIES_META or right_id not in COMPARE_ITER_SERIES_META:
                continue
            left_df = frames.get(left_id)
            right_df = frames.get(right_id)
            if left_df is None or right_df is None:
                continue

            variant_key = _to_str_or_empty(request.get("variant_key", "with_uncertainty")).lower() or "with_uncertainty"
            uncertainty_mode = _to_str_or_empty(request.get("show_uncertainty", "on")).lower() or "on"
            include_uncertainty = uncertainty_mode != "off"
            dual_variant = bool(request.get("dual_variant", False))
            ctx = _compare_iter_pair_context(left_id, right_id)
            title_prefix = f"Compare {ctx['pair_title']}"
            filename_base = f"compare_iteracoes_{ctx['pair_slug']}_{spec['filename_slug']}"
            abs_filename, abs_title = _decorate_plot_variant_output(
                f"{filename_base}.png",
                f"{title_prefix} - {spec['title']}",
                variant_key,
                dual_variant,
            )
            abs_saved = _plot_compare_metric_absolute(
                left_df,
                right_df,
                value_name=value_name,
                y_label=spec["y_label"],
                title=abs_title,
                filename=abs_filename,
                target_dir=target_dir,
                label_left=ctx["left_label"],
                label_right=ctx["right_label"],
                include_uncertainty=include_uncertainty,
            )
            delta_filename, delta_title = _decorate_plot_variant_output(
                f"{filename_base}_delta_pct.png",
                f"{title_prefix} - Delta percentual",
                variant_key,
                dual_variant,
            )
            delta_df = _plot_compare_metric_delta_pct(
                left_df,
                right_df,
                value_name=value_name,
                title=delta_title,
                filename=delta_filename,
                target_dir=target_dir,
                label_left=ctx["left_label"],
                label_right=ctx["right_label"],
                label_line=ctx["line_label"],
                note_text=ctx["note_text"],
                interpret_neg=ctx["interpret_neg"],
                interpret_pos=ctx["interpret_pos"],
                include_uncertainty=include_uncertainty,
            )
            if abs_saved:
                generated_any = True
            if not delta_df.empty:
                generated_any = True
                delta_df = delta_df.copy()
                delta_df.insert(0, "Metrica", metric_id)
                delta_df.insert(1, "Comparacao", ctx["pair_slug"])
                delta_df.insert(2, "Incerteza", "com" if include_uncertainty else "sem")
                export_rows.append(delta_df)

    if not generated_any:
        print("[WARN] compare iteracoes: nao encontrei pares baseline/aditivado/descida/subida validos.")
        return
    _export_compare_iteracoes_metricas_excel(export_rows, target_dir)


def _plot_ethanol_equivalent_consumption_overlay(df: pd.DataFrame, *, plot_dir: Optional[Path] = None) -> None:
    target_dir = PLOTS_DIR if plot_dir is None else plot_dir
    target_dir.mkdir(parents=True, exist_ok=True)

    x_col = "UPD_Power_Bin_kW" if "UPD_Power_Bin_kW" in df.columns else ("UPD_Power_kW" if "UPD_Power_kW" in df.columns else None)
    y_col = "Fuel_E94H6_eq_kg_h"
    if x_col is None or y_col not in df.columns:
        print(
            "[WARN] Plot consumo equivalente EtOH: faltam colunas requeridas "
            f"(x={x_col}, y={y_col if y_col in df.columns else None}). Pulei."
        )
        return

    # E94H6 remains the measured consumption in this equivalent basis:
    # Fuel_E94H6_eq_kg_h = Fuel_mix_kg_h * (EtOH_pct/100) / 0.94.
    # For E94H6, EtOH_pct=94 -> equivalent equals measured.
    blend_specs = [
        ("E94H6", 94.0, 6.0),
        ("E75H25", 75.0, 25.0),
        ("E65H35", 65.0, 35.0),
    ]

    plt.figure()
    any_curve = False
    for label, etoh_pct, h2o_pct in blend_specs:
        m = _blend_mask(df, etoh_pct=etoh_pct, h2o_pct=h2o_pct)
        d = df[m].copy()
        d[x_col] = pd.to_numeric(d[x_col], errors="coerce")
        d[y_col] = pd.to_numeric(d[y_col], errors="coerce")
        d = d.dropna(subset=[x_col, y_col]).sort_values(x_col)
        if d.empty:
            print(f"[WARN] Plot consumo equivalente EtOH: sem dados para {label}.")
            continue
        any_curve = True
        plt.plot(d[x_col], d[y_col], "o-", label=label)

    if not any_curve:
        plt.close()
        print("[WARN] Plot consumo equivalente EtOH: nenhum blend alvo com dados. Pulei.")
        return

    plt.xlabel("Potencia UPD medida (kW, bin 0.1)")
    plt.ylabel("Consumo equivalente E94H6 (kg/h)")
    plt.title("Consumo equivalente de etanol vs potencia UPD (E94H6/E75H25/E65H35)")
    plt.grid(True, which="both", linestyle="--", linewidth=0.5)
    plt.legend()
    outpath = target_dir / "consumo_equiv_etanol_vs_upd_power_overlay.png"
    plt.tight_layout()
    plt.savefig(outpath, dpi=200)
    plt.close()
    print(f"[OK] Salvei {outpath}")


def _plot_ethanol_equivalent_ratio(df: pd.DataFrame, *, plot_dir: Optional[Path] = None) -> None:
    target_dir = PLOTS_DIR if plot_dir is None else plot_dir
    target_dir.mkdir(parents=True, exist_ok=True)

    y_col = "Fuel_E94H6_eq_kg_h"
    if y_col not in df.columns or "Load_kW" not in df.columns:
        print(
            "[WARN] Plot razao consumo equivalente EtOH: faltam colunas requeridas "
            f"(Load_kW={ 'Load_kW' in df.columns }, y={ y_col in df.columns }). Pulei."
        )
        return

    base = df[_blend_mask(df, etoh_pct=94.0, h2o_pct=6.0)].copy()
    base["Load_kW"] = pd.to_numeric(base["Load_kW"], errors="coerce")
    base["UPD_Power_Bin_kW"] = pd.to_numeric(base.get("UPD_Power_Bin_kW", pd.NA), errors="coerce")
    base[y_col] = pd.to_numeric(base[y_col], errors="coerce")
    base = base.dropna(subset=["Load_kW", y_col]).copy()
    if base.empty:
        print("[WARN] Plot razao consumo equivalente EtOH: sem dados base para E94H6. Pulei.")
        return

    plt.figure()
    any_curve = False
    ratio_specs = [
        ("E94H6 / E75H25", 75.0, 25.0, "ratio_pct_e94_over_e75"),
        ("E94H6 / E65H35", 65.0, 35.0, "ratio_pct_e94_over_e65"),
    ]

    for label, etoh_pct, h2o_pct, _ in ratio_specs:
        oth = df[_blend_mask(df, etoh_pct=etoh_pct, h2o_pct=h2o_pct)].copy()
        oth["Load_kW"] = pd.to_numeric(oth["Load_kW"], errors="coerce")
        oth[y_col] = pd.to_numeric(oth[y_col], errors="coerce")
        oth = oth.dropna(subset=["Load_kW", y_col]).copy()
        if oth.empty:
            print(f"[WARN] Plot razao consumo equivalente EtOH: sem dados para {label}.")
            continue

        merged = (
            base[["Load_kW", "UPD_Power_Bin_kW", y_col]]
            .rename(columns={y_col: "cons_eq_e94"})
            .merge(
                oth[["Load_kW", y_col]].rename(columns={y_col: "cons_eq_other"}),
                on="Load_kW",
                how="inner",
            )
        )
        merged["ratio_pct"] = 100.0 * merged["cons_eq_e94"] / merged["cons_eq_other"]
        merged["delta_pct"] = merged["ratio_pct"] - 100.0
        merged = merged.dropna(subset=["delta_pct"]).copy()
        if merged.empty:
            print(f"[WARN] Plot razao consumo equivalente EtOH: sem pares validos para {label}.")
            continue

        x = pd.to_numeric(merged["UPD_Power_Bin_kW"], errors="coerce")
        if x.notna().sum() == 0:
            x = pd.to_numeric(merged["Load_kW"], errors="coerce")
        merged = merged.assign(_x=x).dropna(subset=["_x"]).sort_values("_x")
        if merged.empty:
            print(f"[WARN] Plot razao consumo equivalente EtOH: sem eixo X valido para {label}.")
            continue

        any_curve = True
        plt.plot(merged["_x"], merged["delta_pct"], "o-", label=label)

    if not any_curve:
        plt.close()
        print("[WARN] Plot razao consumo equivalente EtOH: nenhum ratio valido. Pulei.")
        return

    plt.axhline(0.0, color="gray", linestyle="--", linewidth=1.0, label="0% (ref = 100%)")
    plt.xlabel("Potencia UPD medida (kW, bin 0.1)")
    plt.ylabel("Delta percentual de consumo equivalente (%)")
    plt.title("Delta percentual de consumo equivalente (ref=100%)")
    plt.grid(True, which="both", linestyle="--", linewidth=0.5)
    plt.legend()
    outpath = target_dir / "consumo_equiv_etanol_ratio_pct_vs_upd_power.png"
    plt.tight_layout()
    plt.savefig(outpath, dpi=200)
    plt.close()
    print(f"[OK] Salvei {outpath}")


def _plot_nth_e94h6_eq_flow(df: pd.DataFrame, *, plot_dir: Optional[Path] = None) -> None:
    target_dir = PLOTS_DIR if plot_dir is None else plot_dir
    target_dir.mkdir(parents=True, exist_ok=True)

    x_col = "UPD_Power_Bin_kW" if "UPD_Power_Bin_kW" in df.columns else ("UPD_Power_kW" if "UPD_Power_kW" in df.columns else None)
    y_col = "n_th_E94H6_eq_flow_pct"
    if x_col is None or y_col not in df.columns:
        print(
            "[WARN] Plot n_th_E94H6_eq_flow: faltam colunas requeridas "
            f"(x={x_col}, y={y_col if y_col in df.columns else None}). Pulei."
        )
        return

    blend_specs = [
        ("E94H6", 94.0, 6.0, "#1f77b4"),
        ("E75H25", 75.0, 25.0, "#ff7f0e"),
        ("E65H35", 65.0, 35.0, "#2ca02c"),
    ]

    plt.figure()
    any_curve = False
    for label, etoh_pct, h2o_pct, color in blend_specs:
        m = _blend_mask(df, etoh_pct=etoh_pct, h2o_pct=h2o_pct)
        d = df[m].copy()
        d[x_col] = pd.to_numeric(d[x_col], errors="coerce")
        d[y_col] = pd.to_numeric(d[y_col], errors="coerce")
        d = d.dropna(subset=[x_col, y_col]).sort_values(x_col)
        if d.empty:
            print(f"[WARN] Plot n_th_E94H6_eq_flow: sem dados para {label}.")
            continue
        any_curve = True
        plt.plot(d[x_col], d[y_col], "o-", label=label, color=color, linewidth=1.8, markersize=5)

    if not any_curve:
        plt.close()
        print("[WARN] Plot n_th_E94H6_eq_flow: nenhum blend alvo com dados. Pulei.")
        return

    plt.xlim(0.0, 55.0)
    plt.xticks(np.arange(0.0, 55.0 + 1e-12, 5.0).tolist())
    plt.xlabel("Potencia UPD medida (kW, bin 0.1)")
    plt.ylabel("Thermal efficiency (%)")
    plt.title("n_th_E94H6_eq_flow vs potencia UPD (all fuels)")
    plt.grid(True, which="both", linestyle="--", linewidth=0.5)
    _apply_y_tick_step(plt.gca(), 2.0)
    plt.legend()
    outpath = target_dir / "nth_e94h6_eq_flow_vs_upd_power_all.png"
    plt.tight_layout()
    plt.savefig(outpath, dpi=200)
    plt.close()
    print(f"[OK] Salvei {outpath}")


def _plot_nth_lhv_vs_eq6(df: pd.DataFrame, *, plot_dir: Optional[Path] = None) -> None:
    target_dir = PLOTS_DIR if plot_dir is None else plot_dir
    target_dir.mkdir(parents=True, exist_ok=True)

    x_col = "UPD_Power_Bin_kW" if "UPD_Power_Bin_kW" in df.columns else ("UPD_Power_kW" if "UPD_Power_kW" in df.columns else None)
    y_lhv = "n_th_pct"
    y_eq = "n_th_E94H6_eq_flow_pct"
    if x_col is None or y_lhv not in df.columns or y_eq not in df.columns:
        print(
            "[WARN] Plot comparacao 6 n_th: faltam colunas requeridas "
            f"(x={x_col}, y_lhv={y_lhv in df.columns}, y_eq={y_eq in df.columns}). Pulei."
        )
        return

    blend_specs = [
        ("E94H6", 94.0, 6.0, "#1f77b4"),
        ("E75H25", 75.0, 25.0, "#ff7f0e"),
        ("E65H35", 65.0, 35.0, "#2ca02c"),
    ]

    plt.figure()
    any_curve = False
    for label, etoh_pct, h2o_pct, color in blend_specs:
        m = _blend_mask(df, etoh_pct=etoh_pct, h2o_pct=h2o_pct)
        d = df[m].copy()
        d[x_col] = pd.to_numeric(d[x_col], errors="coerce")
        d[y_lhv] = pd.to_numeric(d[y_lhv], errors="coerce")
        d[y_eq] = pd.to_numeric(d[y_eq], errors="coerce")
        d = d.dropna(subset=[x_col]).sort_values(x_col)
        if d.empty:
            print(f"[WARN] Plot comparacao 6 n_th: sem dados para {label}.")
            continue

        d_lhv = d.dropna(subset=[y_lhv])
        if not d_lhv.empty:
            any_curve = True
            plt.plot(
                d_lhv[x_col],
                d_lhv[y_lhv],
                "o-",
                label=f"{label} | n_th_lhv",
                color=color,
                linewidth=1.8,
                markersize=5,
            )

        d_eq = d.dropna(subset=[y_eq])
        if not d_eq.empty:
            any_curve = True
            plt.plot(
                d_eq[x_col],
                d_eq[y_eq],
                "s--",
                label=f"{label} | n_th_E94H6_eq_flow",
                color=color,
                linewidth=1.8,
                markersize=4.5,
            )

    if not any_curve:
        plt.close()
        print("[WARN] Plot comparacao 6 n_th: nenhuma curva valida. Pulei.")
        return

    plt.xlim(0.0, 55.0)
    plt.xticks(np.arange(0.0, 55.0 + 1e-12, 5.0).tolist())
    plt.xlabel("Potencia UPD medida (kW, bin 0.1)")
    plt.ylabel("Thermal efficiency (%)")
    plt.title("Comparacao n_th: LHV da mistura vs E94H6 equivalente (6 curvas)")
    plt.grid(True, which="both", linestyle="--", linewidth=0.5)
    _apply_y_tick_step(plt.gca(), 2.0)
    plt.legend()
    outpath = target_dir / "nth_lhv_vs_e94h6_eq_flow_6curves.png"
    plt.tight_layout()
    plt.savefig(outpath, dpi=200)
    plt.close()
    print(f"[OK] Salvei {outpath}")


def plot_all_fuels(
    df: pd.DataFrame,
    y_col: str,
    yerr_col: Optional[str],
    title: str,
    filename: str,
    y_label: str,
    fixed_y: Optional[Tuple[float, float, float]] = None,
    fixed_y_limits: Optional[Tuple[float, float]] = None,
    y_tick_step: Optional[float] = None,
    fixed_x: Optional[Tuple[float, float, float]] = None,
    x_col: str = "Load_kW",
    x_label: str = "Power (kW)",
    fuels_override: Optional[List[int]] = None,
    series_col: Optional[str] = None,
    plot_dir: Optional[Path] = None,
    y_tol_plus: object = 0.0,
    y_tol_minus: object = 0.0,
) -> bool:
    target_dir = PLOTS_DIR if plot_dir is None else plot_dir
    target_dir.mkdir(parents=True, exist_ok=True)

    plt.figure()
    any_curve = False
    legend_entries = 0
    table_rows: List[Tuple[str, object, object]] = []
    table_rows: List[Tuple[str, object, object]] = []

    for label, d in _series_fuel_plot_groups(df, fuels_override=fuels_override, series_col=series_col):
        d[x_col] = pd.to_numeric(d[x_col], errors="coerce")
        d[y_col] = pd.to_numeric(d[y_col], errors="coerce")
        if yerr_col:
            d[yerr_col] = pd.to_numeric(d[yerr_col], errors="coerce")
            d = d.dropna(subset=[x_col, y_col, yerr_col]).sort_values(x_col)
        else:
            d = d.dropna(subset=[x_col, y_col]).sort_values(x_col)

        if d.empty:
            continue

        for xi, yi in zip(d[x_col].tolist(), d[y_col].tolist()):
            table_rows.append((label or "", xi, yi))

        any_curve = True
        if yerr_col:
            if label:
                plt.errorbar(d[x_col], d[y_col], yerr=d[yerr_col], fmt="o-", capsize=3, label=label)
                legend_entries += 1
            else:
                plt.errorbar(d[x_col], d[y_col], yerr=d[yerr_col], fmt="o-", capsize=3)
        else:
            if label:
                plt.plot(d[x_col], d[y_col], "o-", label=label)
                legend_entries += 1
            else:
                plt.plot(d[x_col], d[y_col], "o-")

    if not any_curve:
        plt.close()
        print(f"[WARN] Sem dados para plot {filename}")
        return False

    if fixed_x is not None:
        xmin, xmax, xstep = fixed_x
        plt.xlim(xmin, xmax)
        try:
            ticks = np.arange(xmin, xmax + 1e-12, xstep).tolist()
            plt.xticks(ticks)
        except Exception:
            pass

    if fixed_y is not None:
        ymin, ymax, ystep = fixed_y
        plt.ylim(ymin, ymax)
        try:
            ticks = np.arange(ymin, ymax + 1e-12, ystep).tolist()
            plt.yticks(ticks)
        except Exception:
            pass
    elif fixed_y_limits is not None:
        ymin, ymax = fixed_y_limits
        plt.ylim(ymin, ymax)

    ax = plt.gca()
    guide_entries = _add_y_tolerance_guides(ax, y_tol_plus=y_tol_plus, y_tol_minus=y_tol_minus)
    if fixed_y is None:
        _apply_y_tick_step(ax, y_tick_step)

    plt.xlabel(x_label)
    plt.ylabel(y_label)
    plt.title(title)
    plt.grid(True, which="both", linestyle="--", linewidth=0.5)
    _add_xy_value_table(ax, table_rows)
    if legend_entries > 0 or guide_entries > 0:
        plt.legend()
    outpath = target_dir / filename
    plt.tight_layout()
    plt.savefig(outpath, dpi=200)
    plt.close()
    print(f"[OK] Salvei {outpath}")
    return True


def plot_all_fuels_xy(
    df: pd.DataFrame,
    x_col: str,
    y_col: str,
    yerr_col: Optional[str],
    title: str,
    filename: str,
    x_label: str,
    y_label: str,
    fixed_y: Optional[Tuple[float, float, float]] = None,
    fixed_y_limits: Optional[Tuple[float, float]] = None,
    y_tick_step: Optional[float] = None,
    fixed_x: Optional[Tuple[float, float, float]] = None,
    fuels_override: Optional[List[int]] = None,
    series_col: Optional[str] = None,
    plot_dir: Optional[Path] = None,
    y_tol_plus: object = 0.0,
    y_tol_minus: object = 0.0,
) -> bool:
    target_dir = PLOTS_DIR if plot_dir is None else plot_dir
    target_dir.mkdir(parents=True, exist_ok=True)

    plt.figure()
    any_curve = False
    legend_entries = 0
    table_rows: List[Tuple[str, object, object]] = []

    for label, d in _series_fuel_plot_groups(df, fuels_override=fuels_override, series_col=series_col):
        d[x_col] = pd.to_numeric(d[x_col], errors="coerce")
        d[y_col] = pd.to_numeric(d[y_col], errors="coerce")
        if yerr_col:
            d[yerr_col] = pd.to_numeric(d[yerr_col], errors="coerce")
            d = d.dropna(subset=[x_col, y_col, yerr_col]).sort_values(x_col)
        else:
            d = d.dropna(subset=[x_col, y_col]).sort_values(x_col)

        if d.empty:
            continue

        for xi, yi in zip(d[x_col].tolist(), d[y_col].tolist()):
            table_rows.append((label or "", xi, yi))

        any_curve = True
        if yerr_col:
            if label:
                plt.errorbar(d[x_col], d[y_col], yerr=d[yerr_col], fmt="o-", capsize=3, label=label)
                legend_entries += 1
            else:
                plt.errorbar(d[x_col], d[y_col], yerr=d[yerr_col], fmt="o-", capsize=3)
        else:
            if label:
                plt.plot(d[x_col], d[y_col], "o-", label=label)
                legend_entries += 1
            else:
                plt.plot(d[x_col], d[y_col], "o-")

    if not any_curve:
        plt.close()
        print(f"[WARN] Sem dados para plot {filename}")
        return False

    if fixed_x is not None:
        xmin, xmax, xstep = fixed_x
        plt.xlim(xmin, xmax)
        try:
            ticks = np.arange(xmin, xmax + 1e-12, xstep).tolist()
            plt.xticks(ticks)
        except Exception:
            pass

    if fixed_y is not None:
        ymin, ymax, ystep = fixed_y
        plt.ylim(ymin, ymax)
        try:
            ticks = np.arange(ymin, ymax + 1e-12, ystep).tolist()
            plt.yticks(ticks)
        except Exception:
            pass
    elif fixed_y_limits is not None:
        ymin, ymax = fixed_y_limits
        plt.ylim(ymin, ymax)

    ax = plt.gca()
    guide_entries = _add_y_tolerance_guides(ax, y_tol_plus=y_tol_plus, y_tol_minus=y_tol_minus)
    if fixed_y is None:
        _apply_y_tick_step(ax, y_tick_step)

    plt.xlabel(x_label)
    plt.ylabel(y_label)
    plt.title(title)
    plt.grid(True, which="both", linestyle="--", linewidth=0.5)
    _add_xy_value_table(ax, table_rows)
    if legend_entries > 0 or guide_entries > 0:
        plt.legend()
    outpath = target_dir / filename
    plt.tight_layout()
    plt.savefig(outpath, dpi=200)
    plt.close()
    print(f"[OK] Salvei {outpath}")
    return True


def _annotate_points_variants(ax, x: np.ndarray, y: np.ndarray, variant: str) -> None:
    for xi, yi in zip(x, y):
        if not np.isfinite(xi) or not np.isfinite(yi):
            continue
        txt = f"{yi:.2f}"
        if variant == "box":
            ax.text(xi, yi, txt, fontsize=8, ha="left", va="bottom",
                    bbox=dict(boxstyle="round,pad=0.25", fc="white", ec="black", lw=0.6))
        elif variant == "tag":
            ax.annotate(txt, xy=(xi, yi), xytext=(6, 6), textcoords="offset points",
                        fontsize=8, ha="left", va="bottom",
                        bbox=dict(boxstyle="round,pad=0.25", fc="white", ec="black", lw=0.6),
                        arrowprops=dict(arrowstyle="->", lw=0.6))
        elif variant == "marker":
            ax.text(xi, yi, txt, fontsize=8, ha="center", va="bottom")
        elif variant == "badge":
            ax.text(xi, yi, txt, fontsize=8, ha="center", va="center",
                    bbox=dict(boxstyle="round,pad=0.22", fc="white", ec="black", lw=0.6, alpha=0.75))
        else:
            ax.text(xi, yi, txt, fontsize=8, ha="left", va="bottom")


def plot_all_fuels_with_value_labels(
    df: pd.DataFrame,
    y_col: str,
    title: str,
    filename: str,
    y_label: str,
    label_variant: str = "box",
    fixed_y: Optional[Tuple[float, float, float]] = None,
    fixed_y_limits: Optional[Tuple[float, float]] = None,
    y_tick_step: Optional[float] = None,
    fixed_x: Optional[Tuple[float, float, float]] = None,
    x_col: str = "Load_kW",
    x_label: str = "Power (kW)",
    fuels_override: Optional[List[int]] = None,
    series_col: Optional[str] = None,
    plot_dir: Optional[Path] = None,
    y_tol_plus: object = 0.0,
    y_tol_minus: object = 0.0,
) -> bool:
    target_dir = PLOTS_DIR if plot_dir is None else plot_dir
    target_dir.mkdir(parents=True, exist_ok=True)

    fig, ax = plt.subplots()
    any_curve = False
    legend_entries = 0
    table_rows: List[Tuple[str, object, object]] = []

    for label, d in _series_fuel_plot_groups(df, fuels_override=fuels_override, series_col=series_col):
        d[x_col] = pd.to_numeric(d[x_col], errors="coerce")
        d[y_col] = pd.to_numeric(d[y_col], errors="coerce")
        d = d.dropna(subset=[x_col, y_col]).sort_values(x_col)

        if d.empty:
            continue

        for xi, yi in zip(d[x_col].tolist(), d[y_col].tolist()):
            table_rows.append((label or "", xi, yi))

        any_curve = True
        if label:
            ax.plot(d[x_col], d[y_col], "o-", label=label)
            legend_entries += 1
        else:
            ax.plot(d[x_col], d[y_col], "o-")

        x = pd.to_numeric(d[x_col], errors="coerce").values.astype(float)
        y = pd.to_numeric(d[y_col], errors="coerce").values.astype(float)
        _annotate_points_variants(ax, x, y, label_variant)

    if not any_curve:
        plt.close(fig)
        print(f"[WARN] Sem dados para plot {filename}")
        return False

    if fixed_x is not None:
        xmin, xmax, xstep = fixed_x
        ax.set_xlim(xmin, xmax)
        try:
            ticks = np.arange(xmin, xmax + 1e-12, xstep).tolist()
            ax.set_xticks(ticks)
        except Exception:
            pass

    if fixed_y is not None:
        ymin, ymax, ystep = fixed_y
        ax.set_ylim(ymin, ymax)
        try:
            ticks = np.arange(ymin, ymax + 1e-12, ystep).tolist()
            ax.set_yticks(ticks)
        except Exception:
            pass
    elif fixed_y_limits is not None:
        ymin, ymax = fixed_y_limits
        ax.set_ylim(ymin, ymax)

    guide_entries = _add_y_tolerance_guides(ax, y_tol_plus=y_tol_plus, y_tol_minus=y_tol_minus)
    if fixed_y is None:
        _apply_y_tick_step(ax, y_tick_step)

    ax.set_xlabel(x_label)
    ax.set_ylabel(y_label)
    ax.set_title(title)
    ax.grid(True, which="both", linestyle="--", linewidth=0.5)
    _add_xy_value_table(ax, table_rows)
    if legend_entries > 0 or guide_entries > 0:
        ax.legend()

    outpath = target_dir / filename
    fig.tight_layout()
    fig.savefig(outpath, dpi=220)
    plt.close(fig)
    print(f"[OK] Salvei {outpath}")
    return True


def _prepare_machine_scenario_plot_df(df: pd.DataFrame) -> Tuple[pd.DataFrame, Optional[str], str]:
    if df is None or df.empty:
        return pd.DataFrame(), None, ""

    x_candidates: List[Tuple[str, bool]] = [
        ("UPD_Power_Bin_kW", False),
        ("UPD_Power_kW", False),
    ]
    x_col_base, mestrado_x_override = _resolve_plot_x_request("Load_kW")
    if x_col_base not in {"UPD_Power_Bin_kW", "UPD_Power_kW"}:
        x_candidates.append((x_col_base, mestrado_x_override))
    if "Load_kW" != x_col_base:
        x_candidates.append(("Load_kW", False))

    x_col = None
    x_label = ""
    for requested, is_runtime_override in x_candidates:
        try:
            x_col = resolve_col(df, requested)
            if requested == "UPD_Power_Bin_kW" or x_col == "UPD_Power_Bin_kW":
                x_label = "Potencia UPD medida (kW, bin 0.1)"
            elif requested == "UPD_Power_kW" or x_col == "UPD_Power_kW":
                x_label = "Potencia UPD medida (kW)"
            else:
                x_label = _runtime_plot_x_label("", "Load_kW", x_col, is_runtime_override)
            break
        except Exception:
            continue

    if x_col is None:
        return pd.DataFrame(), None, ""

    if not x_label:
        x_label = x_col

    fuel_labels = df.get("Fuel_Label", pd.Series(pd.NA, index=df.index, dtype="object"))
    fuel_labels = fuel_labels.where(fuel_labels.notna(), _fuel_blend_labels(df))
    out = df[fuel_labels.eq(SCENARIO_REFERENCE_FUEL_LABEL)].copy()
    if out.empty:
        return pd.DataFrame(), x_col, x_label

    out[x_col] = pd.to_numeric(out[x_col], errors="coerce")
    out = out.dropna(subset=[x_col]).sort_values(x_col)
    return out, x_col, x_label


def _machine_scaled_tick_formatter(divisor: float) -> FuncFormatter:
    return FuncFormatter(lambda value, _pos: f"{(value / divisor):g}")


def _reserve_upper_legend_headroom(ax, *, ratio: float = 0.32) -> None:
    try:
        ymin, ymax = ax.get_ylim()
    except Exception:
        return

    if not (np.isfinite(ymin) and np.isfinite(ymax)):
        return

    span = ymax - ymin
    if not np.isfinite(span) or span <= 0:
        span = max(abs(ymax), abs(ymin), 1.0)

    ax.set_ylim(ymin, ymax + span * ratio)


def _style_machine_scenario_axes(
    fig,
    ax,
    *,
    title: str,
    x_label: str,
    y_label: str,
    y_tick_divisor: Optional[float] = None,
) -> None:
    ax.set_xlim(0.0, 55.0)
    ax.set_xticks(np.arange(0.0, 55.0 + 1e-12, 5.0).tolist())
    ax.set_xlabel(x_label)
    ax.set_ylabel(y_label)
    ax.set_title(title)
    ax.grid(True, which="both", linestyle="--", linewidth=0.5)
    if y_tick_divisor is not None and np.isfinite(y_tick_divisor) and y_tick_divisor > 0 and y_tick_divisor != 1.0:
        ax.yaxis.set_major_formatter(_machine_scaled_tick_formatter(float(y_tick_divisor)))

    handles, labels = ax.get_legend_handles_labels()
    if handles:
        _reserve_upper_legend_headroom(ax)
        ax.legend(
            loc="upper left",
            frameon=True,
        )
    fig.tight_layout()


def _plot_machine_scenario_dual_metric(
    df: pd.DataFrame,
    *,
    diesel_suffix: str,
    ethanol_suffix: str,
    ethanol_u_suffix: Optional[str],
    title: str,
    filename: str,
    y_label: str,
    plot_dir: Optional[Path] = None,
    y_tick_divisor: Optional[float] = None,
) -> None:
    target_dir = PLOTS_DIR if plot_dir is None else plot_dir
    target_dir.mkdir(parents=True, exist_ok=True)

    plot_df, x_col, x_label = _prepare_machine_scenario_plot_df(df)
    if x_col is None or plot_df.empty:
        print(f"[WARN] Sem dados {SCENARIO_REFERENCE_FUEL_LABEL} para plot de cenario {filename}.")
        return

    fig, ax = plt.subplots()
    any_curve = False
    for spec in MACHINE_SCENARIO_SPECS:
        diesel_col = _scenario_machine_col(spec["key"], diesel_suffix)
        ethanol_col = _scenario_machine_col(spec["key"], ethanol_suffix)
        ethanol_u_col = _scenario_machine_col(spec["key"], ethanol_u_suffix) if ethanol_u_suffix else None

        if diesel_col in plot_df.columns:
            d_diesel = plot_df[[x_col, diesel_col]].copy()
            d_diesel[diesel_col] = pd.to_numeric(d_diesel[diesel_col], errors="coerce")
            d_diesel = d_diesel.dropna(subset=[x_col, diesel_col]).sort_values(x_col)
            if not d_diesel.empty:
                any_curve = True
                ax.plot(
                    d_diesel[x_col],
                    d_diesel[diesel_col],
                    "o--",
                    color=spec["color"],
                    linewidth=1.8,
                    markersize=4.5,
                    label=f"{spec['label']} diesel",
                )

        if ethanol_col in plot_df.columns:
            cols = [x_col, ethanol_col]
            if ethanol_u_col and ethanol_u_col in plot_df.columns:
                cols.append(ethanol_u_col)
            d_eth = plot_df[cols].copy()
            d_eth[ethanol_col] = pd.to_numeric(d_eth[ethanol_col], errors="coerce")
            if ethanol_u_col and ethanol_u_col in d_eth.columns:
                d_eth[ethanol_u_col] = pd.to_numeric(d_eth[ethanol_u_col], errors="coerce")
            d_eth = d_eth.dropna(subset=[x_col, ethanol_col]).sort_values(x_col)
            if d_eth.empty:
                continue

            any_curve = True
            if ethanol_u_col and ethanol_u_col in d_eth.columns and d_eth[ethanol_u_col].notna().any():
                ax.errorbar(
                    d_eth[x_col],
                    d_eth[ethanol_col],
                    yerr=d_eth[ethanol_u_col],
                    fmt="o-",
                    capsize=3,
                    color=spec["color"],
                    linewidth=1.8,
                    markersize=4.5,
                    label=f"{spec['label']} {SCENARIO_REFERENCE_FUEL_LABEL}",
                )
            else:
                ax.plot(
                    d_eth[x_col],
                    d_eth[ethanol_col],
                    "o-",
                    color=spec["color"],
                    linewidth=1.8,
                    markersize=4.5,
                    label=f"{spec['label']} {SCENARIO_REFERENCE_FUEL_LABEL}",
                )

    if not any_curve:
        plt.close(fig)
        print(f"[WARN] Cenario {filename}: nenhuma curva valida.")
        return

    _style_machine_scenario_axes(
        fig,
        ax,
        title=title,
        x_label=x_label,
        y_label=y_label,
        y_tick_divisor=y_tick_divisor,
    )
    outpath = target_dir / filename
    fig.savefig(outpath, dpi=200)
    plt.close(fig)
    print(f"[OK] Salvei {outpath}")


def _plot_machine_scenario_single_metric(
    df: pd.DataFrame,
    *,
    value_suffix: str,
    u_suffix: Optional[str],
    title: str,
    filename: str,
    y_label: str,
    plot_dir: Optional[Path] = None,
    y_tick_divisor: Optional[float] = None,
) -> None:
    target_dir = PLOTS_DIR if plot_dir is None else plot_dir
    target_dir.mkdir(parents=True, exist_ok=True)

    plot_df, x_col, x_label = _prepare_machine_scenario_plot_df(df)
    if x_col is None or plot_df.empty:
        print(f"[WARN] Sem dados {SCENARIO_REFERENCE_FUEL_LABEL} para plot de cenario {filename}.")
        return

    fig, ax = plt.subplots()
    any_curve = False
    for spec in MACHINE_SCENARIO_SPECS:
        value_col = _scenario_machine_col(spec["key"], value_suffix)
        u_col = _scenario_machine_col(spec["key"], u_suffix) if u_suffix else None
        cols = [x_col, value_col]
        if u_col and u_col in plot_df.columns:
            cols.append(u_col)

        if value_col not in plot_df.columns:
            continue

        d = plot_df[cols].copy()
        d[value_col] = pd.to_numeric(d[value_col], errors="coerce")
        if u_col and u_col in d.columns:
            d[u_col] = pd.to_numeric(d[u_col], errors="coerce")
        d = d.dropna(subset=[x_col, value_col]).sort_values(x_col)
        if d.empty:
            continue

        any_curve = True
        if u_col and u_col in d.columns and d[u_col].notna().any():
            ax.errorbar(
                d[x_col],
                d[value_col],
                yerr=d[u_col],
                fmt="o-",
                capsize=3,
                color=spec["color"],
                linewidth=1.8,
                markersize=4.5,
                label=spec["label"],
            )
        else:
            ax.plot(
                d[x_col],
                d[value_col],
                "o-",
                color=spec["color"],
                linewidth=1.8,
                markersize=4.5,
                label=spec["label"],
            )

    if not any_curve:
        plt.close(fig)
        print(f"[WARN] Cenario {filename}: nenhuma curva valida.")
        return

    _style_machine_scenario_axes(
        fig,
        ax,
        title=title,
        x_label=x_label,
        y_label=y_label,
        y_tick_divisor=y_tick_divisor,
    )
    outpath = target_dir / filename
    fig.savefig(outpath, dpi=200)
    plt.close(fig)
    print(f"[OK] Salvei {outpath}")


def _plot_machine_scenario_suite(df: pd.DataFrame, *, plot_dir: Optional[Path] = None) -> None:
    _plot_machine_scenario_dual_metric(
        df,
        diesel_suffix="Diesel_Custo_R_h",
        ethanol_suffix="E94H6_Custo_R_h",
        ethanol_u_suffix="U_E94H6_Custo_R_h",
        title="Cenario de maquinas: custo horario diesel vs E94H6",
        filename="scenario_maquinas_custo_r_h_diesel_vs_e94h6.png",
        y_label="Custo horario (R$/h)",
        plot_dir=plot_dir,
    )
    _plot_machine_scenario_single_metric(
        df,
        value_suffix="Economia_R_h",
        u_suffix="U_Economia_R_h",
        title="Cenario de maquinas: economia horaria vs diesel (negativo = economia)",
        filename="scenario_maquinas_economia_r_h_vs_diesel.png",
        y_label="Delta de custo vs diesel (R$/h)",
        plot_dir=plot_dir,
    )
    _plot_machine_scenario_dual_metric(
        df,
        diesel_suffix="Diesel_L_h",
        ethanol_suffix="E94H6_L_h",
        ethanol_u_suffix="U_E94H6_L_h",
        title="Cenario de maquinas: consumo volumetrico diesel vs E94H6",
        filename="scenario_maquinas_consumo_l_h_diesel_vs_e94h6.png",
        y_label="Consumo volumetrico (L/h)",
        plot_dir=plot_dir,
    )
    _plot_machine_scenario_single_metric(
        df,
        value_suffix="E94H6_L_ano",
        u_suffix="U_E94H6_L_ano",
        title="Cenario de maquinas: consumo anual de E94H6",
        filename="scenario_maquinas_consumo_anual_e94h6_l.png",
        y_label="Consumo anual de E94H6 (x10^3 L/ano)",
        plot_dir=plot_dir,
        y_tick_divisor=1000.0,
    )
    _plot_machine_scenario_dual_metric(
        df,
        diesel_suffix="Diesel_Custo_R_ano",
        ethanol_suffix="E94H6_Custo_R_ano",
        ethanol_u_suffix="U_E94H6_Custo_R_ano",
        title="Cenario de maquinas: custo anual diesel vs E94H6",
        filename="scenario_maquinas_custo_anual_diesel_vs_e94h6.png",
        y_label="Custo anual (x10^3 R$/ano)",
        plot_dir=plot_dir,
        y_tick_divisor=1000.0,
    )
    _plot_machine_scenario_single_metric(
        df,
        value_suffix="Economia_R_ano",
        u_suffix="U_Economia_R_ano",
        title="Cenario de maquinas: economia anual vs diesel (negativo = economia)",
        filename="scenario_maquinas_economia_anual_vs_diesel.png",
        y_label="Delta de custo anual vs diesel (x10^3 R$/ano)",
        plot_dir=plot_dir,
        y_tick_divisor=1000.0,
    )


# =========================
# Plots-config dispatcher
# =========================
def _row_enabled(v: object) -> bool:
    if v is None:
        return False
    try:
        if pd.isna(v):
            return False
    except Exception:
        pass
    s = str(v).strip().lower()
    if s in {"1", "true", "yes", "y", "on"}:
        return True
    try:
        return bool(int(float(s)))
    except Exception:
        return False


def _yerr_disabled_token(s: str) -> bool:
    t = str(s or "").strip().lower()
    return t in {"none", "off", "disable", "disabled", "0", "na", "n/a"}


def _plot_uncertainty_mode(v: object) -> str:
    text = _to_str_or_empty(v).lower()
    if not text or text in {"auto", "guess", "default"}:
        return "auto"
    if text in {"0", "false", "no", "off", "disable", "disabled", "none", "na", "n/a"}:
        return "off"
    return "on"


def _plot_uncertainty_flags(row: pd.Series) -> Tuple[bool, bool]:
    with_raw = _to_str_or_empty(row.get("with_uncertainty", "")).lower()
    without_raw = _to_str_or_empty(row.get("without_uncertainty", "")).lower()
    with_flag = with_raw in {"1", "true", "yes", "on", "y", "checked"}
    without_flag = without_raw in {"1", "true", "yes", "on", "y", "checked"}
    with_defined = with_raw in {"1", "0", "true", "false", "yes", "no", "on", "off", "y", "n", "checked", "unchecked"}
    without_defined = without_raw in {"1", "0", "true", "false", "yes", "no", "on", "off", "y", "n", "checked", "unchecked"}

    if not with_defined and not without_defined:
        mode = _to_str_or_empty(row.get("show_uncertainty", "auto")).lower()
        if mode in {"off", "disable", "disabled", "none", "0", "false", "no", "na", "n/a"}:
            return False, True
        if mode in {"both", "all", "dual", "on_off"}:
            return True, True
        return True, False

    if not with_flag and not without_flag:
        return True, False
    return with_flag, without_flag


def _plot_uncertainty_variants(row: pd.Series) -> List[Tuple[str, str, bool]]:
    with_flag, without_flag = _plot_uncertainty_flags(row)
    both_selected = with_flag and without_flag
    variants: List[Tuple[str, str, bool]] = []
    if with_flag:
        variants.append(("with_uncertainty", "on", both_selected))
    if without_flag:
        variants.append(("without_uncertainty", "off", both_selected))
    if not variants:
        variants.append(("with_uncertainty", "on", False))
    return variants


def _decorate_plot_variant_output(filename: str, title: str, variant_key: str, dual_variant: bool) -> Tuple[str, str]:
    if not dual_variant:
        return filename, title

    filename_suffix = "with_uncertainty" if variant_key == "with_uncertainty" else "without_uncertainty"
    title_suffix = "with uncertainty" if variant_key == "with_uncertainty" else "without uncertainty"

    fn = _strip_leading_raw_plot_name(filename)
    if fn.lower().endswith(".png"):
        fn = f"{fn[:-4]}_{filename_suffix}.png"
    elif fn:
        fn = f"{fn}_{filename_suffix}"

    tt = _strip_leading_raw_plot_name(title)
    if tt:
        tt = f"{tt} | {title_suffix}"

    return fn, tt


def _resolve_plot_yerr_col(
    out_df: pd.DataFrame,
    row: pd.Series,
    *,
    y_col: str,
    mappings: dict,
    plot_label: str,
) -> Optional[str]:
    yerr_req = _to_str_or_empty(row.get("yerr_col", ""))
    uncertainty_mode = _plot_uncertainty_mode(row.get("show_uncertainty", "auto"))
    if uncertainty_mode == "off":
        return None

    if yerr_req and not _yerr_disabled_token(yerr_req):
        try:
            return resolve_col(out_df, yerr_req)
        except Exception:
            print(f"[INFO] Plot '{plot_label}': yerr_col '{yerr_req}' nao encontrado. Vou tentar fallback.")

    guessed = _guess_plot_uncertainty_col(out_df, y_col, mappings)
    if guessed:
        print(f"[INFO] Plot '{plot_label}': usando '{guessed}' como incerteza final.")
        return guessed

    if yerr_req and not _yerr_disabled_token(yerr_req):
        print(f"[INFO] Plot '{plot_label}': fallback sem yerr, porque '{yerr_req}' nao existe no output.")
    return None


def _shared_plot_y_limits_for_variants(
    df: pd.DataFrame,
    *,
    x_col: str,
    y_col: str,
    variant_yerr_cols: List[Optional[str]],
    fuels_override: Optional[List[int]] = None,
    series_col: Optional[str] = None,
    y_tol_plus: object = 0.0,
    y_tol_minus: object = 0.0,
) -> Optional[Tuple[float, float]]:
    values: List[float] = []

    for yerr_col in variant_yerr_cols:
        for _, d in _series_fuel_plot_groups(df, fuels_override=fuels_override, series_col=series_col):
            work = d.copy()
            work[x_col] = pd.to_numeric(work[x_col], errors="coerce")
            work[y_col] = pd.to_numeric(work[y_col], errors="coerce")
            if yerr_col:
                work[yerr_col] = pd.to_numeric(work[yerr_col], errors="coerce")
                work = work.dropna(subset=[x_col, y_col, yerr_col]).sort_values(x_col)
            else:
                work = work.dropna(subset=[x_col, y_col]).sort_values(x_col)

            if work.empty:
                continue

            y = pd.to_numeric(work[y_col], errors="coerce")
            values.extend(float(v) for v in y.dropna().tolist() if np.isfinite(v))
            if yerr_col:
                yerr = pd.to_numeric(work[yerr_col], errors="coerce").abs()
                low = y - yerr
                high = y + yerr
                values.extend(float(v) for v in low.dropna().tolist() if np.isfinite(v))
                values.extend(float(v) for v in high.dropna().tolist() if np.isfinite(v))

    tp = _normalize_tol_value(y_tol_plus)
    tm = _normalize_tol_value(y_tol_minus)
    if tp > 0:
        values.append(float(tp))
    if tm > 0:
        values.append(float(-tm))

    finite_values = [float(v) for v in values if np.isfinite(v)]
    if not finite_values:
        return None

    ymin = min(finite_values)
    ymax = max(finite_values)
    if ymax <= ymin:
        span_ref = max(abs(ymax), abs(ymin), 1.0)
        pad = span_ref * 0.05
    else:
        pad = (ymax - ymin) * 0.05
    return ymin - pad, ymax + pad


def _new_plot_run_summary() -> Dict[str, object]:
    return {
        "generated": 0,
        "skipped": 0,
        "disabled": 0,
        "generated_files": [],
        "generated_labels": [],
        "skipped_items": [],
    }


def _merge_plot_run_summary(total: Dict[str, object], partial: Optional[Dict[str, object]]) -> Dict[str, object]:
    if not partial:
        return total
    total["generated"] = int(total.get("generated", 0)) + int(partial.get("generated", 0))
    total["skipped"] = int(total.get("skipped", 0)) + int(partial.get("skipped", 0))
    total["disabled"] = int(total.get("disabled", 0)) + int(partial.get("disabled", 0))
    total.setdefault("generated_files", []).extend(partial.get("generated_files", []))
    total.setdefault("generated_labels", []).extend(partial.get("generated_labels", []))
    total.setdefault("skipped_items", []).extend(partial.get("skipped_items", []))
    return total


def _strip_leading_raw_plot_name(value: object) -> str:
    text = _to_str_or_empty(value)
    if text.lower().startswith("raw_"):
        return text[4:]
    return text


def _derive_filename_for_expansion(template: str, y_col: str) -> str:
    t = _strip_leading_raw_plot_name(template)
    if not t:
        return f"kibox_{_safe_name(y_col)}_vs_power_all.png"
    if "{y}" in t:
        return t.replace("{y}", _safe_name(y_col))
    if t.lower().endswith(".png"):
        stem = t[:-4]
        return f"{stem}_{_safe_name(y_col)}.png"
    return f"{t}_{_safe_name(y_col)}.png"


def _derive_title_for_expansion(template: str, x_col: str, y_col: str) -> str:
    t = _strip_leading_raw_plot_name(template)
    if not t:
        return f"{y_col} vs {x_col} (all fuels)"
    if "{y}" in t or "{x}" in t:
        return t.replace("{y}", y_col).replace("{x}", x_col)
    return t


def _resolve_plot_x_request(x_col_req: str) -> Tuple[str, bool]:
    req = _to_str_or_empty(x_col_req)
    req_norm = norm_key(req) if req else ""
    load_norm = norm_key("Load_kW")
    if is_mestrado_runtime() and (not req or req_norm == load_norm):
        return "UPD_Power_Bin_kW", True
    return ("Load_kW" if not req else req), False


def _runtime_plot_x_label(
    x_label: str,
    x_col_base: str,
    x_col_resolved: str,
    mestrado_override: bool,
) -> str:
    label = _to_str_or_empty(x_label)
    if mestrado_override:
        label_norm = norm_key(label) if label else ""
        x_base_norm = norm_key(x_col_base)
        if not label or label_norm in {
            x_base_norm,
            norm_key("Load_kW"),
            norm_key("Carga (kW)"),
            norm_key("Power (kW)"),
            norm_key("Power"),
            norm_key("Potencia (kW)"),
        }:
            return "Potencia UPD medida (kW, bin 0.1)"
    return label if label else x_col_resolved


def make_plots_from_config(
    out_df: pd.DataFrame,
    plots_df: pd.DataFrame,
    mappings: dict,
    plot_dir: Optional[Path] = None,
    series_col: Optional[str] = None,
) -> Dict[str, object]:
    """
    Config-driven plotting (rev3):
      - Each row defines one plot.
      - plot_type supports:
          * all_fuels_yx      -> plot_all_fuels
          * all_fuels_xy      -> plot_all_fuels_xy
          * all_fuels_labels  -> plot_all_fuels_with_value_labels
          * kibox_all         -> expands into one plot per KIBOX_* column (except KIBOX_N_files)
    Notes:
      - Empty cells from Excel come as NaN; we treat them as empty (no more 'nan' column lookup).
      - Missing yerr_col: INFO (plot still works).
      - Missing required columns (y_col or x_col when required): ERROR and skip that plot.
    """
    if plots_df is None or plots_df.empty:
        print("[WARN] Plots config vazio; nÃ£o gerei plots via planilha.")
        return

    n_ok = 0
    n_skip = 0
    n_disabled = 0
    generated_files: List[str] = []
    generated_labels: List[str] = []
    skipped_items: List[Tuple[str, str]] = []

    for _, r in plots_df.iterrows():
        if not _row_enabled(r.get("enabled", 0)):
            n_disabled += 1
            continue

        plot_type = _to_str_or_empty(r.get("plot_type", ""))
        filename = _strip_leading_raw_plot_name(r.get("filename", ""))
        title = _strip_leading_raw_plot_name(r.get("title", ""))
        plot_label = filename or title or _to_str_or_empty(r.get("y_col", "")) or _to_str_or_empty(r.get("plot_type", "")) or "plot_sem_nome"

        if not plot_type:
            print("[ERROR] Plots row invÃ¡lida: plot_type vazio. Pulei.")
            n_skip += 1
            continue

        x_col_req = _to_str_or_empty(r.get("x_col", ""))
        y_col_req = _to_str_or_empty(r.get("y_col", ""))

        x_label = _to_str_or_empty(r.get("x_label", ""))
        y_label = _to_str_or_empty(r.get("y_label", ""))

        y_axis_unit = _mapping_unit_for_y_col(y_col_req, mappings)
        fixed_x = _parse_axis_spec(r.get("x_min", pd.NA), r.get("x_max", pd.NA), r.get("x_step", pd.NA))
        fixed_y = _parse_axis_spec(
            r.get("y_min", pd.NA),
            r.get("y_max", pd.NA),
            r.get("y_step", pd.NA),
            target_unit=y_axis_unit,
        )
        fixed_y_limits = _parse_axis_limits(
            r.get("y_min", pd.NA),
            r.get("y_max", pd.NA),
            target_unit=y_axis_unit,
        )
        y_tick_step = _parse_axis_value(r.get("y_step", pd.NA), target_unit=y_axis_unit, default=np.nan)
        if not np.isfinite(y_tick_step) or y_tick_step <= 0:
            y_tick_step = None
        if fixed_y is not None:
            y_tick_step = None
        y_tol_plus = _to_float(r.get("y_tol_plus", r.get("tol_plus", 0.0)), 0.0)
        y_tol_minus = _to_float(r.get("y_tol_minus", r.get("tol_minus", 0.0)), 0.0)

        fuels = _parse_csv_list_ints(r.get("filter_h2o_list", pd.NA))
        fuels_override = fuels if fuels is not None else None

        label_variant = _to_str_or_empty(r.get("label_variant", "box")).lower() or "box"

        pt = plot_type.lower().strip()

        # ---------
        # Expansion: Kibox (all columns)
        # ---------
        if pt in {"kibox_all", "all_kibox"}:
            kibox_cols = [c for c in out_df.columns if str(c).startswith("KIBOX_") and c != "KIBOX_N_files"]
            if not kibox_cols:
                print("[WARN] kibox_all: nÃ£o hÃ¡ colunas KIBOX_* no output. Pulei expansÃ£o.")
                n_skip += 1
                continue

            # x column default for kibox_all
            x_col_base, mestrado_x_override = _resolve_plot_x_request(x_col_req)
            try:
                x_col = resolve_col(out_df, x_col_base)
            except Exception as e:
                print(f"[ERROR] kibox_all: x_col '{x_col_base}' nÃ£o encontrado. Pulei expansÃ£o. ({e})")
                n_skip += 1
                continue

            xlab = _runtime_plot_x_label(x_label, x_col_base, x_col, mestrado_x_override)
            seen_filenames: set[str] = set()

            for yc in sorted(kibox_cols):
                fn = _derive_filename_for_expansion(filename, yc)
                fn_key = norm_key(fn)
                if fn_key in seen_filenames:
                    print(f"[INFO] kibox_all: filename duplicado apos normalizacao ('{fn}'). Pulei a expansao de '{yc}'.")
                    continue
                seen_filenames.add(fn_key)
                tt = _derive_title_for_expansion(title, x_col=x_col, y_col=yc)
                ylab = y_label if y_label else yc

                plot_all_fuels(
                    out_df,
                    y_col=yc,
                    yerr_col=None,
                title=tt,
                filename=fn,
                y_label=ylab,
                fixed_y=fixed_y,
                fixed_y_limits=fixed_y_limits,
                y_tick_step=y_tick_step,
                fixed_x=fixed_x,
                x_col=x_col,
                    x_label=xlab,
                    fuels_override=fuels_override,
                    series_col=series_col,
                    plot_dir=plot_dir,
                    y_tol_plus=y_tol_plus,
                    y_tol_minus=y_tol_minus,
                )
                n_ok += 1
            continue

        # ---------
        # Normal plots (one output per row)
        # ---------
        if pt in {"all_fuels_yx", "all_fuels", "all_fuels_y_vs_x"}:
            x_col_base, mestrado_x_override = _resolve_plot_x_request(x_col_req)
            try:
                x_col = resolve_col(out_df, x_col_base)
            except Exception as e:
                print(f"[ERROR] Plot '{filename or title}': x_col '{x_col_base}' nÃ£o encontrado. Pulei. ({e})")
                n_skip += 1
                continue

            if not y_col_req:
                print(f"[ERROR] Plot '{filename or title}': y_col vazio. Pulei.")
                n_skip += 1
                continue
            try:
                y_col = resolve_col(out_df, y_col_req)
            except Exception as e:
                print(f"[ERROR] Plot '{filename or title}': y_col '{y_col_req}' nÃ£o encontrado. Pulei. ({e})")
                n_skip += 1
                continue

            x_label = _runtime_plot_x_label(x_label, x_col_base, x_col, mestrado_x_override)
            if not y_label:
                y_label = y_col
            if not title:
                title = f"{y_col} vs {x_col} (all fuels)"
            if not filename:
                filename = f"{_safe_name(y_col)}_vs_{_safe_name(x_col)}_all.png"

            for variant_key, uncertainty_mode, dual_variant in _plot_uncertainty_variants(r):
                variant_row = r.copy()
                variant_row["show_uncertainty"] = uncertainty_mode
                variant_filename, variant_title = _decorate_plot_variant_output(filename, title, variant_key, dual_variant)
                yerr_col = _resolve_plot_yerr_col(
                    out_df,
                    variant_row,
                    y_col=y_col,
                    mappings=mappings,
                    plot_label=variant_filename or variant_title or y_col,
                )

                plot_all_fuels(
                    out_df,
                    y_col=y_col,
                    yerr_col=yerr_col,
                    title=variant_title,
                    filename=variant_filename,
                    y_label=y_label,
                    fixed_y=fixed_y,
                    fixed_y_limits=fixed_y_limits,
                    y_tick_step=y_tick_step,
                    fixed_x=fixed_x,
                    x_col=x_col,
                    x_label=x_label,
                    fuels_override=fuels_override,
                    series_col=series_col,
                    plot_dir=plot_dir,
                    y_tol_plus=y_tol_plus,
                    y_tol_minus=y_tol_minus,
                )
                n_ok += 1
            continue

        if pt in {"all_fuels_xy", "xy"}:
            if not y_col_req:
                print(f"[ERROR] Plot '{filename or title}': y_col vazio (plot_type=all_fuels_xy). Pulei.")
                n_skip += 1
                continue

            x_col_base, mestrado_x_override = _resolve_plot_x_request(x_col_req)
            try:
                x_col = resolve_col(out_df, x_col_base)
            except Exception as e:
                print(f"[ERROR] Plot '{filename or title}': x_col '{x_col_base}' nÃ£o encontrado. Pulei. ({e})")
                n_skip += 1
                continue

            try:
                y_col = resolve_col(out_df, y_col_req)
            except Exception as e:
                print(f"[ERROR] Plot '{filename or title}': y_col '{y_col_req}' nÃ£o encontrado. Pulei. ({e})")
                n_skip += 1
                continue

            x_label = _runtime_plot_x_label(x_label, x_col_base, x_col, mestrado_x_override)
            if not y_label:
                y_label = y_col
            if not title:
                title = f"{y_col} vs {x_col} (all fuels)"
            if not filename:
                filename = f"{_safe_name(y_col)}_vs_{_safe_name(x_col)}_all.png"

            for variant_key, uncertainty_mode, dual_variant in _plot_uncertainty_variants(r):
                variant_row = r.copy()
                variant_row["show_uncertainty"] = uncertainty_mode
                variant_filename, variant_title = _decorate_plot_variant_output(filename, title, variant_key, dual_variant)
                yerr_col = _resolve_plot_yerr_col(
                    out_df,
                    variant_row,
                    y_col=y_col,
                    mappings=mappings,
                    plot_label=variant_filename or variant_title or y_col,
                )

                plot_all_fuels_xy(
                    out_df,
                    x_col=x_col,
                    y_col=y_col,
                    yerr_col=yerr_col,
                    title=variant_title,
                    filename=variant_filename,
                    x_label=x_label,
                    y_label=y_label,
                    fixed_y=fixed_y,
                    fixed_y_limits=fixed_y_limits,
                    y_tick_step=y_tick_step,
                    fixed_x=fixed_x,
                    fuels_override=fuels_override,
                    series_col=series_col,
                    plot_dir=plot_dir,
                    y_tol_plus=y_tol_plus,
                    y_tol_minus=y_tol_minus,
                )
                n_ok += 1
            continue

        if pt in {"all_fuels_labels", "labels"}:
            x_col_base, mestrado_x_override = _resolve_plot_x_request(x_col_req)
            try:
                x_col = resolve_col(out_df, x_col_base)
            except Exception as e:
                print(f"[ERROR] Plot '{filename or title}': x_col '{x_col_base}' nÃ£o encontrado. Pulei. ({e})")
                n_skip += 1
                continue

            if not y_col_req:
                print(f"[ERROR] Plot '{filename or title}': y_col vazio (plot_type=all_fuels_labels). Pulei.")
                n_skip += 1
                continue
            try:
                y_col = resolve_col(out_df, y_col_req)
            except Exception as e:
                print(f"[ERROR] Plot '{filename or title}': y_col '{y_col_req}' nÃ£o encontrado. Pulei. ({e})")
                n_skip += 1
                continue

            x_label = _runtime_plot_x_label(x_label, x_col_base, x_col, mestrado_x_override)
            if not y_label:
                y_label = y_col
            if not title:
                title = f"{y_col} vs {x_col} (labels)"
            if not filename:
                filename = f"{_safe_name(y_col)}_vs_{_safe_name(x_col)}_labels.png"

            plot_all_fuels_with_value_labels(
                out_df,
                y_col=y_col,
                title=title,
                filename=filename,
                y_label=y_label,
                label_variant=label_variant,
                fixed_y=fixed_y,
                fixed_y_limits=fixed_y_limits,
                y_tick_step=y_tick_step,
                fixed_x=fixed_x,
                x_col=x_col,
                x_label=x_label,
                fuels_override=fuels_override,
                series_col=series_col,
                plot_dir=plot_dir,
                y_tol_plus=y_tol_plus,
                y_tol_minus=y_tol_minus,
            )
            n_ok += 1
            continue

        print(f"[ERROR] Plot '{filename or title}': plot_type '{plot_type}' nÃ£o suportado. Pulei.")
        n_skip += 1

    print(f"[OK] Plots-config: {n_ok} gerados; {n_skip} pulados.")


def make_plots_from_config_with_summary(
    out_df: pd.DataFrame,
    plots_df: pd.DataFrame,
    mappings: dict,
    plot_dir: Optional[Path] = None,
    series_col: Optional[str] = None,
) -> Dict[str, object]:
    summary = _new_plot_run_summary()
    target_dir = PLOTS_DIR if plot_dir is None else plot_dir

    def mark_generated(label: str, filename_value: str) -> None:
        summary["generated"] = int(summary.get("generated", 0)) + 1
        summary.setdefault("generated_labels", []).append(label)
        if filename_value:
            summary.setdefault("generated_files", []).append(str((target_dir / filename_value).resolve()))

    def mark_skipped(label: str, reason: str) -> None:
        summary["skipped"] = int(summary.get("skipped", 0)) + 1
        summary.setdefault("skipped_items", []).append((label, reason))

    if plots_df is None or plots_df.empty:
        print("[WARN] Plots config vazio; não gerei plots via planilha.")
        return summary

    for _, r in plots_df.iterrows():
        if not _row_enabled(r.get("enabled", 0)):
            summary["disabled"] = int(summary.get("disabled", 0)) + 1
            continue

        plot_type = _to_str_or_empty(r.get("plot_type", ""))
        filename = _strip_leading_raw_plot_name(r.get("filename", ""))
        title = _strip_leading_raw_plot_name(r.get("title", ""))
        plot_label = filename or title or _to_str_or_empty(r.get("y_col", "")) or _to_str_or_empty(r.get("plot_type", "")) or "plot_sem_nome"

        if not plot_type:
            print("[ERROR] Plots row inválida: plot_type vazio. Pulei.")
            mark_skipped(plot_label, "plot_type vazio")
            continue

        x_col_req = _to_str_or_empty(r.get("x_col", ""))
        y_col_req = _to_str_or_empty(r.get("y_col", ""))
        x_label = _to_str_or_empty(r.get("x_label", ""))
        y_label = _to_str_or_empty(r.get("y_label", ""))

        y_axis_unit = _mapping_unit_for_y_col(y_col_req, mappings)
        fixed_x = _parse_axis_spec(r.get("x_min", pd.NA), r.get("x_max", pd.NA), r.get("x_step", pd.NA))
        fixed_y = _parse_axis_spec(
            r.get("y_min", pd.NA),
            r.get("y_max", pd.NA),
            r.get("y_step", pd.NA),
            target_unit=y_axis_unit,
        )
        fixed_y_limits = _parse_axis_limits(
            r.get("y_min", pd.NA),
            r.get("y_max", pd.NA),
            target_unit=y_axis_unit,
        )
        y_tick_step = _parse_axis_value(r.get("y_step", pd.NA), target_unit=y_axis_unit, default=np.nan)
        if not np.isfinite(y_tick_step) or y_tick_step <= 0:
            y_tick_step = None
        if fixed_y is not None:
            y_tick_step = None
        y_tol_plus = _to_float(r.get("y_tol_plus", r.get("tol_plus", 0.0)), 0.0)
        y_tol_minus = _to_float(r.get("y_tol_minus", r.get("tol_minus", 0.0)), 0.0)

        fuels = _parse_csv_list_ints(r.get("filter_h2o_list", pd.NA))
        fuels_override = fuels if fuels is not None else None
        label_variant = _to_str_or_empty(r.get("label_variant", "box")).lower() or "box"
        pt = plot_type.lower().strip()

        if pt in {"kibox_all", "all_kibox"}:
            kibox_cols = [c for c in out_df.columns if str(c).startswith("KIBOX_") and c != "KIBOX_N_files"]
            if not kibox_cols:
                print("[WARN] kibox_all: não há colunas KIBOX_* no output. Pulei expansão.")
                mark_skipped(plot_label, "sem colunas KIBOX_* no output")
                continue

            x_col_base, mestrado_x_override = _resolve_plot_x_request(x_col_req)
            try:
                x_col = resolve_col(out_df, x_col_base)
            except Exception as e:
                print(f"[ERROR] kibox_all: x_col '{x_col_base}' não encontrado. Pulei expansão. ({e})")
                mark_skipped(plot_label, f"x_col ausente: {x_col_base}")
                continue

            xlab = _runtime_plot_x_label(x_label, x_col_base, x_col, mestrado_x_override)
            seen_filenames: set[str] = set()
            for yc in sorted(kibox_cols):
                fn = _derive_filename_for_expansion(filename, yc)
                fn_key = norm_key(fn)
                item_label = fn or yc
                if fn_key in seen_filenames:
                    print(f"[INFO] kibox_all: filename duplicado apos normalizacao ('{fn}'). Pulei a expansao de '{yc}'.")
                    mark_skipped(item_label, "filename duplicado após normalização")
                    continue
                seen_filenames.add(fn_key)
                tt = _derive_title_for_expansion(title, x_col=x_col, y_col=yc)
                ylab = y_label if y_label else yc
                ok = plot_all_fuels(
                    out_df,
                    y_col=yc,
                    yerr_col=None,
                    title=tt,
                    filename=fn,
                    y_label=ylab,
                    fixed_y=fixed_y,
                    fixed_y_limits=fixed_y_limits,
                    y_tick_step=y_tick_step,
                    fixed_x=fixed_x,
                    x_col=x_col,
                    x_label=xlab,
                    fuels_override=fuels_override,
                    series_col=series_col,
                    plot_dir=plot_dir,
                    y_tol_plus=y_tol_plus,
                    y_tol_minus=y_tol_minus,
                )
                if ok:
                    mark_generated(item_label, fn)
                else:
                    mark_skipped(item_label, "sem dados válidos para plot")
            continue

        if pt in {"all_fuels_yx", "all_fuels", "all_fuels_y_vs_x"}:
            x_col_base, mestrado_x_override = _resolve_plot_x_request(x_col_req)
            try:
                x_col = resolve_col(out_df, x_col_base)
            except Exception as e:
                print(f"[ERROR] Plot '{filename or title}': x_col '{x_col_base}' não encontrado. Pulei. ({e})")
                mark_skipped(plot_label, f"x_col ausente: {x_col_base}")
                continue

            if not y_col_req:
                print(f"[ERROR] Plot '{filename or title}': y_col vazio. Pulei.")
                mark_skipped(plot_label, "y_col vazio")
                continue
            try:
                y_col = resolve_col(out_df, y_col_req)
            except Exception as e:
                print(f"[ERROR] Plot '{filename or title}': y_col '{y_col_req}' não encontrado. Pulei. ({e})")
                mark_skipped(plot_label, f"y_col ausente: {y_col_req}")
                continue

            x_label = _runtime_plot_x_label(x_label, x_col_base, x_col, mestrado_x_override)
            if not y_label:
                y_label = y_col
            if not title:
                title = f"{y_col} vs {x_col} (all fuels)"
            if not filename:
                filename = f"{_safe_name(y_col)}_vs_{_safe_name(x_col)}_all.png"

            variant_specs: List[Tuple[str, str, Optional[str], str]] = []
            for variant_key, uncertainty_mode, dual_variant in _plot_uncertainty_variants(r):
                variant_row = r.copy()
                variant_row["show_uncertainty"] = uncertainty_mode
                variant_filename, variant_title = _decorate_plot_variant_output(filename, title, variant_key, dual_variant)
                yerr_col = _resolve_plot_yerr_col(
                    out_df,
                    variant_row,
                    y_col=y_col,
                    mappings=mappings,
                    plot_label=variant_filename or variant_title or y_col,
                )
                variant_specs.append((variant_filename, variant_title, yerr_col, variant_key))

            variant_fixed_y_limits = fixed_y_limits
            if fixed_y is None and fixed_y_limits is None and len(variant_specs) > 1:
                shared_limits = _shared_plot_y_limits_for_variants(
                    out_df,
                    x_col=x_col,
                    y_col=y_col,
                    variant_yerr_cols=[spec[2] for spec in variant_specs],
                    fuels_override=fuels_override,
                    series_col=series_col,
                    y_tol_plus=y_tol_plus,
                    y_tol_minus=y_tol_minus,
                )
                if shared_limits is not None:
                    variant_fixed_y_limits = shared_limits

            for variant_filename, variant_title, yerr_col, _variant_key in variant_specs:
                item_label = variant_filename or variant_title or plot_label
                ok = plot_all_fuels(
                    out_df,
                    y_col=y_col,
                    yerr_col=yerr_col,
                    title=variant_title,
                    filename=variant_filename,
                    y_label=y_label,
                    fixed_y=fixed_y,
                    fixed_y_limits=variant_fixed_y_limits,
                    y_tick_step=y_tick_step,
                    fixed_x=fixed_x,
                    x_col=x_col,
                    x_label=x_label,
                    fuels_override=fuels_override,
                    series_col=series_col,
                    plot_dir=plot_dir,
                    y_tol_plus=y_tol_plus,
                    y_tol_minus=y_tol_minus,
                )
                if ok:
                    mark_generated(item_label, variant_filename)
                else:
                    mark_skipped(item_label, "sem dados válidos para plot")
            continue

        if pt in {"all_fuels_xy", "xy"}:
            if not y_col_req:
                print(f"[ERROR] Plot '{filename or title}': y_col vazio (plot_type=all_fuels_xy). Pulei.")
                mark_skipped(plot_label, "y_col vazio")
                continue

            x_col_base, mestrado_x_override = _resolve_plot_x_request(x_col_req)
            try:
                x_col = resolve_col(out_df, x_col_base)
            except Exception as e:
                print(f"[ERROR] Plot '{filename or title}': x_col '{x_col_base}' não encontrado. Pulei. ({e})")
                mark_skipped(plot_label, f"x_col ausente: {x_col_base}")
                continue

            try:
                y_col = resolve_col(out_df, y_col_req)
            except Exception as e:
                print(f"[ERROR] Plot '{filename or title}': y_col '{y_col_req}' não encontrado. Pulei. ({e})")
                mark_skipped(plot_label, f"y_col ausente: {y_col_req}")
                continue

            x_label = _runtime_plot_x_label(x_label, x_col_base, x_col, mestrado_x_override)
            if not y_label:
                y_label = y_col
            if not title:
                title = f"{y_col} vs {x_col} (all fuels)"
            if not filename:
                filename = f"{_safe_name(y_col)}_vs_{_safe_name(x_col)}_all.png"

            variant_specs = []
            for variant_key, uncertainty_mode, dual_variant in _plot_uncertainty_variants(r):
                variant_row = r.copy()
                variant_row["show_uncertainty"] = uncertainty_mode
                variant_filename, variant_title = _decorate_plot_variant_output(filename, title, variant_key, dual_variant)
                yerr_col = _resolve_plot_yerr_col(
                    out_df,
                    variant_row,
                    y_col=y_col,
                    mappings=mappings,
                    plot_label=variant_filename or variant_title or y_col,
                )
                variant_specs.append((variant_filename, variant_title, yerr_col, variant_key))

            variant_fixed_y_limits = fixed_y_limits
            if fixed_y is None and fixed_y_limits is None and len(variant_specs) > 1:
                shared_limits = _shared_plot_y_limits_for_variants(
                    out_df,
                    x_col=x_col,
                    y_col=y_col,
                    variant_yerr_cols=[spec[2] for spec in variant_specs],
                    fuels_override=fuels_override,
                    series_col=series_col,
                    y_tol_plus=y_tol_plus,
                    y_tol_minus=y_tol_minus,
                )
                if shared_limits is not None:
                    variant_fixed_y_limits = shared_limits

            for variant_filename, variant_title, yerr_col, _variant_key in variant_specs:
                item_label = variant_filename or variant_title or plot_label
                ok = plot_all_fuels_xy(
                    out_df,
                    x_col=x_col,
                    y_col=y_col,
                    yerr_col=yerr_col,
                    title=variant_title,
                    filename=variant_filename,
                    x_label=x_label,
                    y_label=y_label,
                    fixed_y=fixed_y,
                    fixed_y_limits=variant_fixed_y_limits,
                    y_tick_step=y_tick_step,
                    fixed_x=fixed_x,
                    fuels_override=fuels_override,
                    series_col=series_col,
                    plot_dir=plot_dir,
                    y_tol_plus=y_tol_plus,
                    y_tol_minus=y_tol_minus,
                )
                if ok:
                    mark_generated(item_label, variant_filename)
                else:
                    mark_skipped(item_label, "sem dados válidos para plot")
            continue

        if pt in {"all_fuels_labels", "labels"}:
            x_col_base, mestrado_x_override = _resolve_plot_x_request(x_col_req)
            try:
                x_col = resolve_col(out_df, x_col_base)
            except Exception as e:
                print(f"[ERROR] Plot '{filename or title}': x_col '{x_col_base}' não encontrado. Pulei. ({e})")
                mark_skipped(plot_label, f"x_col ausente: {x_col_base}")
                continue

            if not y_col_req:
                print(f"[ERROR] Plot '{filename or title}': y_col vazio (plot_type=all_fuels_labels). Pulei.")
                mark_skipped(plot_label, "y_col vazio")
                continue
            try:
                y_col = resolve_col(out_df, y_col_req)
            except Exception as e:
                print(f"[ERROR] Plot '{filename or title}': y_col '{y_col_req}' não encontrado. Pulei. ({e})")
                mark_skipped(plot_label, f"y_col ausente: {y_col_req}")
                continue

            x_label = _runtime_plot_x_label(x_label, x_col_base, x_col, mestrado_x_override)
            if not y_label:
                y_label = y_col
            if not title:
                title = f"{y_col} vs {x_col} (labels)"
            if not filename:
                filename = f"{_safe_name(y_col)}_vs_{_safe_name(x_col)}_labels.png"

            ok = plot_all_fuels_with_value_labels(
                out_df,
                y_col=y_col,
                title=title,
                filename=filename,
                y_label=y_label,
                label_variant=label_variant,
                fixed_y=fixed_y,
                fixed_y_limits=fixed_y_limits,
                y_tick_step=y_tick_step,
                fixed_x=fixed_x,
                x_col=x_col,
                x_label=x_label,
                fuels_override=fuels_override,
                series_col=series_col,
                plot_dir=plot_dir,
                y_tol_plus=y_tol_plus,
                y_tol_minus=y_tol_minus,
            )
            if ok:
                mark_generated(filename or plot_label, filename)
            else:
                mark_skipped(filename or plot_label, "sem dados válidos para plot")
            continue

        print(f"[ERROR] Plot '{filename or title}': plot_type '{plot_type}' não suportado. Pulei.")
        mark_skipped(plot_label, f"plot_type não suportado: {plot_type}")

    print(
        f"[OK] Plots-config: {int(summary.get('generated', 0))} gerados; "
        f"{int(summary.get('skipped', 0))} pulados; "
        f"{int(summary.get('disabled', 0))} desabilitados."
    )
    return summary


def _snapshot_png_files(root: Path) -> set[str]:
    if not root.exists():
        return set()
    return {str(p.resolve()) for p in root.rglob("*.png") if p.is_file()}


def _run_plot_job_with_snapshot(
    label: str,
    func,
    *,
    snapshot_dir: Path,
    no_output_reason: str,
    **kwargs,
) -> Dict[str, object]:
    summary = _new_plot_run_summary()
    before = _snapshot_png_files(snapshot_dir)
    func(**kwargs)
    after = _snapshot_png_files(snapshot_dir)
    created = sorted(after - before)
    if created:
        summary["generated"] = len(created)
        summary["generated_labels"] = [label]
        summary["generated_files"] = created
    else:
        summary["skipped"] = 1
        summary["skipped_items"] = [(label, no_output_reason)]
    return summary


def _top_label_counts(df: pd.DataFrame, mask: pd.Series, *, limit: int = 5) -> str:
    if df is None or df.empty:
        return ""
    valid_mask = mask.fillna(False)
    if not bool(valid_mask.any()):
        return ""
    if "Fuel_Label" not in df.columns:
        return ""
    labels = df.loc[valid_mask, "Fuel_Label"].map(_to_str_or_empty).replace("", "(sem Fuel_Label)")
    counts = labels.value_counts(dropna=False)
    parts = [f"{label} ({int(count)})" for label, count in counts.head(limit).items()]
    return ", ".join(parts)


def _print_processing_summary(
    out: pd.DataFrame,
    plot_out: pd.DataFrame,
    plot_summary: Optional[Dict[str, object]],
) -> None:
    total_points = int(len(out))
    plotted_points = int(len(plot_out)) if isinstance(plot_out, pd.DataFrame) else 0

    airflow_method = out.get("Airflow_Method", pd.Series(pd.NA, index=out.index, dtype="object")).fillna("unavailable")
    airflow_counts = airflow_method.value_counts(dropna=False)
    lambda_counts = out.get("LAMBDA_SOURCE", pd.Series(pd.NA, index=out.index, dtype="object")).fillna("default_1.0").value_counts(dropna=False)
    airflow_parts = [
        f"MAF={int(airflow_counts.get('MAF', 0))}",
        f"fuel+lambda={int(airflow_counts.get('fuel_lambda', 0))}",
        f"fuel+lambda_default={int(airflow_counts.get('fuel_lambda_default', 0))}",
        f"indisponivel={int(airflow_counts.get('unavailable', 0))}",
    ]
    print(
        "[SUMMARY] Airflow: "
        + ", ".join(airflow_parts)
        + f" | lambda medida={int(lambda_counts.get('measured', 0))}, default_1.0={int(lambda_counts.get('default_1.0', 0))}"
    )

    bsfc_valid = int(pd.to_numeric(out.get("BSFC_g_kWh", pd.NA), errors="coerce").notna().sum())
    bsfc_missing = pd.to_numeric(out.get("BSFC_g_kWh", pd.NA), errors="coerce").isna()
    bsfc_labels = _top_label_counts(out, bsfc_missing)
    print(
        f"[SUMMARY] Calculos: BSFC_g_kWh={bsfc_valid}/{total_points}"
        + (f" | faltou por potencia/consumo invalidos em {bsfc_labels}" if bsfc_labels else "")
    )

    for emission_col in [
        "CO2_g_kWh",
        "CO_g_kWh",
        "THC_g_kWh",
        "NOx_as_NO_g_kWh",
        "NOx_as_NO2_g_kWh",
    ]:
        emission_values = pd.to_numeric(out.get(emission_col, pd.NA), errors="coerce")
        valid_count = int(emission_values.notna().sum())
        missing_mask = emission_values.isna()
        missing_labels = _top_label_counts(out, missing_mask)
        print(
            f"[SUMMARY] Calculos: {emission_col}={valid_count}/{total_points}"
            + (f" | faltou por potencia/airflow/gases/H2O_wet_frac invalidos em {missing_labels}" if missing_labels else "")
        )

    ve_valid = int(pd.to_numeric(out.get("ETA_V_pct", pd.NA), errors="coerce").notna().sum())
    ve_missing = pd.to_numeric(out.get("ETA_V_pct", pd.NA), errors="coerce").isna()
    ve_labels = _top_label_counts(out, ve_missing)
    print(
        f"[SUMMARY] Calculos: ETA_V_pct={ve_valid}/{total_points}"
        + (f" | faltou por airflow/RPM/T_admissao invalidos em {ve_labels}" if ve_labels else "")
    )

    consumo_l_valid = int(pd.to_numeric(out.get("Consumo_L_h", pd.NA), errors="coerce").notna().sum())
    density_missing = pd.to_numeric(out.get("Consumo_L_h", pd.NA), errors="coerce").isna() & pd.to_numeric(out.get("Consumo_kg_h_mean_of_windows", pd.NA), errors="coerce").gt(0)
    density_labels = _top_label_counts(out, density_missing)
    print(
        f"[SUMMARY] Calculos: Consumo_L_h={consumo_l_valid}/{total_points}"
        + (f" | faltou por densidade ausente/invalida em {density_labels}" if density_labels else "")
    )

    cost_valid = int(pd.to_numeric(out.get("Custo_R_h", pd.NA), errors="coerce").notna().sum())
    cost_missing = pd.to_numeric(out.get("Custo_R_h", pd.NA), errors="coerce").isna() & pd.to_numeric(out.get("Consumo_kg_h_mean_of_windows", pd.NA), errors="coerce").gt(0)
    cost_labels = _top_label_counts(out, cost_missing)
    print(
        f"[SUMMARY] Calculos: Custo_R_h={cost_valid}/{total_points}"
        + (f" | faltou por custo/densidade do combustivel em {cost_labels}" if cost_labels else "")
    )

    if plot_summary is None:
        return

    skip_counter = Counter(reason for _label, reason in plot_summary.get("skipped_items", []))
    skip_parts = [f"{reason} ({count})" for reason, count in skip_counter.most_common(4)]
    print(
        f"[SUMMARY] Plots: gerados={int(plot_summary.get('generated', 0))}, "
        f"pulados={int(plot_summary.get('skipped', 0))}, "
        f"desabilitados={int(plot_summary.get('disabled', 0))}, "
        f"pontos filtrados={plotted_points}/{total_points}"
        + (f" | motivos: {', '.join(skip_parts)}" if skip_parts else "")
    )


# =========================
# Main
# =========================
def _parse_cli_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Pipeline 29 com configuracao textual e fallback para Excel.")
    parser.add_argument("--config-source", choices=["auto", "text", "excel"], default="auto")
    parser.add_argument("--config-dir", default="", help="Diretorio da configuracao textual do pipeline29.")
    parser.add_argument("--rebuild-text-config", action="store_true", help="Regera a config textual a partir do Excel rev3.")
    parser.add_argument("--config-gui", action="store_true", help="Abre o editor GUI da configuracao textual e sai.")
    parser.add_argument("--skip-config-gui-prompt", action="store_true", help="Nao pergunta se deve abrir a GUI antes do run.")
    parser.add_argument(
        "--plot-scope",
        choices=["all", "unitary", "compare", "none"],
        default="",
        help="Controla quais familias de plots gerar: all, unitary, compare ou none.",
    )
    parser.add_argument(
        "--compare-iter-pairs",
        default="",
        help=(
            "Seleciona os pares do compare_iteracoes. Aceita default, cross_all, all "
            "ou uma lista separada por virgula como baseline_media:aditivado_subida."
        ),
    )
    return parser.parse_args(argv)


def _resolve_plot_scope(cli_value: object) -> str:
    raw = _to_str_or_empty(cli_value).lower()
    if not raw:
        raw = _to_str_or_empty(os.environ.get("PIPELINE29_PLOT_SCOPE", "")).lower()
    if raw not in {"all", "unitary", "compare", "none"}:
        return "all"
    return raw


def _cfg_bool(defaults_cfg: Dict[str, str], key: str, *, default: bool) -> bool:
    cfg = defaults_cfg or {}
    raw = _to_str_or_empty(cfg.get(key, ""))
    if not raw:
        raw = _to_str_or_empty(cfg.get(norm_key(key), ""))
    raw = raw.lower()
    if not raw:
        return default
    return raw in {"1", "true", "yes", "on", "y", "checked"}


def _plot_scope_from_config(defaults_cfg: Dict[str, str], compare_df: Optional[pd.DataFrame]) -> str:
    unitary_enabled = _cfg_bool(defaults_cfg, GUI_COMPARE_ENABLE_UNITARY_KEY, default=True)
    compare_enabled = False
    if compare_df is not None and not compare_df.empty:
        for row in compare_df.to_dict(orient="records"):
            if _row_enabled((row or {}).get("enabled", "")):
                compare_enabled = True
                break
    if unitary_enabled and compare_enabled:
        return "all"
    if unitary_enabled:
        return "unitary"
    if compare_enabled:
        return "compare"
    return "none"


def main(argv: Optional[List[str]] = None) -> None:
    args = _parse_cli_args(argv)
    print(f"[INFO] Base do script: {BASE_DIR}")
    text_config_dir = _choose_text_config_dir(Path(args.config_dir) if args.config_dir else None)
    gui_save_run_exit_code = 1001
    plot_scope_valid_values = {"all", "unitary", "compare", "none"}
    plot_scope_cli_raw = _to_str_or_empty(args.plot_scope).lower()
    plot_scope_env_raw = _to_str_or_empty(os.environ.get("PIPELINE29_PLOT_SCOPE", "")).lower()
    if plot_scope_cli_raw and plot_scope_cli_raw not in plot_scope_valid_values:
        print(
            f"[WARN] --plot-scope invalido ('{plot_scope_cli_raw}'); "
            "vou ignorar o override e usar a configuracao da GUI."
        )
    if plot_scope_env_raw and plot_scope_env_raw not in plot_scope_valid_values:
        print(
            f"[WARN] PIPELINE29_PLOT_SCOPE invalido ('{plot_scope_env_raw}'); "
            "vou ignorar o override de ambiente e usar a configuracao da GUI."
        )
    plot_scope = _resolve_plot_scope(args.plot_scope)
    plot_scope_cli_override = (
        plot_scope_cli_raw in plot_scope_valid_values
        or plot_scope_env_raw in plot_scope_valid_values
    )
    compare_iter_pairs = _resolve_compare_iter_pairs(args.compare_iter_pairs)
    compare_iter_pairs_cli_override = bool(_to_str_or_empty(args.compare_iter_pairs)) or bool(
        _to_str_or_empty(os.environ.get("PIPELINE29_COMPARE_ITER_PAIRS", ""))
    )

    if args.config_gui:
        try:
            from pipeline29_config_gui import PIPELINE29_GUI_SAVE_RUN_EXIT_CODE, launch_config_gui
        except Exception as exc:
            raise RuntimeError(f"Nao consegui abrir a GUI de configuracao do pipeline29: {exc}") from exc
        gui_save_run_exit_code = PIPELINE29_GUI_SAVE_RUN_EXIT_CODE
        gui_exit_code = launch_config_gui(base_dir=BASE_DIR, config_dir=text_config_dir, excel_path=_choose_config_path())
        state = load_gui_state(default_gui_state_path())
        state_config_dir = str(state.get("config_dir", "")).strip()
        if state_config_dir:
            text_config_dir = Path(state_config_dir).expanduser().resolve()
        if gui_exit_code != gui_save_run_exit_code:
            return
        print("[INFO] Save & Run selecionado na GUI; seguindo para o processamento.")

    skip_gui_prompt = norm_key(os.environ.get("PIPELINE29_SKIP_CONFIG_GUI_PROMPT", ""))
    if not args.skip_config_gui_prompt and skip_gui_prompt not in {"1", "true", "yes", "on"}:
        if _prompt_open_config_gui():
            try:
                from pipeline29_config_gui import PIPELINE29_GUI_SAVE_RUN_EXIT_CODE, launch_config_gui
            except Exception as exc:
                print(f"[WARN] Nao consegui abrir a GUI de configuracao do pipeline29: {exc}")
            else:
                gui_save_run_exit_code = PIPELINE29_GUI_SAVE_RUN_EXIT_CODE
                gui_exit_code = launch_config_gui(
                    base_dir=BASE_DIR,
                    config_dir=text_config_dir,
                    excel_path=_choose_config_path(),
                )
                state = load_gui_state(default_gui_state_path())
                state_config_dir = str(state.get("config_dir", "")).strip()
                if state_config_dir:
                    text_config_dir = Path(state_config_dir).expanduser().resolve()
                if gui_exit_code == gui_save_run_exit_code:
                    print("[INFO] Save & Run selecionado na GUI; seguindo para o processamento.")

    config_bundle = load_pipeline29_config_bundle(
        config_source=args.config_source,
        text_config_dir=text_config_dir,
        rebuild_text_config=args.rebuild_text_config,
    )
    mappings = config_bundle.mappings
    instruments_df = config_bundle.instruments_df
    reporting_df = config_bundle.reporting_df
    plots_df = config_bundle.plots_df
    compare_df = config_bundle.compare_df
    data_quality_cfg = config_bundle.data_quality_cfg
    defaults_cfg = config_bundle.defaults_cfg
    compare_iter_requests, compare_request_source = _resolve_compare_iter_requests(
        compare_df,
        fallback_pairs=compare_iter_pairs,
        force_fallback_pairs=compare_iter_pairs_cli_override,
    )
    apply_runtime_path_overrides(defaults_cfg, config_bundle=config_bundle)
    if not plot_scope_cli_override:
        plot_scope = _plot_scope_from_config(defaults_cfg, compare_df)
    config_label = config_bundle.source_path if config_bundle.source_path is not None else text_config_dir
    print(f"[INFO] Config ({config_bundle.source_kind}): {config_label}")
    print(f"[INFO] Entrada LabVIEW/Kibox: {PROCESS_DIR}")
    print(f"[INFO] Saida: {OUT_DIR}")
    if plot_scope_cli_override:
        print(f"[INFO] Escopo de plots (CLI/ambiente): {plot_scope}")
    else:
        print(f"[INFO] Escopo de plots (aba Compare): {plot_scope}")
    if compare_request_source == "gui_compare_tab":
        print(f"[INFO] compare_iteracoes configurado na aba Compare: {len(compare_iter_requests)} selecao(oes) ativa(s).")
    elif compare_request_source == "forced_fallback_pairs":
        print(
            "[INFO] compare_iteracoes forçado por CLI/ambiente: "
            + ", ".join(f"{left_id}:{right_id}" for left_id, right_id in compare_iter_pairs)
        )
    elif compare_request_source == "gui_empty":
        print("[INFO] compare_iteracoes: aba Compare sem linhas habilitadas; compare_iteracoes/ sera pulado.")
    elif compare_request_source == "gui_invalid":
        print("[WARN] compare_iteracoes: aba Compare sem selecoes validas; compare_iteracoes/ sera pulado.")
    else:
        print(
            "[INFO] compare_iteracoes fallback por CLI/ambiente: "
            + ", ".join(f"{left_id}:{right_id}" for left_id, right_id in compare_iter_pairs)
        )
    clear_output_dir(OUT_DIR)
    PLOTS_DIR.mkdir(parents=True, exist_ok=True)

    raw_files = [
        p
        for pattern in ("*.xlsx", "*.csv")
        for p in PROCESS_DIR.rglob(pattern)
        if p.is_file() and not p.name.startswith("~$")
    ]
    metas = [parse_meta(p) for p in raw_files]

    lv_files = [m for m in metas if m.source_type == "LABVIEW" and m.path.suffix.lower() == ".xlsx"]
    kibox_files = [m for m in metas if m.source_type == "KIBOX" and m.path.suffix.lower() == ".csv"]
    motec_files = [m for m in metas if m.source_type == "MOTEC" and m.path.suffix.lower() == ".csv"]

    if not lv_files:
        raise SystemExit(f"NÃ£o achei .xlsx do LabVIEW em {PROCESS_DIR}.")

    missing_comp = [m.basename for m in lv_files if m.composition_parse == "missing_filename"]
    if missing_comp:
        preview = ", ".join(missing_comp[:5])
        suffix = " ..." if len(missing_comp) > 5 else ""
        print(
            f"[INFO] {len(missing_comp)} arquivo(s) sem composiÃ§Ã£o no nome; "
            f"DIES_pct/BIOD_pct/EtOH_pct/H2O_pct ficarÃ£o em branco no output. Exemplos: {preview}{suffix}"
        )

    ambiguous_load = [m.basename for m in lv_files if m.load_parse == "ambiguous_filename"]
    if ambiguous_load:
        preview = ", ".join(ambiguous_load[:5])
        suffix = " ..." if len(ambiguous_load) > 5 else ""
        print(
            f"[INFO] {len(ambiguous_load)} arquivo(s) com mÃºltiplas cargas no nome; "
            f"vou inferir Load_kW pela coluna de carga. Exemplos: {preview}{suffix}"
        )

    missing_load = [m.basename for m in lv_files if m.load_parse == "missing_filename"]
    if missing_load:
        preview = ", ".join(missing_load[:5])
        suffix = " ..." if len(missing_load) > 5 else ""
        print(
            f"[INFO] {len(missing_load)} arquivo(s) sem carga explÃ­cita no nome; "
            f"vou tentar inferir Load_kW pela coluna de carga. Exemplos: {preview}{suffix}"
        )

    print("[INFO] Abrindo filtro de pontos para os plots finais...")
    selected_plot_points = prompt_plot_point_filter_from_metas(lv_files)

    lv_all: List[pd.DataFrame] = []
    for m in lv_files:
        try:
            df_i = read_labview_xlsx(m)
            if not df_i.empty:
                lv_all.append(df_i)
        except Exception as e:
            print(f"[ERROR] Falha lendo LabVIEW {m.path.name}: {e}")

    if not lv_all:
        raise SystemExit("Nenhum arquivo LabVIEW foi lido com sucesso.")

    lv_raw = pd.concat(lv_all, ignore_index=True)
    lv_time_diag = build_time_diagnostics(lv_raw, quality_cfg=data_quality_cfg)
    if not lv_time_diag.empty:
        time_diag_out = lv_time_diag.copy()
        time_diag_out["Iteracao"] = pd.to_numeric(time_diag_out.get("Iteracao", pd.NA), errors="coerce").astype("Int64")
        time_diag_out["_SENTIDO_rank"] = time_diag_out.get("Sentido_Carga", pd.Series([pd.NA] * len(time_diag_out))).map(_sentido_carga_rank).fillna(9)

        diag_first_cols = [c for c in ["Iteracao", "Sentido_Carga", "Load_kW"] if c in time_diag_out.columns]
        if diag_first_cols:
            diag_rest = [c for c in time_diag_out.columns if c not in diag_first_cols]
            time_diag_out = time_diag_out[diag_first_cols + diag_rest].copy()

        time_diag_xlsx = safe_to_excel(
            time_diag_out.sort_values(
                ["Iteracao", "_SENTIDO_rank", "Load_kW", "BaseName", "Index"],
                ascending=[True, True, True, True, True],
                na_position="last",
            ).drop(columns=["_SENTIDO_rank"]).copy(),
            OUT_DIR / "lv_time_diagnostics.xlsx",
        )
        print(f"[OK] DiagnÃ³stico de qualidade por amostra gerado: {time_diag_xlsx}")

        lv_time_summary = summarize_time_diagnostics(lv_time_diag)
        if not lv_time_summary.empty:
            summary_out = add_run_context_columns(lv_time_summary.copy())
            summary_out["Iteracao"] = pd.to_numeric(summary_out.get("Iteracao", pd.NA), errors="coerce").astype("Int64")
            status_rank = {"ERRO": 0, "OK": 1, "NA": 2}
            summary_out["_DQ_rank"] = summary_out["DQ_ERROR"].map(status_rank).fillna(9)
            summary_out["_SMP_rank"] = summary_out["Smp_ERROR"].map(status_rank).fillna(9)
            summary_out["_ACT_rank"] = summary_out["ACT_CTRL_ERRO"].map(status_rank).fillna(9)
            summary_out["_ACT_TRANS_rank"] = summary_out["ACT_CTRL_ERRO_TRANSIENTE"].map(status_rank).fillna(9)
            summary_out["_ECT_rank"] = summary_out.get("ECT_CTRL_ERRO", pd.Series([pd.NA] * len(summary_out))).map(status_rank).fillna(9)
            summary_out["_ECT_TRANS_rank"] = summary_out.get("ECT_CTRL_ERRO_TRANSIENTE", pd.Series([pd.NA] * len(summary_out))).map(status_rank).fillna(9)
            summary_out["_SENTIDO_rank"] = summary_out.get("Sentido_Carga", pd.Series([pd.NA] * len(summary_out))).map(_sentido_carga_rank).fillna(9)

            sum_first_cols = [c for c in ["Iteracao", "Sentido_Carga", "Load_kW"] if c in summary_out.columns]
            if sum_first_cols:
                sum_rest = [c for c in summary_out.columns if c not in sum_first_cols]
                summary_out = summary_out[sum_first_cols + sum_rest].copy()
            time_summary_xlsx = safe_to_excel(
                summary_out.sort_values(
                    [
                        "Iteracao",
                        "_SENTIDO_rank",
                        "Load_kW",
                        "_DQ_rank",
                        "_SMP_rank",
                        "_ACT_rank",
                        "_ACT_TRANS_rank",
                        "_ECT_rank",
                        "_ECT_TRANS_rank",
                        "SourceFolder",
                        "BaseName",
                    ],
                    ascending=[True, True, True, True, True, True, True, True, True, True, True],
                    na_position="last",
                ).drop(
                    columns=[
                        "_DQ_rank",
                        "_SMP_rank",
                        "_ACT_rank",
                        "_ACT_TRANS_rank",
                        "_ECT_rank",
                        "_ECT_TRANS_rank",
                        "_SENTIDO_rank",
                    ]
                ).copy(),
                OUT_DIR / "lv_diagnostics_summay.xlsx",
            )
            print(f"[OK] Resumo geral de qualidade gerado: {time_summary_xlsx}")

        for source_folder, plot_dir, time_group in iter_source_plot_groups(lv_time_diag):
            source_label = source_folder if source_folder else "(raiz PROCESSAR)"
            print(f"[INFO] Gerando plots de delta T em {plot_dir} para {source_label}.")
            plot_time_delta_all_samples(time_group, plot_dir=plot_dir)
            plot_time_delta_by_file(time_group, plot_dir=plot_dir)
    else:
        print("[WARN] Coluna TIME nÃ£o encontrada ou invÃ¡lida; pulei o diagnÃ³stico de delta T.")

    trechos = compute_trechos_stats(lv_raw, instruments_df=instruments_df)
    ponto = compute_ponto_stats(trechos)

    fuel_properties = load_fuel_properties_lookup(config_bundle, defaults_cfg)
    kibox_agg = (
        kibox_aggregate(kibox_files)
        if kibox_files
        else pd.DataFrame(columns=["Load_kW", "DIES_pct", "BIOD_pct", "EtOH_pct", "H2O_pct"])
    )
    motec_ponto = pd.DataFrame(columns=["Load_kW", "DIES_pct", "BIOD_pct", "EtOH_pct", "H2O_pct"])
    if motec_files:
        motec_all: List[pd.DataFrame] = []
        for m in motec_files:
            try:
                df_i = read_motec_csv(m)
                if not df_i.empty:
                    motec_all.append(df_i)
            except Exception as e:
                print(f"[ERROR] Falha lendo MOTEC {m.path.name}: {e}")

        if motec_all:
            motec_raw = pd.concat(motec_all, ignore_index=True)
            motec_trechos = compute_motec_trechos_stats(motec_raw)
            motec_ponto = compute_motec_ponto_stats(motec_trechos)
            print(
                f"[INFO] MOTEC: {len(motec_files)} arquivo(s), "
                f"{len(motec_trechos)} trecho(s) valido(s), "
                f"{len(motec_ponto)} ponto(s) agregado(s)."
            )
        else:
            print("[WARN] Arquivos MOTEC encontrados, mas nenhum foi lido com sucesso.")

    out = build_final_table(
        ponto,
        fuel_properties,
        kibox_agg,
        motec_ponto,
        mappings,
        instruments_df,
        reporting_df,
        defaults_cfg,
    )

    out_xlsx = safe_to_excel(out, OUT_DIR / "lv_kpis_clean.xlsx")
    print(f"[OK] Excel gerado: {out_xlsx}")
    if selected_plot_points is None:
        selected_plot_points = prompt_plot_point_filter(out)
    plot_out = _apply_plot_point_filter(out, selected_plot_points)
    plot_summary_total = _new_plot_run_summary()
    unitary_plot_out = plot_out if plot_scope in {"all", "unitary"} else plot_out.iloc[0:0].copy()
    compare_plot_out = plot_out if plot_scope in {"all", "compare"} else plot_out.iloc[0:0].copy()
    if plot_scope not in {"all", "unitary"}:
        print("[INFO] Escopo de plots sem unitary; pulei plots finais por pasta.")
    if plot_scope not in {"all", "compare"}:
        print("[INFO] Escopo de plots sem compare; pulei compare/ e compare_iteracoes/.")

    for source_folder, plot_dir, out_group in iter_source_plot_groups(unitary_plot_out):
        source_label = source_folder if source_folder else "(raiz PROCESSAR)"
        print(f"[INFO] Gerando plots finais em {plot_dir} para {source_label}.")
        _merge_plot_run_summary(
            plot_summary_total,
            make_plots_from_config_with_summary(out_group, plots_df, mappings=mappings, plot_dir=plot_dir),
        )
        _merge_plot_run_summary(
            plot_summary_total,
            _run_plot_job_with_snapshot(
                "consumo_equiv_etanol_overlay",
                _plot_ethanol_equivalent_consumption_overlay,
                snapshot_dir=plot_dir,
                no_output_reason="faltaram blends alvo do overlay E94H6/E75H25/E65H35",
                df=out_group,
                plot_dir=plot_dir,
            ),
        )
        _merge_plot_run_summary(
            plot_summary_total,
            _run_plot_job_with_snapshot(
                "consumo_equiv_etanol_ratio",
                _plot_ethanol_equivalent_ratio,
                snapshot_dir=plot_dir,
                no_output_reason="faltaram pares válidos para razão de consumo equivalente",
                df=out_group,
                plot_dir=plot_dir,
            ),
        )
        _merge_plot_run_summary(
            plot_summary_total,
            _run_plot_job_with_snapshot(
                "nth_e94h6_eq_flow",
                _plot_nth_e94h6_eq_flow,
                snapshot_dir=plot_dir,
                no_output_reason="faltaram blends alvo para n_th_E94H6_eq_flow",
                df=out_group,
                plot_dir=plot_dir,
            ),
        )
        _merge_plot_run_summary(
            plot_summary_total,
            _run_plot_job_with_snapshot(
                "nth_lhv_vs_e94h6_eq_flow",
                _plot_nth_lhv_vs_eq6,
                snapshot_dir=plot_dir,
                no_output_reason="faltaram curvas válidas para comparação n_th",
                df=out_group,
                plot_dir=plot_dir,
            ),
        )
        _merge_plot_run_summary(
            plot_summary_total,
            _run_plot_job_with_snapshot(
                "machine_scenario_suite",
                _plot_machine_scenario_suite,
                snapshot_dir=plot_dir,
                no_output_reason="faltaram dados do cenário de máquinas baseados em E94H6",
                df=out_group,
                plot_dir=plot_dir,
            ),
        )

    compare_groups = iter_compare_plot_groups(compare_plot_out, root=PLOTS_DIR)
    if compare_groups:
        for compare_key, plot_dir, cmp_group in compare_groups:
            series_vals = sorted(
                str(v).strip()
                for v in cmp_group.get("_COMPARE_SERIES", pd.Series([], dtype="object")).dropna().unique().tolist()
                if str(v).strip()
            )
            series_txt = ", ".join(series_vals) if series_vals else "origens desconhecidas"
            print(f"[INFO] Gerando plots de comparacao em {plot_dir} para '{compare_key}' ({series_txt}).")
            _merge_plot_run_summary(
                plot_summary_total,
                make_plots_from_config_with_summary(
                    cmp_group,
                    plots_df,
                    mappings=mappings,
                    plot_dir=plot_dir,
                    series_col="_COMPARE_SERIES",
                ),
            )
    else:
        print("[INFO] Nenhum par subida/descida detectado para gerar plots em compare/.")

    _merge_plot_run_summary(
        plot_summary_total,
        _run_plot_job_with_snapshot(
            "compare_iteracoes_bl_vs_adtv",
            _plot_compare_iteracoes_bl_vs_adtv,
            snapshot_dir=PLOTS_DIR,
            no_output_reason="faltaram pares baseline/aditivado subida/descida para comparação",
            df=compare_plot_out,
            root_plot_dir=PLOTS_DIR,
            mappings=mappings,
            compare_iter_pairs=compare_iter_pairs,
            compare_requests=compare_iter_requests,
        ),
    )

    if kibox_files:
        print("[INFO] Kibox csv em raw/ detectado. (Histogramas KPEAK continuam fora do workflow por enquanto.)")
    else:
        print("[WARN] Sem Kibox csv em raw/.")

    _print_processing_summary(out, plot_out, plot_summary_total)


if __name__ == "__main__":
    main()
